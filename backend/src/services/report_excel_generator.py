"""
Servicio de Generación de Excel para Reportes (Sprint 2)

Fuente Laravel:
- ReportController.php líneas 2595-2788 (descargaReporteOT)
- ReportController.php líneas 3093-3327 (descargaReporteExcel)
- ReportController.php líneas 10554-10923 (descargaReporteSalaMuestra)
- ReportController.php líneas 10923-11027 (descargaReporteTiempoPrimeraMuestra)

Este módulo genera archivos Excel para los reportes del sistema.
"""
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# Estilos reutilizables
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Colores para semáforo
VERDE_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
AMARILLO_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ROJO_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


def _apply_header_style(cell):
    """Aplica estilo de encabezado a una celda."""
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGNMENT
    cell.border = THIN_BORDER


def _apply_cell_style(cell, is_number=False):
    """Aplica estilo de celda normal."""
    cell.alignment = NUMBER_ALIGNMENT if is_number else CELL_ALIGNMENT
    cell.border = THIN_BORDER


def _auto_adjust_columns(ws):
    """Ajusta automáticamente el ancho de las columnas."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def generar_excel_ots_completadas(
    items: List[Dict[str, Any]],
    resumen: Dict[str, Any],
    titulo: str = "OTs Completadas"
) -> BytesIO:
    """
    Genera Excel de OTs completadas.

    Fuente Laravel: descargaReporteOT (líneas 2595-2788)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "OTs Completadas"

    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Fecha de generación
    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # Resumen
    ws['A4'] = "Resumen"
    ws['A4'].font = Font(bold=True)
    ws['A5'] = f"Total Completadas: {resumen.get('total_completadas', 0)}"
    ws['A6'] = f"Tiempo Promedio: {resumen.get('tiempo_promedio', 0):.1f} días"
    ws['A7'] = f"Tiempo Mínimo: {resumen.get('tiempo_minimo', 0):.1f} días"
    ws['A8'] = f"Tiempo Máximo: {resumen.get('tiempo_maximo', 0):.1f} días"

    # Encabezados de tabla
    headers = ['ID', 'Fecha Creación', 'Fecha Completado', 'Cliente', 'Descripción', 'Tiempo (días)', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        _apply_header_style(cell)

    # Datos
    for row_idx, item in enumerate(items, 11):
        ws.cell(row=row_idx, column=1, value=item.get('id'))
        ws.cell(row=row_idx, column=2, value=item.get('created_at'))
        ws.cell(row=row_idx, column=3, value=item.get('completed_at'))
        ws.cell(row=row_idx, column=4, value=item.get('client_name'))
        ws.cell(row=row_idx, column=5, value=item.get('descripcion'))
        ws.cell(row=row_idx, column=6, value=item.get('tiempo_total'))
        ws.cell(row=row_idx, column=7, value=item.get('estado'))

        for col in range(1, 8):
            _apply_cell_style(ws.cell(row=row_idx, column=col), is_number=(col in [1, 6]))

    _auto_adjust_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_excel_conversion_ot(
    items: List[Dict[str, Any]],
    resumen: Dict[str, Any],
    date_desde: str,
    date_hasta: str
) -> BytesIO:
    """
    Genera Excel de conversión de OTs entre fechas.

    Fuente Laravel: descargaReporteOT (líneas 2595-2788)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Conversión OT"

    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Conversión de OTs: {date_desde} - {date_hasta}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Fecha de generación
    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # Resumen por tipo
    ws['A4'] = "Resumen por Tipo"
    ws['A4'].font = Font(bold=True)

    desarrollo = resumen.get('desarrollo', {})
    arte = resumen.get('arte_material', {})

    ws['A5'] = "Desarrollo:"
    ws['B5'] = f"Total: {desarrollo.get('total', 0)}"
    ws['C5'] = f"Completadas: {desarrollo.get('completadas', 0)}"
    ws['D5'] = f"Porcentaje: {desarrollo.get('porcentaje', 0)}%"

    ws['A6'] = "Arte con Material:"
    ws['B6'] = f"Total: {arte.get('total', 0)}"
    ws['C6'] = f"Completadas: {arte.get('completadas', 0)}"
    ws['D6'] = f"Porcentaje: {arte.get('porcentaje', 0)}%"

    # Encabezados de tabla
    headers = ['ID', 'Tipo', 'Fecha Creación', 'Cliente', 'Descripción', 'Estado', 'Completada']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        _apply_header_style(cell)

    # Datos
    for row_idx, item in enumerate(items, 10):
        ws.cell(row=row_idx, column=1, value=item.get('id'))
        ws.cell(row=row_idx, column=2, value=item.get('tipo_nombre'))
        ws.cell(row=row_idx, column=3, value=item.get('created_at'))
        ws.cell(row=row_idx, column=4, value=item.get('client_name'))
        ws.cell(row=row_idx, column=5, value=item.get('descripcion'))
        ws.cell(row=row_idx, column=6, value=item.get('estado'))
        ws.cell(row=row_idx, column=7, value='Sí' if item.get('completada') else 'No')

        for col in range(1, 8):
            _apply_cell_style(ws.cell(row=row_idx, column=col))

    _auto_adjust_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_excel_ots_activas_area(
    items: List[Dict[str, Any]],
    por_area: List[Dict[str, Any]],
    semaforo_totales: Dict[str, int]
) -> BytesIO:
    """
    Genera Excel de OTs activas por área con semáforo.

    Fuente Laravel: descargaReporteExcel (líneas 3093-3327)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "OTs Activas por Área"

    # Título
    ws.merge_cells('A1:K1')
    ws['A1'] = "OTs Activas por Área"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Resumen semáforo
    ws['A3'] = "Resumen Semáforo:"
    ws['A3'].font = Font(bold=True)

    ws['B3'] = f"Verde (< 5 días): {semaforo_totales.get('verde', 0)}"
    ws['B3'].fill = VERDE_FILL

    ws['C3'] = f"Amarillo (5-10 días): {semaforo_totales.get('amarillo', 0)}"
    ws['C3'].fill = AMARILLO_FILL

    ws['D3'] = f"Rojo (> 10 días): {semaforo_totales.get('rojo', 0)}"
    ws['D3'].fill = ROJO_FILL

    # Resumen por área
    ws['A5'] = "Resumen por Área"
    ws['A5'].font = Font(bold=True)

    area_headers = ['Área', 'Total', 'Verde', 'Amarillo', 'Rojo']
    for col, header in enumerate(area_headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        _apply_header_style(cell)

    for row_idx, area in enumerate(por_area, 7):
        ws.cell(row=row_idx, column=1, value=area.get('area_nombre'))
        ws.cell(row=row_idx, column=2, value=area.get('total'))
        ws.cell(row=row_idx, column=3, value=area.get('verde'))
        ws.cell(row=row_idx, column=4, value=area.get('amarillo'))
        ws.cell(row=row_idx, column=5, value=area.get('rojo'))
        for col in range(1, 6):
            _apply_cell_style(ws.cell(row=row_idx, column=col))

    # Detalle de OTs
    start_row = len(por_area) + 9
    ws.cell(row=start_row, column=1, value="Detalle de OTs")
    ws.cell(row=start_row, column=1).font = Font(bold=True)

    headers = ['ID', 'Fecha Creación', 'Cliente', 'Descripción', 'Área', 'Vendedor',
               'Tipo Solicitud', 'Estado', 'Días Total', 'Días en Área', 'Semáforo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row + 1, column=col, value=header)
        _apply_header_style(cell)

    for row_idx, item in enumerate(items, start_row + 2):
        ws.cell(row=row_idx, column=1, value=item.get('id'))
        ws.cell(row=row_idx, column=2, value=item.get('created_at'))
        ws.cell(row=row_idx, column=3, value=item.get('client_name'))
        ws.cell(row=row_idx, column=4, value=item.get('descripcion'))
        ws.cell(row=row_idx, column=5, value=item.get('area_nombre'))
        ws.cell(row=row_idx, column=6, value=item.get('vendedor_nombre'))
        ws.cell(row=row_idx, column=7, value=item.get('tipo_solicitud'))
        ws.cell(row=row_idx, column=8, value=item.get('estado'))
        ws.cell(row=row_idx, column=9, value=item.get('dias_transcurridos'))
        ws.cell(row=row_idx, column=10, value=item.get('dias_en_area'))
        ws.cell(row=row_idx, column=11, value=item.get('semaforo'))

        for col in range(1, 12):
            cell = ws.cell(row=row_idx, column=col)
            _apply_cell_style(cell)

        # Aplicar color de semáforo
        semaforo = item.get('semaforo')
        semaforo_cell = ws.cell(row=row_idx, column=11)
        if semaforo == 'verde':
            semaforo_cell.fill = VERDE_FILL
        elif semaforo == 'amarillo':
            semaforo_cell.fill = AMARILLO_FILL
        elif semaforo == 'rojo':
            semaforo_cell.fill = ROJO_FILL

    _auto_adjust_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_excel_sala_muestra(
    data: Dict[str, Any],
    titulo: str = "Sala de Muestras"
) -> BytesIO:
    """
    Genera Excel de sala de muestras.

    Fuente Laravel: descargaReporteSalaMuestra (líneas 10554-10923)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sala de Muestras"

    periodo = data.get('periodo', {})
    muestras = data.get('muestras', {})
    stats = data.get('estadisticas_mes', {})
    comparacion = data.get('comparacion_meses', [])

    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Reporte Sala de Muestras - {periodo.get('nombre_mes')} {periodo.get('year')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # Estadísticas generales
    ws['A4'] = "Estadísticas del Mes"
    ws['A4'].font = Font(bold=True)

    ws['A5'] = "OTs en Sala de Muestras:"
    ws['B5'] = data.get('ots_en_sala', 0)

    ws['A6'] = "Muestras en Proceso:"
    ws['B6'] = muestras.get('en_proceso', 0)

    ws['A7'] = "Muestras Pendiente Entrega:"
    ws['B7'] = muestras.get('pendiente_entrega', 0)

    ws['A8'] = "Muestras Cortadas (mes):"
    ws['B8'] = muestras.get('cortadas_mes', 0)

    ws['A9'] = "Muestras Terminadas (mes):"
    ws['B9'] = muestras.get('terminadas_mes', 0)

    ws['A10'] = "Total OTs procesadas:"
    ws['B10'] = stats.get('total_ots', 0)

    ws['A11'] = "Tiempo Promedio (días):"
    ws['B11'] = stats.get('tiempo_promedio_dias', 0)

    # Comparación últimos 5 meses
    ws['A13'] = "Comparación Últimos 5 Meses"
    ws['A13'].font = Font(bold=True)

    for col, mes_data in enumerate(comparacion, 1):
        cell = ws.cell(row=14, column=col, value=mes_data.get('nombre'))
        _apply_header_style(cell)
        ws.cell(row=15, column=col, value=mes_data.get('total'))
        _apply_cell_style(ws.cell(row=15, column=col))

    # Comparación anual
    comp_anual = data.get('comparacion_anual', {})
    ws['A17'] = "Comparación Anual"
    ws['A17'].font = Font(bold=True)

    anio_ant = comp_anual.get('anio_anterior', {})
    anio_act = comp_anual.get('anio_actual', {})

    ws['A18'] = f"Año {anio_ant.get('year')}:"
    ws['B18'] = f"OTs: {anio_ant.get('total_ots', 0)}"
    ws['C18'] = f"Tiempo Prom: {anio_ant.get('tiempo_promedio', 0)} días"

    ws['A19'] = f"Año {anio_act.get('year')}:"
    ws['B19'] = f"OTs: {anio_act.get('total_ots', 0)}"
    ws['C19'] = f"Tiempo Prom: {anio_act.get('tiempo_promedio', 0)} días"

    _auto_adjust_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_excel_tiempo_primera_muestra(
    items: List[Dict[str, Any]],
    promedio_dias: float,
    minimo_dias: float,
    maximo_dias: float
) -> BytesIO:
    """
    Genera Excel de tiempo primera muestra.

    Fuente Laravel: descargaReporteTiempoPrimeraMuestra (líneas 10923-11027)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tiempo Primera Muestra"

    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = "Tiempo hasta Primera Muestra"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    # Resumen
    ws['A4'] = "Estadísticas"
    ws['A4'].font = Font(bold=True)
    ws['A5'] = f"Promedio: {promedio_dias:.1f} días"
    ws['A6'] = f"Mínimo: {minimo_dias:.1f} días"
    ws['A7'] = f"Máximo: {maximo_dias:.1f} días"
    ws['A8'] = f"Total OTs: {len(items)}"

    # Encabezados
    headers = ['ID', 'Cliente', 'Descripción', 'Fecha Creación', 'Primera Muestra', 'Días']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        _apply_header_style(cell)

    # Datos
    for row_idx, item in enumerate(items, 11):
        ws.cell(row=row_idx, column=1, value=item.get('id'))
        ws.cell(row=row_idx, column=2, value=item.get('client_name'))
        ws.cell(row=row_idx, column=3, value=item.get('descripcion'))
        ws.cell(row=row_idx, column=4, value=item.get('created_at'))
        ws.cell(row=row_idx, column=5, value=item.get('primera_muestra_at'))
        ws.cell(row=row_idx, column=6, value=item.get('dias_hasta_muestra'))

        for col in range(1, 7):
            _apply_cell_style(ws.cell(row=row_idx, column=col))

    _auto_adjust_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_excel_anulaciones(
    items: List[Dict[str, Any]],
    por_mes: List[Dict[str, Any]]
) -> BytesIO:
    """
    Genera Excel de anulaciones.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Anulaciones"

    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = "Reporte de Anulaciones"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A3'] = f"Total Anulaciones: {len(items)}"

    # Resumen por mes
    if por_mes:
        ws['A5'] = "Por Mes"
        ws['A5'].font = Font(bold=True)
        for row_idx, mes_data in enumerate(por_mes, 6):
            ws.cell(row=row_idx, column=1, value=mes_data.get('mes'))
            ws.cell(row=row_idx, column=2, value=mes_data.get('cantidad'))

    # Encabezados
    start_row = len(por_mes) + 8 if por_mes else 6
    headers = ['ID', 'Fecha', 'Cliente', 'Descripción', 'Motivo', 'Usuario']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        _apply_header_style(cell)

    # Datos
    for row_idx, item in enumerate(items, start_row + 1):
        ws.cell(row=row_idx, column=1, value=item.get('id'))
        ws.cell(row=row_idx, column=2, value=item.get('fecha'))
        ws.cell(row=row_idx, column=3, value=item.get('client_name'))
        ws.cell(row=row_idx, column=4, value=item.get('descripcion'))
        ws.cell(row=row_idx, column=5, value=item.get('motivo'))
        ws.cell(row=row_idx, column=6, value=item.get('usuario'))

        for col in range(1, 7):
            _apply_cell_style(ws.cell(row=row_idx, column=col))

    _auto_adjust_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_excel_rechazos(
    items: List[Dict[str, Any]],
    total_rechazos: int,
    por_area: List[Dict[str, Any]]
) -> BytesIO:
    """
    Genera Excel de rechazos por mes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rechazos"

    # Título
    ws.merge_cells('A1:D1')
    ws['A1'] = "Reporte de Rechazos por Mes"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A3'] = f"Total Rechazos: {total_rechazos}"

    # Resumen por área
    ws['A5'] = "Rechazos por Área"
    ws['A5'].font = Font(bold=True)

    area_headers = ['Área', 'Cantidad']
    for col, header in enumerate(area_headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        _apply_header_style(cell)

    for row_idx, area in enumerate(por_area, 7):
        ws.cell(row=row_idx, column=1, value=area.get('area'))
        ws.cell(row=row_idx, column=2, value=area.get('cantidad'))
        for col in range(1, 3):
            _apply_cell_style(ws.cell(row=row_idx, column=col))

    # Detalle por mes
    start_row = len(por_area) + 9
    ws.cell(row=start_row, column=1, value="Detalle por Mes")
    ws.cell(row=start_row, column=1).font = Font(bold=True)

    headers = ['Mes', 'Área', 'Total Rechazos']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row + 1, column=col, value=header)
        _apply_header_style(cell)

    for row_idx, item in enumerate(items, start_row + 2):
        ws.cell(row=row_idx, column=1, value=item.get('mes'))
        ws.cell(row=row_idx, column=2, value=item.get('area_nombre'))
        ws.cell(row=row_idx, column=3, value=item.get('total_rechazos'))
        for col in range(1, 4):
            _apply_cell_style(ws.cell(row=row_idx, column=col))

    _auto_adjust_columns(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
