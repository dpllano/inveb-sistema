"""
Generador de Excel SAP - INVEB Cascade Service
===============================================

Genera archivos Excel para integración con sistema SAP.

Fuente Laravel: WorkOrderExcelController.php (3,319 líneas)
- descargarExcelSap() - línea 1030
- descargarExcelSapSemielaborado() - línea 2448
- descargarReporteExcel() - línea 390
"""
from io import BytesIO
from typing import Dict, Any, Optional, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import pymysql


def format_decimal_sap(value, decimals: int = 4) -> str:
    """
    Formatea un número para SAP (usa coma como separador decimal).
    Fuente Laravel: number_format_unlimited_precision()
    """
    if value is None or value == '' or value == 'N/A':
        return 'N/A'

    try:
        num = float(str(value).replace(',', '.'))
        formatted = f"{num:.{decimals}f}".replace('.', ',')
        return formatted
    except (ValueError, TypeError):
        return 'N/A'


def get_sap_data_from_ot(ot: Dict[str, Any], connection) -> Dict[str, List]:
    """
    Genera el array de datos para Excel SAP desde una OT.

    Fuente Laravel: WorkOrderExcelController@array_data_excel_sap (líneas 1076-2446)

    El formato es: {codigo_sap: [descripcion, valor]}
    """
    with connection.cursor() as cursor:
        # Obtener datos relacionados

        # Cliente
        client = None
        if ot.get('client_id'):
            cursor.execute("SELECT codigo, nombre FROM clients WHERE id = %s", (ot['client_id'],))
            client = cursor.fetchone()

        # Cartón
        carton = None
        if ot.get('carton_id'):
            cursor.execute("SELECT codigo, descripcion, gramaje, espesor FROM cartones WHERE id = %s", (ot['carton_id'],))
            carton = cursor.fetchone()

        # Tipo de producto
        product_type = None
        if ot.get('product_type_id'):
            cursor.execute("SELECT descripcion, codigo_sap FROM product_types WHERE id = %s", (ot['product_type_id'],))
            product_type = cursor.fetchone()

        # Jerarquía SAP
        jerarquia_sap = None
        if ot.get('subsubhiearchy_id'):
            cursor.execute("SELECT jerarquia_sap FROM subsubhierarchies WHERE id = %s", (ot['subsubhiearchy_id'],))
            row = cursor.fetchone()
            if row:
                jerarquia_sap = row.get('jerarquia_sap')

        # Material
        material = None
        material_code = ''
        if ot.get('material_id'):
            cursor.execute("SELECT codigo, descripcion FROM materials WHERE id = %s", (ot['material_id'],))
            material = cursor.fetchone()
            if material:
                material_code = material.get('codigo', '')

        # Número de material SAP
        numero_material = ''
        if ot.get('material_id') and material_code:
            numero_material = f"GE1{material_code}"

        # Semielaborado
        semielaborado = ''
        if ot.get('material_id') and material_code:
            if ot.get('product_type_id') not in [16, 21]:
                semielaborado = f"GE2{material_code}"

        # Creador/Vendedor
        vendedor = None
        if ot.get('creador_id'):
            cursor.execute("""
                SELECT CONCAT(nombre, ' ', apellido) as nombre, nombre_sap
                FROM users WHERE id = %s
            """, (ot['creador_id'],))
            vendedor = cursor.fetchone()

        # Estilo
        style = None
        if ot.get('style_id'):
            cursor.execute("SELECT glosa, grupo_materiales FROM styles WHERE id = %s", (ot['style_id'],))
            style = cursor.fetchone()

        # Impresión
        impresion_desc = None
        if ot.get('impresion'):
            cursor.execute("SELECT descripcion FROM impresiones WHERE id = %s", (ot.get('impresion'),))
            row = cursor.fetchone()
            if row:
                impresion_desc = row.get('descripcion')

        # Rayado
        rayado = None
        if ot.get('rayado_type_id'):
            cursor.execute("SELECT descripcion FROM rayados WHERE id = %s", (ot['rayado_type_id'],))
            rayado = cursor.fetchone()

        # Colores (1-7)
        colores = {}
        for i in range(1, 8):
            color_id = ot.get(f'color_{i}')
            if color_id:
                cursor.execute("SELECT codigo FROM colors WHERE id = %s", (color_id,))
                row = cursor.fetchone()
                colores[i] = row.get('codigo') if row else None
            else:
                colores[i] = None

        # Armado
        armado = None
        if ot.get('armado_id'):
            cursor.execute("SELECT descripcion FROM armados WHERE id = %s", (ot['armado_id'],))
            armado = cursor.fetchone()

        # Tipo pallet
        tipo_pallet = None
        if ot.get('tipo_pallet_id'):
            cursor.execute("SELECT codigo, descripcion FROM pallet_types WHERE id = %s", (ot['tipo_pallet_id'],))
            tipo_pallet = cursor.fetchone()

        # QA / Certificado
        qa = None
        if ot.get('qa_id'):
            cursor.execute("SELECT descripcion FROM pallet_qas WHERE id = %s", (ot['qa_id'],))
            qa = cursor.fetchone()

        # Prepicado
        prepicado = None
        if ot.get('prepicado_id'):
            cursor.execute("SELECT descripcion FROM precut_types WHERE id = %s", (ot['prepicado_id'],))
            prepicado = cursor.fetchone()

        # Protección
        protection = None
        if ot.get('protection_type_id'):
            cursor.execute("SELECT descripcion FROM protection_types WHERE id = %s", (ot['protection_type_id'],))
            protection = cursor.fetchone()

        # Pallet status
        pallet_status = None
        if ot.get('pallet_status_type_id'):
            cursor.execute("SELECT descripcion FROM pallet_status_types WHERE id = %s", (ot['pallet_status_type_id'],))
            pallet_status = cursor.fetchone()

        # Tipo cinta
        tipo_cinta = None
        if ot.get('tipo_cinta'):
            cursor.execute("SELECT codigo, descripcion FROM tipos_cintas WHERE id = %s", (ot['tipo_cinta'],))
            tipo_cinta = cursor.fetchone()

        # Centro y almacén (desde grupo_plantas)
        centro = ''
        num_almacen = ''
        if ot.get('planta_id'):
            cursor.execute("""
                SELECT centro, num_almacen FROM grupo_plantas
                WHERE planta_id = %s AND active = 1 LIMIT 1
            """, (ot['planta_id'],))
            gp = cursor.fetchone()
            if gp:
                centro = gp.get('centro', '')
                num_almacen = gp.get('num_almacen', '')

        # Matriz
        matriz = None
        if ot.get('matriz_id'):
            cursor.execute("SELECT material FROM matrices WHERE id = %s", (ot['matriz_id'],))
            matriz = cursor.fetchone()

    # Calcular valores derivados
    recorte_caracteristico = format_decimal_sap(ot.get('recorte_caracteristico'))
    recorte_adicional = format_decimal_sap(ot.get('recorte_adicional')) if ot.get('recorte_adicional') else '0,0000'

    # Área producto
    area_producto = 0
    if ot.get('area_producto_calculo') and ot['area_producto_calculo'] > 0:
        area_producto = int(float(ot['area_producto_calculo']) * 1000000)

    # Consumo cinta
    consumo_cinta = ''
    if ot.get('cintas_x_caja') and ot.get('largura_hm'):
        consumo_cinta = ot['largura_hm'] * ot['cintas_x_caja']

    # Total golpes
    total_golpes = 0
    if ot.get('golpes_ancho') and ot.get('golpes_largo'):
        total_golpes = ot['golpes_ancho'] * ot['golpes_largo']

    # Construir array de datos SAP
    # Formato: {codigo_sap: [descripcion_campo, valor]}
    array_data = {
        "STDPD": ['Mat. Configurable', product_type.get('codigo_sap') if product_type else None],
        "MATNR": ['Número de Material', numero_material],
        "EN_PLANCHA_SEMI": ['Número de Semielaborado', semielaborado],
        "RMMG1_REF-MATNR": ['Material Modelo', None],  # Modelo material
        "MAKTX": ['Descripción Comercial', material.get('descripcion') if material else None],
        "WERKS": ['Centro', centro],
        "LGORT": ['Almacén', '001'],
        "VKORG": ['Organiz. Ventas', None],  # Org ventas
        "VTWEG": ['Canal distrib.', None],  # Canal
        "LGNUM": ['Numero Almacen', num_almacen],
        "MARA-BISMT": ['N° Material antiguo', ot['id']],
        "SPART": ['Sector', None],  # Sector
        "PRDHA": ['Jerarquia', jerarquia_sap],
        "BRGEW": ['Peso bruto (G1)', format_decimal_sap(ot.get('peso_bruto'))],
        "NTGEW": ['Peso neto (G1)', format_decimal_sap(ot.get('peso_neto'))],
        "VOLUM": ['Volumen (G1)', int(round(ot.get('volumen_unitario', 0))) if ot.get('volumen_unitario') else None],
        "NORMT": ['Denom.estándar', ot.get('codigo_producto_cliente')],
        "UMREN": ['UMA Área (m2)', int(ot.get('uma_area', 0)) if ot.get('uma_area') else 0],
        "UMREN01": ['UMA Peso (kg)', int(ot.get('uma_peso', 0)) if ot.get('uma_peso') else 0],
        "KONDM": ['Gr.imputación mat.', None],  # Grupo imputación material
        "PRODH": ['Jerarquía productos', jerarquia_sap],
        "MVGR1": ['Grupo materiales 1', None],  # Grupo materiales 1
        "MVGR2": ['Grupo de material 2', None],  # Grupo materiales 2
        "MVKE-MVGR3": ['Grupo de material 3', style.get('grupo_materiales') if style else None],
        "MARC-PRCTR": ['CeBe', ''],  # CeBe
        "MARC-AUSSS": ['Rechazo conjunto (%)', 0],  # Rechazo conjunto
        "MARC-WEBAZ": ['Tmpo.tratamiento EM', 0],  # Tiempo tratamiento
        "MARC-BASMG": ['Cantidad base', 1000],  # Fijo según cliente
        "MBEW-KOSGR": ['Grupo gastos gral.', ''],
        "EN_OT": ['Numero OT', ot['id']],
        "EN_CODCLI": ['Cliente', client.get('codigo') if client else None],
        "EN_CODVEN": ['Vendedor', vendedor.get('nombre_sap') or vendedor.get('nombre') if vendedor else None],
        "EN_LARGO": ['Largo Interior (MM)', ot.get('interno_largo')],
        "EN_ANCHO": ['Ancho Interior (MM)', ot.get('interno_ancho')],
        "EN_ALTO": ['Alto Interior (MM)', ot.get('interno_alto')],
        "EN_LARGURA": ['Largura HM (MM)', ot.get('largura_hm')],
        "EN_ANCHURA": ['Anchura HM (MM)', ot.get('anchura_hm')],
        "EN_LARGOEXT": ['Largo Exterior (MM)', ot.get('externo_largo')],
        "EN_ANCHOEXT": ['Ancho Exterior (MM)', ot.get('externo_ancho')],
        "EN_ALTOEXT": ['Alto Exterior (MM)', ot.get('externo_alto')],
        "EN_CARTON": ['Cartón', carton.get('codigo') if carton else None],
        "EN_TIPO": ['Tipo de Producto', product_type.get('descripcion') if product_type else None],
        "EN_ESTILO": ['Estilo de Producto', style.get('glosa') if style else None],
        "EN_CARCTRIST_ESTILO": ['Caracteristicas Adicionales', ot.get('caracteristicas_adicionales') or 'No Aplica'],
        "EN_C1_R1": ['Rayado C1/R1 (MM)', ot.get('rayado_c1r1')],
        "EN_R1_R2": ['Rayado R1/R2 (MM)', ot.get('rayado_r1_r2')],
        "EN_R2_C2": ['Rayado R2/C2 (MM)', ot.get('rayado_r2_c2')],
        "EN_RAYADO": ['Tipo de Rayado', rayado.get('descripcion') if rayado else None],
        "EN_TIPOIMPRESION": ['Tipo Impresión', impresion_desc],
        "EN_COLORES": ['Número de Colores', ot.get('numero_colores')],
        "EN_PRUEBACOLOR": ['Prueba de Color', 'Si' if ot.get('prueba_color') == 1 else 'No' if ot.get('prueba_color') == 0 else None],
        "EN_CARACTERISTICO": ['Recorte Característico (M2)', recorte_caracteristico],
        "EN_ADICIONAL": ['Recorte Adicional (M2)', recorte_adicional],
        "EN_CAD": ['Plano CAD', ot.get('cad')],
        "EN_AREA_INTERIOR_PERIMETRO": ['Area Producto (M2)', area_producto],
        "EN_ESTADO_PALETIZAD0": ['Estado de Palletizado', pallet_status.get('descripcion') if pallet_status else None],
        "EN_TIPO_PALET_GE": ['Tipo de Pallet', tipo_pallet.get('codigo') if tipo_pallet else None],
        "EN_TRATAMIENTO_PALET": ['Tratamiento de Pallet', 'SI' if ot.get('pallet_treatment') == 1 else 'NO' if ot.get('pallet_treatment') == 0 else None],
        "EN_CAJAS_POR_PALET": ['Nro Cajas por Pallet', ot.get('cajas_por_pallet')],
        "EN_PLACAS_POR_PALET": ['Nro Placas por Pallet', ot.get('placas_por_pallet')],
        "EN_PROTECCION": ['Proteccion', protection.get('descripcion') if protection else None],
        "EN_CERTIFICADO": ['Certificado Calidad', qa.get('descripcion') if qa else None],
        "EN_RESISTENCIA_MT": ['BCT MIN (LB)', ot.get('bct_min_lb')],
        "EN_JERARQUIA": ['Jerarquia', jerarquia_sap],
        "EN_COD_PT_CLI": ['Código Producto Cliente', ot.get('codigo_producto_cliente')],
        "EN_C1_CINTA1": ['Distancia corte 1 a cinta 1', ot.get('distancia_cinta_1')],
        "EN_C1_CINTA2": ['Distancia corte 1 a cinta 2', ot.get('distancia_cinta_2')],
        "EN_C1_CINTA3": ['Distancia corte 1 a cinta 3', ot.get('distancia_cinta_3')],
        "EN_C1_CINTA4": ['Distancia corte 1 a cinta 4', ot.get('distancia_cinta_4')],
        "EN_C1_CINTA5": ['Distancia corte 1 a cinta 5', ot.get('distancia_cinta_5')],
        "EN_C1_CINTA6": ['Distancia corte 1 a cinta 6', ot.get('distancia_cinta_6')],
        "EN_TIPO_CINTA": ['TIPO DE CINTA', tipo_cinta.get('descripcion') if tipo_cinta else None],
        "EN_CORTE_LINER": ['Cantidad Cinta por CAJA', ot.get('cintas_x_caja')],
        "EN_CINTA": ['Codigo Cinta', tipo_cinta.get('codigo') if tipo_cinta else None],
        "EN_CONSUMO_CINTA": ['Consumo cinta', consumo_cinta],
        "EN_SELLO": ['Etiqueta FSC/PRODUCTO FSC', {2: 'NO', 5: 'No', 3: 'FACTURACION Y LOGO', 4: 'FACTURACION Y LOGO', 6: 'SOLO FACTURACION'}.get(ot.get('fsc'))],
        "EN_ORIENTACION": ['Orientación Placa', 0 if ot.get('orientacion_placa') == 0 else 90 if ot.get('orientacion_placa') == 1 else None],
        "EN_CARACT_ADICION": ['Características Adicionales', prepicado.get('descripcion') if prepicado else None],
        "EN_TIPO_CIERRE": ['Tipo de Pegado', {0: 'No Aplica', 2: 'Pegado Interno', 3: 'Pegado Externo', 4: 'Pegado 3 Puntos', 5: 'Pegado 4 Puntos'}.get(ot.get('pegado_terminacion'))],
        "EN_ARMADO": ['Armado', armado.get('descripcion') if armado else None],
        "EN_SENTIDO_DE_ARMADO": ['Sentido Armado', {1: 'No aplica', 2: 'Ancho a la Derecha', 3: 'Ancho a la Izquierda', 4: 'Largo a la Izquierda', 5: 'Largo a la Derecha'}.get(ot.get('sentido_armado'))],
        "EN_GRAMAJE_G_M2": ['Gramaje (G/m2)', ot.get('gramaje')],
        "EN_PESO_G": ['Peso (G)', ot.get('peso')],
        "EN_ESPESOR_MM": ['Espesor Caja (mm)', ot.get('espesor_caja')],
        "EN_ECT_MINIMO": ['ECT Minimo (LB/PULG2)', format_decimal_sap(ot.get('ect'))],
        "EN_FCT_MINIMO": ['FCT Minimo (LB/PULG22)', format_decimal_sap(ot.get('fct'))],
        "EN_COBB_INT_MAX": ['Cobb INT. (2 Min.) Max.', ot.get('cobb_interior')],
        "EN_COBB_EXT_MAX": ['Cobb EXT. (2 Min.) Max.', ot.get('cobb_exterior')],
        "EN_FLEXION_ALETA": ['Flexion de Aleta (N)', ot.get('flexion_aleta')],
        "EN_MULLEN": ['Mullen (LB/PULG2)', format_decimal_sap(ot.get('mullen'))],
        "EN_RESISTENCIA_STD": ['Resistencia mínima (Humeda)', ot.get('bct_humedo_lb')],
        "EN_DST_BPI": ['DST (BPI)', ot.get('dst')],
        "EN_ESPESOR_PL": ['Espesor Placa (mm)', ot.get('espesor_placa')],
        "EN_POROSIDAD": ['Porosidad (SEG)', ot.get('porosidad')],
        "EN_BRILLO": ['Brillo (%)', ot.get('brillo')],
        "EN_CANT_LARGO": ['Golpes al Largo', ot.get('golpes_largo')],
        "EN_CANT_ANCHO": ['Golpes al Ancho', ot.get('golpes_ancho')],
        "EN_TOTAL_GOL_MATRIZ": ['Golpes Total', total_golpes],
        "EN_LARGURA_HC": ['Largura HC (MM)', ot.get('largura_hc')],
        "EN_ANCHURA_HC": ['Anchura HC (MM)', ot.get('anchura_hc')],
        "EN_COLOR_COMP_1": ['Código Color 1 (INTERIOR TyR)', colores.get(1)],
        "EN_CONSUMO_1": ['Gramos Color 1 (INTERIOR TyR)', format_decimal_sap(ot.get('consumo_color_1'))],
        "EN_COLOR_COMP_2": ['Código Color 2', colores.get(2)],
        "EN_CONSUMO_2": ['Gramos Color 2', format_decimal_sap(ot.get('consumo_color_2'))],
        "EN_COLOR_COMP_3": ['Código Color 3', colores.get(3)],
        "EN_CONSUMO_3": ['Gramos Color 3', format_decimal_sap(ot.get('consumo_color_3'))],
        "EN_COLOR_COMP_4": ['Código Color 4', colores.get(4)],
        "EN_CONSUMO_4": ['Gramos Color 4', format_decimal_sap(ot.get('consumo_color_4'))],
        "EN_COLOR_COMP_5": ['Código Color 5', colores.get(5)],
        "EN_CONSUMO_5": ['Gramos Color 5', format_decimal_sap(ot.get('consumo_color_5'))],
        "EN_COLOR_COMP_6": ['Código Color 6', colores.get(6)],
        "EN_CONSUMO_6": ['Gramos Color 6', format_decimal_sap(ot.get('consumo_color_6'))],
        "EN_COLOR_COMP_7": ['Código Color 7', colores.get(7)],
        "EN_CONSUMO_7": ['Gramos Color 7', format_decimal_sap(ot.get('consumo_color_7'))],
        "EN_CLISSE_1": ['Clisse 1', f"ENC{material.get('codigo')}" if material and material.get('codigo') else None],
        "EN_CLISSE_2": ['Clisse 2', ''],
        "EN_MATRIZ_1": ['Matriz 1', matriz.get('material') if matriz else None],
        "MAKT-MAKTX": ['URL FT', f"\\\\pro-eeii05\\ftecnica\\{numero_material}_FT.pdf" if numero_material else ''],
    }

    return array_data


def generar_excel_sap(ot_id: int, connection) -> BytesIO:
    """
    Genera archivo Excel SAP para una OT.

    Fuente Laravel: WorkOrderExcelController@descargarExcelSap (líneas 1030-1074)
    """
    with connection.cursor() as cursor:
        # Obtener OT completa
        cursor.execute("""
            SELECT wo.*
            FROM work_orders wo
            WHERE wo.id = %s
        """, (ot_id,))
        ot = cursor.fetchone()

        if not ot:
            raise ValueError(f"OT {ot_id} no encontrada")

    # Generar datos SAP
    sap_data = get_sap_data_from_ot(ot, connection)

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Excel SAP OT {ot_id}"

    # Estilos
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    ws['A1'] = 'Código SAP'
    ws['B1'] = 'Campo'
    ws['C1'] = 'Valor'

    for col in ['A', 'B', 'C']:
        ws[f'{col}1'].font = header_font_white
        ws[f'{col}1'].fill = header_fill
        ws[f'{col}1'].border = border
        ws[f'{col}1'].alignment = Alignment(horizontal='center')

    # Datos (formato vertical como en Laravel)
    row = 2
    for codigo_sap, (descripcion, valor) in sap_data.items():
        ws[f'A{row}'] = codigo_sap
        ws[f'B{row}'] = descripcion
        ws[f'C{row}'] = valor if valor is not None else ''

        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = border

        row += 1

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50

    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_excel_sap_semielaborado(ot_id: int, connection) -> BytesIO:
    """
    Genera archivo Excel SAP Semielaborado para una OT.

    Fuente Laravel: WorkOrderExcelController@descargarExcelSapSemielaborado (línea 2448)

    Similar a generar_excel_sap pero con campos específicos para semielaborados.
    """
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT wo.*
            FROM work_orders wo
            WHERE wo.id = %s
        """, (ot_id,))
        ot = cursor.fetchone()

        if not ot:
            raise ValueError(f"OT {ot_id} no encontrada")

    # Generar datos SAP (mismo que SAP normal, filtrado para semielaborado)
    sap_data = get_sap_data_from_ot(ot, connection)

    # Para semielaborado, solo incluimos campos relevantes
    campos_semielaborado = [
        "EN_PLANCHA_SEMI", "MATNR", "MAKTX", "WERKS", "LGORT", "LGNUM",
        "EN_OT", "EN_CODCLI", "EN_CODVEN", "EN_CARTON", "EN_TIPO",
        "EN_LARGURA", "EN_ANCHURA", "EN_LARGURA_HC", "EN_ANCHURA_HC",
        "EN_CANT_LARGO", "EN_CANT_ANCHO", "EN_TOTAL_GOL_MATRIZ",
        "BRGEW", "NTGEW", "EN_GRAMAJE_G_M2"
    ]

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Semielaborado OT {ot_id}"

    # Estilos
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    ws['A1'] = 'Código SAP'
    ws['B1'] = 'Campo'
    ws['C1'] = 'Valor'

    for col in ['A', 'B', 'C']:
        ws[f'{col}1'].font = header_font_white
        ws[f'{col}1'].fill = header_fill
        ws[f'{col}1'].border = border
        ws[f'{col}1'].alignment = Alignment(horizontal='center')

    # Datos filtrados para semielaborado
    row = 2
    for codigo_sap in campos_semielaborado:
        if codigo_sap in sap_data:
            descripcion, valor = sap_data[codigo_sap]
            ws[f'A{row}'] = codigo_sap
            ws[f'B{row}'] = descripcion
            ws[f'C{row}'] = valor if valor is not None else ''

            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border

            row += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 40

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_reporte_excel_ot(ot_id: int, connection) -> BytesIO:
    """
    Genera reporte Excel COMPLETO de una OT.

    Fuente Laravel: WorkOrderExcelController@descargarReporteExcel (linea 390)
    Fuente Laravel: array_data_excel() (lineas 436-857) - 130+ campos

    BRECHA 3: Implementacion completa para paridad con Laravel.
    """
    with connection.cursor() as cursor:
        # Query principal con TODAS las relaciones
        cursor.execute("""
            SELECT
                wo.*,
                c.nombre as client_name,
                c.codigo as client_codigo,
                CONCAT(u.nombre, ' ', u.apellido) as creador_nombre,
                u.nombre_sap as vendedor_nombre_sap,
                pt.descripcion as product_type_name,
                car.codigo as carton_codigo,
                car.descripcion as carton_descripcion,
                car.peso as carton_peso,
                car.onda as carton_onda,
                s.nombre as estado_nombre,
                p.descripcion as proceso_nombre,
                st.glosa as style_glosa,
                arm.descripcion as armado_descripcion,
                ray.descripcion as rayado_descripcion,
                imp.descripcion as impresion_descripcion,
                cov_int.descripcion as coverage_internal_desc,
                cov_ext.descripcion as coverage_external_desc,
                pal_t.descripcion as pallet_type_desc,
                pal_pat.descripcion as pallet_patron_desc,
                pal_prot.descripcion as pallet_protection_desc,
                pal_box.descripcion as pallet_box_qty_desc,
                pal_tag.descripcion as pallet_tag_format_desc,
                pal_qa.descripcion as pallet_qa_desc,
                pal_st.descripcion as pallet_status_desc,
                prot.descripcion as protection_desc,
                prec.descripcion as precut_desc,
                ssh.jerarquia_sap,
                mat.codigo as material_codigo,
                mat.descripcion as material_descripcion,
                mat_ref.codigo as material_ref_codigo,
                maq.servicio as maquila_servicio,
                tip_cinta.descripcion as tipo_cinta_desc,
                matriz1.material as matriz_1_material,
                matriz2.material as matriz_2_material,
                matriz3.material as matriz_3_material,
                col1.descripcion as color_1_desc, col1.codigo as color_1_codigo,
                col2.descripcion as color_2_desc, col2.codigo as color_2_codigo,
                col3.descripcion as color_3_desc, col3.codigo as color_3_codigo,
                col4.descripcion as color_4_desc, col4.codigo as color_4_codigo,
                col5.descripcion as color_5_desc, col5.codigo as color_5_codigo,
                col6.descripcion as color_6_desc, col6.codigo as color_6_codigo,
                col7.descripcion as color_7_desc, col7.codigo as color_7_codigo
            FROM work_orders wo
            LEFT JOIN clients c ON wo.client_id = c.id
            LEFT JOIN users u ON wo.creador_id = u.id
            LEFT JOIN product_types pt ON wo.product_type_id = pt.id
            LEFT JOIN cartones car ON wo.carton_id = car.id
            LEFT JOIN (
                SELECT m1.work_order_id, m1.state_id
                FROM managements m1
                INNER JOIN (
                    SELECT work_order_id, MAX(id) as max_id
                    FROM managements
                    GROUP BY work_order_id
                ) m2 ON m1.work_order_id = m2.work_order_id AND m1.id = m2.max_id
            ) latest ON wo.id = latest.work_order_id
            LEFT JOIN states s ON latest.state_id = s.id
            LEFT JOIN processes p ON wo.process_id = p.id
            LEFT JOIN styles st ON wo.style_id = st.id
            LEFT JOIN armados arm ON wo.armado_id = arm.id
            LEFT JOIN rayados ray ON wo.rayado_type_id = ray.id
            LEFT JOIN impresiones imp ON wo.impresion = imp.id
            LEFT JOIN coverages cov_int ON wo.coverage_internal_id = cov_int.id
            LEFT JOIN coverages cov_ext ON wo.coverage_external_id = cov_ext.id
            LEFT JOIN pallet_types pal_t ON wo.tipo_pallet_id = pal_t.id
            LEFT JOIN pallet_patrons pal_pat ON wo.pallet_patron_id = pal_pat.id
            LEFT JOIN pallet_protections pal_prot ON wo.pallet_protection_id = pal_prot.id
            LEFT JOIN pallet_box_quantities pal_box ON wo.pallet_box_quantity_id = pal_box.id
            LEFT JOIN pallet_tag_formats pal_tag ON wo.pallet_tag_format_id = pal_tag.id
            LEFT JOIN pallet_qas pal_qa ON wo.qa_id = pal_qa.id
            LEFT JOIN pallet_status_types pal_st ON wo.pallet_status_type_id = pal_st.id
            LEFT JOIN protection_types prot ON wo.protection_type_id = prot.id
            LEFT JOIN precut_types prec ON wo.prepicado_id = prec.id
            LEFT JOIN subsubhierarchies ssh ON wo.subsubhiearchy_id = ssh.id
            LEFT JOIN materials mat ON wo.material_id = mat.id
            LEFT JOIN materials mat_ref ON wo.material_referencia_id = mat_ref.id
            LEFT JOIN maquila_servicios maq ON wo.maquila_servicio_id = maq.id
            LEFT JOIN tipos_cintas tip_cinta ON wo.tipo_cinta = tip_cinta.id
            LEFT JOIN matrices matriz1 ON wo.matriz_id = matriz1.id
            LEFT JOIN matrices matriz2 ON wo.matriz_id_2 = matriz2.id
            LEFT JOIN matrices matriz3 ON wo.matriz_id_3 = matriz3.id
            LEFT JOIN colors col1 ON wo.color_1 = col1.id
            LEFT JOIN colors col2 ON wo.color_2 = col2.id
            LEFT JOIN colors col3 ON wo.color_3 = col3.id
            LEFT JOIN colors col4 ON wo.color_4 = col4.id
            LEFT JOIN colors col5 ON wo.color_5 = col5.id
            LEFT JOIN colors col6 ON wo.color_6 = col6.id
            LEFT JOIN colors col7 ON wo.color_7 = col7.id
            WHERE wo.id = %s
        """, (ot_id,))
        ot = cursor.fetchone()

        if not ot:
            raise ValueError(f"OT {ot_id} no encontrada")

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Reporte OT {ot_id}"

    # Estilos
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10)
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    row = 1

    def add_section(title: str):
        nonlocal row
        row += 1
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        return row

    def add_field(label: str, value):
        nonlocal row
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = value if value is not None else ''
        row += 1

    def format_si_no(val):
        if val == 1:
            return "Si"
        elif val == 0:
            return "No"
        return None

    def format_decimal_4(val):
        if val is None:
            return 'N/A'
        try:
            return f"{float(val):.4f}".replace('.', ',')
        except:
            return 'N/A'

    # TITULO
    ws['A1'] = f'REPORTE DE CATALOGACION - OT #{ot_id}'
    ws['A1'].font = title_font
    ws.merge_cells('A1:B1')
    row = 2

    # SECCION 1: INFORMACION GENERAL
    add_section('INFORMACION GENERAL')
    add_field('OT', ot['id'])
    add_field('Numero de Material', ot.get('material_codigo'))
    add_field('Descripcion Comercial', ot.get('material_descripcion'))
    add_field('Cliente', ot.get('client_codigo'))
    add_field('Vendedor', ot.get('vendedor_nombre_sap') or ot.get('creador_nombre'))

    # SECCION 2: MEDIDAS
    add_section('MEDIDAS')
    add_field('Largo Interior (MM)', ot.get('interno_largo'))
    add_field('Ancho Interior (MM)', ot.get('interno_ancho'))
    add_field('Alto Interior (MM)', ot.get('interno_alto'))
    add_field('Largura HM (MM)', ot.get('largura_hm'))
    add_field('Anchura HM (MM)', ot.get('anchura_hm'))
    add_field('Largo Exterior (MM)', ot.get('externo_largo'))
    add_field('Ancho Exterior (MM)', ot.get('externo_ancho'))
    add_field('Alto Exterior (MM)', ot.get('externo_alto'))

    # SECCION 3: CARTON Y PRODUCTO
    add_section('CARTON Y PRODUCTO')
    add_field('Carton', ot.get('carton_codigo'))
    add_field('Tipo de Producto (Tipo Item)', ot.get('product_type_name'))
    add_field('Estilo de Producto', ot.get('style_glosa'))
    add_field('Caracteristicas Estilo', ot.get('caracteristicas_adicionales'))
    add_field('Gramaje', ot.get('carton_peso'))
    add_field('Onda', ot.get('carton_onda'))
    add_field('Proceso', ot.get('proceso_nombre'))
    carton_color_map = {1: "Cafe", 2: "Blanco"}
    add_field('Color del Carton', carton_color_map.get(ot.get('carton_color')))

    # SECCION 4: RAYADO
    add_section('RAYADO')
    add_field('Rayado C1/R1 (MM)', ot.get('rayado_c1r1'))
    add_field('Rayado R1/R2 (MM)', ot.get('rayado_r1_r2'))
    add_field('Rayado R2/C2 (MM)', ot.get('rayado_r2_c2'))
    add_field('Tipo de Rayado', ot.get('rayado_descripcion'))

    # SECCION 5: IMPRESION
    add_section('IMPRESION')
    add_field('Tipo Impresion', ot.get('impresion_descripcion'))
    add_field('Numero de Colores', ot.get('numero_colores'))
    add_field('Prueba de Color', format_si_no(ot.get('prueba_color')))
    add_field('Recorte Caracteristico (M2)', format_decimal_4(ot.get('recorte_caracteristico')))
    add_field('Recorte Adicional (M2)', format_decimal_4(ot.get('recorte_adicional')))
    add_field('Plano CAD', ot.get('cad'))

    # SECCION 6: PALLETIZADO
    add_section('PALLETIZADO')
    add_field('Estado de Palletizado', ot.get('pallet_status_desc'))
    add_field('Tipo de Pallet', ot.get('pallet_type_desc'))
    add_field('Tratamiento de Pallet', format_si_no(ot.get('pallet_treatment')))
    add_field('Nro Cajas por Pallet', ot.get('cajas_por_pallet'))
    add_field('Nro Placas por Pallet', ot.get('placas_por_pallet'))
    add_field('Patron Carga Pallet', ot.get('pallet_patron_desc'))
    patron_zuncho_map = {1: "2x0", 2: "2x1", 3: "2x2"}
    add_field('Patron Zuncho Bulto', patron_zuncho_map.get(ot.get('patron_zuncho_bulto')))
    add_field('Proteccion', ot.get('protection_desc'))
    add_field('Patron Zuncho Pallet', patron_zuncho_map.get(ot.get('patron_zuncho')))
    add_field('Proteccion Pallet', ot.get('pallet_protection_desc'))
    add_field('Nro Cajas por Paquete', ot.get('pallet_box_qty_desc'))
    add_field('Patron Zuncho Paquete', ot.get('patron_zuncho_paquete'))
    add_field('Termocontraible', format_si_no(ot.get('termocontraible')))
    add_field('Nro Cajas por Unitizados', ot.get('paquetes_por_unitizado'))
    add_field('Nro Unitizados por Pallet', ot.get('unitizado_por_pallet'))
    add_field('Tipo Formato Etiqueta Pallet', ot.get('pallet_tag_format_desc'))
    add_field('Nro Etiqueta Pallet', ot.get('numero_etiquetas'))
    add_field('Certificado Calidad', ot.get('pallet_qa_desc'))

    # SECCION 7: ESPECIFICACIONES TECNICAS
    add_section('ESPECIFICACIONES TECNICAS')
    add_field('BCT MIN (LB)', ot.get('bct_min_lb'))
    add_field('Unidad Medida BCT', 'Libras F')
    add_field('Gramaje (G/m2)', ot.get('gramaje'))
    add_field('Peso (G)', ot.get('peso'))
    add_field('Espesor Caja (mm)', ot.get('espesor_caja'))
    add_field('ECT Minimo (LB/PULG2)', format_decimal_4(ot.get('ect')))
    add_field('FCT Minimo (LB/PULG2)', format_decimal_4(ot.get('fct')))
    add_field('Cobb INT. (2 Min.) Max.', ot.get('cobb_interior'))
    add_field('Cobb EXT. (2 Min.) Max.', ot.get('cobb_exterior'))
    add_field('Flexion de Aleta (N)', ot.get('flexion_aleta'))
    add_field('Mullen (LB/PULG2)', format_decimal_4(ot.get('mullen')))
    add_field('Incision Rayado Long.[N]', ot.get('incision_rayado_longitudinal'))
    add_field('Incision Rayado Transv.[N]', ot.get('incision_rayado_vertical'))
    add_field('DST (BPI)', ot.get('dst'))
    add_field('Espesor Placa (mm)', ot.get('espesor_placa'))
    add_field('Porosidad (SEG)', ot.get('porosidad'))
    add_field('Brillo (%)', ot.get('brillo'))
    add_field('Rigidez 4 Puntos Longitudinal (N/MM)', ot.get('rigidez_4_ptos_long'))
    add_field('Rigidez 4 Puntos Transversal (N/MM)', ot.get('rigidez_4_ptos_transv'))
    add_field('Angulo de Deslizamiento-Tapa Exterior', ot.get('angulo_deslizamiento_tapa_exterior'))
    add_field('Angulo de Deslizamiento-Tapa Interior', ot.get('angulo_deslizamiento_tapa_interior'))
    add_field('Resistencia al Frote', ot.get('resistencia_frote'))
    add_field('Contenido Reciclado (%)', ot.get('contenido_reciclado'))

    # SECCION 8: TERMINACIONES
    add_section('TERMINACIONES')
    pegado_map = {0: "No Aplica", 2: "Pegado Interno", 3: "Pegado Externo", 4: "Pegado 3 Puntos", 5: "Pegado 4 Puntos"}
    add_field('Tipo de Pegado', pegado_map.get(ot.get('pegado_terminacion')))
    add_field('Armado', ot.get('armado_descripcion'))
    sentido_map = {1: "No aplica", 2: "Ancho a la Derecha", 3: "Ancho a la Izquierda", 4: "Largo a la Izquierda", 5: "Largo a la Derecha"}
    add_field('Sentido Armado', sentido_map.get(ot.get('sentido_armado')))
    add_field('Maquila', format_si_no(ot.get('maquila')))
    add_field('Servicio de Maquila', ot.get('maquila_servicio'))
    add_field('Recubrimiento Interno', ot.get('coverage_internal_desc'))
    add_field('Recubrimiento Externo', ot.get('coverage_external_desc'))

    # SECCION 9: OTROS DATOS
    add_section('OTROS DATOS')
    add_field('Tipo Camion', ot.get('tipo_camion'))
    add_field('Restricciones Especiales', ot.get('restriccion_especial'))
    add_field('Horario Recepcion', ot.get('horario_recepcion'))
    add_field('Jerarquia', ot.get('jerarquia_sap'))
    add_field('Codigo Producto Cliente', ot.get('codigo_producto_cliente'))
    add_field('Para uso de Programa Z', ot.get('uso_programa_z'))
    fsc_map = {0: "NO", 2: "NO", 3: "FACTURACION Y LOGO", 4: "FACTURACION Y LOGO", 5: "No", 6: "SOLO FACTURACION"}
    add_field('Etiqueta FSC', fsc_map.get(ot.get('fsc')))
    orientacion_map = {0: "0", 1: "90"}
    add_field('Orientacion Placa', orientacion_map.get(ot.get('orientacion_placa')))
    add_field('Caracteristicas Adicionales', ot.get('precut_desc'))
    add_field('Impresion de Borde', ot.get('impresion_borde'))
    add_field('Impresion Sobre Rayado', ot.get('impresion_sobre_rayado'))

    # SECCION 10: REFERENCIAS Y OBSERVACIONES
    add_section('REFERENCIAS Y OBSERVACIONES')
    add_field('Observaciones', ot.get('observacion'))
    add_field('Referencia Material', ot.get('material_ref_codigo'))
    add_field('Bloqueo Referencia', "SI" if ot.get('bloqueo_referencia') == 1 else "NO")

    # SECCION 11: LISTA DE MATERIALES
    add_section('LISTA DE MATERIALES')
    add_field('Golpes al Largo', ot.get('golpes_largo'))
    add_field('Golpes al Ancho', ot.get('golpes_ancho'))
    add_field('Largura HC (MM)', ot.get('largura_hc'))
    add_field('Anchura HC (MM)', ot.get('anchura_hc'))

    # Colores 1-7
    for i in range(1, 8):
        color_desc = ot.get(f'color_{i}_desc')
        color_codigo = ot.get(f'color_{i}_codigo')
        cm2_clisse = ot.get(f'cm2_clisse_color_{i}')
        consumo = ot.get(f'consumo{i}')
        suffix = " (INTERIOR TyR)" if i == 1 else ""
        add_field(f'Nombre Color {i}{suffix}', color_desc)
        add_field(f'Codigo Color {i}{suffix}', color_codigo)
        add_field(f'Clisse Cm2 {i}{suffix}', cm2_clisse)
        add_field(f'Gramos Color {i}{suffix}', consumo)

    add_field('Total clisse cm2', ot.get('total_cm2_clisse'))

    # SECCION 12: MATRICES Y CLISSE
    add_section('MATRICES Y CLISSE')
    clisse = f"ENC{ot.get('material_codigo')}" if ot.get('material_codigo') else None
    add_field('Clisse1', clisse)
    add_field('Matriz 1', ot.get('matriz_1_material'))
    add_field('Matriz 2', ot.get('matriz_2_material'))
    add_field('Matriz 3', ot.get('matriz_3_material'))

    # SECCION 13: CINTA
    add_section('CINTA')
    add_field('Distancia corte 1 a cinta 1', ot.get('distancia_cinta_1'))
    add_field('Distancia corte 1 a cinta 2', ot.get('distancia_cinta_2'))
    add_field('Distancia corte 1 a cinta 3', ot.get('distancia_cinta_3'))
    add_field('Distancia corte 1 a cinta 4', ot.get('distancia_cinta_4'))
    add_field('Distancia corte 1 a cinta 5', ot.get('distancia_cinta_5'))
    add_field('Distancia corte 1 a cinta 6', ot.get('distancia_cinta_6'))
    add_field('Tipo de Cinta', ot.get('tipo_cinta_desc'))
    add_field('Cantidad Cinta por CAJA', ot.get('cintas_x_caja'))

    # Calcular consumo cinta
    consumo_cinta = None
    if ot.get('cintas_x_caja') and ot.get('largura_hm'):
        try:
            consumo_cinta = float(ot.get('largura_hm')) * float(ot.get('cintas_x_caja'))
        except:
            pass
    add_field('Consumo cinta', consumo_cinta)

    # Ajustar anchos
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 60

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
