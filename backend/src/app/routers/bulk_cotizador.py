"""
Router de Carga Masiva para Tablas del Cotizador - INVEB
Soporta Excel/CSV con validaciones especificas por tabla
"""
from datetime import datetime
from typing import Optional, List, Dict, Any
from io import BytesIO
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import pymysql

from ..config import get_settings
from .auth import get_current_user

router = APIRouter(prefix="/bulk-cotizador", tags=["Carga Masiva Cotizador"])

settings = get_settings()

def get_db_connection():
    return pymysql.connect(
        host=settings.DB_HOST,
        user=settings.DB_USER,
        password=settings.DB_PASSWORD,
        database=settings.DB_NAME,
        port=settings.DB_PORT,
        cursorclass=pymysql.cursors.DictCursor,
        charset='utf8mb4'
    )


# =============================================
# SCHEMAS
# =============================================

class BulkResult(BaseModel):
    total_filas: int
    insertados: int
    actualizados: int
    errores: int
    items_nuevos: List[Dict[str, Any]] = []
    items_actualizados: List[Dict[str, Any]] = []
    items_error: List[Dict[str, Any]] = []


class TablaPlantilla(BaseModel):
    nombre: str
    columnas: List[str]
    descripcion: str


# =============================================
# CONFIGURACION DE TABLAS
# =============================================

TABLAS_COTIZADOR = {
    "cartones": {
        "table": "cartons",
        "display_name": "Cartones Corrugados",
        "columnas": [
            "codigo", "onda", "color_tapa_exterior", "tipo", "ect_min", "espesor",
            "peso", "peso_total", "tolerancia_gramaje_real", "contenido_cordillera",
            "contenido_reciclado", "porocidad_gurley", "cobb_int", "cobb_ext",
            "recubrimiento", "codigo_tapa_interior", "codigo_onda_1", "codigo_onda_1_2",
            "codigo_tapa_media", "codigo_onda_2", "codigo_tapa_exterior",
            "desperdicio", "excepcion", "carton_muestra", "alta_grafica",
            "provisional", "carton_original", "buin", "tiltil", "osorno", "active"
        ],
        "key_field": "codigo",
        "ondas_validas": ["C", "CB", "CE", "B", "BE", "E", "P", "P-BC", "EB", "EC", "BC"],
        "tipos_validos": ["SIMPLES", "DOBLES", "DOBLE MONOTAPA", "POWER PLY", "SIMPLE EMPLACADO"],
        "colores_validos": ["blanco", "cafe"]
    },
    "papeles": {
        "table": "papers",
        "display_name": "Papeles",
        "columnas": ["codigo", "gramaje", "precio"],
        "key_field": "codigo"
    },
    "fletes": {
        "table": "fletes",
        "display_name": "Fletes",
        "columnas": ["planta", "ciudad", "costo_clp_pallet"],
        "key_field": None,  # Combinacion planta + ciudad
        "lookup_planta": True,
        "lookup_ciudad": True
    },
    "merma_corrugadoras": {
        "table": "merma_corrugadoras",
        "display_name": "Mermas Corrugadoras",
        "columnas": ["planta", "carton", "porcentaje_merma_corrugadora"],
        "key_field": None,
        "lookup_planta": True,
        "lookup_carton": True
    },
    "merma_convertidoras": {
        "table": "merma_convertidoras",
        "display_name": "Mermas Convertidoras",
        "columnas": ["planta", "proceso", "rubro", "porcentaje_merma_convertidora"],
        "key_field": None,
        "lookup_planta": True,
        "lookup_proceso": True,
        "lookup_rubro": True
    },
    "tarifario": {
        "table": "tarifario",
        "display_name": "Tarifarios",
        "columnas": ["mercado", "tipo_cliente", "carton_frecuente", "planta", "estacionalidad", "porcentaje_margen"],
        "key_field": None
    },
    "consumo_adhesivos": {
        "table": "consumo_adhesivos",
        "display_name": "Consumo Adhesivos",
        "columnas": ["planta", "onda", "adhesivo_corrugado", "adhesivo_powerply"],
        "key_field": None,
        "lookup_planta": True
    },
    "consumo_energias": {
        "table": "consumo_energias",
        "display_name": "Consumo Energias",
        "columnas": ["planta", "proceso", "consumo_kwh_mm2"],
        "key_field": None,
        "lookup_planta": True,
        "lookup_proceso": True
    },
    "factores_ondas": {
        "table": "factores_ondas",
        "display_name": "Factores de Ondas",
        "columnas": ["planta", "onda", "factor_onda"],
        "key_field": None,
        "lookup_planta": True
    },
    "factores_desarrollos": {
        "table": "factores_desarrollos",
        "display_name": "Factores de Desarrollos",
        "columnas": ["externo_largo", "externo_ancho", "externo_alto", "d1", "d2", "dh", "caja_entera", "tipo_onda", "onda_id"],
        "key_field": None,
        "lookup_onda": True
    },
    "factores_seguridads": {
        "table": "factores_seguridads",
        "display_name": "Factores de Seguridad",
        "columnas": ["rubro", "envase", "factor_seguridad"],
        "key_field": None,
        "lookup_rubro": True,
        "lookup_envase": True
    },
    "maquila_servicios": {
        "table": "maquila_servicios",
        "display_name": "Maquilas/Servicios",
        "columnas": ["servicio", "precio_clp_caja", "product_type_id", "active"],
        "key_field": "servicio"
    },
    "plantas": {
        "table": "plantas",
        "display_name": "Plantas",
        "columnas": ["nombre", "ancho_corrugadora", "trim_corrugadora"],
        "key_field": "nombre"
    },
    "tipo_ondas": {
        "table": "tipo_ondas",
        "display_name": "Tipos de Ondas",
        "columnas": ["onda", "espesor_promedio", "espesor_maximo", "espesor_minimo"],
        "key_field": "onda"
    },
    "variables_cotizador": {
        "table": "variables_cotizadors",
        "display_name": "Variables del Cotizador",
        "columnas": [
            "esq_perdida_papel", "esq_perdida_adhesivo", "esq_recorte_esquineros",
            "iva", "tasa_mensual_credito", "dias_financiamiento_credito"
        ],
        "key_field": None  # Solo hay un registro
    },
    "carton_esquineros": {
        "table": "carton_esquineros",
        "display_name": "Cartones Esquineros",
        "columnas": ["codigo", "resistencia", "espesor", "ancho_esquinero", "alta_grafica", "active"],
        "key_field": "codigo"
    },
    "insumos_palletizados": {
        "table": "insumos_palletizados",
        "display_name": "Insumos Paletizados",
        "columnas": ["insumo", "precio", "active"],
        "key_field": "insumo"
    },
    "tarifario_margens": {
        "table": "tarifario_margens",
        "display_name": "Tarifario Margenes",
        "columnas": [
            "rubro", "indice_complejidad", "tipo_cliente",
            "volumen_negociacion_minimo_2", "volumen_negociacion_maximo_2", "margen_minimo_usd_mm2"
        ],
        "key_field": None,
        "lookup_rubro": True
    }
}


# =============================================
# HELPERS
# =============================================

def get_lookups(conn):
    """Obtiene diccionarios de lookup para validaciones"""
    cursor = conn.cursor()
    lookups = {}

    # Plantas
    cursor.execute("SELECT id, nombre FROM plantas")
    lookups["plantas"] = {r["nombre"]: r["id"] for r in cursor.fetchall()}

    # Ciudades
    cursor.execute("SELECT id, ciudad FROM ciudades_fletes")
    lookups["ciudades"] = {r["ciudad"]: r["id"] for r in cursor.fetchall()}

    # Cartones
    cursor.execute("SELECT id, codigo FROM cartons")
    lookups["cartones"] = {r["codigo"]: r["id"] for r in cursor.fetchall()}

    # Procesos
    cursor.execute("SELECT id, descripcion FROM processes")
    lookups["procesos"] = {r["descripcion"]: r["id"] for r in cursor.fetchall()}

    # Rubros
    cursor.execute("SELECT id, descripcion FROM rubros")
    lookups["rubros"] = {r["descripcion"]: r["id"] for r in cursor.fetchall()}

    # Envases
    cursor.execute("SELECT id, descripcion FROM envases")
    lookups["envases"] = {r["descripcion"]: r["id"] for r in cursor.fetchall()}

    # Ondas
    cursor.execute("SELECT id, onda FROM tipo_ondas")
    lookups["ondas"] = {r["onda"]: r["id"] for r in cursor.fetchall()}

    # Papeles
    cursor.execute("SELECT id, codigo FROM papers")
    lookups["papeles"] = {str(r["codigo"]): r["id"] for r in cursor.fetchall()}
    lookups["papeles"]["0"] = 0  # Valor nulo

    cursor.close()
    return lookups


def parse_excel_file(file_content: bytes, columnas: List[str]) -> List[Dict[str, Any]]:
    """Parsea archivo Excel y retorna lista de diccionarios"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")

    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Primera fila son headers
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    result = []
    for row_idx, row in enumerate(rows[1:], start=2):
        # Verificar si la fila esta vacia
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        row_data = {"_linea": row_idx}
        for col_idx, header in enumerate(headers):
            if col_idx < len(row):
                value = row[col_idx]
                # Normalizar header para mapear a columna
                normalized = header.replace(" ", "_").lower()
                row_data[normalized] = value

        result.append(row_data)

    return result


# =============================================
# ENDPOINTS
# =============================================

@router.get("/tablas")
async def listar_tablas_disponibles(
    current_user: dict = Depends(get_current_user)
):
    """Lista las tablas disponibles para carga masiva"""
    return [
        {
            "key": key,
            "nombre": config["display_name"],
            "columnas": config["columnas"]
        }
        for key, config in TABLAS_COTIZADOR.items()
    ]


@router.get("/plantilla/{tabla_key}")
async def descargar_plantilla_excel(
    tabla_key: str,
    current_user: dict = Depends(get_current_user)
):
    """Descarga plantilla Excel para carga masiva de una tabla"""
    if tabla_key not in TABLAS_COTIZADOR:
        raise HTTPException(status_code=404, detail=f"Tabla '{tabla_key}' no encontrada")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")

    config = TABLAS_COTIZADOR[tabla_key]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config["display_name"][:31]  # Max 31 chars para nombre de hoja

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    for col_idx, col_name in enumerate(config["columnas"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(col_name) + 2, 12)

    # Guardar en buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"plantilla_{tabla_key}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/cargar/{tabla_key}", response_model=BulkResult)
async def cargar_archivo_excel(
    tabla_key: str,
    archivo: UploadFile = File(...),
    modo: str = Query("preview", description="'preview' para validar, 'ejecutar' para guardar"),
    current_user: dict = Depends(get_current_user)
):
    """
    Carga masiva desde archivo Excel.
    - modo='preview': Solo valida y muestra cambios sin guardar
    - modo='ejecutar': Valida y guarda los cambios
    """
    if tabla_key not in TABLAS_COTIZADOR:
        raise HTTPException(status_code=404, detail=f"Tabla '{tabla_key}' no encontrada")

    config = TABLAS_COTIZADOR[tabla_key]

    # Validar extension
    if not archivo.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos Excel (.xlsx, .xls) o CSV")

    # Leer archivo
    content = await archivo.read()
    rows = parse_excel_file(content, config["columnas"])

    if not rows:
        raise HTTPException(status_code=400, detail="El archivo esta vacio o no tiene datos")

    conn = get_db_connection()
    cursor = conn.cursor()
    lookups = get_lookups(conn)

    result = BulkResult(
        total_filas=len(rows),
        insertados=0,
        actualizados=0,
        errores=0
    )

    try:
        for row in rows:
            linea = row.get("_linea", 0)
            errores = []

            # Preparar datos segun tipo de tabla
            data = {}

            if tabla_key == "cartones":
                data, errores = _procesar_carton(row, config, lookups)
            elif tabla_key == "papeles":
                data, errores = _procesar_simple(row, ["codigo", "gramaje", "precio"])
            elif tabla_key == "fletes":
                data, errores = _procesar_fletes(row, lookups)
            elif tabla_key == "merma_corrugadoras":
                data, errores = _procesar_merma_corrugadoras(row, lookups)
            elif tabla_key == "merma_convertidoras":
                data, errores = _procesar_merma_convertidoras(row, lookups)
            elif tabla_key == "factores_ondas":
                data, errores = _procesar_factores_ondas(row, lookups)
            elif tabla_key == "consumo_adhesivos":
                data, errores = _procesar_consumo_adhesivos(row, lookups)
            elif tabla_key == "consumo_energias":
                data, errores = _procesar_consumo_energias(row, lookups)
            else:
                # Procesamiento generico
                data, errores = _procesar_generico(row, config)

            if errores:
                result.errores += 1
                result.items_error.append({
                    "linea": linea,
                    "errores": errores,
                    "datos": {k: v for k, v in row.items() if not k.startswith("_")}
                })
                continue

            # Verificar si existe (UPDATE) o es nuevo (INSERT)
            key_field = config.get("key_field")
            exists = False
            existing_id = None

            if key_field and key_field in data:
                cursor.execute(
                    f"SELECT id FROM {config['table']} WHERE {key_field} = %s",
                    (data[key_field],)
                )
                existing = cursor.fetchone()
                if existing:
                    exists = True
                    existing_id = existing["id"]

            if exists:
                result.actualizados += 1
                item_info = {"linea": linea, "id": existing_id, **data}
                result.items_actualizados.append(item_info)

                if modo == "ejecutar":
                    # UPDATE
                    set_clause = ", ".join([f"{k} = %s" for k in data.keys()])
                    values = list(data.values()) + [existing_id]
                    cursor.execute(
                        f"UPDATE {config['table']} SET {set_clause}, updated_at = NOW() WHERE id = %s",
                        values
                    )
            else:
                result.insertados += 1
                result.items_nuevos.append({"linea": linea, **data})

                if modo == "ejecutar":
                    # INSERT
                    data["created_at"] = datetime.now()
                    data["updated_at"] = datetime.now()
                    columns = ", ".join(data.keys())
                    placeholders = ", ".join(["%s"] * len(data))
                    cursor.execute(
                        f"INSERT INTO {config['table']} ({columns}) VALUES ({placeholders})",
                        list(data.values())
                    )

        if modo == "ejecutar":
            conn.commit()

        return result

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error en carga masiva: {str(e)}")
    finally:
        cursor.close()
        conn.close()


@router.get("/descargar/{tabla_key}")
async def descargar_datos_actuales(
    tabla_key: str,
    current_user: dict = Depends(get_current_user)
):
    """Descarga los datos actuales de una tabla en formato Excel"""
    if tabla_key not in TABLAS_COTIZADOR:
        raise HTTPException(status_code=404, detail=f"Tabla '{tabla_key}' no encontrada")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")

    config = TABLAS_COTIZADOR[tabla_key]

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute(f"SELECT * FROM {config['table']}")
        rows = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config["display_name"][:31]

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    columns = config["columnas"]
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(col_name) + 2, 12)

    # Datos
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Guardar en buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{tabla_key}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =============================================
# FUNCIONES DE PROCESAMIENTO
# =============================================

def _procesar_carton(row: Dict, config: Dict, lookups: Dict) -> tuple:
    """Procesa una fila de cartones con validaciones especificas"""
    errores = []
    data = {}

    # Validar codigo
    codigo = str(row.get("codigo", "")).strip()
    if not codigo:
        errores.append("Codigo requerido")
    else:
        data["codigo"] = codigo

    # Validar onda
    onda = str(row.get("onda", "")).strip().upper()
    if onda not in config.get("ondas_validas", []):
        errores.append(f"Onda invalida: {onda}")
    else:
        data["onda"] = onda

    # Validar color_tapa_exterior
    color = str(row.get("color_tapa_exterior", "")).strip().lower()
    if color not in config.get("colores_validos", []):
        errores.append(f"Color invalido: {color}")
    else:
        data["color_tapa_exterior"] = color

    # Validar tipo
    tipo = str(row.get("tipo", "")).strip().upper()
    if tipo not in config.get("tipos_validos", []):
        errores.append(f"Tipo invalido: {tipo}")
    else:
        data["tipo"] = tipo

    # Campos numericos
    for field in ["ect_min", "espesor", "peso", "peso_total", "tolerancia_gramaje_real",
                  "contenido_cordillera", "contenido_reciclado", "porocidad_gurley",
                  "cobb_int", "cobb_ext"]:
        value = row.get(field)
        if value is not None and value != "":
            try:
                data[field] = float(value)
            except (ValueError, TypeError):
                errores.append(f"{field} debe ser numerico")

    # Validar codigos de papeles
    for field in ["codigo_tapa_interior", "codigo_onda_1", "codigo_onda_1_2",
                  "codigo_tapa_media", "codigo_onda_2", "codigo_tapa_exterior"]:
        value = str(row.get(field, "")).strip()
        if value and value != "0":
            if value not in lookups["papeles"]:
                errores.append(f"{field} no existe: {value}")
            else:
                data[field] = value
        else:
            data[field] = "0"

    # Campos opcionales
    for field in ["recubrimiento", "desperdicio", "excepcion", "carton_muestra",
                  "alta_grafica", "provisional", "carton_original"]:
        value = row.get(field)
        if value is not None and str(value).strip():
            data[field] = str(value).strip()

    # Active
    active = row.get("active", 1)
    try:
        data["active"] = int(active) if active is not None else 1
    except:
        data["active"] = 1

    return data, errores


def _procesar_simple(row: Dict, campos: List[str]) -> tuple:
    """Procesamiento simple para tablas con pocos campos"""
    errores = []
    data = {}

    for campo in campos:
        value = row.get(campo)
        if value is not None and str(value).strip():
            data[campo] = value
        else:
            errores.append(f"{campo} requerido")

    return data, errores


def _procesar_fletes(row: Dict, lookups: Dict) -> tuple:
    """Procesa fletes con lookup de planta y ciudad"""
    errores = []
    data = {}

    # Lookup planta
    planta = str(row.get("planta", "")).strip()
    if planta not in lookups["plantas"]:
        errores.append(f"Planta no encontrada: {planta}")
    else:
        data["planta_id"] = lookups["plantas"][planta]

    # Lookup ciudad
    ciudad = str(row.get("ciudad", "")).strip()
    if ciudad not in lookups["ciudades"]:
        errores.append(f"Ciudad no encontrada: {ciudad}")
    else:
        data["ciudad_id"] = lookups["ciudades"][ciudad]

    # Costo
    costo = row.get("costo_clp_pallet")
    try:
        data["costo_clp_pallet"] = float(costo)
    except:
        errores.append("costo_clp_pallet debe ser numerico")

    return data, errores


def _procesar_merma_corrugadoras(row: Dict, lookups: Dict) -> tuple:
    """Procesa mermas corrugadoras con lookup"""
    errores = []
    data = {}

    # Lookup planta
    planta = str(row.get("planta", "")).strip()
    if planta not in lookups["plantas"]:
        errores.append(f"Planta no encontrada: {planta}")
    else:
        data["planta_id"] = lookups["plantas"][planta]

    # Lookup carton
    carton = str(row.get("carton", "")).strip()
    if carton not in lookups["cartones"]:
        errores.append(f"Carton no encontrado: {carton}")
    else:
        data["carton_id"] = lookups["cartones"][carton]

    # Porcentaje
    porcentaje = row.get("porcentaje_merma_corrugadora")
    try:
        data["porcentaje_merma_corrugadora"] = float(porcentaje)
    except:
        errores.append("porcentaje_merma_corrugadora debe ser numerico")

    return data, errores


def _procesar_merma_convertidoras(row: Dict, lookups: Dict) -> tuple:
    """Procesa mermas convertidoras con lookup"""
    errores = []
    data = {}

    # Lookup planta
    planta = str(row.get("planta", "")).strip()
    if planta not in lookups["plantas"]:
        errores.append(f"Planta no encontrada: {planta}")
    else:
        data["planta_id"] = lookups["plantas"][planta]

    # Lookup proceso
    proceso = str(row.get("proceso", "")).strip()
    if proceso not in lookups["procesos"]:
        errores.append(f"Proceso no encontrado: {proceso}")
    else:
        data["process_id"] = lookups["procesos"][proceso]

    # Lookup rubro
    rubro = str(row.get("rubro", "")).strip()
    if rubro not in lookups["rubros"]:
        errores.append(f"Rubro no encontrado: {rubro}")
    else:
        data["rubro_id"] = lookups["rubros"][rubro]

    # Porcentaje
    porcentaje = row.get("porcentaje_merma_convertidora")
    try:
        data["porcentaje_merma_convertidora"] = float(porcentaje)
    except:
        errores.append("porcentaje_merma_convertidora debe ser numerico")

    return data, errores


def _procesar_factores_ondas(row: Dict, lookups: Dict) -> tuple:
    """Procesa factores de ondas con lookup"""
    errores = []
    data = {}

    # Lookup planta
    planta = str(row.get("planta", "")).strip()
    if planta not in lookups["plantas"]:
        errores.append(f"Planta no encontrada: {planta}")
    else:
        data["planta_id"] = lookups["plantas"][planta]

    # Onda
    onda = str(row.get("onda", "")).strip().upper()
    data["onda"] = onda

    # Factor
    factor = row.get("factor_onda")
    try:
        data["factor_onda"] = float(factor)
    except:
        errores.append("factor_onda debe ser numerico")

    return data, errores


def _procesar_consumo_adhesivos(row: Dict, lookups: Dict) -> tuple:
    """Procesa consumo adhesivos con lookup"""
    errores = []
    data = {}

    # Lookup planta
    planta = str(row.get("planta", "")).strip()
    if planta not in lookups["plantas"]:
        errores.append(f"Planta no encontrada: {planta}")
    else:
        data["planta_id"] = lookups["plantas"][planta]

    # Onda
    data["onda"] = str(row.get("onda", "")).strip().upper()

    # Consumos
    for field in ["adhesivo_corrugado", "adhesivo_powerply"]:
        value = row.get(field)
        try:
            data[field] = float(value) if value else 0
        except:
            errores.append(f"{field} debe ser numerico")

    return data, errores


def _procesar_consumo_energias(row: Dict, lookups: Dict) -> tuple:
    """Procesa consumo energias con lookup"""
    errores = []
    data = {}

    # Lookup planta
    planta = str(row.get("planta", "")).strip()
    if planta not in lookups["plantas"]:
        errores.append(f"Planta no encontrada: {planta}")
    else:
        data["planta_id"] = lookups["plantas"][planta]

    # Lookup proceso
    proceso = str(row.get("proceso", "")).strip()
    if proceso not in lookups["procesos"]:
        errores.append(f"Proceso no encontrado: {proceso}")
    else:
        data["process_id"] = lookups["procesos"][proceso]

    # Consumo
    consumo = row.get("consumo_kwh_mm2")
    try:
        data["consumo_kwh_mm2"] = float(consumo)
    except:
        errores.append("consumo_kwh_mm2 debe ser numerico")

    return data, errores


def _procesar_generico(row: Dict, config: Dict) -> tuple:
    """Procesamiento generico para tablas sin logica especial"""
    errores = []
    data = {}

    for col in config["columnas"]:
        value = row.get(col)
        if value is not None:
            # Intentar convertir a numero si parece numerico
            if isinstance(value, (int, float)):
                data[col] = value
            else:
                str_value = str(value).strip()
                if str_value:
                    data[col] = str_value

    return data, errores
