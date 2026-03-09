"""
Sprint D: Mantenedor Masivo - Cargas Excel para 24 tablas.

Basado en: MantenedorController.php (5778 líneas)

Cada tabla tiene 2 endpoints:
- POST /import - Importar desde Excel
- GET /descargar-excel - Descargar plantilla/datos

Tablas implementadas:
1. Cartones Corrugados
2. Cartones Esquineros
3. Papeles
4. Fletes
5. Mermas Corrugadoras
6. Mermas Convertidoras
7. Paletizados
8. Insumos Paletizados
9. Tarifarios
10. Consumo Adhesivo
11. Consumo Adhesivo Pegado
12. Consumo Energia
13. Matrices
14. Materiales
15. Factores Desarrollo
16. Factores Seguridad
17. Factores Onda
18. Maquilas
19. Ondas
20. Plantas
21. Variables
22. Margenes Minimos
23. Porcentajes Margenes
24. Mano Obra Mantencion
"""

from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
from pydantic import BaseModel
from datetime import datetime
import pymysql
import pymysql.cursors
import openpyxl
from openpyxl import Workbook
from io import BytesIO
import os

router = APIRouter(prefix="/masivos", tags=["Mantenedores Masivos"])


def get_db_connection():
    """Conexión a MySQL."""
    return pymysql.connect(
        host=os.getenv("MYSQL_HOST", "localhost"),
        user=os.getenv("MYSQL_USER", "root"),
        password=os.getenv("MYSQL_PASSWORD", ""),
        database=os.getenv("MYSQL_DATABASE", "envases_ot"),
        port=int(os.getenv("MYSQL_PORT", 3306)),
        cursorclass=pymysql.cursors.DictCursor,
        charset='utf8mb4'
    )


# =============================================
# 1. CARTONES CORRUGADOS
# Basado en: MantenedorController líneas 60-354
# =============================================

CARTONES_CORRUGADOS_COLUMNS = [
    "codigo", "onda", "color_tapa_exterior", "tipo", "ect_min", "espesor",
    "peso", "peso_total", "tolerancia_gramaje_real", "contenido_cordillera",
    "contenido_reciclado", "porocidad_gurley", "cobb_int", "cobb_ext",
    "recubrimiento", "codigo_tapa_interior", "codigo_onda_1", "codigo_onda_1_2",
    "codigo_tapa_media", "codigo_onda_2", "codigo_tapa_exterior", "desperdicio",
    "excepcion", "carton_muestra", "alta_grafica", "buin", "osorno", "tiltil",
    "provisional", "carton_original", "active"
]

ONDAS_VALIDAS = ["C", "CB", "CE", "B", "BE", "E", "P", "P-BC", "EB", "EC", "BC"]
TIPOS_CARTON_VALIDOS = ["SIMPLES", "DOBLES", "DOBLE MONOTAPA", "POWER PLY", "SIMPLE EMPLACADO"]
COLORES_VALIDOS = ["blanco", "cafe"]


@router.post("/cartones-corrugados/import")
async def import_cartones_corrugados(
    archivo: UploadFile = File(...),
    proceso: str = Query("validacion", description="validacion o cargaCompleta")
):
    """
    Importa cartones corrugados desde Excel.

    Basado en: MantenedorController@importCartons (líneas 60-354)

    Columnas requeridas:
    - codigo, onda, color_tapa_exterior, tipo, ect_min, espesor, peso, peso_total
    - tolerancia_gramaje_real, contenido_cordillera, contenido_reciclado
    - porocidad_gurley, cobb_int, cobb_ext, recubrimiento
    - codigo_tapa_interior, codigo_onda_1, codigo_onda_1_2, codigo_tapa_media
    - codigo_onda_2, codigo_tapa_exterior, desperdicio, excepcion
    - carton_muestra, alta_grafica, buin, osorno, tiltil, provisional
    - carton_original, active

    Args:
        archivo: Archivo Excel (.xlsx, .xls)
        proceso: "validacion" para preview, "cargaCompleta" para guardar

    Returns:
        Resumen de cartones procesados, actualizados y errores
    """
    # Validar extensión
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido. Use .xlsx o .xls")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Obtener papeles válidos para validación de códigos
        cursor.execute("SELECT id, codigo FROM papers WHERE active = 1")
        papeles = {str(row["codigo"]): row["id"] for row in cursor.fetchall()}
        papeles["0"] = 0  # Nulo permitido

        # Leer Excel
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        # Obtener headers
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        cartones_nuevos = []
        cartones_actualizados = []
        cartones_inactivados = []
        cartones_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            # Crear diccionario con valores
            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            # Validaciones según Laravel (líneas 104-156)
            motivos = []

            codigo = str(row_data.get("codigo", "")).strip()
            if not codigo:
                motivos.append("Codigo vacío")

            onda = str(row_data.get("onda", "")).strip()
            if onda not in ONDAS_VALIDAS:
                motivos.append(f"Onda inválida: {onda}")

            color_tapa = str(row_data.get("color_tapa_exterior", "")).strip().lower()
            if color_tapa not in COLORES_VALIDOS:
                motivos.append(f"Color tapa exterior inválido: {color_tapa}")

            tipo = str(row_data.get("tipo", "")).strip().upper()
            if tipo not in TIPOS_CARTON_VALIDOS:
                motivos.append(f"Tipo inválido: {tipo}")

            # Validar códigos de papeles
            for campo in ["codigo_tapa_interior", "codigo_onda_1", "codigo_onda_1_2",
                         "codigo_tapa_media", "codigo_onda_2", "codigo_tapa_exterior"]:
                valor = str(row_data.get(campo, "0")).strip()
                if valor and valor != "0" and valor not in papeles:
                    motivos.append(f"{campo} no existe: {valor}")

            # Validar numéricos
            campos_numericos = [
                "ect_min", "espesor", "peso", "peso_total", "tolerancia_gramaje_real",
                "contenido_cordillera", "contenido_reciclado", "porocidad_gurley",
                "cobb_int", "cobb_ext"
            ]
            for campo in campos_numericos:
                valor = row_data.get(campo)
                if valor is not None and valor != "":
                    try:
                        float(valor)
                    except (ValueError, TypeError):
                        motivos.append(f"{campo} no es numérico: {valor}")

            if motivos:
                cartones_error.append({
                    "linea": row_num,
                    "codigo": codigo,
                    "motivos": motivos
                })
                continue

            # Verificar si existe
            cursor.execute("SELECT id, active FROM cardboards WHERE codigo = %s", (codigo,))
            carton_existente = cursor.fetchone()

            # Procesar plantas (Buin=1, Tiltil=2, Osorno=3)
            plantas = []
            if row_data.get("buin") and str(row_data.get("buin")) not in ["", "0"]:
                plantas.append("1")
            if row_data.get("tiltil") and str(row_data.get("tiltil")) not in ["", "0"]:
                plantas.append("2")
            if row_data.get("osorno") and str(row_data.get("osorno")) not in ["", "0"]:
                plantas.append("3")
            planta_id = ",".join(sorted(plantas)) if plantas else None

            carton_data = {
                "codigo": codigo,
                "onda": onda,
                "color_tapa_exterior": row_data.get("color_tapa_exterior"),
                "tipo": tipo,
                "ect_min": float(row_data.get("ect_min") or 0),
                "espesor": float(row_data.get("espesor") or 0),
                "peso": float(row_data.get("peso") or 0),
                "peso_total": float(row_data.get("peso_total") or 0),
                "tolerancia_gramaje_real": float(row_data.get("tolerancia_gramaje_real") or 0),
                "contenido_cordillera": float(row_data.get("contenido_cordillera") or 0),
                "contenido_reciclado": float(row_data.get("contenido_reciclado") or 0),
                "porocidad_gurley": float(row_data.get("porocidad_gurley") or 0),
                "cobb_int": float(row_data.get("cobb_int") or 0),
                "cobb_ext": float(row_data.get("cobb_ext") or 0),
                "recubrimiento": row_data.get("recubrimiento"),
                "codigo_tapa_interior": row_data.get("codigo_tapa_interior"),
                "codigo_onda_1": row_data.get("codigo_onda_1"),
                "codigo_onda_1_2": row_data.get("codigo_onda_1_2"),
                "codigo_tapa_media": row_data.get("codigo_tapa_media"),
                "codigo_onda_2": row_data.get("codigo_onda_2"),
                "codigo_tapa_exterior": row_data.get("codigo_tapa_exterior"),
                "desperdicio": row_data.get("desperdicio"),
                "excepcion": row_data.get("excepcion"),
                "carton_muestra": row_data.get("carton_muestra"),
                "alta_grafica": row_data.get("alta_grafica"),
                "provisional": row_data.get("provisional"),
                "carton_original": row_data.get("carton_original"),
                "planta_id": planta_id,
                "active": int(row_data.get("active") or 1),
                "orden": row_num
            }

            if proceso == "cargaCompleta":
                if carton_existente:
                    # Verificar inactivación
                    if carton_existente["active"] == 1 and carton_data["active"] == 0:
                        cartones_inactivados.append({"codigo": codigo, "linea": row_num})
                    else:
                        cartones_actualizados.append({"codigo": codigo, "linea": row_num})

                    # UPDATE
                    cursor.execute("""
                        UPDATE cardboards SET
                            onda = %s, color_tapa_exterior = %s, tipo = %s,
                            ect_min = %s, espesor = %s, peso = %s, peso_total = %s,
                            tolerancia_gramaje_real = %s, contenido_cordillera = %s,
                            contenido_reciclado = %s, porocidad_gurley = %s,
                            cobb_int = %s, cobb_ext = %s, recubrimiento = %s,
                            codigo_tapa_interior = %s, codigo_onda_1 = %s,
                            codigo_onda_1_2 = %s, codigo_tapa_media = %s,
                            codigo_onda_2 = %s, codigo_tapa_exterior = %s,
                            desperdicio = %s, excepcion = %s, carton_muestra = %s,
                            alta_grafica = %s, provisional = %s, carton_original = %s,
                            planta_id = %s, active = %s, orden = %s, updated_at = NOW()
                        WHERE codigo = %s
                    """, (
                        carton_data["onda"], carton_data["color_tapa_exterior"], carton_data["tipo"],
                        carton_data["ect_min"], carton_data["espesor"], carton_data["peso"],
                        carton_data["peso_total"], carton_data["tolerancia_gramaje_real"],
                        carton_data["contenido_cordillera"], carton_data["contenido_reciclado"],
                        carton_data["porocidad_gurley"], carton_data["cobb_int"], carton_data["cobb_ext"],
                        carton_data["recubrimiento"], carton_data["codigo_tapa_interior"],
                        carton_data["codigo_onda_1"], carton_data["codigo_onda_1_2"],
                        carton_data["codigo_tapa_media"], carton_data["codigo_onda_2"],
                        carton_data["codigo_tapa_exterior"], carton_data["desperdicio"],
                        carton_data["excepcion"], carton_data["carton_muestra"],
                        carton_data["alta_grafica"], carton_data["provisional"],
                        carton_data["carton_original"], carton_data["planta_id"],
                        carton_data["active"], carton_data["orden"], codigo
                    ))
                else:
                    # INSERT
                    cursor.execute("""
                        INSERT INTO cardboards (
                            codigo, onda, color_tapa_exterior, tipo, ect_min, espesor,
                            peso, peso_total, tolerancia_gramaje_real, contenido_cordillera,
                            contenido_reciclado, porocidad_gurley, cobb_int, cobb_ext,
                            recubrimiento, codigo_tapa_interior, codigo_onda_1, codigo_onda_1_2,
                            codigo_tapa_media, codigo_onda_2, codigo_tapa_exterior, desperdicio,
                            excepcion, carton_muestra, alta_grafica, provisional, carton_original,
                            planta_id, active, orden, created_at, updated_at
                        ) VALUES (
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW()
                        )
                    """, (
                        carton_data["codigo"], carton_data["onda"], carton_data["color_tapa_exterior"],
                        carton_data["tipo"], carton_data["ect_min"], carton_data["espesor"],
                        carton_data["peso"], carton_data["peso_total"], carton_data["tolerancia_gramaje_real"],
                        carton_data["contenido_cordillera"], carton_data["contenido_reciclado"],
                        carton_data["porocidad_gurley"], carton_data["cobb_int"], carton_data["cobb_ext"],
                        carton_data["recubrimiento"], carton_data["codigo_tapa_interior"],
                        carton_data["codigo_onda_1"], carton_data["codigo_onda_1_2"],
                        carton_data["codigo_tapa_media"], carton_data["codigo_onda_2"],
                        carton_data["codigo_tapa_exterior"], carton_data["desperdicio"],
                        carton_data["excepcion"], carton_data["carton_muestra"],
                        carton_data["alta_grafica"], carton_data["provisional"],
                        carton_data["carton_original"], carton_data["planta_id"],
                        carton_data["active"], carton_data["orden"]
                    ))
                    cartones_nuevos.append({"codigo": codigo, "linea": row_num, "id": cursor.lastrowid})
            else:
                # Solo validación
                if carton_existente:
                    if carton_existente["active"] == 1 and carton_data["active"] == 0:
                        cartones_inactivados.append({"codigo": codigo, "linea": row_num})
                    else:
                        cartones_actualizados.append({"codigo": codigo, "linea": row_num})
                else:
                    cartones_nuevos.append({"codigo": codigo, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(cartones_error) == 0,
            "mensaje": "Archivo procesado exitosamente",
            "proceso": proceso,
            "cartones_nuevos": cartones_nuevos,
            "cartones_actualizados": cartones_actualizados,
            "cartones_inactivados": cartones_inactivados,
            "cartones_error": cartones_error,
            "totales": {
                "nuevos": len(cartones_nuevos),
                "actualizados": len(cartones_actualizados),
                "inactivados": len(cartones_inactivados),
                "errores": len(cartones_error)
            }
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/cartones-corrugados/descargar-excel")
async def descargar_excel_cartones_corrugados():
    """
    Descarga Excel con cartones corrugados.

    Basado en: MantenedorController@descargar_excel_cartones_corrugados (líneas 356-470)
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT * FROM cardboards
            WHERE tipo != 'ESQUINEROS' OR tipo IS NULL
            ORDER BY ISNULL(orden), orden ASC
        """)
        cartones = cursor.fetchall()

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cartones Corrugados"

        # Headers según Laravel (líneas 361-394)
        headers = [
            "ID", "codigo", "onda", "color_tapa_exterior", "tipo", "ect_min",
            "espesor", "peso", "peso_total", "tolerancia_gramaje_real",
            "contenido_cordillera", "contenido_reciclado", "porocidad_gurley",
            "cobb_int", "cobb_ext", "recubrimiento", "codigo_tapa_interior",
            "codigo_onda_1", "codigo_onda_1_2", "codigo_tapa_media",
            "codigo_onda_2", "codigo_tapa_exterior", "desperdicio", "excepcion",
            "carton_muestra", "alta_grafica", "Buin", "Osorno", "Tiltil",
            "provisional", "carton_original", "active"
        ]
        ws.append(headers)

        # Datos
        for carton in cartones:
            # Procesar plantas (según Laravel líneas 399-423)
            buin, tiltil, osorno = 0, 0, 0
            planta_id = carton.get("planta_id") or ""
            if planta_id:
                plantas = planta_id.split(",")
                if "1" in plantas:
                    buin = 1
                if "2" in plantas:
                    tiltil = 1
                if "3" in plantas:
                    osorno = 1

            ws.append([
                carton.get("id"),
                carton.get("codigo"),
                carton.get("onda"),
                carton.get("color_tapa_exterior"),
                carton.get("tipo"),
                carton.get("ect_min"),
                carton.get("espesor"),
                carton.get("peso"),
                carton.get("peso_total"),
                carton.get("tolerancia_gramaje_real"),
                carton.get("contenido_cordillera"),
                carton.get("contenido_reciclado"),
                carton.get("porocidad_gurley"),
                carton.get("cobb_int"),
                carton.get("cobb_ext"),
                carton.get("recubrimiento"),
                carton.get("codigo_tapa_interior"),
                carton.get("codigo_onda_1"),
                carton.get("codigo_onda_1_2"),
                carton.get("codigo_tapa_media"),
                carton.get("codigo_onda_2"),
                carton.get("codigo_tapa_exterior"),
                carton.get("desperdicio"),
                carton.get("excepcion"),
                carton.get("carton_muestra"),
                carton.get("alta_grafica"),
                buin,
                osorno,
                tiltil,
                carton.get("provisional"),
                carton.get("carton_original"),
                carton.get("active")
            ])

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Listado_Cartones_Corrugados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 2. CARTONES ESQUINEROS
# Basado en: MantenedorController líneas 485-750
# =============================================

@router.post("/cartones-esquineros/import")
async def import_cartones_esquineros(
    archivo: UploadFile = File(...),
    proceso: str = Query("validacion")
):
    """
    Importa cartones esquineros desde Excel.

    Basado en: MantenedorController@importCartonesEsquineros (líneas 485-748)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        esquineros_nuevos = []
        esquineros_actualizados = []
        esquineros_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []
            codigo = str(row_data.get("codigo", "")).strip()
            if not codigo:
                motivos.append("Codigo vacío")

            # Validar numéricos
            for campo in ["gramaje", "espesor", "ancho", "largo", "precio"]:
                valor = row_data.get(campo)
                if valor is not None and valor != "":
                    try:
                        float(valor)
                    except (ValueError, TypeError):
                        motivos.append(f"{campo} no es numérico")

            if motivos:
                esquineros_error.append({
                    "linea": row_num,
                    "codigo": codigo,
                    "motivos": motivos
                })
                continue

            cursor.execute("SELECT id FROM carton_esquineros WHERE codigo = %s", (codigo,))
            existente = cursor.fetchone()

            if proceso == "cargaCompleta":
                if existente:
                    cursor.execute("""
                        UPDATE carton_esquineros SET
                            descripcion = %s, gramaje = %s, espesor = %s,
                            ancho = %s, largo = %s, precio = %s,
                            active = %s, orden = %s, updated_at = NOW()
                        WHERE codigo = %s
                    """, (
                        row_data.get("descripcion"),
                        float(row_data.get("gramaje") or 0),
                        float(row_data.get("espesor") or 0),
                        float(row_data.get("ancho") or 0),
                        float(row_data.get("largo") or 0),
                        float(row_data.get("precio") or 0),
                        int(row_data.get("active") or 1),
                        row_num,
                        codigo
                    ))
                    esquineros_actualizados.append({"codigo": codigo, "linea": row_num})
                else:
                    cursor.execute("""
                        INSERT INTO carton_esquineros (
                            codigo, descripcion, gramaje, espesor, ancho, largo,
                            precio, active, orden, created_at, updated_at
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                    """, (
                        codigo,
                        row_data.get("descripcion"),
                        float(row_data.get("gramaje") or 0),
                        float(row_data.get("espesor") or 0),
                        float(row_data.get("ancho") or 0),
                        float(row_data.get("largo") or 0),
                        float(row_data.get("precio") or 0),
                        int(row_data.get("active") or 1),
                        row_num
                    ))
                    esquineros_nuevos.append({"codigo": codigo, "linea": row_num, "id": cursor.lastrowid})
            else:
                if existente:
                    esquineros_actualizados.append({"codigo": codigo, "linea": row_num})
                else:
                    esquineros_nuevos.append({"codigo": codigo, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(esquineros_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "esquineros_nuevos": esquineros_nuevos,
            "esquineros_actualizados": esquineros_actualizados,
            "esquineros_error": esquineros_error
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/cartones-esquineros/descargar-excel")
async def descargar_excel_cartones_esquineros():
    """
    Descarga Excel con cartones esquineros.

    Basado en: MantenedorController@descargar_excel_cartones_esquineros (líneas 750-827)
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT * FROM carton_esquineros
            ORDER BY ISNULL(orden), orden ASC
        """)
        esquineros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Cartones Esquineros"

        headers = ["ID", "codigo", "descripcion", "gramaje", "espesor", "ancho", "largo", "precio", "active"]
        ws.append(headers)

        for esq in esquineros:
            ws.append([
                esq.get("id"),
                esq.get("codigo"),
                esq.get("descripcion"),
                esq.get("gramaje"),
                esq.get("espesor"),
                esq.get("ancho"),
                esq.get("largo"),
                esq.get("precio"),
                esq.get("active")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Listado_Cartones_Esquineros_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 3. PAPELES
# Basado en: MantenedorController líneas 828-1039
# =============================================

@router.post("/papeles/import")
async def import_papeles(
    archivo: UploadFile = File(...),
    proceso: str = Query("validacion")
):
    """
    Importa papeles desde Excel.

    Basado en: MantenedorController@importPapeles (líneas 828-993)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        papeles_nuevos = []
        papeles_actualizados = []
        papeles_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []
            codigo = str(row_data.get("codigo", "")).strip()
            if not codigo:
                motivos.append("Codigo vacío")

            # Validar numéricos
            for campo in ["gramaje", "precio", "ancho", "rct", "cmt"]:
                valor = row_data.get(campo)
                if valor is not None and valor != "":
                    try:
                        float(valor)
                    except (ValueError, TypeError):
                        motivos.append(f"{campo} no es numérico")

            if motivos:
                papeles_error.append({
                    "linea": row_num,
                    "codigo": codigo,
                    "motivos": motivos
                })
                continue

            cursor.execute("SELECT id FROM papers WHERE codigo = %s", (codigo,))
            existente = cursor.fetchone()

            if proceso == "cargaCompleta":
                if existente:
                    cursor.execute("""
                        UPDATE papers SET
                            descripcion = %s, gramaje = %s, precio = %s,
                            ancho = %s, rct = %s, cmt = %s, tipo = %s,
                            proveedor = %s, active = %s, updated_at = NOW()
                        WHERE codigo = %s
                    """, (
                        row_data.get("descripcion"),
                        float(row_data.get("gramaje") or 0),
                        float(row_data.get("precio") or 0),
                        float(row_data.get("ancho") or 0),
                        float(row_data.get("rct") or 0),
                        float(row_data.get("cmt") or 0),
                        row_data.get("tipo"),
                        row_data.get("proveedor"),
                        int(row_data.get("active") or 1),
                        codigo
                    ))
                    papeles_actualizados.append({"codigo": codigo, "linea": row_num})
                else:
                    cursor.execute("""
                        INSERT INTO papers (
                            codigo, descripcion, gramaje, precio, ancho,
                            rct, cmt, tipo, proveedor, active, created_at, updated_at
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                    """, (
                        codigo,
                        row_data.get("descripcion"),
                        float(row_data.get("gramaje") or 0),
                        float(row_data.get("precio") or 0),
                        float(row_data.get("ancho") or 0),
                        float(row_data.get("rct") or 0),
                        float(row_data.get("cmt") or 0),
                        row_data.get("tipo"),
                        row_data.get("proveedor"),
                        int(row_data.get("active") or 1)
                    ))
                    papeles_nuevos.append({"codigo": codigo, "linea": row_num, "id": cursor.lastrowid})
            else:
                if existente:
                    papeles_actualizados.append({"codigo": codigo, "linea": row_num})
                else:
                    papeles_nuevos.append({"codigo": codigo, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(papeles_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "papeles_nuevos": papeles_nuevos,
            "papeles_actualizados": papeles_actualizados,
            "papeles_error": papeles_error
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/papeles/descargar-excel")
async def descargar_excel_papeles():
    """
    Descarga Excel con papeles.

    Basado en: MantenedorController@descargar_excel_papeles (líneas 994-1039)
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT * FROM papers ORDER BY codigo")
        papeles = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Papeles"

        headers = ["ID", "codigo", "descripcion", "gramaje", "precio", "ancho", "rct", "cmt", "tipo", "proveedor", "active"]
        ws.append(headers)

        for papel in papeles:
            ws.append([
                papel.get("id"),
                papel.get("codigo"),
                papel.get("descripcion"),
                papel.get("gramaje"),
                papel.get("precio"),
                papel.get("ancho"),
                papel.get("rct"),
                papel.get("cmt"),
                papel.get("tipo"),
                papel.get("proveedor"),
                papel.get("active")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Listado_Papeles_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 4. FLETES
# Basado en: MantenedorController líneas 1040-1253
# =============================================

@router.post("/fletes/import")
async def import_fletes(
    archivo: UploadFile = File(...),
    proceso: str = Query("validacion")
):
    """
    Importa fletes desde Excel.

    Basado en: MantenedorController@importFletes (líneas 1040-1204)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        fletes_nuevos = []
        fletes_actualizados = []
        fletes_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []
            ciudad = str(row_data.get("ciudad", "")).strip()
            if not ciudad:
                motivos.append("Ciudad vacía")

            # Validar numéricos
            for campo in ["valor_flete", "distancia"]:
                valor = row_data.get(campo)
                if valor is not None and valor != "":
                    try:
                        float(valor)
                    except (ValueError, TypeError):
                        motivos.append(f"{campo} no es numérico")

            if motivos:
                fletes_error.append({
                    "linea": row_num,
                    "ciudad": ciudad,
                    "motivos": motivos
                })
                continue

            cursor.execute("SELECT id FROM ciudades_fletes WHERE ciudad = %s", (ciudad,))
            existente = cursor.fetchone()

            if proceso == "cargaCompleta":
                if existente:
                    cursor.execute("""
                        UPDATE ciudades_fletes SET
                            region = %s, valor_flete = %s, distancia = %s,
                            planta_id = %s, active = %s, updated_at = NOW()
                        WHERE ciudad = %s
                    """, (
                        row_data.get("region"),
                        float(row_data.get("valor_flete") or 0),
                        float(row_data.get("distancia") or 0),
                        row_data.get("planta_id"),
                        int(row_data.get("active") or 1),
                        ciudad
                    ))
                    fletes_actualizados.append({"ciudad": ciudad, "linea": row_num})
                else:
                    cursor.execute("""
                        INSERT INTO ciudades_fletes (
                            ciudad, region, valor_flete, distancia,
                            planta_id, active, created_at, updated_at
                        ) VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
                    """, (
                        ciudad,
                        row_data.get("region"),
                        float(row_data.get("valor_flete") or 0),
                        float(row_data.get("distancia") or 0),
                        row_data.get("planta_id"),
                        int(row_data.get("active") or 1)
                    ))
                    fletes_nuevos.append({"ciudad": ciudad, "linea": row_num, "id": cursor.lastrowid})
            else:
                if existente:
                    fletes_actualizados.append({"ciudad": ciudad, "linea": row_num})
                else:
                    fletes_nuevos.append({"ciudad": ciudad, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(fletes_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "fletes_nuevos": fletes_nuevos,
            "fletes_actualizados": fletes_actualizados,
            "fletes_error": fletes_error
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/fletes/descargar-excel")
async def descargar_excel_fletes():
    """
    Descarga Excel con fletes.

    Basado en: MantenedorController@descargar_excel_fletes (líneas 1205-1253)
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT * FROM ciudades_fletes ORDER BY ciudad")
        fletes = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Fletes"

        headers = ["ID", "ciudad", "region", "valor_flete", "distancia", "planta_id", "active"]
        ws.append(headers)

        for flete in fletes:
            ws.append([
                flete.get("id"),
                flete.get("ciudad"),
                flete.get("region"),
                flete.get("valor_flete"),
                flete.get("distancia"),
                flete.get("planta_id"),
                flete.get("active")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Listado_Fletes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 5. MERMAS CORRUGADORAS
# Basado en: MantenedorController líneas 1254-1450
# =============================================

@router.post("/mermas-corrugadoras/import")
async def import_mermas_corrugadoras(
    archivo: UploadFile = File(...),
    proceso: str = Query("validacion")
):
    """
    Importa mermas corrugadoras desde Excel.

    Basado en: MantenedorController@importMermasCorrugadoras (líneas 1254-1406)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        mermas_nuevas = []
        mermas_actualizadas = []
        mermas_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            # Validar numéricos
            for campo in ["ancho_min", "ancho_max", "porcentaje_merma"]:
                valor = row_data.get(campo)
                if valor is None or valor == "":
                    motivos.append(f"{campo} vacío")
                else:
                    try:
                        float(valor)
                    except (ValueError, TypeError):
                        motivos.append(f"{campo} no es numérico")

            if motivos:
                mermas_error.append({
                    "linea": row_num,
                    "motivos": motivos
                })
                continue

            ancho_min = float(row_data.get("ancho_min") or 0)
            ancho_max = float(row_data.get("ancho_max") or 0)
            planta_id = row_data.get("planta_id")

            cursor.execute("""
                SELECT id FROM mermas_corrugadoras
                WHERE ancho_min = %s AND ancho_max = %s AND planta_id = %s
            """, (ancho_min, ancho_max, planta_id))
            existente = cursor.fetchone()

            if proceso == "cargaCompleta":
                if existente:
                    cursor.execute("""
                        UPDATE mermas_corrugadoras SET
                            porcentaje_merma = %s, active = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (
                        float(row_data.get("porcentaje_merma") or 0),
                        int(row_data.get("active") or 1),
                        existente["id"]
                    ))
                    mermas_actualizadas.append({"id": existente["id"], "linea": row_num})
                else:
                    cursor.execute("""
                        INSERT INTO mermas_corrugadoras (
                            ancho_min, ancho_max, porcentaje_merma,
                            planta_id, active, created_at, updated_at
                        ) VALUES (%s, %s, %s, %s, %s, NOW(), NOW())
                    """, (
                        ancho_min, ancho_max,
                        float(row_data.get("porcentaje_merma") or 0),
                        planta_id,
                        int(row_data.get("active") or 1)
                    ))
                    mermas_nuevas.append({"id": cursor.lastrowid, "linea": row_num})
            else:
                if existente:
                    mermas_actualizadas.append({"id": existente["id"], "linea": row_num})
                else:
                    mermas_nuevas.append({"linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(mermas_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "mermas_nuevas": mermas_nuevas,
            "mermas_actualizadas": mermas_actualizadas,
            "mermas_error": mermas_error
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/mermas-corrugadoras/descargar-excel")
async def descargar_excel_mermas_corrugadoras():
    """Descarga Excel con mermas corrugadoras."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT * FROM mermas_corrugadoras ORDER BY ancho_min")
        mermas = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Mermas Corrugadoras"

        headers = ["ID", "ancho_min", "ancho_max", "porcentaje_merma", "planta_id", "active"]
        ws.append(headers)

        for merma in mermas:
            ws.append([
                merma.get("id"),
                merma.get("ancho_min"),
                merma.get("ancho_max"),
                merma.get("porcentaje_merma"),
                merma.get("planta_id"),
                merma.get("active")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Mermas_Corrugadoras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 6. MERMAS CONVERTIDORAS
# Basado en: MantenedorController líneas 1451-1654
# =============================================

@router.post("/mermas-convertidoras/import")
async def import_mermas_convertidoras(
    archivo: UploadFile = File(...),
    proceso: str = Query("validacion")
):
    """
    Importa mermas convertidoras desde Excel.

    Basado en: MantenedorController@importMermasConvertidoras (líneas 1451-1610)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        mermas_nuevas = []
        mermas_actualizadas = []
        mermas_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []
            process_id = row_data.get("process_id")
            if not process_id:
                motivos.append("process_id vacío")

            for campo in ["porcentaje_merma"]:
                valor = row_data.get(campo)
                if valor is not None and valor != "":
                    try:
                        float(valor)
                    except (ValueError, TypeError):
                        motivos.append(f"{campo} no es numérico")

            if motivos:
                mermas_error.append({"linea": row_num, "motivos": motivos})
                continue

            cursor.execute("SELECT id FROM mermas_convertidoras WHERE process_id = %s", (process_id,))
            existente = cursor.fetchone()

            if proceso == "cargaCompleta":
                if existente:
                    cursor.execute("""
                        UPDATE mermas_convertidoras SET
                            porcentaje_merma = %s, active = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (
                        float(row_data.get("porcentaje_merma") or 0),
                        int(row_data.get("active") or 1),
                        existente["id"]
                    ))
                    mermas_actualizadas.append({"id": existente["id"], "linea": row_num})
                else:
                    cursor.execute("""
                        INSERT INTO mermas_convertidoras (
                            process_id, porcentaje_merma, active, created_at, updated_at
                        ) VALUES (%s, %s, %s, NOW(), NOW())
                    """, (
                        process_id,
                        float(row_data.get("porcentaje_merma") or 0),
                        int(row_data.get("active") or 1)
                    ))
                    mermas_nuevas.append({"id": cursor.lastrowid, "linea": row_num})
            else:
                if existente:
                    mermas_actualizadas.append({"linea": row_num})
                else:
                    mermas_nuevas.append({"linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(mermas_error) == 0,
            "proceso": proceso,
            "mermas_nuevas": mermas_nuevas,
            "mermas_actualizadas": mermas_actualizadas,
            "mermas_error": mermas_error
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/mermas-convertidoras/descargar-excel")
async def descargar_excel_mermas_convertidoras():
    """Descarga Excel con mermas convertidoras."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT mc.*, p.descripcion as proceso_nombre
            FROM mermas_convertidoras mc
            LEFT JOIN processes p ON mc.process_id = p.id
            ORDER BY mc.process_id
        """)
        mermas = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Mermas Convertidoras"

        headers = ["ID", "process_id", "proceso_nombre", "porcentaje_merma", "active"]
        ws.append(headers)

        for merma in mermas:
            ws.append([
                merma.get("id"),
                merma.get("process_id"),
                merma.get("proceso_nombre"),
                merma.get("porcentaje_merma"),
                merma.get("active")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Mermas_Convertidoras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 7-24: RESTO DE TABLAS (Continuación)
# Se implementan siguiendo el mismo patrón
# =============================================

# 7. PALETIZADOS
# Basado en: MantenedorController líneas 1655-1780
PALETIZADOS_CAMPOS_NUMERICOS = [
    "tarima_nacional", "tarima_exportacion", "liston_nacional", "liston_exportacion",
    "tabla_tarima", "stretch_film", "sellos", "zunchos", "fundas", "cordel_y_clavos", "maquila"
]

@router.post("/paletizados/import")
async def import_paletizados(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa paletizados desde Excel.

    Basado en: MantenedorController@importPaletizados (líneas 1655-1780)

    Columnas requeridas:
    - tipo_palletizado (string, requerido)
    - tarima_nacional, tarima_exportacion, liston_nacional, liston_exportacion (int)
    - tabla_tarima, stretch_film, sellos, zunchos, fundas, cordel_y_clavos, maquila (int)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido. Solo se permiten archivos .xlsx o .xls")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        paletizados_actualizados = []
        paletizados_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            # Validar tipo_palletizado (requerido)
            tipo_palletizado = str(row_data.get("tipo_palletizado", "")).strip()
            if not tipo_palletizado:
                motivos.append("tipo_palletizado vacío")

            # Validar campos numéricos
            valores_numericos = {}
            for campo in PALETIZADOS_CAMPOS_NUMERICOS:
                valor = row_data.get(campo)
                if valor is not None and valor != "":
                    try:
                        valores_numericos[campo] = int(float(valor))
                    except (ValueError, TypeError):
                        motivos.append(f"{campo} no es numérico")
                else:
                    motivos.append(f"{campo} vacío")

            if motivos:
                paletizados_error.append({
                    "linea": row_num,
                    "tipo_palletizado": tipo_palletizado,
                    "motivos": motivos
                })
                continue

            # Buscar por ID
            id_registro = row_data.get("id")
            if id_registro:
                cursor.execute("SELECT id FROM detalle_precio_palletizados WHERE id = %s", (id_registro,))
                existente = cursor.fetchone()

                if proceso == "cargaCompleta" and existente:
                    cursor.execute("""
                        UPDATE detalle_precio_palletizados SET
                            tipo_palletizado = %s, tarima_nacional = %s, tarima_exportacion = %s,
                            liston_nacional = %s, liston_exportacion = %s, tabla_tarima = %s,
                            stretch_film = %s, sellos = %s, zunchos = %s, fundas = %s,
                            cordel_y_clavos = %s, maquila = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (
                        tipo_palletizado,
                        valores_numericos.get("tarima_nacional", 0),
                        valores_numericos.get("tarima_exportacion", 0),
                        valores_numericos.get("liston_nacional", 0),
                        valores_numericos.get("liston_exportacion", 0),
                        valores_numericos.get("tabla_tarima", 0),
                        valores_numericos.get("stretch_film", 0),
                        valores_numericos.get("sellos", 0),
                        valores_numericos.get("zunchos", 0),
                        valores_numericos.get("fundas", 0),
                        valores_numericos.get("cordel_y_clavos", 0),
                        valores_numericos.get("maquila", 0),
                        id_registro
                    ))
                    paletizados_actualizados.append({"id": id_registro, "linea": row_num})
                elif existente:
                    paletizados_actualizados.append({"id": id_registro, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(paletizados_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "paletizados_actualizados": paletizados_actualizados,
            "paletizados_error": paletizados_error,
            "totales": {
                "actualizados": len(paletizados_actualizados),
                "errores": len(paletizados_error)
            }
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()

@router.get("/paletizados/descargar-excel")
async def descargar_excel_paletizados():
    """
    Descarga paletizados en Excel.

    Basado en: MantenedorController@descargar_excel_paletizados (líneas 1781-1826)
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT id, tipo_palletizado, tarima_nacional, tarima_exportacion,
                   liston_nacional, liston_exportacion, tabla_tarima,
                   stretch_film, sellos, zunchos, fundas, cordel_y_clavos, maquila
            FROM detalle_precio_palletizados
            ORDER BY id
        """)
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Paletizados"

        # Headers según Laravel
        ws.append([
            "ID", "Tipo Palletizado", "Tarima Nacional", "Tarima Exportacion",
            "Liston Nacional", "Liston Exportacion", "Tabla Tarima",
            "Stretch Film", "Sellos", "Zunchos", "Fundas", "Cordel y Clavos", "Maquila"
        ])

        for reg in registros:
            ws.append([
                reg.get("id"),
                reg.get("tipo_palletizado"),
                reg.get("tarima_nacional"),
                reg.get("tarima_exportacion"),
                reg.get("liston_nacional"),
                reg.get("liston_exportacion"),
                reg.get("tabla_tarima"),
                reg.get("stretch_film"),
                reg.get("sellos"),
                reg.get("zunchos"),
                reg.get("fundas"),
                reg.get("cordel_y_clavos"),
                reg.get("maquila")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Paletizados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 8. INSUMOS PALETIZADOS
# Basado en: MantenedorController líneas 1842-1951
# =============================================

@router.post("/insumos-paletizados/import")
async def import_insumos_paletizados(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa insumos paletizados desde Excel.

    Basado en: MantenedorController@importInsumosPaletizados (líneas 1842-1951)

    Columnas:
    - id (para actualizar)
    - insumo (string)
    - precio (numérico, requerido)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        insumos_actualizados = []
        insumos_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            # Validar precio (numérico requerido)
            precio = row_data.get("precio")
            if precio is None or precio == "":
                motivos.append("precio vacío")
            else:
                try:
                    precio = float(precio)
                except (ValueError, TypeError):
                    motivos.append("precio no es numérico")

            if motivos:
                insumos_error.append({
                    "linea": row_num,
                    "motivos": motivos
                })
                continue

            # Buscar por ID
            id_registro = row_data.get("id")
            if id_registro:
                cursor.execute("SELECT id FROM insumos_palletizados WHERE id = %s", (id_registro,))
                existente = cursor.fetchone()

                if proceso == "cargaCompleta" and existente:
                    cursor.execute("""
                        UPDATE insumos_palletizados SET
                            precio = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (precio, id_registro))
                    insumos_actualizados.append({"id": id_registro, "linea": row_num})
                elif existente:
                    insumos_actualizados.append({"id": id_registro, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(insumos_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "insumos_actualizados": insumos_actualizados,
            "insumos_error": insumos_error,
            "totales": {
                "actualizados": len(insumos_actualizados),
                "errores": len(insumos_error)
            }
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/insumos-paletizados/descargar-excel")
async def descargar_excel_insumos_paletizados():
    """
    Descarga insumos paletizados en Excel.

    Basado en: MantenedorController@descargar_excel_insumos_paletizados (líneas 1952-1978)
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT id, insumo, precio FROM insumos_palletizados ORDER BY id")
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Insumos_Paletizados"

        ws.append(["ID", "Insumo", "Precio"])

        for reg in registros:
            ws.append([reg.get("id"), reg.get("insumo"), reg.get("precio")])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Insumos_Paletizados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 9. TARIFARIOS
# Basado en: MantenedorController líneas 1992-2137
# =============================================

@router.post("/tarifarios/import")
async def import_tarifarios(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa tarifarios desde Excel.

    Basado en: MantenedorController@importTarifarios (líneas 1992-2137)

    Columnas:
    - codigo (string, clave para buscar)
    - active (0 o 1)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        tarifarios_actualizados = []
        tarifarios_inactivados = []
        tarifarios_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            codigo = str(row_data.get("codigo", "")).strip()
            if not codigo:
                motivos.append("codigo vacío")

            if motivos:
                tarifarios_error.append({"linea": row_num, "codigo": codigo, "motivos": motivos})
                continue

            cursor.execute("SELECT id, active FROM tarifarios WHERE codigo = %s", (codigo,))
            existente = cursor.fetchone()

            if existente:
                active_nuevo = int(row_data.get("active") or 1)
                active_original = existente.get("active", 1)

                if proceso == "cargaCompleta":
                    cursor.execute("""
                        UPDATE tarifarios SET active = %s, orden = %s, updated_at = NOW()
                        WHERE codigo = %s
                    """, (active_nuevo, row_num, codigo))

                    if active_original == 1 and active_nuevo == 0:
                        tarifarios_inactivados.append({"codigo": codigo, "linea": row_num})
                    else:
                        tarifarios_actualizados.append({"codigo": codigo, "linea": row_num})
                else:
                    if active_original == 1 and active_nuevo == 0:
                        tarifarios_inactivados.append({"codigo": codigo, "linea": row_num})
                    else:
                        tarifarios_actualizados.append({"codigo": codigo, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(tarifarios_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "tarifarios_actualizados": tarifarios_actualizados,
            "tarifarios_inactivados": tarifarios_inactivados,
            "tarifarios_error": tarifarios_error,
            "totales": {
                "actualizados": len(tarifarios_actualizados),
                "inactivados": len(tarifarios_inactivados),
                "errores": len(tarifarios_error)
            }
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/tarifarios/descargar-excel")
async def descargar_excel_tarifarios():
    """Descarga tarifarios en Excel."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT id, codigo, descripcion, active FROM tarifarios ORDER BY orden, id")
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Tarifarios"

        ws.append(["ID", "Codigo", "Descripcion", "Active"])

        for reg in registros:
            ws.append([reg.get("id"), reg.get("codigo"), reg.get("descripcion"), reg.get("active")])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Tarifarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 10. CONSUMO ADHESIVO
# Basado en: MantenedorController líneas 2178-2314
# =============================================

ONDAS_CONSUMO_ADHESIVO = ["C", "B", "E"]

@router.post("/consumo-adhesivo/import")
async def import_consumo_adhesivo(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa consumo adhesivo desde Excel.

    Basado en: MantenedorController@importConsumoAdhesivo (líneas 2178-2314)

    Columnas:
    - id (para actualizar)
    - onda (C, B, E - requerido)
    - adhesivo_corrugado (numérico, requerido)
    - adhesivo_powerply (numérico, requerido)
    - planta (nombre de planta - se valida contra tabla plantas)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Cargar plantas para validación FK
        cursor.execute("SELECT id, nombre FROM plantas")
        plantas = {row["nombre"]: row["id"] for row in cursor.fetchall()}

        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        consumos_actualizados = []
        consumos_inactivados = []
        consumos_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            # Validar onda
            onda = str(row_data.get("onda", "")).strip().upper()
            if onda not in ONDAS_CONSUMO_ADHESIVO:
                motivos.append(f"onda inválida (debe ser C, B o E)")

            # Validar adhesivo_corrugado
            adhesivo_corrugado = row_data.get("adhesivo_corrugado")
            if adhesivo_corrugado is None or adhesivo_corrugado == "":
                motivos.append("adhesivo_corrugado vacío")
            else:
                try:
                    adhesivo_corrugado = int(float(adhesivo_corrugado))
                except (ValueError, TypeError):
                    motivos.append("adhesivo_corrugado no es numérico")

            # Validar adhesivo_powerply
            adhesivo_powerply = row_data.get("adhesivo_powerply")
            if adhesivo_powerply is None or adhesivo_powerply == "":
                motivos.append("adhesivo_powerply vacío")
            else:
                try:
                    adhesivo_powerply = int(float(adhesivo_powerply))
                except (ValueError, TypeError):
                    motivos.append("adhesivo_powerply no es numérico")

            # Validar planta (FK)
            planta_nombre = str(row_data.get("planta", "")).strip()
            planta_id = plantas.get(planta_nombre)
            if not planta_id:
                motivos.append(f"planta '{planta_nombre}' no existe")

            if motivos:
                consumos_error.append({"linea": row_num, "motivos": motivos})
                continue

            # Buscar por ID
            id_registro = row_data.get("id")
            if id_registro:
                cursor.execute("SELECT id, active FROM consumo_adhesivos WHERE id = %s", (id_registro,))
                existente = cursor.fetchone()

                if existente:
                    active_nuevo = int(row_data.get("active") or 1)
                    active_original = existente.get("active", 1)

                    if proceso == "cargaCompleta":
                        cursor.execute("""
                            UPDATE consumo_adhesivos SET
                                onda = %s, adhesivo_corrugado = %s, adhesivo_powerply = %s,
                                planta_id = %s, active = %s, updated_at = NOW()
                            WHERE id = %s
                        """, (onda, adhesivo_corrugado, adhesivo_powerply, planta_id, active_nuevo, id_registro))

                    if active_original == 1 and active_nuevo == 0:
                        consumos_inactivados.append({"id": id_registro, "linea": row_num})
                    else:
                        consumos_actualizados.append({"id": id_registro, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(consumos_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "consumo_adhesivos_actualizados": consumos_actualizados,
            "consumo_adhesivos_inactivados": consumos_inactivados,
            "consumo_adhesivos_error": consumos_error,
            "totales": {
                "actualizados": len(consumos_actualizados),
                "inactivados": len(consumos_inactivados),
                "errores": len(consumos_error)
            }
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/consumo-adhesivo/descargar-excel")
async def descargar_excel_consumo_adhesivo():
    """Descarga consumo adhesivo en Excel con nombre de planta."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT ca.id, p.nombre as planta, ca.onda, ca.adhesivo_corrugado, ca.adhesivo_powerply
            FROM consumo_adhesivos ca
            LEFT JOIN plantas p ON ca.planta_id = p.id
            ORDER BY ca.id
        """)
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Consumo_Adhesivo"

        ws.append(["ID", "Planta", "Onda", "Adhesivo Corrugado", "Adhesivo Powerply"])

        for reg in registros:
            ws.append([
                reg.get("id"), reg.get("planta"), reg.get("onda"),
                reg.get("adhesivo_corrugado"), reg.get("adhesivo_powerply")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Consumo_Adhesivo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 11. CONSUMO ADHESIVO PEGADO
# Basado en: MantenedorController líneas 2359-2516
# =============================================

@router.post("/consumo-adhesivo-pegado/import")
async def import_consumo_adhesivo_pegado(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa consumo adhesivo pegado desde Excel.

    Basado en: MantenedorController@importConsumoAdhesivoPegado (líneas 2359-2516)

    Columnas:
    - id (para actualizar)
    - valor (numérico, requerido)
    - planta_id (FK a plantas)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Cargar plantas para validación FK
        cursor.execute("SELECT id FROM plantas")
        plantas_ids = {row["id"] for row in cursor.fetchall()}

        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        consumos_actualizados = []
        consumos_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            # Validar valor (numérico)
            valor = row_data.get("valor")
            if valor is None or valor == "":
                motivos.append("valor vacío")
            else:
                try:
                    valor = float(valor)
                except (ValueError, TypeError):
                    motivos.append("valor no es numérico")

            # Validar planta_id (FK)
            planta_id = row_data.get("planta_id")
            if planta_id:
                try:
                    planta_id = int(planta_id)
                    if planta_id not in plantas_ids:
                        motivos.append(f"planta_id {planta_id} no existe")
                except (ValueError, TypeError):
                    motivos.append("planta_id no es numérico")

            if motivos:
                consumos_error.append({"linea": row_num, "motivos": motivos})
                continue

            # Buscar por ID
            id_registro = row_data.get("id")
            if id_registro:
                cursor.execute("SELECT id FROM consumo_adhesivos_pegados WHERE id = %s", (id_registro,))
                existente = cursor.fetchone()

                if proceso == "cargaCompleta" and existente:
                    cursor.execute("""
                        UPDATE consumo_adhesivos_pegados SET
                            valor = %s, planta_id = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (valor, planta_id, id_registro))
                    consumos_actualizados.append({"id": id_registro, "linea": row_num})
                elif existente:
                    consumos_actualizados.append({"id": id_registro, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(consumos_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "consumo_adhesivos_pegados_actualizados": consumos_actualizados,
            "consumo_adhesivos_pegados_error": consumos_error,
            "totales": {
                "actualizados": len(consumos_actualizados),
                "errores": len(consumos_error)
            }
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/consumo-adhesivo-pegado/descargar-excel")
async def descargar_excel_consumo_adhesivo_pegado():
    """Descarga consumo adhesivo pegado en Excel."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT cap.id, cap.process_id, cap.valor, cap.planta_id
            FROM consumo_adhesivos_pegados cap
            ORDER BY cap.id
        """)
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Consumo_Adhesivo_Pegado"

        ws.append(["ID", "Process ID", "Valor", "Planta ID"])

        for reg in registros:
            ws.append([reg.get("id"), reg.get("process_id"), reg.get("valor"), reg.get("planta_id")])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Consumo_Adhesivo_Pegado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 12. CONSUMO ENERGIA
# Basado en: MantenedorController líneas 2559-2686
# =============================================

@router.post("/consumo-energia/import")
async def import_consumo_energia(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa consumo energía desde Excel.

    Basado en: MantenedorController@importConsumoEnergia (líneas 2559-2686)

    Columnas:
    - id (para actualizar)
    - descripcion (string)
    - valor (numérico, requerido)
    - planta_id (FK a plantas)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Cargar plantas para validación FK
        cursor.execute("SELECT id FROM plantas")
        plantas_ids = {row["id"] for row in cursor.fetchall()}

        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        consumos_actualizados = []
        consumos_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            # Validar valor (numérico)
            valor = row_data.get("valor")
            if valor is None or valor == "":
                motivos.append("valor vacío")
            else:
                try:
                    valor = float(valor)
                except (ValueError, TypeError):
                    motivos.append("valor no es numérico")

            # Validar planta_id (FK)
            planta_id = row_data.get("planta_id")
            if planta_id:
                try:
                    planta_id = int(planta_id)
                    if planta_id not in plantas_ids:
                        motivos.append(f"planta_id {planta_id} no existe")
                except (ValueError, TypeError):
                    motivos.append("planta_id no es numérico")

            if motivos:
                consumos_error.append({"linea": row_num, "motivos": motivos})
                continue

            # Buscar por ID
            id_registro = row_data.get("id")
            if id_registro:
                cursor.execute("SELECT id FROM consumo_energias WHERE id = %s", (id_registro,))
                existente = cursor.fetchone()

                if proceso == "cargaCompleta" and existente:
                    cursor.execute("""
                        UPDATE consumo_energias SET
                            valor = %s, planta_id = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (valor, planta_id, id_registro))
                    consumos_actualizados.append({"id": id_registro, "linea": row_num})
                elif existente:
                    consumos_actualizados.append({"id": id_registro, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(consumos_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "consumo_energias_actualizados": consumos_actualizados,
            "consumo_energias_error": consumos_error,
            "totales": {
                "actualizados": len(consumos_actualizados),
                "errores": len(consumos_error)
            }
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/consumo-energia/descargar-excel")
async def descargar_excel_consumo_energia():
    """Descarga consumo energía en Excel."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT ce.id, ce.descripcion, ce.valor, ce.unidad, ce.planta_id
            FROM consumo_energias ce
            ORDER BY ce.id
        """)
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Consumo_Energia"

        ws.append(["ID", "Descripcion", "Valor", "Unidad", "Planta ID"])

        for reg in registros:
            ws.append([
                reg.get("id"), reg.get("descripcion"), reg.get("valor"),
                reg.get("unidad"), reg.get("planta_id")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Consumo_Energia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 13. MATRICES
# Basado en: MantenedorController líneas 2722-2894
# =============================================

@router.post("/matrices/import")
async def import_matrices(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa matrices desde Excel.

    Basado en: MantenedorController@importMatrices (líneas 2722-2894)

    Columnas:
    - codigo (string, requerido)
    - descripcion (string)
    - ancho, largo (numéricos)
    - proceso_id, planta_id (FK)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Cargar FKs para validación
        cursor.execute("SELECT id FROM plantas")
        plantas_ids = {row["id"] for row in cursor.fetchall()}

        cursor.execute("SELECT id FROM processes")
        procesos_ids = {row["id"] for row in cursor.fetchall()}

        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        matrices_nuevas = []
        matrices_actualizadas = []
        matrices_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            codigo = str(row_data.get("codigo", "")).strip()
            if not codigo:
                motivos.append("codigo vacío")

            # Validar numéricos
            ancho = row_data.get("ancho")
            largo = row_data.get("largo")
            if ancho is not None and ancho != "":
                try:
                    ancho = float(ancho)
                except (ValueError, TypeError):
                    motivos.append("ancho no es numérico")
            else:
                ancho = None

            if largo is not None and largo != "":
                try:
                    largo = float(largo)
                except (ValueError, TypeError):
                    motivos.append("largo no es numérico")
            else:
                largo = None

            # Validar FKs
            proceso_id = row_data.get("proceso_id")
            planta_id = row_data.get("planta_id")

            if proceso_id:
                try:
                    proceso_id = int(proceso_id)
                    if proceso_id not in procesos_ids:
                        motivos.append(f"proceso_id {proceso_id} no existe")
                except (ValueError, TypeError):
                    motivos.append("proceso_id no es numérico")

            if planta_id:
                try:
                    planta_id = int(planta_id)
                    if planta_id not in plantas_ids:
                        motivos.append(f"planta_id {planta_id} no existe")
                except (ValueError, TypeError):
                    motivos.append("planta_id no es numérico")

            if motivos:
                matrices_error.append({"linea": row_num, "codigo": codigo, "motivos": motivos})
                continue

            cursor.execute("SELECT id FROM matrices WHERE codigo = %s", (codigo,))
            existente = cursor.fetchone()

            if proceso == "cargaCompleta":
                if existente:
                    cursor.execute("""
                        UPDATE matrices SET
                            descripcion = %s, ancho = %s, largo = %s,
                            proceso_id = %s, planta_id = %s, updated_at = NOW()
                        WHERE codigo = %s
                    """, (
                        row_data.get("descripcion"),
                        ancho, largo, proceso_id, planta_id, codigo
                    ))
                    matrices_actualizadas.append({"codigo": codigo, "linea": row_num})
                else:
                    cursor.execute("""
                        INSERT INTO matrices (codigo, descripcion, ancho, largo, proceso_id, planta_id, active, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, %s, %s, 1, NOW(), NOW())
                    """, (codigo, row_data.get("descripcion"), ancho, largo, proceso_id, planta_id))
                    matrices_nuevas.append({"codigo": codigo, "linea": row_num, "id": cursor.lastrowid})
            else:
                if existente:
                    matrices_actualizadas.append({"codigo": codigo, "linea": row_num})
                else:
                    matrices_nuevas.append({"codigo": codigo, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(matrices_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "matrices_nuevas": matrices_nuevas,
            "matrices_actualizadas": matrices_actualizadas,
            "matrices_error": matrices_error,
            "totales": {
                "nuevas": len(matrices_nuevas),
                "actualizadas": len(matrices_actualizadas),
                "errores": len(matrices_error)
            }
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/matrices/descargar-excel")
async def descargar_excel_matrices():
    """Descarga matrices en Excel."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT m.id, m.codigo, m.descripcion, m.ancho, m.largo, m.proceso_id, m.planta_id, m.active
            FROM matrices m
            ORDER BY m.id
        """)
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Matrices"

        ws.append(["ID", "Codigo", "Descripcion", "Ancho", "Largo", "Proceso ID", "Planta ID", "Active"])

        for reg in registros:
            ws.append([
                reg.get("id"), reg.get("codigo"), reg.get("descripcion"),
                reg.get("ancho"), reg.get("largo"), reg.get("proceso_id"),
                reg.get("planta_id"), reg.get("active")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Matrices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 14. MATERIALES
# Basado en: MantenedorController líneas 2985-3262
# =============================================

@router.post("/materiales/import")
async def import_materiales(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa materiales desde Excel.

    Basado en: MantenedorController@importMateriales (líneas 2985-3262)

    Columnas:
    - codigo (string, requerido)
    - descripcion (string)
    - client_id, carton_id, cad_id (FK)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Cargar FKs para validación
        cursor.execute("SELECT id FROM clients")
        clientes_ids = {row["id"] for row in cursor.fetchall()}

        cursor.execute("SELECT id FROM cardboards")
        cartones_ids = {row["id"] for row in cursor.fetchall()}

        cursor.execute("SELECT id FROM cads")
        cads_ids = {row["id"] for row in cursor.fetchall()}

        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        materiales_nuevos = []
        materiales_actualizados = []
        materiales_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            codigo = str(row_data.get("codigo", "")).strip()
            if not codigo:
                motivos.append("codigo vacío")

            # Validar FKs
            client_id = row_data.get("client_id")
            carton_id = row_data.get("carton_id")
            cad_id = row_data.get("cad_id")

            if client_id:
                try:
                    client_id = int(client_id)
                    if client_id not in clientes_ids:
                        motivos.append(f"client_id {client_id} no existe")
                except (ValueError, TypeError):
                    motivos.append("client_id no es numérico")

            if carton_id:
                try:
                    carton_id = int(carton_id)
                    if carton_id not in cartones_ids:
                        motivos.append(f"carton_id {carton_id} no existe")
                except (ValueError, TypeError):
                    motivos.append("carton_id no es numérico")

            if cad_id:
                try:
                    cad_id = int(cad_id)
                    if cad_id not in cads_ids:
                        motivos.append(f"cad_id {cad_id} no existe")
                except (ValueError, TypeError):
                    motivos.append("cad_id no es numérico")

            if motivos:
                materiales_error.append({"linea": row_num, "codigo": codigo, "motivos": motivos})
                continue

            cursor.execute("SELECT id FROM materials WHERE codigo = %s", (codigo,))
            existente = cursor.fetchone()

            if proceso == "cargaCompleta":
                if existente:
                    cursor.execute("""
                        UPDATE materials SET
                            descripcion = %s, client_id = %s, carton_id = %s, cad_id = %s, updated_at = NOW()
                        WHERE codigo = %s
                    """, (row_data.get("descripcion"), client_id, carton_id, cad_id, codigo))
                    materiales_actualizados.append({"codigo": codigo, "linea": row_num})
                else:
                    cursor.execute("""
                        INSERT INTO materials (codigo, descripcion, client_id, carton_id, cad_id, active, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, %s, 1, NOW(), NOW())
                    """, (codigo, row_data.get("descripcion"), client_id, carton_id, cad_id))
                    materiales_nuevos.append({"codigo": codigo, "linea": row_num, "id": cursor.lastrowid})
            else:
                if existente:
                    materiales_actualizados.append({"codigo": codigo, "linea": row_num})
                else:
                    materiales_nuevos.append({"codigo": codigo, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(materiales_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "materiales_nuevos": materiales_nuevos,
            "materiales_actualizados": materiales_actualizados,
            "materiales_error": materiales_error,
            "totales": {
                "nuevos": len(materiales_nuevos),
                "actualizados": len(materiales_actualizados),
                "errores": len(materiales_error)
            }
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/materiales/descargar-excel")
async def descargar_excel_materiales():
    """Descarga materiales en Excel."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT m.id, m.codigo, m.descripcion, m.client_id, m.carton_id, m.cad_id, m.active
            FROM materials m
            ORDER BY m.id
        """)
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Materiales"

        ws.append(["ID", "Codigo", "Descripcion", "Client ID", "Carton ID", "CAD ID", "Active"])

        for reg in registros:
            ws.append([
                reg.get("id"), reg.get("codigo"), reg.get("descripcion"),
                reg.get("client_id"), reg.get("carton_id"), reg.get("cad_id"), reg.get("active")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Materiales_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 15. FACTORES DESARROLLO
# Basado en: MantenedorController líneas 3534-3673
# =============================================

@router.post("/factores-desarrollo/import")
async def import_factores_desarrollo(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa factores desarrollo desde Excel.

    Columnas:
    - id (para actualizar)
    - descripcion (string)
    - factor (numérico, requerido)
    - tipo (string)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        factores_actualizados = []
        factores_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            # Validar factor (numérico)
            factor = row_data.get("factor")
            if factor is None or factor == "":
                motivos.append("factor vacío")
            else:
                try:
                    factor = float(factor)
                except (ValueError, TypeError):
                    motivos.append("factor no es numérico")

            if motivos:
                factores_error.append({"linea": row_num, "motivos": motivos})
                continue

            id_registro = row_data.get("id")
            if id_registro:
                cursor.execute("SELECT id FROM factores_desarrollos WHERE id = %s", (id_registro,))
                existente = cursor.fetchone()

                if proceso == "cargaCompleta" and existente:
                    cursor.execute("""
                        UPDATE factores_desarrollos SET
                            descripcion = %s, factor = %s, tipo = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (row_data.get("descripcion"), factor, row_data.get("tipo"), id_registro))
                    factores_actualizados.append({"id": id_registro, "linea": row_num})
                elif existente:
                    factores_actualizados.append({"id": id_registro, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(factores_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "factores_actualizados": factores_actualizados,
            "factores_error": factores_error,
            "totales": {"actualizados": len(factores_actualizados), "errores": len(factores_error)}
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/factores-desarrollo/descargar-excel")
async def descargar_excel_factores_desarrollo():
    """Descarga factores desarrollo en Excel."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT id, descripcion, factor, tipo, active FROM factores_desarrollos ORDER BY id")
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Factores_Desarrollo"

        ws.append(["ID", "Descripcion", "Factor", "Tipo", "Active"])

        for reg in registros:
            ws.append([reg.get("id"), reg.get("descripcion"), reg.get("factor"), reg.get("tipo"), reg.get("active")])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Factores_Desarrollo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 16. FACTORES SEGURIDAD
# Basado en: MantenedorController líneas 3725-3846
# =============================================

@router.post("/factores-seguridad/import")
async def import_factores_seguridad(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa factores seguridad desde Excel.

    Columnas:
    - id (para actualizar)
    - descripcion (string)
    - factor (numérico, requerido)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        factores_actualizados = []
        factores_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            factor = row_data.get("factor")
            if factor is None or factor == "":
                motivos.append("factor vacío")
            else:
                try:
                    factor = float(factor)
                except (ValueError, TypeError):
                    motivos.append("factor no es numérico")

            if motivos:
                factores_error.append({"linea": row_num, "motivos": motivos})
                continue

            id_registro = row_data.get("id")
            if id_registro:
                cursor.execute("SELECT id FROM factores_seguridads WHERE id = %s", (id_registro,))
                existente = cursor.fetchone()

                if proceso == "cargaCompleta" and existente:
                    cursor.execute("""
                        UPDATE factores_seguridads SET descripcion = %s, factor = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (row_data.get("descripcion"), factor, id_registro))
                    factores_actualizados.append({"id": id_registro, "linea": row_num})
                elif existente:
                    factores_actualizados.append({"id": id_registro, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(factores_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "factores_actualizados": factores_actualizados,
            "factores_error": factores_error,
            "totales": {"actualizados": len(factores_actualizados), "errores": len(factores_error)}
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/factores-seguridad/descargar-excel")
async def descargar_excel_factores_seguridad():
    """Descarga factores seguridad en Excel."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT id, descripcion, factor, active FROM factores_seguridads ORDER BY id")
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Factores_Seguridad"

        ws.append(["ID", "Descripcion", "Factor", "Active"])

        for reg in registros:
            ws.append([reg.get("id"), reg.get("descripcion"), reg.get("factor"), reg.get("active")])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Factores_Seguridad_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 17. FACTORES ONDA
# Basado en: MantenedorController líneas 3888-4000
# =============================================

@router.post("/factores-onda/import")
async def import_factores_onda(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa factores onda desde Excel.

    Columnas:
    - codigo (string, requerido)
    - descripcion (string)
    - factor_onda (numérico, requerido)
    - altura (numérico)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        ondas_nuevas = []
        ondas_actualizadas = []
        ondas_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            codigo = str(row_data.get("codigo", "")).strip()
            if not codigo:
                motivos.append("codigo vacío")

            factor_onda = row_data.get("factor_onda")
            if factor_onda is None or factor_onda == "":
                motivos.append("factor_onda vacío")
            else:
                try:
                    factor_onda = float(factor_onda)
                except (ValueError, TypeError):
                    motivos.append("factor_onda no es numérico")

            altura = row_data.get("altura")
            if altura is not None and altura != "":
                try:
                    altura = float(altura)
                except (ValueError, TypeError):
                    motivos.append("altura no es numérico")
            else:
                altura = None

            if motivos:
                ondas_error.append({"linea": row_num, "codigo": codigo, "motivos": motivos})
                continue

            cursor.execute("SELECT id FROM tipo_ondas WHERE codigo = %s", (codigo,))
            existente = cursor.fetchone()

            if proceso == "cargaCompleta":
                if existente:
                    cursor.execute("""
                        UPDATE tipo_ondas SET descripcion = %s, factor_onda = %s, altura = %s, updated_at = NOW()
                        WHERE codigo = %s
                    """, (row_data.get("descripcion"), factor_onda, altura, codigo))
                    ondas_actualizadas.append({"codigo": codigo, "linea": row_num})
                else:
                    cursor.execute("""
                        INSERT INTO tipo_ondas (codigo, descripcion, factor_onda, altura, active, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, 1, NOW(), NOW())
                    """, (codigo, row_data.get("descripcion"), factor_onda, altura))
                    ondas_nuevas.append({"codigo": codigo, "linea": row_num, "id": cursor.lastrowid})
            else:
                if existente:
                    ondas_actualizadas.append({"codigo": codigo, "linea": row_num})
                else:
                    ondas_nuevas.append({"codigo": codigo, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(ondas_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "ondas_nuevas": ondas_nuevas,
            "ondas_actualizadas": ondas_actualizadas,
            "ondas_error": ondas_error,
            "totales": {"nuevas": len(ondas_nuevas), "actualizadas": len(ondas_actualizadas), "errores": len(ondas_error)}
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/factores-onda/descargar-excel")
async def descargar_excel_factores_onda():
    """Descarga factores onda en Excel."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT id, codigo, descripcion, factor_onda, altura, active FROM tipo_ondas ORDER BY id")
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Factores_Onda"

        ws.append(["ID", "Codigo", "Descripcion", "Factor Onda", "Altura", "Active"])

        for reg in registros:
            ws.append([
                reg.get("id"), reg.get("codigo"), reg.get("descripcion"),
                reg.get("factor_onda"), reg.get("altura"), reg.get("active")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Factores_Onda_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 18. MAQUILAS
# Basado en: MantenedorController líneas 4044-4168
# =============================================

@router.post("/maquilas/import")
async def import_maquilas(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa maquilas desde Excel.

    Columnas:
    - descripcion (string, requerido)
    - precio (numérico, requerido)
    - tipo (string)
    - planta_id (FK)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Cargar plantas para validación FK
        cursor.execute("SELECT id FROM plantas")
        plantas_ids = {row["id"] for row in cursor.fetchall()}

        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        maquilas_actualizadas = []
        maquilas_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            descripcion = str(row_data.get("descripcion", "")).strip()
            if not descripcion:
                motivos.append("descripcion vacía")

            precio = row_data.get("precio")
            if precio is None or precio == "":
                motivos.append("precio vacío")
            else:
                try:
                    precio = float(precio)
                except (ValueError, TypeError):
                    motivos.append("precio no es numérico")

            planta_id = row_data.get("planta_id")
            if planta_id:
                try:
                    planta_id = int(planta_id)
                    if planta_id not in plantas_ids:
                        motivos.append(f"planta_id {planta_id} no existe")
                except (ValueError, TypeError):
                    motivos.append("planta_id no es numérico")

            if motivos:
                maquilas_error.append({"linea": row_num, "descripcion": descripcion, "motivos": motivos})
                continue

            id_registro = row_data.get("id")
            if id_registro:
                cursor.execute("SELECT id FROM maquila_servicios WHERE id = %s", (id_registro,))
                existente = cursor.fetchone()

                if proceso == "cargaCompleta" and existente:
                    cursor.execute("""
                        UPDATE maquila_servicios SET
                            descripcion = %s, precio = %s, tipo = %s, planta_id = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (descripcion, precio, row_data.get("tipo"), planta_id, id_registro))
                    maquilas_actualizadas.append({"id": id_registro, "linea": row_num})
                elif existente:
                    maquilas_actualizadas.append({"id": id_registro, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(maquilas_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "maquilas_actualizadas": maquilas_actualizadas,
            "maquilas_error": maquilas_error,
            "totales": {"actualizadas": len(maquilas_actualizadas), "errores": len(maquilas_error)}
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/maquilas/descargar-excel")
async def descargar_excel_maquilas():
    """Descarga maquilas en Excel."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT id, descripcion, precio, tipo, planta_id, active FROM maquila_servicios ORDER BY id")
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Maquilas"

        ws.append(["ID", "Descripcion", "Precio", "Tipo", "Planta ID", "Active"])

        for reg in registros:
            ws.append([
                reg.get("id"), reg.get("descripcion"), reg.get("precio"),
                reg.get("tipo"), reg.get("planta_id"), reg.get("active")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Maquilas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 19. ONDAS
# Basado en: MantenedorController líneas 4220-4331
# Nota: Usa la misma tabla tipo_ondas que Factores Onda
# =============================================

@router.post("/ondas/import")
async def import_ondas(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """Importa ondas (alias de factores onda)."""
    return await import_factores_onda(archivo, proceso)


@router.get("/ondas/descargar-excel")
async def descargar_excel_ondas():
    """Descarga ondas (alias de factores onda)."""
    return await descargar_excel_factores_onda()


# =============================================
# 20. PLANTAS
# Basado en: MantenedorController líneas 4375-4627
# =============================================

@router.post("/plantas/import")
async def import_plantas(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa plantas desde Excel.

    Columnas:
    - nombre (string, requerido)
    - codigo (string)
    - direccion (string)
    - telefono (string)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        plantas_nuevas = []
        plantas_actualizadas = []
        plantas_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            nombre = str(row_data.get("nombre", "")).strip()
            if not nombre:
                motivos.append("nombre vacío")

            if motivos:
                plantas_error.append({"linea": row_num, "nombre": nombre, "motivos": motivos})
                continue

            cursor.execute("SELECT id FROM plantas WHERE nombre = %s", (nombre,))
            existente = cursor.fetchone()

            if proceso == "cargaCompleta":
                if existente:
                    cursor.execute("""
                        UPDATE plantas SET
                            codigo = %s, direccion = %s, telefono = %s, updated_at = NOW()
                        WHERE nombre = %s
                    """, (row_data.get("codigo"), row_data.get("direccion"), row_data.get("telefono"), nombre))
                    plantas_actualizadas.append({"nombre": nombre, "linea": row_num})
                else:
                    cursor.execute("""
                        INSERT INTO plantas (nombre, codigo, direccion, telefono, active, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, 1, NOW(), NOW())
                    """, (nombre, row_data.get("codigo"), row_data.get("direccion"), row_data.get("telefono")))
                    plantas_nuevas.append({"nombre": nombre, "linea": row_num, "id": cursor.lastrowid})
            else:
                if existente:
                    plantas_actualizadas.append({"nombre": nombre, "linea": row_num})
                else:
                    plantas_nuevas.append({"nombre": nombre, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(plantas_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "plantas_nuevas": plantas_nuevas,
            "plantas_actualizadas": plantas_actualizadas,
            "plantas_error": plantas_error,
            "totales": {"nuevas": len(plantas_nuevas), "actualizadas": len(plantas_actualizadas), "errores": len(plantas_error)}
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/plantas/descargar-excel")
async def descargar_excel_plantas():
    """Descarga plantas en Excel."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT id, nombre, codigo, direccion, telefono, active FROM plantas ORDER BY id")
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Plantas"

        ws.append(["ID", "Nombre", "Codigo", "Direccion", "Telefono", "Active"])

        for reg in registros:
            ws.append([
                reg.get("id"), reg.get("nombre"), reg.get("codigo"),
                reg.get("direccion"), reg.get("telefono"), reg.get("active")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Plantas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 21. VARIABLES
# Basado en: MantenedorController líneas 4762-4993
# =============================================

@router.post("/variables/import")
async def import_variables(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa variables cotizador desde Excel.

    Columnas:
    - nombre (string, requerido)
    - valor (numérico, requerido)
    - descripcion (string)
    - tipo (string)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        variables_actualizadas = []
        variables_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            nombre = str(row_data.get("nombre", "")).strip()
            if not nombre:
                motivos.append("nombre vacío")

            valor = row_data.get("valor")
            if valor is None or valor == "":
                motivos.append("valor vacío")
            else:
                try:
                    valor = float(valor)
                except (ValueError, TypeError):
                    motivos.append("valor no es numérico")

            if motivos:
                variables_error.append({"linea": row_num, "nombre": nombre, "motivos": motivos})
                continue

            cursor.execute("SELECT id FROM variables_cotizadors WHERE nombre = %s", (nombre,))
            existente = cursor.fetchone()

            if proceso == "cargaCompleta" and existente:
                cursor.execute("""
                    UPDATE variables_cotizadors SET
                        valor = %s, descripcion = %s, tipo = %s, updated_at = NOW()
                    WHERE nombre = %s
                """, (valor, row_data.get("descripcion"), row_data.get("tipo"), nombre))
                variables_actualizadas.append({"nombre": nombre, "linea": row_num})
            elif existente:
                variables_actualizadas.append({"nombre": nombre, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(variables_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "variables_actualizadas": variables_actualizadas,
            "variables_error": variables_error,
            "totales": {"actualizadas": len(variables_actualizadas), "errores": len(variables_error)}
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/variables/descargar-excel")
async def descargar_excel_variables():
    """Descarga variables cotizador en Excel."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT id, nombre, valor, descripcion, tipo, active FROM variables_cotizadors ORDER BY id")
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Variables"

        ws.append(["ID", "Nombre", "Valor", "Descripcion", "Tipo", "Active"])

        for reg in registros:
            ws.append([
                reg.get("id"), reg.get("nombre"), reg.get("valor"),
                reg.get("descripcion"), reg.get("tipo"), reg.get("active")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Variables_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 22. MARGENES MINIMOS
# Basado en: MantenedorController líneas 5204-5396
# =============================================

@router.post("/margenes-minimos/import")
async def import_margenes_minimos(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa márgenes mínimos desde Excel.

    Columnas:
    - id (para actualizar)
    - descripcion (string)
    - margen (numérico, requerido)
    - rubro_id (FK)
    - planta_id (FK)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Cargar FKs
        cursor.execute("SELECT id FROM plantas")
        plantas_ids = {row["id"] for row in cursor.fetchall()}

        cursor.execute("SELECT id FROM rubros")
        rubros_ids = {row["id"] for row in cursor.fetchall()}

        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        margenes_actualizados = []
        margenes_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            margen = row_data.get("margen")
            if margen is None or margen == "":
                motivos.append("margen vacío")
            else:
                try:
                    margen = float(margen)
                except (ValueError, TypeError):
                    motivos.append("margen no es numérico")

            # Validar FKs
            planta_id = row_data.get("planta_id")
            if planta_id:
                try:
                    planta_id = int(planta_id)
                    if planta_id not in plantas_ids:
                        motivos.append(f"planta_id {planta_id} no existe")
                except (ValueError, TypeError):
                    motivos.append("planta_id no es numérico")

            rubro_id = row_data.get("rubro_id")
            if rubro_id:
                try:
                    rubro_id = int(rubro_id)
                    if rubros_ids and rubro_id not in rubros_ids:
                        motivos.append(f"rubro_id {rubro_id} no existe")
                except (ValueError, TypeError):
                    motivos.append("rubro_id no es numérico")

            if motivos:
                margenes_error.append({"linea": row_num, "motivos": motivos})
                continue

            id_registro = row_data.get("id")
            if id_registro:
                cursor.execute("SELECT id FROM margenes_minimos WHERE id = %s", (id_registro,))
                existente = cursor.fetchone()

                if proceso == "cargaCompleta" and existente:
                    cursor.execute("""
                        UPDATE margenes_minimos SET
                            descripcion = %s, margen = %s, rubro_id = %s, planta_id = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (row_data.get("descripcion"), margen, rubro_id, planta_id, id_registro))
                    margenes_actualizados.append({"id": id_registro, "linea": row_num})
                elif existente:
                    margenes_actualizados.append({"id": id_registro, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(margenes_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "margenes_actualizados": margenes_actualizados,
            "margenes_error": margenes_error,
            "totales": {"actualizados": len(margenes_actualizados), "errores": len(margenes_error)}
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/margenes-minimos/descargar-excel")
async def descargar_excel_margenes_minimos():
    """Descarga márgenes mínimos en Excel."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT id, descripcion, margen, rubro_id, planta_id, active FROM margenes_minimos ORDER BY id")
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Margenes_Minimos"

        ws.append(["ID", "Descripcion", "Margen", "Rubro ID", "Planta ID", "Active"])

        for reg in registros:
            ws.append([
                reg.get("id"), reg.get("descripcion"), reg.get("margen"),
                reg.get("rubro_id"), reg.get("planta_id"), reg.get("active")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Margenes_Minimos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 23. PORCENTAJES MARGENES
# Basado en: MantenedorController líneas 5430-5585
# =============================================

@router.post("/porcentajes-margenes/import")
async def import_porcentajes_margenes(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa porcentajes márgenes desde Excel.

    Columnas:
    - id (para actualizar)
    - descripcion (string)
    - porcentaje (numérico, requerido)
    - clasificacion_id (FK)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        porcentajes_actualizados = []
        porcentajes_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            porcentaje = row_data.get("porcentaje")
            if porcentaje is None or porcentaje == "":
                motivos.append("porcentaje vacío")
            else:
                try:
                    porcentaje = float(porcentaje)
                except (ValueError, TypeError):
                    motivos.append("porcentaje no es numérico")

            if motivos:
                porcentajes_error.append({"linea": row_num, "motivos": motivos})
                continue

            id_registro = row_data.get("id")
            if id_registro:
                cursor.execute("SELECT id FROM porcentaje_margens WHERE id = %s", (id_registro,))
                existente = cursor.fetchone()

                if proceso == "cargaCompleta" and existente:
                    cursor.execute("""
                        UPDATE porcentaje_margens SET
                            descripcion = %s, porcentaje = %s, clasificacion_id = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (row_data.get("descripcion"), porcentaje, row_data.get("clasificacion_id"), id_registro))
                    porcentajes_actualizados.append({"id": id_registro, "linea": row_num})
                elif existente:
                    porcentajes_actualizados.append({"id": id_registro, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(porcentajes_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "porcentajes_actualizados": porcentajes_actualizados,
            "porcentajes_error": porcentajes_error,
            "totales": {"actualizados": len(porcentajes_actualizados), "errores": len(porcentajes_error)}
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/porcentajes-margenes/descargar-excel")
async def descargar_excel_porcentajes_margenes():
    """Descarga porcentajes márgenes en Excel."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT id, descripcion, porcentaje, clasificacion_id, active FROM porcentaje_margens ORDER BY id")
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Porcentajes_Margenes"

        ws.append(["ID", "Descripcion", "Porcentaje", "Clasificacion ID", "Active"])

        for reg in registros:
            ws.append([
                reg.get("id"), reg.get("descripcion"), reg.get("porcentaje"),
                reg.get("clasificacion_id"), reg.get("active")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Porcentajes_Margenes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# 24. MANO OBRA MANTENCION
# Basado en: MantenedorController líneas 5586-5741
# =============================================

@router.post("/mano-obra-mantencion/import")
async def import_mano_obra_mantencion(archivo: UploadFile = File(...), proceso: str = Query("validacion")):
    """
    Importa mano obra mantención desde Excel.

    Columnas:
    - id (para actualizar)
    - descripcion (string)
    - valor (numérico, requerido)
    - unidad (string)
    - planta_id (FK)
    """
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Cargar plantas
        cursor.execute("SELECT id FROM plantas")
        plantas_ids = {row["id"] for row in cursor.fetchall()}

        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        mano_obra_actualizados = []
        mano_obra_error = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            motivos = []

            valor = row_data.get("valor")
            if valor is None or valor == "":
                motivos.append("valor vacío")
            else:
                try:
                    valor = float(valor)
                except (ValueError, TypeError):
                    motivos.append("valor no es numérico")

            planta_id = row_data.get("planta_id")
            if planta_id:
                try:
                    planta_id = int(planta_id)
                    if planta_id not in plantas_ids:
                        motivos.append(f"planta_id {planta_id} no existe")
                except (ValueError, TypeError):
                    motivos.append("planta_id no es numérico")

            if motivos:
                mano_obra_error.append({"linea": row_num, "motivos": motivos})
                continue

            id_registro = row_data.get("id")
            if id_registro:
                cursor.execute("SELECT id FROM mano_obra_mantencions WHERE id = %s", (id_registro,))
                existente = cursor.fetchone()

                if proceso == "cargaCompleta" and existente:
                    cursor.execute("""
                        UPDATE mano_obra_mantencions SET
                            descripcion = %s, valor = %s, unidad = %s, planta_id = %s, updated_at = NOW()
                        WHERE id = %s
                    """, (row_data.get("descripcion"), valor, row_data.get("unidad"), planta_id, id_registro))
                    mano_obra_actualizados.append({"id": id_registro, "linea": row_num})
                elif existente:
                    mano_obra_actualizados.append({"id": id_registro, "linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(mano_obra_error) == 0,
            "mensaje": "Archivo procesado",
            "proceso": proceso,
            "mano_obra_actualizados": mano_obra_actualizados,
            "mano_obra_error": mano_obra_error,
            "totales": {"actualizados": len(mano_obra_actualizados), "errores": len(mano_obra_error)}
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


@router.get("/mano-obra-mantencion/descargar-excel")
async def descargar_excel_mano_obra_mantencion():
    """Descarga mano obra mantención en Excel."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT id, descripcion, valor, unidad, planta_id, active FROM mano_obra_mantencions ORDER BY id")
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Mano_Obra_Mantencion"

        ws.append(["ID", "Descripcion", "Valor", "Unidad", "Planta ID", "Active"])

        for reg in registros:
            ws.append([
                reg.get("id"), reg.get("descripcion"), reg.get("valor"),
                reg.get("unidad"), reg.get("planta_id"), reg.get("active")
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Mano_Obra_Mantencion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# FUNCIONES AUXILIARES
# =============================================

async def _import_tabla_simple(archivo: UploadFile, proceso: str, tabla: str, campos: List[str]):
    """Función auxiliar para importar tablas simples."""
    filename = archivo.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(status_code=400, detail="Formato inválido")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        content = await archivo.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip().replace(" ", "_"))
            else:
                headers.append(None)

        registros_nuevos = []
        registros_actualizados = []
        registros_error = []

        campo_clave = campos[0]  # Primer campo como clave

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            row_data = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_data[header] = row[i]

            clave = row_data.get(campo_clave)
            if not clave:
                registros_error.append({"linea": row_num, "error": f"{campo_clave} vacío"})
                continue

            cursor.execute(f"SELECT id FROM {tabla} WHERE {campo_clave} = %s", (clave,))
            existente = cursor.fetchone()

            if proceso == "cargaCompleta":
                if existente:
                    set_clause = ", ".join([f"{c} = %s" for c in campos[1:]])
                    values = [row_data.get(c) for c in campos[1:]]
                    values.append(existente["id"])
                    cursor.execute(f"UPDATE {tabla} SET {set_clause}, updated_at = NOW() WHERE id = %s", values)
                    registros_actualizados.append({"id": existente["id"], "linea": row_num})
                else:
                    columns = ", ".join(campos + ["created_at", "updated_at"])
                    placeholders = ", ".join(["%s"] * len(campos) + ["NOW()", "NOW()"])
                    values = [row_data.get(c) for c in campos]
                    cursor.execute(f"INSERT INTO {tabla} ({columns}) VALUES ({placeholders})", values)
                    registros_nuevos.append({"id": cursor.lastrowid, "linea": row_num})
            else:
                if existente:
                    registros_actualizados.append({"linea": row_num})
                else:
                    registros_nuevos.append({"linea": row_num})

        if proceso == "cargaCompleta":
            conn.commit()

        return {
            "success": len(registros_error) == 0,
            "tabla": tabla,
            "proceso": proceso,
            "nuevos": registros_nuevos,
            "actualizados": registros_actualizados,
            "errores": registros_error
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


async def _descargar_tabla_simple(tabla: str, titulo: str, columnas: List[str]):
    """Función auxiliar para descargar tablas simples."""
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        columns_sql = ", ".join(columnas)
        cursor.execute(f"SELECT {columns_sql} FROM {tabla} ORDER BY id")
        registros = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = titulo

        ws.append(columnas)

        for reg in registros:
            ws.append([reg.get(col) for col in columnas])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{titulo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()
