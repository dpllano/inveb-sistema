"""
Router de Órdenes de Trabajo - INVEB Cascade Service
Lee y escribe datos en MySQL de Laravel para el dashboard React.
Actualizado: Fixed table column issues for form-options-complete
Sprint G: Agregado middleware de roles (FASE 2)
"""
from datetime import datetime
from typing import Optional, List
from fastapi import APIRouter, HTTPException, status, Query, Depends, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel, Field, EmailStr
import pymysql
import jwt
import os
import json
import openpyxl
from io import BytesIO

from app.config import get_settings
from app.utils.working_hours import WorkingHoursCalculator
from services.work_order_calculations import calcular_todos

# Sprint G: Middleware de roles y constantes
from app.constants import (
    Roles, Areas, ROLES_ADMIN, ROLES_CREAR_OT,
    can_see_all_ots, can_create_ot, is_admin, get_user_area
)
from app.middleware.roles import require_roles, require_crear_ot, get_user_filter_query

# Sprint M: Bitácora de cambios
from app.services.bitacora_service import (
    get_bitacora_service,
    UserData,
    comparar_campos_ot,
)

# Sprint M: Push Notifications FCM
from app.services.fcm_service import get_fcm_service, TipoNotificacion

router = APIRouter(prefix="/work-orders", tags=["Órdenes de Trabajo"])
security = HTTPBearer(auto_error=False)
settings = get_settings()


def get_current_user_id(credentials: HTTPAuthorizationCredentials = Depends(security)) -> int:
    """Extrae el user_id del token JWT."""
    if not credentials:
        raise HTTPException(status_code=401, detail="Token no proporcionado")
    try:
        payload = jwt.decode(
            credentials.credentials,
            settings.JWT_SECRET_KEY,
            algorithms=[settings.JWT_ALGORITHM]
        )
        return int(payload.get("sub"))
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalido")


def get_current_user_with_role(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    """Extrae user_id y role_id del token JWT."""
    if not credentials:
        raise HTTPException(status_code=401, detail="Token no proporcionado")
    try:
        payload = jwt.decode(
            credentials.credentials,
            settings.JWT_SECRET_KEY,
            algorithms=[settings.JWT_ALGORITHM]
        )
        return {
            "user_id": int(payload.get("sub")),
            "role_id": int(payload.get("role_id", 0))
        }
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalido")


# Schemas
class WorkOrderListItem(BaseModel):
    """Item de la lista de OTs."""
    id: int
    created_at: str
    client_name: str
    descripcion: str
    canal: Optional[str]
    item_tipo: Optional[str]
    estado: str
    estado_abrev: str
    creador_nombre: str
    current_area_id: Optional[int]
    ultimo_cambio_area: Optional[str]  # Fecha cuando la OT entró al área actual
    # Tiempos por área (días laborales - 9.5h/día como Laravel)
    tiempo_total: Optional[float]
    tiempo_venta: Optional[float]
    tiempo_desarrollo: Optional[float]
    tiempo_muestra: Optional[float]
    tiempo_diseno: Optional[float]
    tiempo_externo: Optional[float]
    tiempo_precatalogacion: Optional[float]
    tiempo_catalogacion: Optional[float]
    # Extras
    tipo_solicitud: Optional[int]
    cad: Optional[str]
    carton: Optional[str]
    material_codigo: Optional[str]


class WorkOrderListResponse(BaseModel):
    """Respuesta paginada de OTs."""
    items: List[WorkOrderListItem]
    total: int
    page: int
    page_size: int
    total_pages: int


class FilterOptions(BaseModel):
    """Opciones para los filtros."""
    estados: List[dict]
    areas: List[dict]
    canales: List[dict]
    clientes: List[dict]
    vendedores: List[dict]
    plantas: List[dict]
    impresiones: List[dict]
    procesos: List[dict]
    estilos: List[dict]
    fsc: List[dict]
    org_ventas: List[dict]


class WorkOrderCreate(BaseModel):
    """Schema para crear una OT."""
    # Datos Comerciales (requeridos)
    client_id: int = Field(..., description="ID del cliente")
    descripcion: str = Field(..., max_length=40, description="Descripcion del producto")
    tipo_solicitud: int = Field(..., description="Tipo de solicitud (1-7)")
    canal_id: int = Field(..., description="ID del canal")

    # Datos Comerciales (opcionales)
    org_venta_id: Optional[int] = None
    subsubhierarchy_id: Optional[int] = None
    nombre_contacto: Optional[str] = None
    email_contacto: Optional[EmailStr] = None
    telefono_contacto: Optional[str] = None
    volumen_venta_anual: Optional[int] = None
    usd: Optional[int] = None
    oc: Optional[int] = None
    codigo_producto: Optional[str] = None
    dato_sub_cliente: Optional[str] = None
    instalacion_cliente: Optional[int] = None

    # Solicitante (checkboxes)
    analisis: Optional[int] = 0
    plano: Optional[int] = 0
    muestra: Optional[int] = 0
    datos_cotizar: Optional[int] = 0
    boceto: Optional[int] = 0
    nuevo_material: Optional[int] = 0
    prueba_industrial: Optional[int] = 0
    numero_muestras: Optional[int] = None

    # Referencia Material
    reference_type: Optional[int] = None
    reference_id: Optional[int] = None

    # Caracteristicas (Cascade)
    product_type_id: Optional[int] = None
    impresion: Optional[int] = None
    fsc: Optional[str] = None
    cinta: Optional[int] = None
    coverage_internal_id: Optional[int] = None
    coverage_external_id: Optional[int] = None
    carton_color: Optional[int] = None
    carton_id: Optional[int] = None
    cad_id: Optional[int] = None
    cad: Optional[str] = None
    style_id: Optional[int] = None

    # Medidas
    interno_largo: Optional[int] = None
    interno_ancho: Optional[int] = None
    interno_alto: Optional[int] = None
    externo_largo: Optional[int] = None
    externo_ancho: Optional[int] = None
    externo_alto: Optional[int] = None

    # Terminaciones
    process_id: Optional[int] = None
    armado_id: Optional[int] = None
    sentido_armado: Optional[int] = None
    tipo_sentido_onda: Optional[str] = None

    # Colores
    numero_colores: Optional[int] = None
    color_1_id: Optional[int] = None
    color_2_id: Optional[int] = None
    color_3_id: Optional[int] = None
    color_4_id: Optional[int] = None
    color_5_id: Optional[int] = None

    # Desarrollo
    peso_contenido_caja: Optional[int] = None
    autosoportante: Optional[int] = None
    envase_id: Optional[int] = None
    cantidad: Optional[int] = None
    observacion: Optional[str] = None

    # Planta objetivo
    planta_id: Optional[int] = None

    # Tipo de OT Especial (Sprint 1 - OTs Especiales)
    # Fuente Laravel: WorkOrderController.php líneas 799, 1250, 1731, 2007
    # Valores: 1=Licitación, 2=Ficha Técnica, 3=Estudio Benchmarking, null=OT Normal
    ajuste_area_desarrollo: Optional[int] = Field(
        None,
        description="Tipo de ajuste área desarrollo: 1=Licitación, 2=Ficha Técnica, 3=Estudio Benchmarking"
    )

    # ===========================================================================
    # Checkboxes OTs Especiales - Licitación (ajuste_area_desarrollo=1)
    # Fuente Laravel: ficha-form-licitacion.blade.php líneas 235-376
    # ===========================================================================
    check_entregadas_todas: Optional[int] = Field(0, description="Muestras entregadas: Todas")
    check_entregadas_algunas: Optional[int] = Field(0, description="Muestras entregadas: Algunas")
    cantidad_entregadas_algunas: Optional[int] = Field(None, description="Cantidad muestras entregadas (si algunas)")
    cantidad_item_licitacion: Optional[int] = Field(None, description="Cantidad items licitación")
    fecha_maxima_entrega_licitacion: Optional[str] = Field(None, description="Fecha máxima entrega licitación YYYY-MM-DD")

    # ===========================================================================
    # Checkboxes OTs Especiales - Ficha Técnica (ajuste_area_desarrollo=2)
    # Fuente Laravel: ficha-form-ficha-tecnica.blade.php
    # ===========================================================================
    check_ficha_simple: Optional[int] = Field(0, description="Tipo ficha: Simple")
    check_ficha_doble: Optional[int] = Field(0, description="Tipo ficha: Completa")

    # ===========================================================================
    # Checkboxes OTs Especiales - Estudio Benchmarking (ajuste_area_desarrollo=3)
    # Fuente Laravel: ficha-form-estudio-bench.blade.php líneas 145-310
    # Solo visible para Ingeniero (role=6) o JefeDesarrollo (role=5)
    # ===========================================================================
    # Campos adicionales Benchmarking
    cantidad_estudio_bench: Optional[int] = Field(None, description="Cantidad para estudio benchmarking")
    fecha_maxima_entrega_estudio: Optional[str] = Field(None, description="Fecha máxima entrega estudio YYYY-MM-DD")
    detalle_estudio_bench: Optional[str] = Field(None, description="JSON con detalles del estudio benchmarking")

    # Checkboxes Ensayos Caja - 17 checkboxes
    check_estudio_bct: Optional[int] = Field(0, description="BCT (lbf)")
    check_estudio_ect: Optional[int] = Field(0, description="ECT (lb/in)")
    check_estudio_bct_humedo: Optional[int] = Field(0, description="BCT en Húmedo (lbf)")
    check_estudio_flat: Optional[int] = Field(0, description="Flat Crush (lb/in)")
    check_estudio_humedad: Optional[int] = Field(0, description="Humedad (%)")
    check_estudio_porosidad_ext: Optional[int] = Field(0, description="Porosidad Exterior Gurley")
    check_estudio_espesor: Optional[int] = Field(0, description="Espesor (mm)")
    check_estudio_cera: Optional[int] = Field(0, description="Cera")
    check_estudio_porosidad_int: Optional[int] = Field(0, description="Porosidad Interior Gurley")
    check_estudio_flexion_fondo: Optional[int] = Field(0, description="Flexión de Fondo")
    check_estudio_gramaje: Optional[int] = Field(0, description="Gramaje (gr/mt2)")
    check_estudio_composicion_papeles: Optional[int] = Field(0, description="Composición Papeles")
    check_estudio_cobb_interno: Optional[int] = Field(0, description="Cobb Interno")
    check_estudio_cobb_externo: Optional[int] = Field(0, description="Cobb Externo")
    check_estudio_flexion_4_puntos: Optional[int] = Field(0, description="Flexión 4 Puntos")
    check_estudio_medidas: Optional[int] = Field(0, description="Medidas")
    check_estudio_impresion: Optional[int] = Field(0, description="Impresión")

    # Fórmula McKee (Issue 25)
    # Fuente Laravel: ot-creation.js líneas 3620-3810
    largo_mckee: Optional[int] = None
    ancho_mckee: Optional[int] = None
    alto_mckee: Optional[int] = None
    perimetro_mckee: Optional[int] = None
    carton_id_mckee: Optional[int] = None
    ect_mckee: Optional[int] = None
    espesor_mckee: Optional[float] = None
    bct_lib_mckee: Optional[int] = None
    bct_kilos_mckee: Optional[int] = None
    fecha_mckee: Optional[str] = None
    aplicar_mckee: Optional[int] = None


class WorkOrderCreateResponse(BaseModel):
    """Respuesta al crear OT."""
    id: int
    message: str


def get_mysql_connection():
    """Crea conexión a MySQL de Laravel."""
    try:
        connection = pymysql.connect(
            host=settings.LARAVEL_MYSQL_HOST,
            port=settings.LARAVEL_MYSQL_PORT,
            user=settings.LARAVEL_MYSQL_USER,
            password=settings.LARAVEL_MYSQL_PASSWORD,
            database=settings.LARAVEL_MYSQL_DATABASE,
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor
        )
        return connection
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=f"Error conectando a base de datos Laravel: {str(e)}"
        )


@router.get("/", response_model=WorkOrderListResponse)
async def list_work_orders(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    # Filtros
    id_ot: Optional[int] = None,
    date_desde: Optional[str] = None,
    date_hasta: Optional[str] = None,
    client_id: Optional[List[int]] = Query(None),
    estado_id: Optional[List[int]] = Query(None),
    area_id: Optional[List[int]] = Query(None),
    canal_id: Optional[List[int]] = Query(None),
    vendedor_id: Optional[List[int]] = Query(None),
    cad: Optional[str] = None,
    carton: Optional[str] = None,
    material: Optional[str] = None,
    descripcion: Optional[str] = None,
    planta_id: Optional[List[int]] = Query(None),
    tipo_solicitud: Optional[List[int]] = Query(None),
    # Autenticación para filtro por rol
    current_user: dict = Depends(get_current_user_with_role),
):
    """
    Lista órdenes de trabajo con filtros y paginación.

    REGLA DE NEGOCIO (replica Laravel):
    - Vendedores (role_id=4) y Vendedores Externos (role_id=19) solo ven OTs que ellos crearon
    - Otros roles ven todas las OTs según sus filtros
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Query base con tiempos almacenados (sin cálculo real-time)
            # work_space_id: 1=Ventas, 2=Desarrollo, 3=Diseño Gráfico, 4=Catalogación, 5=Precatalogación, 6=Muestra
            # NOTA: El cálculo real-time se hace en Python usando get_working_hours()
            # para replicar exactamente el comportamiento de Laravel
            base_query = """
                SELECT
                    wo.id,
                    wo.created_at,
                    wo.descripcion,
                    wo.tipo_solicitud,
                    wo.current_area_id,
                    wo.ultimo_cambio_area,
                    c.nombre as client_name,
                    CONCAT(u.nombre, ' ', u.apellido) as creador_nombre,
                    ch.nombre as canal,
                    pt.descripcion as item_tipo,
                    COALESCE(s.nombre, 'Proceso de Ventas') as estado,
                    COALESCE(s.abreviatura, 'PV') as estado_abrev,
                    cad.cad as cad_codigo,
                    cart.codigo as carton_codigo,
                    mat.codigo as material_codigo,
                    -- Tiempos almacenados por área (horas)
                    COALESCE((SELECT SUM(duracion_segundos) / 3600.0 FROM managements
                        WHERE work_order_id = wo.id AND management_type_id = 1 AND work_space_id = 1 AND mostrar = 1), 0) as horas_venta_stored,
                    COALESCE((SELECT SUM(duracion_segundos) / 3600.0 FROM managements
                        WHERE work_order_id = wo.id AND management_type_id = 1 AND work_space_id = 2 AND mostrar = 1), 0) as horas_desarrollo_stored,
                    COALESCE((SELECT SUM(duracion_segundos) / 3600.0 FROM managements
                        WHERE work_order_id = wo.id AND management_type_id = 1 AND work_space_id = 3 AND mostrar = 1), 0) as horas_diseno_stored,
                    COALESCE((SELECT SUM(duracion_segundos) / 3600.0 FROM managements
                        WHERE work_order_id = wo.id AND management_type_id = 1 AND work_space_id = 4 AND mostrar = 1), 0) as horas_catalogacion_stored,
                    COALESCE((SELECT SUM(duracion_segundos) / 3600.0 FROM managements
                        WHERE work_order_id = wo.id AND management_type_id = 1 AND work_space_id = 5 AND mostrar = 1), 0) as horas_precatalogacion_stored,
                    COALESCE((SELECT SUM(duracion_segundos) / 3600.0 FROM managements
                        WHERE work_order_id = wo.id AND management_type_id = 1 AND work_space_id = 6 AND mostrar = 1), 0) as horas_muestra_stored,
                    -- Tiempo total almacenado (horas)
                    COALESCE((SELECT SUM(duracion_segundos) / 3600.0 FROM managements
                        WHERE work_order_id = wo.id AND management_type_id = 1 AND mostrar = 1), 0) as horas_total_stored
                FROM work_orders wo
                LEFT JOIN clients c ON wo.client_id = c.id
                LEFT JOIN users u ON wo.creador_id = u.id
                LEFT JOIN canals ch ON wo.canal_id = ch.id
                LEFT JOIN product_types pt ON wo.product_type_id = pt.id
                LEFT JOIN (
                    SELECT work_order_id, state_id
                    FROM managements
                    WHERE id IN (SELECT MAX(id) FROM managements GROUP BY work_order_id)
                ) m ON wo.id = m.work_order_id
                LEFT JOIN states s ON m.state_id = s.id
                LEFT JOIN cads cad ON wo.cad_id = cad.id
                LEFT JOIN cartons cart ON wo.carton_id = cart.id
                LEFT JOIN materials mat ON wo.material_id = mat.id
                WHERE wo.active = 1
            """

            params = []

            # REGLA DE NEGOCIO: Vendedores (role_id=4) y Vendedores Externos (role_id=19)
            # solo ven las OTs que ellos crearon (replica comportamiento de Laravel)
            user_role_id = current_user.get("role_id", 0)
            user_id = current_user.get("user_id", 0)

            if user_role_id in [4, 19]:  # Vendedor o Vendedor Externo
                base_query += " AND wo.creador_id = %s"
                params.append(user_id)

            # REGLA DE NEGOCIO: Ingenieros, Diseñadores y Catalogadores
            # solo ven OTs donde están asignados (replica comportamiento de Laravel línea 118)
            #
            # CORREGIDO 2026-02-23: Laravel tenía bug donde Constants.php decía:
            #   - Constants::Catalogador = 10, Constants::Precatalogador = 12
            # Pero la BD real tiene:
            #   - role_id 10 = "Precatalogador", role_id 12 = "Catalogador"
            #
            # Usamos valores REALES de la BD:
            # - 6 = Ingeniero
            # - 8 = Diseñador
            # - 12 = Catalogador (el real según BD)
            ROLES_FILTRAR_POR_ASIGNACION = [6, 8, 12]
            if user_role_id in ROLES_FILTRAR_POR_ASIGNACION:
                base_query += """
                    AND wo.id IN (
                        SELECT work_order_id
                        FROM user_work_orders
                        WHERE user_id = %s
                    )
                """
                params.append(user_id)

            # REGLA DE NEGOCIO: Filtrado por área según rol (replica Laravel líneas 312-315)
            # Diseñador/Jefe Diseño (7, 8): solo ven OTs en área 3 (Diseño)
            # Precatalogador/JefePrecatalogador/Catalogador/JefeCatalogador (9,10,11,12):
            # solo ven OTs en áreas 4,5 (Precatalogación, Catalogación)
            ROLES_AREA_DISENO = [7, 8]  # Jefe Diseño, Diseñador
            ROLES_AREA_CATALOGACION = [9, 10, 11, 12]  # Jefe Precatalog, Precatalog, Jefe Catalog, Catalogador

            if user_role_id in ROLES_AREA_DISENO:
                base_query += " AND wo.current_area_id = 3"
            elif user_role_id in ROLES_AREA_CATALOGACION:
                base_query += " AND wo.current_area_id IN (4, 5)"

            # REGLA DE NEGOCIO Sprint G: VendedorExterno (role_id=19) solo ve cliente ID=8
            # Fuente Laravel: WorkOrderController.php líneas 657-661
            if user_role_id == 19:  # VendedorExterno
                base_query += " AND wo.client_id = 8"

            # Aplicar filtros
            if id_ot:
                base_query += " AND wo.id = %s"
                params.append(id_ot)

            if date_desde:
                base_query += " AND wo.created_at >= %s"
                params.append(date_desde)

            if date_hasta:
                base_query += " AND wo.created_at <= %s"
                params.append(date_hasta + " 23:59:59")

            if client_id:
                placeholders = ','.join(['%s'] * len(client_id))
                base_query += f" AND wo.client_id IN ({placeholders})"
                params.extend(client_id)

            if estado_id:
                placeholders = ','.join(['%s'] * len(estado_id))
                base_query += f" AND m.state_id IN ({placeholders})"
                params.extend(estado_id)

            if area_id:
                placeholders = ','.join(['%s'] * len(area_id))
                base_query += f" AND wo.current_area_id IN ({placeholders})"
                params.extend(area_id)

            if canal_id:
                placeholders = ','.join(['%s'] * len(canal_id))
                base_query += f" AND wo.canal_id IN ({placeholders})"
                params.extend(canal_id)

            if vendedor_id:
                placeholders = ','.join(['%s'] * len(vendedor_id))
                base_query += f" AND wo.creador_id IN ({placeholders})"
                params.extend(vendedor_id)

            if cad:
                base_query += " AND cad.codigo LIKE %s"
                params.append(f"%{cad}%")

            if carton:
                base_query += " AND cart.codigo LIKE %s"
                params.append(f"%{carton}%")

            if material:
                base_query += " AND mat.codigo LIKE %s"
                params.append(f"%{material}%")

            if descripcion:
                base_query += " AND wo.descripcion LIKE %s"
                params.append(f"%{descripcion}%")

            if planta_id:
                placeholders = ','.join(['%s'] * len(planta_id))
                base_query += f" AND wo.planta_id IN ({placeholders})"
                params.extend(planta_id)

            if tipo_solicitud:
                placeholders = ','.join(['%s'] * len(tipo_solicitud))
                base_query += f" AND wo.tipo_solicitud IN ({placeholders})"
                params.extend(tipo_solicitud)

            # Contar total
            count_query = f"SELECT COUNT(*) as total FROM ({base_query}) as subquery"
            cursor.execute(count_query, params)
            total = cursor.fetchone()['total']

            # Ordenar y paginar
            base_query += " ORDER BY wo.id DESC LIMIT %s OFFSET %s"
            offset = (page - 1) * page_size
            params.extend([page_size, offset])

            cursor.execute(base_query, params)
            rows = cursor.fetchall()

            # Inicializar calculador de horas laborales (replica Laravel get_working_hours)
            wh_calculator = WorkingHoursCalculator(connection)
            now = datetime.now()
            HOURS_PER_DAY = 9.5  # Horas laborales por día (igual que Laravel)

            # Mapeo área -> campo de tiempo
            area_to_field = {
                1: 'horas_venta_stored',
                2: 'horas_desarrollo_stored',
                3: 'horas_diseno_stored',
                4: 'horas_catalogacion_stored',
                5: 'horas_precatalogacion_stored',
                6: 'horas_muestra_stored',
            }

            # Transformar resultados con cálculo de tiempo real-time
            items = []
            for row in rows:
                # Obtener horas almacenadas
                horas_venta = float(row['horas_venta_stored'] or 0)
                horas_desarrollo = float(row['horas_desarrollo_stored'] or 0)
                horas_diseno = float(row['horas_diseno_stored'] or 0)
                horas_catalogacion = float(row['horas_catalogacion_stored'] or 0)
                horas_precatalogacion = float(row['horas_precatalogacion_stored'] or 0)
                horas_muestra = float(row['horas_muestra_stored'] or 0)
                horas_total = float(row['horas_total_stored'] or 0)

                # Si la OT está en un área activa, calcular tiempo real usando get_working_hours
                current_area_id = row.get('current_area_id')
                ultimo_cambio = row.get('ultimo_cambio_area')

                if current_area_id and ultimo_cambio:
                    # Calcular horas laborales desde ultimo_cambio_area hasta ahora
                    # Esto replica exactamente Laravel get_working_hours()
                    horas_realtime = wh_calculator.get_working_hours(ultimo_cambio, now)

                    # Sumar al área correspondiente
                    if current_area_id == 1:
                        horas_venta += horas_realtime
                    elif current_area_id == 2:
                        horas_desarrollo += horas_realtime
                    elif current_area_id == 3:
                        horas_diseno += horas_realtime
                    elif current_area_id == 4:
                        horas_catalogacion += horas_realtime
                    elif current_area_id == 5:
                        horas_precatalogacion += horas_realtime
                    elif current_area_id == 6:
                        horas_muestra += horas_realtime

                    # Sumar al total también
                    horas_total += horas_realtime

                # Convertir horas a días laborales (9.5 horas/día)
                items.append(WorkOrderListItem(
                    id=row['id'],
                    created_at=row['created_at'].strftime('%d/%m/%y') if row['created_at'] else '',
                    client_name=row['client_name'] or '',
                    descripcion=row['descripcion'] or '',
                    canal=row['canal'],
                    item_tipo=row['item_tipo'],
                    estado=row['estado'],
                    estado_abrev=row['estado_abrev'],
                    creador_nombre=row['creador_nombre'] or '',
                    tipo_solicitud=row['tipo_solicitud'],
                    current_area_id=current_area_id,
                    ultimo_cambio_area=ultimo_cambio.strftime('%d/%m/%y') if ultimo_cambio else None,
                    cad=row['cad_codigo'],
                    carton=row['carton_codigo'],
                    material_codigo=row['material_codigo'],
                    # Tiempos en días laborales (horas / 9.5)
                    tiempo_total=round(horas_total / HOURS_PER_DAY, 1),
                    tiempo_venta=round(horas_venta / HOURS_PER_DAY, 1),
                    tiempo_desarrollo=round(horas_desarrollo / HOURS_PER_DAY, 1),
                    tiempo_muestra=round(horas_muestra / HOURS_PER_DAY, 1),
                    tiempo_diseno=round(horas_diseno / HOURS_PER_DAY, 1),
                    tiempo_externo=0,  # No hay área específica para externo
                    tiempo_precatalogacion=round(horas_precatalogacion / HOURS_PER_DAY, 1),
                    tiempo_catalogacion=round(horas_catalogacion / HOURS_PER_DAY, 1),
                ))

            total_pages = (total + page_size - 1) // page_size

            return WorkOrderListResponse(
                items=items,
                total=total,
                page=page,
                page_size=page_size,
                total_pages=total_pages
            )
    finally:
        connection.close()


@router.get("/filter-options", response_model=FilterOptions)
async def get_filter_options():
    """
    Obtiene todas las opciones para los filtros del dashboard.
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            options = {}

            # Estados (usa 'status' texto, no 'active')
            cursor.execute("SELECT id, nombre, abreviatura FROM states WHERE status = 'active' ORDER BY id")
            options['estados'] = cursor.fetchall()

            # Áreas (work_spaces usa 'status' texto)
            cursor.execute("SELECT id, nombre FROM work_spaces WHERE status = 'active' ORDER BY id")
            options['areas'] = cursor.fetchall()

            # Canales
            cursor.execute("SELECT id, nombre FROM canals WHERE active = 1 ORDER BY nombre")
            options['canales'] = cursor.fetchall()

            # Clientes (top 500 más usados) - ordenados por codigo
            cursor.execute("""
                SELECT DISTINCT c.id, c.nombre as nombre, c.codigo
                FROM clients c
                INNER JOIN work_orders wo ON c.id = wo.client_id
                WHERE c.active = 1
                ORDER BY c.codigo ASC
                LIMIT 500
            """)
            options['clientes'] = cursor.fetchall()

            # Vendedores (usuarios que pueden crear OT)
            cursor.execute("""
                SELECT id, CONCAT(nombre, ' ', apellido) as nombre
                FROM users
                WHERE active = 1 AND role_id IN (2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12)
                ORDER BY nombre
            """)
            options['vendedores'] = cursor.fetchall()

            # Plantas (no tiene columna active)
            cursor.execute("SELECT id, nombre FROM plantas ORDER BY nombre")
            options['plantas'] = cursor.fetchall()

            # Impresiones (tabla impresion - usa 'status' no 'active')
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM impresion
                WHERE status = 1
                ORDER BY id
            """)
            options['impresiones'] = cursor.fetchall()

            # Procesos - Issue 47: Filtrar por type='EV' y ordenar por 'orden' (igual que Laravel)
            cursor.execute("SELECT id, descripcion as nombre FROM processes WHERE active = 1 AND type = 'EV' ORDER BY orden ASC")
            options['procesos'] = cursor.fetchall()

            # Estilos
            cursor.execute("SELECT id, glosa as nombre FROM styles WHERE active = 1 ORDER BY glosa")
            options['estilos'] = cursor.fetchall()

            # FSC
            cursor.execute("SELECT codigo as id, descripcion as nombre FROM fsc WHERE active = 1 ORDER BY descripcion")
            options['fsc'] = cursor.fetchall()

            # Organizaciones de venta (tabla no existe - hardcoded)
            options['org_ventas'] = [
                {"id": 1, "nombre": "Nacional"},
                {"id": 2, "nombre": "Exportación"}
            ]

            return FilterOptions(**options)
    finally:
        connection.close()


# =============================================
# FORM OPTIONS COMPLETE - debe estar antes de /{ot_id}
# =============================================
# Nota: La implementación está más abajo, pero la ruta debe definirse aquí
# para que FastAPI no intente parsear "form-options-complete" como un int

@router.get("/form-options-complete")
async def get_form_options_complete_route():
    """
    Obtiene TODAS las opciones necesarias para el formulario de crear/editar OT.
    Redirige a la implementación completa.
    """
    return await _get_form_options_complete_impl()


# =============================================
# RUTAS ESTÁTICAS - deben estar antes de /{ot_id}
# =============================================
# Estas rutas se definen aquí para que FastAPI no intente
# parsear sus paths como integers

@router.get("/pending-approval")
async def get_pending_approval_route(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(get_current_user_with_role)
):
    """Lista OTs pendientes de aprobación según rol del usuario."""
    return await _get_pending_approval_impl(page, page_size, current_user["user_id"], current_user["role_id"])


@router.get("/pending-assignment")
async def get_pending_assignment_route(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    asignado: Optional[str] = Query(None, description="SI o NO"),
    tipo_solicitud: Optional[str] = None,
    canal_id: Optional[int] = None,
    vendedor_id: Optional[int] = None,
    estado_id: Optional[int] = None,
    date_desde: Optional[str] = None,
    date_hasta: Optional[str] = None,
    user_id: int = Depends(get_current_user_id)
):
    """Redirige a la implementación de pending-assignment."""
    return await _get_pending_assignment_impl(
        page, page_size, asignado, tipo_solicitud,
        canal_id, vendedor_id, estado_id, date_desde, date_hasta, user_id
    )


@router.get("/professionals")
async def get_professionals_route(user_id: int = Depends(get_current_user_id)):
    """Redirige a la implementación de professionals."""
    return await _get_professionals_impl(user_id)


@router.get("/{ot_id}")
async def get_work_order(ot_id: int):
    """
    Obtiene detalle de una OT específica.

    Incluye valores calculados (áreas, pesos, consumos) según fórmulas de Laravel.
    Fuente: app/WorkOrder.php líneas 346-880
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            sql = """
                SELECT
                    wo.*,
                    c.nombre as client_name,
                    c.codigo as client_codigo,
                    CONCAT(u.nombre, ' ', u.apellido) as creador_nombre,
                    ch.nombre as canal_nombre,
                    pt.descripcion as product_type,
                    cad.cad as cad_codigo,
                    cart.codigo as carton_codigo,
                    cart.peso as carton_gramaje,
                    cart.espesor as carton_espesor,
                    cart.tipo as carton_tipo,
                    mat.codigo as material_codigo,
                    s.nombre as estado_nombre,
                    s.abreviatura as estado_abrev,
                    ws.nombre as area_actual,
                    proc.descripcion as proceso_descripcion
                FROM work_orders wo
                LEFT JOIN clients c ON wo.client_id = c.id
                LEFT JOIN users u ON wo.creador_id = u.id
                LEFT JOIN canals ch ON wo.canal_id = ch.id
                LEFT JOIN product_types pt ON wo.product_type_id = pt.id
                LEFT JOIN cads cad ON wo.cad_id = cad.id
                LEFT JOIN cartons cart ON wo.carton_id = cart.id
                LEFT JOIN materials mat ON wo.material_id = mat.id
                LEFT JOIN work_spaces ws ON wo.current_area_id = ws.id
                LEFT JOIN processes proc ON wo.process_id = proc.id
                LEFT JOIN (
                    SELECT work_order_id, state_id
                    FROM managements
                    WHERE id IN (SELECT MAX(id) FROM managements GROUP BY work_order_id)
                ) m ON wo.id = m.work_order_id
                LEFT JOIN states s ON m.state_id = s.id
                WHERE wo.id = %s
            """
            cursor.execute(sql, (ot_id,))
            ot = cursor.fetchone()

            if not ot:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"OT {ot_id} no encontrada"
                )

            # Convertir datetime a string
            for key, value in ot.items():
                if isinstance(value, datetime):
                    ot[key] = value.isoformat()

            # Calcular valores derivados si hay datos suficientes
            # Fuente Laravel: app/WorkOrder.php líneas 346-880
            if ot.get('carton_gramaje') and ot.get('carton_tipo'):
                try:
                    work_order_data = {
                        "largura_hm": ot.get("largura_hm"),
                        "anchura_hm": ot.get("anchura_hm"),
                        "golpes_largo": ot.get("golpes_largo"),
                        "golpes_ancho": ot.get("golpes_ancho"),
                        "proceso_descripcion": ot.get("proceso_descripcion"),
                        "area_producto": ot.get("area_producto"),
                        "separacion_golpes_largo": ot.get("separacion_golpes_largo"),
                        "separacion_golpes_ancho": ot.get("separacion_golpes_ancho"),
                        "impresion_1": ot.get("impresion_1"),
                        "impresion_2": ot.get("impresion_2"),
                        "impresion_3": ot.get("impresion_3"),
                        "impresion_4": ot.get("impresion_4"),
                        "impresion_5": ot.get("impresion_5"),
                        "impresion_6": ot.get("impresion_6"),
                        "impresion_7": ot.get("impresion_7"),
                        "porcentanje_barniz_uv": ot.get("porcentanje_barniz_uv"),
                        "impresion_color_interno": ot.get("impresion_color_interno"),
                        "longitud_pegado": ot.get("longitud_pegado"),
                        "percentage_coverage_internal": ot.get("percentage_coverage_internal"),
                        "coverage_internal_id": ot.get("coverage_internal_id"),
                        "percentage_coverage_external": ot.get("percentage_coverage_external"),
                        "coverage_external_id": ot.get("coverage_external_id"),
                        "porcentaje_barniz_interior": ot.get("porcentaje_barniz_interior"),
                        "recorte_adicional": ot.get("recorte_adicional"),
                    }
                    carton_data = {
                        "tipo": ot.get("carton_tipo"),
                        "peso": ot.get("carton_gramaje"),
                        "espesor": ot.get("carton_espesor"),
                    }
                    calculos = calcular_todos(work_order_data, carton_data)
                    ot["calculos"] = calculos
                except Exception as e:
                    # Si falla el cálculo, agregar error pero no fallar la respuesta
                    ot["calculos"] = {"error": str(e)}

            return ot
    finally:
        connection.close()


@router.post("/", response_model=WorkOrderCreateResponse)
async def create_work_order(
    data: WorkOrderCreate,
    user: dict = Depends(require_crear_ot())
):
    """
    Crea una nueva orden de trabajo.

    El usuario autenticado se asigna como creador y la OT inicia en area 1 (Ventas).

    Sprint G: Requiere rol autorizado para crear OTs.
    Roles permitidos: Admin, SuperAdmin, JefeVenta, Vendedor, VendedorExterno,
                      JefeDesarrollo, Ingeniero, JefeDiseño, Diseñador

    Restricciones por rol (Laravel WorkOrderController.php líneas 898-904):
    - Vendedor Externo: Solo tipos 1 (Desarrollo Completo) y 5 (Arte con Material)
    - Vendedor Externo: Solo cliente ID=8
    - Área Desarrollo: Solo tipos 1, 3, 6, 7 (no puede 5: Arte con Material)
    """
    user_id = user["id"]
    role_id = user.get("role_id", 0)

    # VALIDACIÓN: Restricciones por rol
    # Constantes
    ROLE_VENDEDOR_EXTERNO = 19
    TIPOS_VENDEDOR_EXTERNO = [1, 5]  # Desarrollo Completo, Arte con Material
    CLIENTE_VENDEDOR_EXTERNO = 8
    AREA_DESARROLLO = 2
    TIPOS_AREA_DESARROLLO = [1, 3, 6, 7]  # No incluye 5 (Arte con Material)

    # Vendedor Externo: Solo puede crear tipos 1 y 5
    if role_id == ROLE_VENDEDOR_EXTERNO:
        if data.tipo_solicitud not in TIPOS_VENDEDOR_EXTERNO:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"Vendedor Externo solo puede crear OTs de tipo: Desarrollo Completo (1) o Arte con Material (5)"
            )
        # Vendedor Externo: Solo puede usar cliente ID=8
        if data.client_id != CLIENTE_VENDEDOR_EXTERNO:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"Vendedor Externo solo puede crear OTs para el cliente autorizado"
            )

    # Usuarios de Área Desarrollo: No pueden crear tipo 5 (Arte con Material)
    # Roles de área Desarrollo: JefeDesarrollo (5), Ingeniero (6)
    if role_id in [5, 6]:
        if data.tipo_solicitud not in TIPOS_AREA_DESARROLLO:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"Usuarios de Desarrollo no pueden crear OTs de tipo Arte con Material (5)"
            )

    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Construir campos e insertar OT
            insert_fields = {
                # Campos requeridos
                'client_id': data.client_id,
                'descripcion': data.descripcion,
                'tipo_solicitud': data.tipo_solicitud,
                'canal_id': data.canal_id,
                'creador_id': user_id,
                'current_area_id': 1,  # Inicia en Ventas
                'active': 1,
                'ultimo_cambio_area': now,  # Fecha de entrada al área actual
                'created_at': now,
                'updated_at': now,
                # Campo observacion requerido por BD (sin default para TEXT)
                'observacion': data.observacion or '',
            }

            # Agregar campos opcionales si tienen valor
            optional_fields = {
                'org_venta_id': data.org_venta_id,
                'subsubhierarchy_id': data.subsubhierarchy_id,
                'nombre_contacto': data.nombre_contacto,
                'email_contacto': data.email_contacto,
                'telefono_contacto': data.telefono_contacto,
                'volumen_venta_anual': data.volumen_venta_anual,
                'usd': data.usd,
                'oc': data.oc,
                'codigo_producto': data.codigo_producto,
                'dato_sub_cliente': data.dato_sub_cliente,
                'instalacion_cliente': data.instalacion_cliente,
                # Solicitante
                'analisis': data.analisis,
                'plano': data.plano,
                'muestra': data.muestra,
                'datos_cotizar': data.datos_cotizar,
                'boceto': data.boceto,
                'nuevo_material': data.nuevo_material,
                'prueba_industrial': data.prueba_industrial,
                'numero_muestras': data.numero_muestras,
                # Referencia
                'reference_type': data.reference_type,
                'reference_id': data.reference_id,
                # Caracteristicas (Cascade)
                'product_type_id': data.product_type_id,
                'impresion': data.impresion,
                'fsc': data.fsc,
                'cinta': data.cinta,
                'coverage_internal_id': data.coverage_internal_id,
                'coverage_external_id': data.coverage_external_id,
                'carton_color': data.carton_color,
                'carton_id': data.carton_id,
                'cad_id': data.cad_id,
                'cad': data.cad,
                'style_id': data.style_id,
                # Medidas
                'interno_largo': data.interno_largo,
                'interno_ancho': data.interno_ancho,
                'interno_alto': data.interno_alto,
                'externo_largo': data.externo_largo,
                'externo_ancho': data.externo_ancho,
                'externo_alto': data.externo_alto,
                # Terminaciones
                'process_id': data.process_id,
                'armado_id': data.armado_id,
                'sentido_armado': data.sentido_armado,
                'tipo_sentido_onda': data.tipo_sentido_onda,
                # Colores
                'numero_colores': data.numero_colores,
                'color_1_id': data.color_1_id,
                'color_2_id': data.color_2_id,
                'color_3_id': data.color_3_id,
                'color_4_id': data.color_4_id,
                'color_5_id': data.color_5_id,
                # Desarrollo
                'peso_contenido_caja': data.peso_contenido_caja,
                'autosoportante': data.autosoportante,
                'envase_id': data.envase_id,
                'cantidad': data.cantidad,
                # observacion movido a campos requeridos
                # Planta
                'planta_id': data.planta_id,
                # OTs Especiales (Sprint 1)
                # Fuente Laravel: WorkOrderController.php líneas 1250, 1731, 2007
                'ajuste_area_desarrollo': data.ajuste_area_desarrollo,
                # Checkboxes Licitación
                'check_entregadas_todas': data.check_entregadas_todas,
                'check_entregadas_algunas': data.check_entregadas_algunas,
                'cantidad_entregadas_algunas': data.cantidad_entregadas_algunas,
                'cantidad_item_licitacion': data.cantidad_item_licitacion,
                'fecha_maxima_entrega_licitacion': data.fecha_maxima_entrega_licitacion,
                # Checkboxes Ficha Técnica
                'check_ficha_simple': data.check_ficha_simple,
                'check_ficha_doble': data.check_ficha_doble,
                # Campos y Checkboxes Benchmarking
                'cantidad_estudio_bench': data.cantidad_estudio_bench,
                'fecha_maxima_entrega_estudio': data.fecha_maxima_entrega_estudio,
                'detalle_estudio_bench': data.detalle_estudio_bench,
                'check_estudio_bct': data.check_estudio_bct,
                'check_estudio_ect': data.check_estudio_ect,
                'check_estudio_bct_humedo': data.check_estudio_bct_humedo,
                'check_estudio_flat': data.check_estudio_flat,
                'check_estudio_humedad': data.check_estudio_humedad,
                'check_estudio_porosidad_ext': data.check_estudio_porosidad_ext,
                'check_estudio_espesor': data.check_estudio_espesor,
                'check_estudio_cera': data.check_estudio_cera,
                'check_estudio_porosidad_int': data.check_estudio_porosidad_int,
                'check_estudio_flexion_fondo': data.check_estudio_flexion_fondo,
                'check_estudio_gramaje': data.check_estudio_gramaje,
                'check_estudio_composicion_papeles': data.check_estudio_composicion_papeles,
                'check_estudio_cobb_interno': data.check_estudio_cobb_interno,
                'check_estudio_cobb_externo': data.check_estudio_cobb_externo,
                'check_estudio_flexion_4_puntos': data.check_estudio_flexion_4_puntos,
                'check_estudio_medidas': data.check_estudio_medidas,
                'check_estudio_impresion': data.check_estudio_impresion,
            }

            for field, value in optional_fields.items():
                if value is not None:
                    insert_fields[field] = value

            # Construir query dinámico
            columns = ', '.join(insert_fields.keys())
            placeholders = ', '.join(['%s'] * len(insert_fields))
            values = list(insert_fields.values())

            sql = f"INSERT INTO work_orders ({columns}) VALUES ({placeholders})"
            cursor.execute(sql, values)
            ot_id = cursor.lastrowid

            # Crear registro en managements (estado inicial)
            mgmt_sql = """
                INSERT INTO managements
                (work_order_id, work_space_id, state_id, user_id, management_type_id, created_at, updated_at)
                VALUES (%s, 1, 1, %s, 1, %s, %s)
            """
            cursor.execute(mgmt_sql, (ot_id, user_id, now, now))

            # Asignar usuario a la OT (tiempo_inicial = 0 segundos para nueva OT)
            user_wo_sql = """
                INSERT INTO user_work_orders (user_id, work_order_id, area_id, tiempo_inicial, created_at, updated_at)
                VALUES (%s, %s, 1, 0, %s, %s)
            """
            cursor.execute(user_wo_sql, (user_id, ot_id, now, now))

            connection.commit()

            # Sprint M: Registrar inserción en bitácora
            try:
                # Obtener datos del usuario para bitácora
                cursor.execute("""
                    SELECT nombre, apellido, rut, role_id
                    FROM users WHERE id = %s
                """, (user_id,))
                user_row = cursor.fetchone()

                if user_row:
                    user_data = UserData(
                        nombre=user_row['nombre'] or '',
                        apellido=user_row.get('apellido') or '',
                        rut=user_row.get('rut') or '',
                        role_id=user_row.get('role_id') or 0
                    )

                    # Registrar inserción con los datos insertados
                    bitacora_service = get_bitacora_service()
                    bitacora_service.registrar_insercion(
                        work_order_id=ot_id,
                        user_id=user_id,
                        user_data=user_data,
                        datos_insertados=insert_fields,
                        ip_solicitud=None,  # Se puede agregar Request dependency
                        url=f"/work-orders/{ot_id}"
                    )
            except Exception as e_bitacora:
                # No fallar la creación si la bitácora falla
                print(f"Error al registrar bitácora: {e_bitacora}")

            return WorkOrderCreateResponse(
                id=ot_id,
                message=f"Orden de trabajo {ot_id} creada exitosamente"
            )

    except pymysql.Error as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al crear OT: {str(e)}"
        )
    finally:
        connection.close()


class WorkOrderUpdate(BaseModel):
    """Schema para actualizar una OT."""
    # Datos Comerciales
    client_id: Optional[int] = None
    descripcion: Optional[str] = Field(None, max_length=40)
    tipo_solicitud: Optional[int] = None
    canal_id: Optional[int] = None
    org_venta_id: Optional[int] = None
    subsubhierarchy_id: Optional[int] = None
    nombre_contacto: Optional[str] = None
    email_contacto: Optional[EmailStr] = None
    telefono_contacto: Optional[str] = None
    volumen_venta_anual: Optional[int] = None
    usd: Optional[int] = None
    oc: Optional[int] = None
    codigo_producto: Optional[str] = None
    dato_sub_cliente: Optional[str] = None
    instalacion_cliente: Optional[int] = None
    # Solicitante
    analisis: Optional[int] = None
    plano: Optional[int] = None
    muestra: Optional[int] = None
    datos_cotizar: Optional[int] = None
    boceto: Optional[int] = None
    nuevo_material: Optional[int] = None
    prueba_industrial: Optional[int] = None
    numero_muestras: Optional[int] = None
    # Referencia
    reference_type: Optional[int] = None
    reference_id: Optional[int] = None
    # Caracteristicas (Cascade)
    product_type_id: Optional[int] = None
    impresion: Optional[int] = None
    fsc: Optional[str] = None
    cinta: Optional[int] = None
    coverage_internal_id: Optional[int] = None
    coverage_external_id: Optional[int] = None
    carton_color: Optional[int] = None
    carton_id: Optional[int] = None
    cad_id: Optional[int] = None
    cad: Optional[str] = None
    style_id: Optional[int] = None
    # Medidas
    interno_largo: Optional[int] = None
    interno_ancho: Optional[int] = None
    interno_alto: Optional[int] = None
    externo_largo: Optional[int] = None
    externo_ancho: Optional[int] = None
    externo_alto: Optional[int] = None
    # Terminaciones
    process_id: Optional[int] = None
    armado_id: Optional[int] = None
    sentido_armado: Optional[int] = None
    tipo_sentido_onda: Optional[str] = None
    # Colores
    numero_colores: Optional[int] = None
    color_1_id: Optional[int] = None
    color_2_id: Optional[int] = None
    color_3_id: Optional[int] = None
    color_4_id: Optional[int] = None
    color_5_id: Optional[int] = None
    # Desarrollo
    peso_contenido_caja: Optional[int] = None
    autosoportante: Optional[int] = None
    envase_id: Optional[int] = None
    cantidad: Optional[int] = None
    observacion: Optional[str] = None
    # Planta
    planta_id: Optional[int] = None

    # OTs Especiales
    ajuste_area_desarrollo: Optional[int] = None
    # Checkboxes Licitación
    check_entregadas_todas: Optional[int] = None
    check_entregadas_algunas: Optional[int] = None
    cantidad_entregadas_algunas: Optional[int] = None
    cantidad_item_licitacion: Optional[int] = None
    fecha_maxima_entrega_licitacion: Optional[str] = None
    # Checkboxes Ficha Técnica
    check_ficha_simple: Optional[int] = None
    check_ficha_doble: Optional[int] = None
    # Campos y Checkboxes Benchmarking
    cantidad_estudio_bench: Optional[int] = None
    fecha_maxima_entrega_estudio: Optional[str] = None
    detalle_estudio_bench: Optional[str] = None
    check_estudio_bct: Optional[int] = None
    check_estudio_ect: Optional[int] = None
    check_estudio_bct_humedo: Optional[int] = None
    check_estudio_flat: Optional[int] = None
    check_estudio_humedad: Optional[int] = None
    check_estudio_porosidad_ext: Optional[int] = None
    check_estudio_espesor: Optional[int] = None
    check_estudio_cera: Optional[int] = None
    check_estudio_porosidad_int: Optional[int] = None
    check_estudio_flexion_fondo: Optional[int] = None
    check_estudio_gramaje: Optional[int] = None
    check_estudio_composicion_papeles: Optional[int] = None
    check_estudio_cobb_interno: Optional[int] = None
    check_estudio_cobb_externo: Optional[int] = None
    check_estudio_flexion_4_puntos: Optional[int] = None
    check_estudio_medidas: Optional[int] = None
    check_estudio_impresion: Optional[int] = None

    # Fórmula McKee (Issue 25)
    largo_mckee: Optional[int] = None
    ancho_mckee: Optional[int] = None
    alto_mckee: Optional[int] = None
    perimetro_mckee: Optional[int] = None
    carton_id_mckee: Optional[int] = None
    ect_mckee: Optional[int] = None
    espesor_mckee: Optional[float] = None
    bct_lib_mckee: Optional[int] = None
    bct_kilos_mckee: Optional[int] = None
    fecha_mckee: Optional[str] = None
    aplicar_mckee: Optional[int] = None


class WorkOrderUpdateResponse(BaseModel):
    """Respuesta al actualizar OT."""
    id: int
    message: str


@router.put("/{ot_id}", response_model=WorkOrderUpdateResponse)
async def update_work_order(
    ot_id: int,
    data: WorkOrderUpdate,
    user_id: int = Depends(get_current_user_id)
):
    """
    Actualiza una orden de trabajo existente.
    Sprint M: Registra cambios en bitácora.
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Verificar que la OT existe y obtener valores actuales para bitácora
            cursor.execute("SELECT * FROM work_orders WHERE id = %s AND active = 1", (ot_id,))
            ot_actual = cursor.fetchone()
            if not ot_actual:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"OT {ot_id} no encontrada"
                )

            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Construir campos a actualizar (solo los que tienen valor)
            update_fields = {'updated_at': now}

            # Iterar sobre todos los campos del modelo
            for field_name, field_value in data.model_dump(exclude_unset=True).items():
                if field_value is not None:
                    update_fields[field_name] = field_value

            if len(update_fields) <= 1:  # Solo updated_at
                return WorkOrderUpdateResponse(
                    id=ot_id,
                    message="No hay campos para actualizar"
                )

            # Construir query dinámico
            set_clause = ', '.join([f"{k} = %s" for k in update_fields.keys()])
            values = list(update_fields.values())
            values.append(ot_id)

            sql = f"UPDATE work_orders SET {set_clause} WHERE id = %s"
            cursor.execute(sql, values)

            connection.commit()

            # Sprint M: Registrar modificación en bitácora
            try:
                # Obtener datos del usuario para bitácora
                cursor.execute("""
                    SELECT nombre, apellido, rut, role_id
                    FROM users WHERE id = %s
                """, (user_id,))
                user_row = cursor.fetchone()

                if user_row:
                    user_data = UserData(
                        nombre=user_row['nombre'] or '',
                        apellido=user_row.get('apellido') or '',
                        rut=user_row.get('rut') or '',
                        role_id=user_row.get('role_id') or 0
                    )

                    # Comparar valores antiguos con nuevos
                    datos_modificados = comparar_campos_ot(
                        ot_actual=dict(ot_actual),
                        ot_nuevo=update_fields
                    )

                    if datos_modificados:
                        bitacora_service = get_bitacora_service()
                        bitacora_service.registrar_modificacion(
                            work_order_id=ot_id,
                            user_id=user_id,
                            user_data=user_data,
                            datos_modificados=datos_modificados,
                            ip_solicitud=None,
                            url=f"/work-orders/{ot_id}"
                        )
            except Exception as e_bitacora:
                # No fallar la actualización si la bitácora falla
                print(f"Error al registrar bitácora: {e_bitacora}")

            return WorkOrderUpdateResponse(
                id=ot_id,
                message=f"Orden de trabajo {ot_id} actualizada exitosamente"
            )

    except pymysql.Error as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al actualizar OT: {str(e)}"
        )
    finally:
        connection.close()


# ========== GESTION DE OT (Workflow) ==========

class ManagementHistoryItem(BaseModel):
    """Item del historial de gestion."""
    id: int
    work_space: str
    state: str
    user_name: str
    observation: Optional[str]
    created_at: str


class ManagementHistoryResponse(BaseModel):
    """Respuesta con historial de gestion."""
    ot_id: int
    current_area: str
    current_state: str
    history: List[ManagementHistoryItem]


class WorkflowOptions(BaseModel):
    """Opciones disponibles para transicion."""
    areas: List[dict]
    states: List[dict]
    management_types: List[dict]
    motives: Optional[List[dict]] = None  # Sprint N: Motivos de rechazo
    proveedores: Optional[List[dict]] = None  # Sprint N: Proveedores externos


class TransitionRequest(BaseModel):
    """Solicitud de transicion de OT."""
    management_type_id: int = Field(..., description="ID del tipo de gestion (1=Cambio Estado, 2=Consulta, 3=Archivo)")
    work_space_id: Optional[int] = Field(None, description="ID del area destino (requerido para Cambio de Estado)")
    state_id: Optional[int] = Field(None, description="ID del nuevo estado (requerido para Cambio de Estado)")
    observation: Optional[str] = Field(None, max_length=500, description="Observacion")
    motive_id: Optional[int] = Field(None, description="Sprint N: ID del motivo de rechazo (cuando state_id=12)")
    proveedor_id: Optional[int] = Field(None, description="Sprint N: ID del proveedor externo (management_type 9 o 10)")


class TransitionResponse(BaseModel):
    """Respuesta de transicion."""
    id: int
    message: str
    new_area: str
    new_state: str


@router.get("/{ot_id}/management", response_model=ManagementHistoryResponse)
async def get_management_history(ot_id: int):
    """
    Obtiene el historial de gestion de una OT.
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Verificar que la OT existe
            cursor.execute("""
                SELECT wo.id, ws.nombre as area_actual, s.nombre as estado_actual
                FROM work_orders wo
                LEFT JOIN work_spaces ws ON wo.current_area_id = ws.id
                LEFT JOIN (
                    SELECT work_order_id, state_id
                    FROM managements
                    WHERE id IN (SELECT MAX(id) FROM managements GROUP BY work_order_id)
                ) latest ON wo.id = latest.work_order_id
                LEFT JOIN states s ON latest.state_id = s.id
                WHERE wo.id = %s AND wo.active = 1
            """, (ot_id,))
            ot = cursor.fetchone()

            if not ot:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"OT {ot_id} no encontrada"
                )

            # Obtener historial de managements
            cursor.execute("""
                SELECT
                    m.id,
                    ws.nombre as work_space,
                    s.nombre as state,
                    CONCAT(u.nombre, ' ', u.apellido) as user_name,
                    m.observacion,
                    m.created_at
                FROM managements m
                LEFT JOIN work_spaces ws ON m.work_space_id = ws.id
                LEFT JOIN states s ON m.state_id = s.id
                LEFT JOIN users u ON m.user_id = u.id
                WHERE m.work_order_id = %s
                ORDER BY m.created_at DESC
            """, (ot_id,))
            history_raw = cursor.fetchall()

            history = []
            for item in history_raw:
                history.append(ManagementHistoryItem(
                    id=item['id'],
                    work_space=item['work_space'] or 'N/A',
                    state=item['state'] or 'N/A',
                    user_name=item['user_name'] or 'Sistema',
                    observation=item['observacion'],
                    created_at=item['created_at'].isoformat() if item['created_at'] else ''
                ))

            return ManagementHistoryResponse(
                ot_id=ot_id,
                current_area=ot['area_actual'] or 'Sin asignar',
                current_state=ot['estado_actual'] or 'Sin estado',
                history=history
            )

    finally:
        connection.close()


@router.get("/{ot_id}/workflow-options", response_model=WorkflowOptions)
async def get_workflow_options(
    ot_id: int,
    current_user: dict = Depends(get_current_user_with_role)
):
    """
    Obtiene las opciones disponibles para transicion de una OT.
    Filtra management_types y estados basado en el rol del usuario y el area actual de la OT.

    Regla de negocio (de Laravel):
    - Si el area actual de la OT == work_space_id del rol del usuario: puede hacer Cambio de Estado
    - Si no: solo puede hacer Consulta y Archivo
    - Estados filtrados segun el area del usuario (replica comportamiento Laravel)
    """
    # Constantes de areas (work_space_id)
    AREA_VENTAS = 1
    AREA_DESARROLLO = 2
    AREA_DISENO = 3
    AREA_PRECATALOGACION = 4
    AREA_CATALOGACION = 5
    AREA_MUESTRAS = 6

    # IDs de management_types
    MGMT_CAMBIO_ESTADO = 1
    MGMT_CONSULTA = 2
    MGMT_ARCHIVO = 3

    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Obtener info de la OT incluyendo tipo_solicitud
            cursor.execute("""
                SELECT id, current_area_id, tipo_solicitud
                FROM work_orders WHERE id = %s AND active = 1
            """, (ot_id,))
            ot = cursor.fetchone()

            if not ot:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"OT {ot_id} no encontrada"
                )

            ot_current_area = ot.get('current_area_id') or AREA_VENTAS
            tipo_solicitud = ot.get('tipo_solicitud')

            # Obtener role y work_space_id del usuario
            role_id = current_user.get('role_id')

            cursor.execute("""
                SELECT r.id, r.nombre, r.work_space_id
                FROM roles r WHERE r.id = %s
            """, (role_id,))
            role = cursor.fetchone()

            user_work_space_id = role.get('work_space_id') if role else None

            # Determinar si el usuario tiene permisos para cambiar estado
            can_change_state = False

            if user_work_space_id is None:
                # Gerente o Admin (sin area asignada) puede todo
                can_change_state = True
            elif user_work_space_id == ot_current_area:
                # El area del rol coincide con el area actual de la OT
                can_change_state = True
            elif user_work_space_id == AREA_CATALOGACION and ot_current_area in [AREA_CATALOGACION, AREA_PRECATALOGACION]:
                # Catalogador tambien puede en Precatalogacion
                can_change_state = True

            # Filtrar management_types
            if can_change_state:
                allowed_mgmt_types = [MGMT_CAMBIO_ESTADO, MGMT_CONSULTA, MGMT_ARCHIVO]
            else:
                allowed_mgmt_types = [MGMT_CONSULTA, MGMT_ARCHIVO]

            # Obtener management_types filtrados
            placeholders = ','.join(['%s'] * len(allowed_mgmt_types))
            cursor.execute(f"""
                SELECT id, nombre FROM management_types
                WHERE id IN ({placeholders})
                ORDER BY id
            """, tuple(allowed_mgmt_types))
            management_types = cursor.fetchall()

            # Areas de trabajo
            cursor.execute("SELECT id, nombre FROM work_spaces WHERE status = 'active' ORDER BY id")
            areas = cursor.fetchall()

            # =============================================
            # FILTRAR ESTADOS SEGUN AREA DEL USUARIO (replica Laravel)
            # Estados:
            # 1=Proceso Ventas, 2=Diseño Estructural, 3=Laboratorio, 4=Muestra,
            # 5=Diseño Gráfico, 6=Cálculo Paletizado, 7=Catalogación, 8=Terminada,
            # 9=Perdida, 10=Consulta Cliente, 11=Anulada, 12=Rechazada, 13=Entregado,
            # 14=Espera OC, 15=Falta definición Cliente, 16=VB Cliente,
            # 17=Sala Muestras, 18=Muestras Listas, 20=Hibernación, 21=Cotización
            # =============================================
            states_by_area = []

            if user_work_space_id is None:
                # Gerente/Admin: todos los estados activos
                states_by_area = None  # Sin filtro
            elif user_work_space_id == AREA_VENTAS:
                # Vendedor: estados completos (caso general cuando ya fue enviado a desarrollo)
                # Laravel: [2, 5, 6, 7, 9, 10, 11, 14, 15, 16, 20, 21]
                states_by_area = [2, 5, 6, 7, 9, 10, 11, 14, 15, 16, 20, 21]
            elif user_work_space_id == AREA_DESARROLLO:
                # Ingeniero: [1, 3, 5, 6, 7, 12, 16, 17] (caso normal)
                # 1=Ventas, 3=Lab, 5=DG, 6=Calc, 7=Cat, 12=Rechazada, 16=VBC, 17=Sala
                states_by_area = [1, 3, 5, 6, 7, 12, 16, 17]
            elif user_work_space_id == AREA_DISENO:
                # Disenador: [1, 2, 7, 12, 16]
                # 1=Ventas, 2=DE, 7=Cat, 12=Rechazada, 16=VBC
                states_by_area = [1, 2, 7, 12, 16]
                # Si tipo_solicitud != 1 (no es Desarrollo Completo), agregar Entregado (13)
                if tipo_solicitud and tipo_solicitud != 1:
                    states_by_area.append(13)
            elif user_work_space_id == AREA_PRECATALOGACION:
                # Precatalogador: [1, 2, 5, 8, 12]
                # 1=Ventas, 2=DE, 5=DG, 8=Terminada, 12=Rechazada
                states_by_area = [1, 2, 5, 8, 12]
            elif user_work_space_id == AREA_CATALOGACION:
                # Catalogador: depende del area actual de la OT
                if ot_current_area == AREA_PRECATALOGACION:
                    # Si OT en Precatalogación: [1, 2, 5, 8, 12]
                    states_by_area = [1, 2, 5, 8, 12]
                else:
                    # Si OT en Catalogación: [1, 2, 5, 7, 12]
                    states_by_area = [1, 2, 5, 7, 12]
            elif user_work_space_id == AREA_MUESTRAS:
                # Tecnico Muestras: [12, 18, 22]
                # 12=Rechazada, 18=Muestras Listas, 22=Muestra Devuelta
                states_by_area = [12, 18, 22]
            else:
                # Otros: estados basicos
                states_by_area = [1, 2, 7, 12]

            # Obtener estados filtrados
            if states_by_area:
                placeholders = ','.join(['%s'] * len(states_by_area))
                cursor.execute(f"""
                    SELECT id, nombre, abreviatura FROM states
                    WHERE status = 'active' AND id IN ({placeholders})
                    ORDER BY id
                """, tuple(states_by_area))
            else:
                cursor.execute("SELECT id, nombre, abreviatura FROM states WHERE status = 'active' ORDER BY id")
            states = cursor.fetchall()

            # =============================================
            # Sprint N: Cargar motivos y proveedores
            # Fuente Laravel: ManagementController.php línea 287
            # Los motivos de rechazo están HARDCODEADOS en Laravel, no en tabla
            # =============================================

            # Motivos de rechazo (hardcoded según Laravel ManagementController.php)
            motives = [
                {"id": 1, "nombre": "Falta Información"},
                {"id": 2, "nombre": "Información Erronea"},
                {"id": 3, "nombre": "Falta Muestra Física"},
                {"id": 4, "nombre": "Formato Imagen Inadecuado"},
                {"id": 5, "nombre": "Medida Erronea"},
                {"id": 6, "nombre": "Descripción de Producto"},
                {"id": 7, "nombre": "Plano mal Acotado"},
                {"id": 8, "nombre": "Error de Digitación"},
                {"id": 9, "nombre": "Error tipo Sustrato"},
                {"id": 10, "nombre": "No viable por restricciones"},
                {"id": 11, "nombre": "Falta CAD para corte"},
                {"id": 12, "nombre": "Falta OT Chileexpress"},
                {"id": 13, "nombre": "Falta OT Laboratorio"},
            ]

            # Proveedores externos (tabla proveedors - opcional)
            proveedores = []
            try:
                cursor.execute("""
                    SELECT id, nombre
                    FROM proveedors
                    WHERE active = 1 OR active IS NULL
                    ORDER BY nombre
                """)
                proveedores = cursor.fetchall()
            except Exception:
                pass  # Tabla no existe, devolver lista vacía

            return WorkflowOptions(
                areas=areas,
                states=states,
                management_types=management_types,
                motives=motives,
                proveedores=proveedores
            )

    finally:
        connection.close()


@router.post("/{ot_id}/transition", response_model=TransitionResponse)
async def transition_work_order(
    ot_id: int,
    data: TransitionRequest,
    user_id: int = Depends(get_current_user_id)
):
    """
    Realiza una transicion de estado/area en una OT.
    Tipos de gestion:
    - 1: Cambio de Estado (requiere work_space_id y state_id)
    - 2: Consulta (solo registra observacion)
    - 3: Archivo (archiva la OT)

    Calcula duracion_segundos usando horas laborales (replica Laravel):
    - Obtiene la fecha del ultimo cambio (mayor entre ultimo_cambio_area y ultimo management)
    - Calcula horas laborales con WorkingHoursCalculator
    - Multiplica por 3600 para obtener segundos
    """
    # Constantes de tipos de gestion
    MGMT_CAMBIO_ESTADO = 1
    MGMT_CONSULTA = 2
    MGMT_ARCHIVO = 3

    # Estados que no acumulan tiempo (replica Laravel líneas 838-840)
    ESTADOS_SIN_TIEMPO = [8, 9, 11, 13, 20]  # Terminado, Perdido, Anulado, Entregado, Hibernación
    STATE_COTIZACION = 21  # Estado cotización también tiene tiempo 0

    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Verificar que la OT existe y obtener datos actuales incluyendo ultimo_cambio_area
            cursor.execute("""
                SELECT
                    wo.id,
                    wo.current_area_id,
                    wo.ultimo_cambio_area,
                    ws.nombre as area_actual,
                    s.nombre as estado_actual,
                    m.state_id as ultimo_state_id,
                    m.created_at as ultimo_mgmt_created
                FROM work_orders wo
                LEFT JOIN work_spaces ws ON wo.current_area_id = ws.id
                LEFT JOIN managements m ON wo.id = m.work_order_id
                LEFT JOIN states s ON m.state_id = s.id
                WHERE wo.id = %s AND wo.active = 1
                ORDER BY m.id DESC LIMIT 1
            """, (ot_id,))
            ot = cursor.fetchone()

            if not ot:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"OT {ot_id} no encontrada"
                )

            now = datetime.now()
            now_str = now.strftime('%Y-%m-%d %H:%M:%S')
            new_area_name = ot.get('area_actual') or 'N/A'
            new_state_name = ot.get('estado_actual') or 'N/A'
            message = ""

            # Calcular duracion_segundos usando horas laborales (replica Laravel)
            duracion_segundos = 0
            ultimo_state_id = ot.get('ultimo_state_id')

            # Determinar fecha de inicio para el calculo (replica Laravel líneas 813-824)
            ultimo_cambio_area = ot.get('ultimo_cambio_area')
            ultimo_mgmt_created = ot.get('ultimo_mgmt_created')

            if ultimo_cambio_area is None and ultimo_mgmt_created is None:
                # OT nueva sin historial, tiempo = 0
                fecha_inicio = now
            elif ultimo_cambio_area is None:
                fecha_inicio = ultimo_mgmt_created
            elif ultimo_mgmt_created is None:
                fecha_inicio = ultimo_cambio_area
            elif ultimo_cambio_area > ultimo_mgmt_created:
                fecha_inicio = ultimo_cambio_area
            else:
                fecha_inicio = ultimo_mgmt_created

            # Caso especial: si ultimo estado es Cotización, usar ultimo_cambio_area
            if ultimo_state_id == STATE_COTIZACION and ultimo_cambio_area:
                fecha_inicio = ultimo_cambio_area

            # Calcular horas laborales y convertir a segundos
            if fecha_inicio and fecha_inicio < now:
                try:
                    calculator = WorkingHoursCalculator()
                    horas_laborales = calculator.get_working_hours(fecha_inicio, now)
                    duracion_segundos = int(horas_laborales * 3600)
                except Exception as e:
                    # Si falla el calculo, usar diferencia simple en segundos
                    duracion_segundos = int((now - fecha_inicio).total_seconds())

            # Procesar segun tipo de gestion
            if data.management_type_id == MGMT_CAMBIO_ESTADO:
                # Validar que state_id sea requerido
                if not data.state_id:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Para Cambio de Estado se requiere state_id"
                    )

                # Constantes de estados especiales
                STATE_SALA_MUESTRAS = 17
                STATE_MUESTRAS_LISTAS = 18
                STATE_RECHAZADA = 12
                AREA_MUESTRAS = 6
                AREA_DESARROLLO = 2

                # Determinar work_space_id automaticamente si no se proporciona
                effective_work_space_id = data.work_space_id

                # Mapeo de estados a areas (replica logica Laravel)
                state_to_area_map = {
                    1: 1,   # Proceso de Ventas -> Area Ventas
                    2: 2,   # Proceso de Diseño Estructural -> Area Desarrollo
                    3: 2,   # Laboratorio -> Area Desarrollo
                    5: 3,   # Proceso de Diseño Grafico -> Area Diseño
                    6: 2,   # Proceso de Calculo Paletizado -> Area Desarrollo
                    7: 5,   # Proceso de Catalogacion -> Area Catalogacion
                    17: 6,  # Sala de Muestras -> Area Muestras
                    18: 2,  # Muestras Listas -> Area Desarrollo
                }

                if not effective_work_space_id:
                    # Si no se proporciono area, obtenerla del mapeo o de la tabla states
                    if data.state_id in state_to_area_map:
                        effective_work_space_id = state_to_area_map[data.state_id]
                    else:
                        # Intentar obtener work_space_id de la tabla states
                        cursor.execute("SELECT work_space_id FROM states WHERE id = %s", (data.state_id,))
                        state_area = cursor.fetchone()
                        if state_area and state_area.get('work_space_id'):
                            effective_work_space_id = state_area['work_space_id']
                        else:
                            # Usar el area actual de la OT como fallback
                            effective_work_space_id = ot.get('current_area_id') or 1

                # Logica especial para Sala de Muestras (state_id = 17)
                if data.state_id == STATE_SALA_MUESTRAS:
                    # Verificar que existan muestras en la OT
                    cursor.execute("SELECT COUNT(*) as count FROM muestras WHERE work_order_id = %s", (ot_id,))
                    muestras_count = cursor.fetchone()
                    if not muestras_count or muestras_count['count'] == 0:
                        raise HTTPException(
                            status_code=status.HTTP_400_BAD_REQUEST,
                            detail="Debes ingresar al menos una muestra para enviar a Sala de Muestras"
                        )

                    # Forzar area = 6 (Sala de Muestras)
                    effective_work_space_id = AREA_MUESTRAS

                # Sprint N: Preparar motive_id y consulted_work_space_id para rechazos
                # Fuente Laravel: ManagementController.php línea 847-850
                motive_id = data.motive_id if data.state_id == STATE_RECHAZADA else None
                consulted_work_space_id = effective_work_space_id if data.state_id == STATE_RECHAZADA else None

                # Calcular duracion_segundos final (replica Laravel líneas 827-840)
                # Si el nuevo estado es Cotización (21), tiempo = 0
                # Si el estado anterior es uno de los estados finales, tiempo = 0
                duracion_final = duracion_segundos
                if data.state_id == STATE_COTIZACION:
                    duracion_final = 0
                if ultimo_state_id and ultimo_state_id in ESTADOS_SIN_TIEMPO:
                    duracion_final = 0

                # IMPORTANTE: work_space_id es el área ANTERIOR (donde estaba la OT antes)
                # NO el área destino. Replica Laravel línea 806: $gestion->work_space_id = $ot->current_area_id;
                area_anterior = ot.get('current_area_id') or 1

                # Crear registro de management con duracion_segundos
                cursor.execute("""
                    INSERT INTO managements
                    (work_order_id, work_space_id, state_id, management_type_id, user_id, observacion,
                     motive_id, consulted_work_space_id, duracion_segundos, mostrar, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 1, %s, %s)
                """, (ot_id, area_anterior, data.state_id, MGMT_CAMBIO_ESTADO, user_id,
                      data.observation, motive_id, consulted_work_space_id, duracion_final, now_str, now_str))

                # Actualizar current_area_id de la OT
                cursor.execute("""
                    UPDATE work_orders SET current_area_id = %s, ultimo_cambio_area = %s, updated_at = %s WHERE id = %s
                """, (effective_work_space_id, now_str, now_str, ot_id))

                # Logica especial post-transicion para Sala de Muestras
                if data.state_id == STATE_SALA_MUESTRAS:
                    # Actualizar muestras a estado "En Proceso" (1) e iniciar tiempo en sala
                    cursor.execute("""
                        UPDATE muestras
                        SET estado = 1, inicio_sala_corte = %s, updated_at = %s
                        WHERE work_order_id = %s AND estado IN (0, 1)
                    """, (now_str, now_str, ot_id))

                # Obtener nombres para respuesta
                cursor.execute("SELECT nombre FROM work_spaces WHERE id = %s", (effective_work_space_id,))
                area_row = cursor.fetchone()
                cursor.execute("SELECT nombre FROM states WHERE id = %s", (data.state_id,))
                state_row = cursor.fetchone()

                new_area_name = area_row['nombre'] if area_row else 'N/A'
                new_state_name = state_row['nombre'] if state_row else 'N/A'
                message = f"OT {ot_id} transicionada exitosamente"

            elif data.management_type_id == MGMT_CONSULTA:
                # Solo registrar consulta (sin cambiar estado)
                cursor.execute("""
                    INSERT INTO managements
                    (work_order_id, work_space_id, state_id, management_type_id, user_id, observacion, created_at, updated_at)
                    VALUES (%s, %s, NULL, %s, %s, %s, %s, %s)
                """, (ot_id, ot.get('current_area_id'), MGMT_CONSULTA, user_id, data.observation, now_str, now_str))

                message = f"Consulta registrada en OT {ot_id}"

            elif data.management_type_id == MGMT_ARCHIVO:
                # Archivar OT - Solo registrar gestión (igual que Laravel ManagementController.php)
                # Laravel no tiene columna 'archived', solo guarda el registro de management
                cursor.execute("""
                    INSERT INTO managements
                    (work_order_id, work_space_id, state_id, management_type_id, user_id, observacion, created_at, updated_at)
                    VALUES (%s, %s, NULL, %s, %s, %s, %s, %s)
                """, (ot_id, ot.get('current_area_id'), MGMT_ARCHIVO, user_id, data.observation or 'OT archivada', now_str, now_str))

                message = f"OT {ot_id} archivada exitosamente"

            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Tipo de gestion {data.management_type_id} no valido"
                )

            connection.commit()

            # Sprint M: Enviar notificación push al creador si corresponde
            # Fuente: ManagementController.php líneas 862, 977, 1174, 1183
            try:
                AREA_VENTAS = 1
                STATE_RECHAZADA = 12

                # Obtener token del creador de la OT
                cursor.execute("""
                    SELECT u.token_push_mobile, u.id as creador_id
                    FROM work_orders wo
                    JOIN users u ON wo.creador_id = u.id
                    WHERE wo.id = %s
                """, (ot_id,))
                creador = cursor.fetchone()

                if creador and creador.get('token_push_mobile'):
                    token = creador['token_push_mobile']
                    fcm_service = get_fcm_service()

                    # Determinar tipo de notificación según flujo
                    if data.management_type_id == MGMT_CAMBIO_ESTADO:
                        # Si va a área de Ventas (devolución o rechazo)
                        if effective_work_space_id == AREA_VENTAS:
                            if data.state_id == STATE_RECHAZADA:
                                # Rechazo
                                fcm_service.notificar_rechazo(ot_id, token)
                            else:
                                # Devolución
                                fcm_service.notificar_devolucion(ot_id, token)

                    elif data.management_type_id == MGMT_CONSULTA:
                        # Consulta al área de Ventas
                        if ot.get('current_area_id') == AREA_VENTAS:
                            fcm_service.notificar_consulta(ot_id, token)

            except Exception as e_fcm:
                # No fallar la transición si FCM falla
                print(f"Error al enviar notificación FCM: {e_fcm}")

            return TransitionResponse(
                id=ot_id,
                message=message,
                new_area=new_area_name,
                new_state=new_state_name
            )

    except HTTPException:
        connection.rollback()
        raise
    except pymysql.Error as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al transicionar OT: {str(e)}"
        )
    finally:
        connection.close()


# =============================================
# APROBACIÓN DE OTs
# =============================================

class ApprovalListItem(BaseModel):
    """Item de la lista de OTs pendientes de aprobación."""
    id: int
    created_at: str
    client_name: str
    descripcion: str
    canal: Optional[str]
    item_tipo: Optional[str]
    estado: str
    estado_abrev: str
    creador_nombre: str


class ApprovalListResponse(BaseModel):
    """Respuesta paginada de OTs pendientes de aprobación."""
    items: List[ApprovalListItem]
    total: int
    page: int
    page_size: int
    total_pages: int


class ApprovalActionResponse(BaseModel):
    """Respuesta de acción de aprobar/rechazar."""
    id: int
    message: str
    new_state: str


async def _get_pending_approval_impl(
    page: int,
    page_size: int,
    user_id: int,
    role_id: int
):
    """
    Implementación: Lista OTs pendientes de aprobación según rol.

    Lógica de Laravel (WorkOrderController.php líneas 10555-10561):
    - Jefe de Venta (role_id=3): OTs donde aprobacion_jefe_venta = 1
    - Jefe de Desarrollo (role_id=5): OTs donde aprobacion_jefe_venta = 2 AND aprobacion_jefe_desarrollo = 1

    Estados de aprobación:
    - 0: Sin asignar
    - 1: Pendiente de aprobación
    - 2: Aprobado
    - 3: Rechazado
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Construir WHERE según rol (igual que Laravel)
            # role_id 3 = Jefe de Venta
            # role_id 5 = Jefe de Desarrollo
            if role_id == 3:  # Jefe de Venta
                where_clause = "wo.aprobacion_jefe_venta = 1"
            elif role_id == 5:  # Jefe de Desarrollo
                where_clause = "wo.aprobacion_jefe_venta = 2 AND wo.aprobacion_jefe_desarrollo = 1"
            else:
                # Otros roles: mostrar todas las pendientes (admin, etc)
                where_clause = "(wo.aprobacion_jefe_venta = 1 OR (wo.aprobacion_jefe_venta = 2 AND wo.aprobacion_jefe_desarrollo = 1))"

            # Contar total
            count_sql = f"""
                SELECT COUNT(*) as total
                FROM work_orders wo
                WHERE wo.active = 1
                AND {where_clause}
            """
            cursor.execute(count_sql)
            total = cursor.fetchone()['total']

            # Calcular paginación
            total_pages = (total + page_size - 1) // page_size if total > 0 else 0
            offset = (page - 1) * page_size

            # Obtener OTs pendientes de aprobación
            sql = f"""
                SELECT
                    wo.id,
                    wo.created_at,
                    wo.descripcion,
                    c.nombre as client_name,
                    CONCAT(u.nombre, ' ', u.apellido) as creador_nombre,
                    ch.nombre as canal,
                    pt.descripcion as item_tipo,
                    COALESCE(s.nombre, 'Proceso de Ventas') as estado,
                    COALESCE(s.abreviatura, 'PV') as estado_abrev,
                    wo.aprobacion_jefe_venta,
                    wo.aprobacion_jefe_desarrollo
                FROM work_orders wo
                LEFT JOIN clients c ON wo.client_id = c.id
                LEFT JOIN users u ON wo.creador_id = u.id
                LEFT JOIN canals ch ON wo.canal_id = ch.id
                LEFT JOIN product_types pt ON wo.product_type_id = pt.id
                LEFT JOIN (
                    SELECT m1.work_order_id, m1.state_id
                    FROM managements m1
                    INNER JOIN (
                        SELECT work_order_id, MAX(created_at) as max_date
                        FROM managements
                        GROUP BY work_order_id
                    ) m2 ON m1.work_order_id = m2.work_order_id AND m1.created_at = m2.max_date
                ) last_mgmt ON wo.id = last_mgmt.work_order_id
                LEFT JOIN states s ON last_mgmt.state_id = s.id
                WHERE wo.active = 1
                AND {where_clause}
                ORDER BY wo.created_at DESC
                LIMIT %s OFFSET %s
            """
            cursor.execute(sql, (page_size, offset))
            rows = cursor.fetchall()

            items = []
            for row in rows:
                items.append(ApprovalListItem(
                    id=row['id'],
                    created_at=row['created_at'].strftime('%Y-%m-%d') if row['created_at'] else '',
                    client_name=row['client_name'] or 'Sin cliente',
                    descripcion=row['descripcion'] or '',
                    canal=row['canal'],
                    item_tipo=row['item_tipo'],
                    estado=row['estado'],
                    estado_abrev=row['estado_abrev'],
                    creador_nombre=row['creador_nombre'] or 'Sin creador'
                ))

            return ApprovalListResponse(
                items=items,
                total=total,
                page=page,
                page_size=page_size,
                total_pages=total_pages
            )

    finally:
        connection.close()


@router.put("/{ot_id}/approve", response_model=ApprovalActionResponse)
async def approve_work_order(
    ot_id: int,
    current_user: dict = Depends(get_current_user_with_role)
):
    """
    Aprueba una OT pendiente según el rol del usuario.

    Lógica de Laravel (WorkOrderController.php líneas 10567-10574):
    - Jefe de Venta (role_id=3): actualiza aprobacion_jefe_venta = 2
    - Jefe de Desarrollo (role_id=5): actualiza aprobacion_jefe_desarrollo = 2
    """
    user_id = current_user["user_id"]
    role_id = current_user["role_id"]

    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Verificar que la OT existe
            cursor.execute("""
                SELECT id, aprobacion_jefe_venta, aprobacion_jefe_desarrollo
                FROM work_orders
                WHERE id = %s AND active = 1
            """, (ot_id,))
            ot = cursor.fetchone()

            if not ot:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"OT {ot_id} no encontrada"
                )

            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Actualizar según rol (igual que Laravel)
            if role_id == 3:  # Jefe de Venta
                if ot['aprobacion_jefe_venta'] == 2:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"OT {ot_id} ya fue aprobada por Jefe de Venta"
                    )
                cursor.execute("""
                    UPDATE work_orders
                    SET aprobacion_jefe_venta = 2, updated_at = %s
                    WHERE id = %s
                """, (now, ot_id))
                approval_type = "Jefe de Venta"

            elif role_id == 5:  # Jefe de Desarrollo
                if ot['aprobacion_jefe_desarrollo'] == 2:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"OT {ot_id} ya fue aprobada por Jefe de Desarrollo"
                    )
                cursor.execute("""
                    UPDATE work_orders
                    SET aprobacion_jefe_desarrollo = 2, updated_at = %s
                    WHERE id = %s
                """, (now, ot_id))
                approval_type = "Jefe de Desarrollo"
            else:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="No tiene permisos para aprobar OTs"
                )

            connection.commit()

            return ApprovalActionResponse(
                id=ot_id,
                message=f"OT {ot_id} aprobada por {approval_type}",
                new_state="Aprobado"
            )

    except pymysql.Error as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al aprobar OT: {str(e)}"
        )
    finally:
        connection.close()


@router.put("/{ot_id}/reject", response_model=ApprovalActionResponse)
async def reject_work_order(
    ot_id: int,
    current_user: dict = Depends(get_current_user_with_role)
):
    """
    Rechaza una OT pendiente según el rol del usuario.

    Lógica de Laravel (WorkOrderController.php líneas 10577-10584):
    - Jefe de Venta (role_id=3): actualiza aprobacion_jefe_venta = 3
    - Jefe de Desarrollo (role_id=5): actualiza aprobacion_jefe_desarrollo = 3
    """
    user_id = current_user["user_id"]
    role_id = current_user["role_id"]

    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Verificar que la OT existe
            cursor.execute("""
                SELECT id, aprobacion_jefe_venta, aprobacion_jefe_desarrollo
                FROM work_orders
                WHERE id = %s AND active = 1
            """, (ot_id,))
            ot = cursor.fetchone()

            if not ot:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"OT {ot_id} no encontrada"
                )

            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Actualizar según rol (igual que Laravel)
            if role_id == 3:  # Jefe de Venta
                cursor.execute("""
                    UPDATE work_orders
                    SET aprobacion_jefe_venta = 3, updated_at = %s
                    WHERE id = %s
                """, (now, ot_id))
                rejection_type = "Jefe de Venta"

            elif role_id == 5:  # Jefe de Desarrollo
                cursor.execute("""
                    UPDATE work_orders
                    SET aprobacion_jefe_desarrollo = 3, updated_at = %s
                    WHERE id = %s
                """, (now, ot_id))
                rejection_type = "Jefe de Desarrollo"
            else:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="No tiene permisos para rechazar OTs"
                )

            connection.commit()

            return ApprovalActionResponse(
                id=ot_id,
                message=f"OT {ot_id} rechazada por {rejection_type}",
                new_state="Rechazado"
            )

    except pymysql.Error as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al rechazar OT: {str(e)}"
        )
    finally:
        connection.close()


# =============================================
# ASIGNACIONES DE OTs
# =============================================

class AssignmentListItem(BaseModel):
    """Item de la lista de OTs pendientes de asignacion."""
    id: int
    created_at: str
    client_name: str
    vendedor_nombre: str
    tipo_solicitud: str
    canal: Optional[str]
    jerarquia_1: Optional[str]
    jerarquia_2: Optional[str]
    jerarquia_3: Optional[str]
    cad: Optional[str]
    profesional_asignado: Optional[str]
    dias_sin_asignar: int


class AssignmentListResponse(BaseModel):
    """Respuesta de lista de OTs pendientes de asignacion."""
    items: List[AssignmentListItem]
    total: int
    page: int
    page_size: int
    total_pages: int


class AssignProfessionalRequest(BaseModel):
    """Request para asignar profesional a OT."""
    profesional_id: int
    area_id: Optional[int] = None  # Area del asignador (opcional, se detecta automaticamente)
    observacion: Optional[str] = None  # Mensaje opcional para la asignacion
    generar_notificacion: bool = True  # Si se debe generar notificacion


class AssignmentActionResponse(BaseModel):
    """Respuesta de accion de asignacion."""
    id: int
    message: str
    profesional_nombre: str
    es_reasignacion: bool = False
    notificacion_creada: bool = False


async def _get_pending_assignment_impl(
    page: int,
    page_size: int,
    asignado: Optional[str],
    tipo_solicitud: Optional[str],
    canal_id: Optional[int],
    vendedor_id: Optional[int],
    estado_id: Optional[int],
    date_desde: Optional[str],
    date_hasta: Optional[str],
    user_id: int,
):
    """
    Implementación: Lista OTs pendientes de asignacion.
    Usado por jefes de area para asignar profesionales.
    Replica logica de Laravel: filtra por current_area_id y user_work_orders.
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Obtener el area (work_space_id) del usuario actual
            cursor.execute("""
                SELECT r.work_space_id as area_id
                FROM users u
                JOIN roles r ON u.role_id = r.id
                WHERE u.id = %s
            """, [user_id])
            user_row = cursor.fetchone()
            if not user_row or not user_row.get("area_id"):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Usuario no tiene area asignada"
                )
            user_area_id = user_row["area_id"]

            # Estados activos por defecto (igual que Laravel)
            estados_activos = [1, 2, 3, 4, 5, 6, 7, 10, 12, 14, 15, 16, 17, 18]

            # Query base - obtiene profesional asignado desde user_work_orders
            query = """
                SELECT
                    ot.id,
                    ot.created_at,
                    c.nombre as client_name,
                    CONCAT(u.nombre, ' ', u.apellido) as vendedor_nombre,
                    ot.tipo_solicitud,
                    cn.nombre as canal,
                    h1.descripcion as jerarquia_1,
                    h2.descripcion as jerarquia_2,
                    h3.descripcion as jerarquia_3,
                    ot.cad,
                    (
                        SELECT CONCAT(pu.nombre, ' ', pu.apellido)
                        FROM user_work_orders uwo
                        JOIN users pu ON uwo.user_id = pu.id
                        WHERE uwo.work_order_id = ot.id AND uwo.area_id = %s AND uwo.active = 1
                        LIMIT 1
                    ) as profesional_asignado,
                    DATEDIFF(NOW(), ot.created_at) as dias_sin_asignar
                FROM work_orders ot
                LEFT JOIN clients c ON ot.client_id = c.id
                LEFT JOIN users u ON ot.creador_id = u.id
                LEFT JOIN canals cn ON ot.canal_id = cn.id
                LEFT JOIN subsubhierarchies h3 ON ot.subsubhierarchy_id = h3.id
                LEFT JOIN subhierarchies h2 ON h3.subhierarchy_id = h2.id
                LEFT JOIN hierarchies h1 ON h2.hierarchy_id = h1.id
                WHERE ot.active = 1
            """
            params = [user_area_id]

            # Filtro de asignado basado en user_work_orders
            if asignado == "SI":
                # OTs que YA tienen asignacion en el area del jefe
                query += """
                    AND EXISTS (
                        SELECT 1 FROM user_work_orders uwo
                        WHERE uwo.work_order_id = ot.id
                        AND uwo.area_id = %s
                        AND uwo.active = 1
                    )
                """
                params.append(user_area_id)
            else:
                # Por defecto (NO o null): OTs SIN asignacion en el area Y que estan en el area
                # Caso especial: Catalogadores manejan areas 4 y 5
                if user_area_id in [4, 5]:
                    query += """
                        AND NOT EXISTS (
                            SELECT 1 FROM user_work_orders uwo
                            WHERE uwo.work_order_id = ot.id
                            AND uwo.area_id = %s
                            AND uwo.active = 1
                        )
                        AND ot.current_area_id IN (4, 5)
                    """
                    params.append(user_area_id)  # Para NOT EXISTS
                else:
                    query += """
                        AND NOT EXISTS (
                            SELECT 1 FROM user_work_orders uwo
                            WHERE uwo.work_order_id = ot.id
                            AND uwo.area_id = %s
                            AND uwo.active = 1
                        )
                        AND ot.current_area_id = %s
                    """
                    params.append(user_area_id)  # Para NOT EXISTS
                    params.append(user_area_id)  # Para current_area_id

            if tipo_solicitud:
                query += " AND ot.tipo_solicitud = %s"
                params.append(tipo_solicitud)

            if canal_id:
                query += " AND ot.canal_id = %s"
                params.append(canal_id)

            if vendedor_id:
                query += " AND ot.creador_id = %s"
                params.append(vendedor_id)

            # Filtro por estado usando managements (igual que Laravel)
            if estado_id:
                query += """
                    AND EXISTS (
                        SELECT 1 FROM managements m
                        WHERE m.work_order_id = ot.id
                        AND m.management_type_id = 1
                        AND m.state_id = %s
                        AND m.id = (
                            SELECT MAX(m2.id) FROM managements m2
                            WHERE m2.work_order_id = ot.id AND m2.management_type_id = 1
                        )
                    )
                """
                params.append(estado_id)
            else:
                # Por defecto filtra por estados activos
                estados_placeholder = ','.join(['%s'] * len(estados_activos))
                query += f"""
                    AND EXISTS (
                        SELECT 1 FROM managements m
                        WHERE m.work_order_id = ot.id
                        AND m.management_type_id = 1
                        AND m.state_id IN ({estados_placeholder})
                        AND m.id = (
                            SELECT MAX(m2.id) FROM managements m2
                            WHERE m2.work_order_id = ot.id AND m2.management_type_id = 1
                        )
                    )
                """
                params.extend(estados_activos)

            if date_desde:
                query += " AND ot.created_at >= %s"
                params.append(date_desde)

            if date_hasta:
                query += " AND ot.created_at <= %s"
                params.append(date_hasta + " 23:59:59")

            # Contar total
            count_query = f"SELECT COUNT(*) as total FROM ({query}) as subq"
            cursor.execute(count_query, params)
            total = cursor.fetchone()["total"]

            # Paginacion - ordenar por más antigua primero (ASC) como en Laravel
            query += " ORDER BY ot.created_at ASC LIMIT %s OFFSET %s"
            offset = (page - 1) * page_size
            params.extend([page_size, offset])

            cursor.execute(query, params)
            rows = cursor.fetchall()

            # Mapeo de tipo_solicitud
            tipo_map = {
                1: "Desarrollo Nuevo",
                2: "Repeticion",
                3: "Modificacion",
                4: "Cotizacion",
            }

            items = []
            for row in rows:
                items.append(AssignmentListItem(
                    id=row["id"],
                    created_at=row["created_at"].strftime("%Y-%m-%d %H:%M:%S") if row["created_at"] else "",
                    client_name=row["client_name"] or "",
                    vendedor_nombre=row["vendedor_nombre"] or "",
                    tipo_solicitud=tipo_map.get(row["tipo_solicitud"], "Desconocido") if row["tipo_solicitud"] else "-",
                    canal=row["canal"],
                    jerarquia_1=row["jerarquia_1"],
                    jerarquia_2=row["jerarquia_2"],
                    jerarquia_3=row["jerarquia_3"],
                    cad=row["cad"],
                    profesional_asignado=row["profesional_asignado"],
                    dias_sin_asignar=row["dias_sin_asignar"] or 0
                ))

            total_pages = (total + page_size - 1) // page_size

            return AssignmentListResponse(
                items=items,
                total=total,
                page=page,
                page_size=page_size,
                total_pages=total_pages
            )

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener OTs pendientes de asignacion: {str(e)}"
        )
    finally:
        connection.close()


@router.put("/{ot_id}/assign", response_model=AssignmentActionResponse)
async def assign_professional(
    ot_id: int,
    data: AssignProfessionalRequest,
    user_id: int = Depends(get_current_user_id)
):
    """
    Asigna un profesional a una OT.
    - Registra en user_work_orders la asignacion
    - Genera notificacion si corresponde
    - Soporta observacion/mensaje
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Verificar que la OT existe
            cursor.execute(
                "SELECT id, ultimo_cambio_area FROM work_orders WHERE id = %s AND active = 1",
                (ot_id,)
            )
            ot = cursor.fetchone()
            if not ot:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"OT {ot_id} no encontrada"
                )

            # Verificar que el profesional existe
            cursor.execute(
                "SELECT id, nombre, apellido FROM users WHERE id = %s",
                (data.profesional_id,)
            )
            profesional = cursor.fetchone()
            if not profesional:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Profesional {data.profesional_id} no encontrado"
                )

            # Obtener area del usuario que asigna (work_space_id del rol)
            area_id = data.area_id
            if not area_id:
                cursor.execute(
                    "SELECT r.work_space_id as area_id FROM users u JOIN roles r ON u.role_id = r.id WHERE u.id = %s",
                    (user_id,)
                )
                area_row = cursor.fetchone()
                if area_row and area_row.get("area_id"):
                    area_id = area_row["area_id"]
                else:
                    area_id = 1  # Default

            # Verificar si ya existe asignacion para esta OT y area
            cursor.execute(
                "SELECT id, user_id FROM user_work_orders WHERE work_order_id = %s AND area_id = %s",
                (ot_id, area_id)
            )
            existing = cursor.fetchone()

            es_reasignacion = False
            motivo = "Asignado"

            if existing:
                # Reasignacion - actualizar registro existente
                cursor.execute(
                    """UPDATE user_work_orders
                       SET user_id = %s, updated_at = NOW()
                       WHERE id = %s""",
                    (data.profesional_id, existing["id"])
                )
                es_reasignacion = True
                motivo = "Reasignado"
            else:
                # Nueva asignacion - calcular tiempo inicial
                tiempo_inicial = 0
                if ot.get("ultimo_cambio_area"):
                    cursor.execute(
                        "SELECT TIMESTAMPDIFF(SECOND, %s, NOW()) as diff",
                        (ot["ultimo_cambio_area"],)
                    )
                    diff_row = cursor.fetchone()
                    if diff_row:
                        tiempo_inicial = diff_row["diff"] or 0

                cursor.execute(
                    """INSERT INTO user_work_orders (work_order_id, user_id, area_id, tiempo_inicial, created_at, updated_at)
                       VALUES (%s, %s, %s, %s, NOW(), NOW())""",
                    (ot_id, data.profesional_id, area_id, tiempo_inicial)
                )

            # Generar notificacion si corresponde
            notificacion_creada = False
            if data.generar_notificacion and user_id != data.profesional_id:
                cursor.execute(
                    """INSERT INTO notifications (work_order_id, user_id, generador_id, motivo, observacion, active, created_at, updated_at)
                       VALUES (%s, %s, %s, %s, %s, 1, NOW(), NOW())""",
                    (ot_id, data.profesional_id, user_id, motivo, data.observacion or "")
                )
                notificacion_creada = True

            connection.commit()

            profesional_nombre = f"{profesional['nombre']} {profesional['apellido']}"

            return AssignmentActionResponse(
                id=ot_id,
                message=f"{'Reasignado' if es_reasignacion else 'Asignado'} a {profesional_nombre}",
                profesional_nombre=profesional_nombre,
                es_reasignacion=es_reasignacion,
                notificacion_creada=notificacion_creada
            )

    except pymysql.Error as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al asignar profesional: {str(e)}"
        )
    finally:
        connection.close()


async def _get_professionals_impl(user_id: int):
    """
    Implementacion: Lista profesionales disponibles para asignacion.
    Replica logica de Laravel: subordinados (role_id + 1) + jefe actual.
    Ejemplo: Jefe Diseño (7) ve Diseñadores (8) + él mismo.
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Obtener el role_id del usuario actual
            cursor.execute(
                "SELECT role_id FROM users WHERE id = %s",
                [user_id]
            )
            user_row = cursor.fetchone()
            if not user_row:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Usuario no encontrado"
                )
            current_role_id = user_row["role_id"]
            subordinate_role_id = current_role_id + 1

            # Obtener profesionales: subordinados (role_id + 1) + el jefe actual
            cursor.execute("""
                SELECT u.id, u.nombre, u.apellido, r.nombre as rol
                FROM users u
                LEFT JOIN roles r ON u.role_id = r.id
                WHERE u.active = 1
                AND (u.role_id = %s OR u.id = %s)
                ORDER BY u.nombre, u.apellido
            """, [subordinate_role_id, user_id])
            rows = cursor.fetchall()

            return [
                {
                    "id": row["id"],
                    "nombre": f"{row['nombre']} {row['apellido']}",
                    "rol": row["rol"]
                }
                for row in rows
            ]

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener profesionales: {str(e)}"
        )
    finally:
        connection.close()


# =============================================
# OPCIONES DEL FORMULARIO OT (COMPLETO)
# =============================================

class FormOptionsComplete(BaseModel):
    """Todas las opciones necesarias para el formulario de crear/editar OT."""
    # Catálogos principales
    clients: List[dict]
    canals: List[dict]
    vendedores: List[dict]
    org_ventas: List[dict]
    plantas: List[dict]
    # Catálogos de producto
    product_types: List[dict]
    cads: List[dict]
    cartons: List[dict]
    styles: List[dict]
    colors: List[dict]
    envases: List[dict]
    matrices: List[dict]  # Matrices para Sección 7
    # Catálogos de procesos
    processes: List[dict]
    armados: List[dict]
    pegados: List[dict]
    sentidos_armado: List[dict]
    impresiones: List[dict]
    fsc: List[dict]
    # Catálogos de materiales
    materials: List[dict]
    recubrimientos: List[dict]
    coverages_internal: List[dict]
    coverages_external: List[dict]
    # Catálogos de referencia
    reference_types: List[dict]
    design_types: List[dict]
    bloqueo_referencia: List[dict]  # Hardcoded como en Laravel
    indicador_facturacion: List[dict]  # Hardcoded como en Laravel
    # Catálogos de calidad
    trazabilidad: List[dict]
    tipo_cinta: List[dict]
    pallet_types: List[dict]
    salas_corte: List[dict]
    pallet_qas: List[dict]  # Certificado de Calidad
    pallet_tag_formats: List[dict]  # Formato Etiqueta Pallet
    # Jerarquías
    hierarchies: List[dict]
    subhierarchies: List[dict]
    subsubhierarchies: List[dict]
    # Otros
    tipos_solicitud: List[dict]
    maquila_servicios: List[dict]
    comunas: List[dict]
    pais_referencia: List[dict]
    # Configuración
    secuencia_operacional: List[dict]
    # Sección 13 - Datos para Desarrollo
    product_type_developing: List[dict]
    food_types: List[dict]
    expected_uses: List[dict]
    recycled_uses: List[dict]
    # Opciones para Muestras
    tipos_pegado_muestra: List[dict]
    cartons_muestra: List[dict]
    destinatarios_muestra: List[dict]
    class_substance_packeds: List[dict]
    transportation_ways: List[dict]
    target_markets: List[dict]


async def _get_form_options_complete_impl() -> FormOptionsComplete:
    """
    Implementación de form-options-complete.
    La ruta está definida más arriba para evitar conflicto con /{ot_id}.
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            options = {}

            # ========== CATÁLOGOS PRINCIPALES ==========

            # Clientes activos - ordenados por codigo
            cursor.execute("""
                SELECT id, nombre, codigo, rut
                FROM clients
                WHERE active = 1
                ORDER BY codigo ASC
                LIMIT 2000
            """)
            options['clients'] = cursor.fetchall()

            # Canales
            cursor.execute("SELECT id, nombre FROM canals WHERE active = 1 ORDER BY nombre")
            options['canals'] = cursor.fetchall()

            # Vendedores (usuarios con roles de venta)
            cursor.execute("""
                SELECT id, CONCAT(nombre, ' ', apellido) as nombre
                FROM users
                WHERE active = 1 AND role_id IN (3, 4, 19)
                ORDER BY nombre
            """)
            options['vendedores'] = cursor.fetchall()

            # Organizaciones de venta (tabla no existe - hardcoded)
            options['org_ventas'] = [
                {"id": 1, "nombre": "2020 OV NACIONAL PLANTA PLACILLA"},
                {"id": 2, "nombre": "2000 OV EXPORTACIONES"}
            ]

            # Plantas (no tiene columna codigo)
            cursor.execute("SELECT id, nombre FROM plantas ORDER BY nombre")
            options['plantas'] = cursor.fetchall()

            # ========== CATÁLOGOS DE PRODUCTO ==========

            # Tipos de producto
            cursor.execute("""
                SELECT id, descripcion as nombre, codigo
                FROM product_types
                WHERE active = 1
                ORDER BY descripcion
            """)
            options['product_types'] = cursor.fetchall()

            # CADs (no tiene columna descripcion)
            cursor.execute("""
                SELECT id, cad as codigo
                FROM cads
                WHERE active = 1
                ORDER BY cad
            """)
            options['cads'] = cursor.fetchall()

            # Cartones (no tiene columna descripcion, usar codigo y onda)
            cursor.execute("""
                SELECT id, codigo, onda
                FROM cartons
                WHERE active = 1
                ORDER BY codigo
            """)
            options['cartons'] = cursor.fetchall()

            # Estilos
            cursor.execute("""
                SELECT id, glosa as nombre, codigo
                FROM styles
                WHERE active = 1
                ORDER BY glosa
            """)
            options['styles'] = cursor.fetchall()

            # Colores (usa descripcion, no nombre)
            cursor.execute("""
                SELECT id, descripcion as nombre, codigo
                FROM colors
                WHERE active = 1
                ORDER BY descripcion
            """)
            options['colors'] = cursor.fetchall()

            # Envases
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM envases
                WHERE active = 1
                ORDER BY descripcion
            """)
            options['envases'] = cursor.fetchall()

            # Matrices (para Sección 7)
            cursor.execute("""
                SELECT id, plano_cad as nombre, tipo_matriz
                FROM matrices
                WHERE active = 1
                ORDER BY id DESC
                LIMIT 500
            """)
            options['matrices'] = cursor.fetchall()

            # ========== CATÁLOGOS DE PROCESOS ==========

            # Procesos - Fuente: WorkOrderController.php línea 685
            # NOTA: Laravel filtra por type='EV' pero los datos actuales tienen timestamps en type
            # Se muestra todos los procesos activos hasta que se corrijan los datos
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM processes
                WHERE active = 1
                ORDER BY orden ASC, id ASC
            """)
            options['processes'] = cursor.fetchall()

            # Armados
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM armados
                WHERE active = 1
                ORDER BY descripcion
            """)
            options['armados'] = cursor.fetchall()

            # Pegados (Tipo de pegado)
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM pegados
                WHERE active = 1
                ORDER BY id
            """)
            options['pegados'] = cursor.fetchall()

            # Sentidos de armado (hardcoded como en Laravel)
            options['sentidos_armado'] = [
                {"id": 1, "nombre": "No aplica"},
                {"id": 2, "nombre": "Ancho a la Derecha"},
                {"id": 3, "nombre": "Ancho a la Izquierda"},
                {"id": 4, "nombre": "Largo a la Izquierda"},
                {"id": 5, "nombre": "Largo a la Derecha"},
            ]

            # Impresiones (tabla impresion - usa status)
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM impresion
                WHERE status = 1
                ORDER BY id
            """)
            options['impresiones'] = cursor.fetchall()

            # FSC
            cursor.execute("""
                SELECT codigo as id, descripcion as nombre
                FROM fsc
                WHERE active = 1
                ORDER BY descripcion
            """)
            options['fsc'] = cursor.fetchall()

            # ========== CATÁLOGOS DE MATERIALES ==========

            # Materiales (solo últimos 500 por performance)
            cursor.execute("""
                SELECT id, codigo, descripcion
                FROM materials
                WHERE active = 1
                ORDER BY codigo DESC
                LIMIT 500
            """)
            options['materials'] = cursor.fetchall()

            # Recubrimientos
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM recubrimiento_types
                WHERE active = 1
                ORDER BY descripcion
            """)
            options['recubrimientos'] = cursor.fetchall()

            # Coberturas internas (usa status)
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM coverage_internal
                WHERE status = 1
                ORDER BY descripcion
            """)
            options['coverages_internal'] = cursor.fetchall()

            # Coberturas externas (usa status)
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM coverage_external
                WHERE status = 1
                ORDER BY descripcion
            """)
            options['coverages_external'] = cursor.fetchall()

            # ========== CATÁLOGOS DE REFERENCIA ==========

            # Tipos de referencia (usa codigo como value, como Laravel)
            # Issue 16: Quitadas opciones SI/NO (codigos 0 y 1) segun requerimiento
            cursor.execute("""
                SELECT codigo as id, descripcion as nombre
                FROM reference_types
                WHERE active = 1 AND codigo NOT IN (0, 1)
                ORDER BY codigo
            """)
            options['reference_types'] = cursor.fetchall()

            # Tipos de diseño
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM design_types
                WHERE active = 1
                ORDER BY descripcion
            """)
            options['design_types'] = cursor.fetchall()

            # Bloqueo Referencia (hardcoded como en Laravel)
            options['bloqueo_referencia'] = [
                {"id": 1, "nombre": "Si"},
                {"id": 0, "nombre": "No"}
            ]

            # Indicador Facturación D.E. (hardcoded como en Laravel)
            options['indicador_facturacion'] = [
                {"id": 1, "nombre": "RRP"},
                {"id": 2, "nombre": "E-Commerce"},
                {"id": 3, "nombre": "Esquineros"},
                {"id": 4, "nombre": "Geometría"},
                {"id": 5, "nombre": "Participación nuevo Mercado"},
                {"id": 7, "nombre": "Innovación"},
                {"id": 8, "nombre": "Sustentabilidad"},
                {"id": 9, "nombre": "Automatización"},
                {"id": 10, "nombre": "No Aplica"},
                {"id": 11, "nombre": "Ahorro"}
            ]

            # ========== CATÁLOGOS DE CALIDAD ==========

            # Trazabilidad (usa status)
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM trazabilidad
                WHERE status = 1
                ORDER BY descripcion
            """)
            options['trazabilidad'] = cursor.fetchall()

            # Issue 42: Tipos de cinta desde tabla tipos_cintas (Laravel TipoCinta.php)
            cursor.execute("""
                SELECT id, descripcion as nombre, codigo
                FROM tipos_cintas
                WHERE active = 1
                ORDER BY descripcion
            """)
            options['tipo_cinta'] = cursor.fetchall()

            # Tipos de pallet
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM pallet_types
                WHERE active = 1
                ORDER BY descripcion
            """)
            options['pallet_types'] = cursor.fetchall()

            # Salas de corte
            cursor.execute("""
                SELECT id, nombre
                FROM salas_cortes
                WHERE deleted = 0
                ORDER BY nombre
            """)
            options['salas_corte'] = cursor.fetchall()

            # Certificado de Calidad (pallet_qas)
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM pallet_qas
                WHERE active = 1
                ORDER BY descripcion
            """)
            options['pallet_qas'] = cursor.fetchall()

            # Formato Etiqueta Pallet (pallet_tag_formats)
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM pallet_tag_formats
                WHERE active = 1
                ORDER BY descripcion
            """)
            options['pallet_tag_formats'] = cursor.fetchall()

            # ========== JERARQUÍAS ==========

            # Jerarquías nivel 1
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM hierarchies
                WHERE active = 1
                ORDER BY descripcion
            """)
            options['hierarchies'] = cursor.fetchall()

            # Subjerarquías nivel 2
            cursor.execute("""
                SELECT id, descripcion as nombre, hierarchy_id
                FROM subhierarchies
                WHERE active = 1
                ORDER BY descripcion
            """)
            options['subhierarchies'] = cursor.fetchall()

            # Subsubjerarquías nivel 3
            cursor.execute("""
                SELECT id, descripcion as nombre, subhierarchy_id
                FROM subsubhierarchies
                WHERE active = 1
                ORDER BY descripcion
            """)
            options['subsubhierarchies'] = cursor.fetchall()

            # ========== OTROS CATÁLOGOS ==========

            # Tipos de solicitud (hardcoded como en Laravel)
            options['tipos_solicitud'] = [
                {"id": 1, "nombre": "Desarrollo Completo"},
                {"id": 4, "nombre": "Cotiza sin CAD"},
                {"id": 2, "nombre": "Cotiza con CAD"},
                {"id": 3, "nombre": "Muestra con CAD"},
                {"id": 7, "nombre": "OT Proyectos Innovación"},
                {"id": 5, "nombre": "Arte con Material"},
                {"id": 6, "nombre": "Otras Solicitudes Desarrollo"},
            ]

            # Maquila servicios
            # NOTA: Laravel filtra por active=1 pero los datos actuales tienen valores no-booleanos
            # Se muestran todos hasta que se corrijan los datos
            cursor.execute("""
                SELECT id, servicio as nombre
                FROM maquila_servicios
                ORDER BY servicio
            """)
            options['maquila_servicios'] = cursor.fetchall()

            # Comunas (tabla no existe en esta BD, retornar lista vacía)
            options['comunas'] = []

            # País referencia
            cursor.execute("""
                SELECT id, name as nombre
                FROM paises
                WHERE active = 1
                ORDER BY name
            """)
            options['pais_referencia'] = cursor.fetchall()

            # Secuencia operacional
            cursor.execute("""
                SELECT id, codigo, descripcion, nombre_corto, planta_id
                FROM secuencias_operacionales
                WHERE active = 1 AND deleted = 0
                ORDER BY codigo
            """)
            options['secuencia_operacional'] = cursor.fetchall()

            # ========== SECCIÓN 13 - DATOS PARA DESARROLLO ==========

            # Tipo de producto para desarrollo
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM product_type_developing
                WHERE deleted = 0
                ORDER BY id
            """)
            options['product_type_developing'] = cursor.fetchall()

            # Tipos de alimento
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM food_types
                WHERE deleted = 0
                ORDER BY id
            """)
            options['food_types'] = cursor.fetchall()

            # Uso previsto
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM expected_uses
                ORDER BY id
            """)
            options['expected_uses'] = cursor.fetchall()

            # Uso reciclado
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM recycled_uses
                ORDER BY id
            """)
            options['recycled_uses'] = cursor.fetchall()

            # Clase sustancia a embalar
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM class_substance_packeds
                WHERE deleted = 0
                ORDER BY id
            """)
            options['class_substance_packeds'] = cursor.fetchall()

            # Medio de transporte
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM transportation_ways
                WHERE deleted = 0
                ORDER BY id
            """)
            options['transportation_ways'] = cursor.fetchall()

            # Mercado destino
            cursor.execute("""
                SELECT id, descripcion as nombre
                FROM target_markets
                WHERE deleted = 0
                ORDER BY id
            """)
            options['target_markets'] = cursor.fetchall()

            # ========== OPCIONES PARA MUESTRAS ==========
            # Tipos de pegado para muestras (hardcoded como en Laravel muestras-ot.blade.php línea 475)
            options['tipos_pegado_muestra'] = [
                {"id": 1, "nombre": "Sin Pegar"},
                {"id": 2, "nombre": "Pegado Flexo Interior"},
                {"id": 3, "nombre": "Pegado Flexo Exterior"},
                {"id": 4, "nombre": "Pegado Diecutter"},
                {"id": 5, "nombre": "Pegado Cajas Fruta"},
                {"id": 6, "nombre": "Pegado con Cinta"},
                {"id": 7, "nombre": "Sin Pegar con Cinta"},
            ]

            # Cartones muestra (solo cartones marcados como carton_muestra=1)
            cursor.execute("""
                SELECT id, codigo as nombre
                FROM cartons
                WHERE carton_muestra = 1
                ORDER BY codigo
            """)
            options['cartons_muestra'] = cursor.fetchall()

            # Destinatarios de muestra (hardcoded como en Laravel muestras-ot.blade.php línea 511)
            options['destinatarios_muestra'] = [
                {"id": 1, "nombre": "Retira Ventas VB"},
                {"id": 2, "nombre": "Retira Diseñador VB"},
                {"id": 3, "nombre": "Envío Laboratorio"},
                {"id": 4, "nombre": "Envío Cliente VB"},
                {"id": 5, "nombre": "Retira Diseñador Revisión"},
            ]

            return FormOptionsComplete(**options)

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener opciones del formulario: {str(e)}"
        )
    finally:
        connection.close()


# =============================================
# DUPLICAR OT
# =============================================

class DuplicateOTResponse(BaseModel):
    """Respuesta al duplicar OT."""
    id: int
    original_id: int
    message: str


@router.post("/{ot_id}/duplicate", response_model=DuplicateOTResponse)
async def duplicate_work_order(
    ot_id: int,
    user_id: int = Depends(get_current_user_id)
):
    """
    Duplica una OT existente con toda su información.
    La nueva OT inicia en estado inicial y área de Ventas.
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Obtener OT original
            cursor.execute("""
                SELECT * FROM work_orders WHERE id = %s AND active = 1
            """, (ot_id,))
            original = cursor.fetchone()

            if not original:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"OT {ot_id} no encontrada"
                )

            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Campos a excluir de la copia
            exclude_fields = {'id', 'created_at', 'updated_at', 'aprobado', 'material_id'}

            # Construir campos para la nueva OT
            new_fields = {}
            for key, value in original.items():
                if key not in exclude_fields:
                    new_fields[key] = value

            # Sobrescribir algunos valores
            new_fields['creador_id'] = user_id
            new_fields['current_area_id'] = 1  # Inicia en Ventas
            new_fields['active'] = 1
            new_fields['created_at'] = now
            new_fields['updated_at'] = now
            new_fields['aprobado'] = 0  # No aprobada

            # Agregar referencia a OT original en descripción
            new_fields['descripcion'] = f"[Copia OT-{ot_id}] {original.get('descripcion', '')}"[:40]

            # Construir query de inserción
            columns = ', '.join(new_fields.keys())
            placeholders = ', '.join(['%s'] * len(new_fields))
            values = list(new_fields.values())

            sql = f"INSERT INTO work_orders ({columns}) VALUES ({placeholders})"
            cursor.execute(sql, values)
            new_ot_id = cursor.lastrowid

            # Crear registro inicial en managements
            cursor.execute("""
                INSERT INTO managements
                (work_order_id, work_space_id, state_id, user_id, management_type_id, observacion, created_at, updated_at)
                VALUES (%s, 1, 1, %s, 1, %s, %s, %s)
            """, (new_ot_id, user_id, f"Duplicado de OT-{ot_id}", now, now))

            connection.commit()

            return DuplicateOTResponse(
                id=new_ot_id,
                original_id=ot_id,
                message=f"OT {ot_id} duplicada exitosamente como OT {new_ot_id}"
            )

    except pymysql.Error as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al duplicar OT: {str(e)}"
        )
    finally:
        connection.close()


# =============================================================================
# ENDPOINT CAD - Issue 26 y 45-46: Cargar datos de CAD seleccionado
# =============================================================================

class CADDetailsResponse(BaseModel):
    """Respuesta con detalles de un CAD para cargar en el formulario."""
    id: int
    cad: str
    # Medidas interiores
    interno_largo: int
    interno_ancho: int
    interno_alto: int
    # Medidas exteriores
    externo_largo: int
    externo_ancho: int
    externo_alto: int
    # Otros datos del CAD
    area_producto: float
    largura_hm: int
    anchura_hm: int
    largura_hc: int
    anchura_hc: int
    area_hm: int
    area_hc_unitario: int
    rayado_c1r1: float
    rayado_r1_r2: float
    rayado_r2_c2: float
    recorte_caracteristico: float
    recorte_adicional: float
    veces_item: int


@router.get("/cad/{cad_id}", response_model=CADDetailsResponse)
async def get_cad_details(cad_id: int):
    """
    Obtiene los detalles de un CAD específico.
    Issue 26: Cargar datos al seleccionar CAD.
    Issues 45-46: Cargar medidas interiores y exteriores.
    """
    connection = get_laravel_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, cad, interno_largo, interno_ancho, interno_alto,
                       externo_largo, externo_ancho, externo_alto,
                       area_producto, largura_hm, anchura_hm, largura_hc, anchura_hc,
                       area_hm, area_hc_unitario, rayado_c1r1, rayado_r1_r2, rayado_r2_c2,
                       recorte_caracteristico, recorte_adicional, veces_item
                FROM cads
                WHERE id = %s AND active = 1
            """, (cad_id,))
            cad = cursor.fetchone()

            if not cad:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"CAD con id {cad_id} no encontrado"
                )

            return CADDetailsResponse(
                id=cad['id'],
                cad=cad['cad'],
                interno_largo=cad['interno_largo'] or 0,
                interno_ancho=cad['interno_ancho'] or 0,
                interno_alto=cad['interno_alto'] or 0,
                externo_largo=cad['externo_largo'] or 0,
                externo_ancho=cad['externo_ancho'] or 0,
                externo_alto=cad['externo_alto'] or 0,
                area_producto=float(cad['area_producto'] or 0),
                largura_hm=cad['largura_hm'] or 0,
                anchura_hm=cad['anchura_hm'] or 0,
                largura_hc=cad['largura_hc'] or 0,
                anchura_hc=cad['anchura_hc'] or 0,
                area_hm=cad['area_hm'] or 0,
                area_hc_unitario=cad['area_hc_unitario'] or 0,
                rayado_c1r1=float(cad['rayado_c1r1'] or 0),
                rayado_r1_r2=float(cad['rayado_r1_r2'] or 0),
                rayado_r2_c2=float(cad['rayado_r2_c2'] or 0),
                recorte_caracteristico=float(cad['recorte_caracteristico'] or 0),
                recorte_adicional=float(cad['recorte_adicional'] or 0),
                veces_item=cad['veces_item'] or 0
            )

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener CAD: {str(e)}"
        )
    finally:
        connection.close()


# =============================================================================
# ENDPOINT CARTON - Issue 25: Fórmula McKee - Obtener datos de cartón
# Fuente Laravel: WorkOrderController@getCarton línea 3737-3764
# =============================================================================

class CartonDetailsResponse(BaseModel):
    """Respuesta con detalles de un Cartón para cálculo McKee."""
    id: int
    codigo: Optional[str] = None
    nombre: Optional[str] = None
    ect_min: Optional[int] = None
    espesor: Optional[float] = None


@router.get("/carton/{carton_id}", response_model=CartonDetailsResponse)
async def get_carton_details(carton_id: int):
    """
    Obtiene los detalles de un Cartón específico.

    Issue 25: Fórmula McKee necesita ECT y Espesor del cartón.
    Fuente Laravel: /getCarton en routes/web.php
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, codigo, descripcion, ect_min, espesor
                FROM cartons
                WHERE id = %s AND active = 1
            """, (carton_id,))

            carton = cursor.fetchone()

            if not carton:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Cartón con id {carton_id} no encontrado"
                )

            return CartonDetailsResponse(
                id=carton['id'],
                codigo=carton.get('codigo'),
                nombre=carton.get('descripcion'),
                ect_min=carton.get('ect_min'),
                espesor=float(carton['espesor']) if carton.get('espesor') else None
            )

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener Cartón: {str(e)}"
        )
    finally:
        connection.close()


# =============================================
# SPRINT B: GRUPO 2 - CASCADAS Y DATOS
# Basado en: WorkOrderController.php líneas 9969-10290
# =============================================

@router.get("/matriz")
async def get_matriz(cad_id: Optional[int] = None):
    """
    Obtiene matrices disponibles para un CAD específico.

    Basado en: WorkOrderController@getMatriz (líneas 9969-9979)

    Args:
        cad_id: ID del CAD para filtrar matrices

    Returns:
        Lista de matrices que coinciden con el plano_cad del CAD
    """
    if not cad_id:
        return {"matrices": []}

    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Primero obtener el código CAD
            cursor.execute(
                "SELECT cad FROM cads WHERE id = %s AND active = 1",
                (cad_id,)
            )
            cad = cursor.fetchone()

            if not cad:
                return {"matrices": []}

            # Buscar matrices con ese plano_cad
            cursor.execute("""
                SELECT id, material, plano_cad, cuchillas
                FROM matrizs
                WHERE active = 1 AND plano_cad = %s
                ORDER BY material
            """, (cad['cad'],))

            matrices = cursor.fetchall()

            return {
                "cad_id": cad_id,
                "cad_codigo": cad['cad'],
                "total": len(matrices),
                "matrices": matrices
            }

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener matrices: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/matriz/{matriz_id}/data")
async def get_matriz_data(matriz_id: int):
    """
    Obtiene los datos detallados de una matriz específica.

    Basado en: WorkOrderController@getMatrizData (líneas 9980-9990)

    Args:
        matriz_id: ID de la matriz

    Returns:
        Datos completos de la matriz incluyendo cuchillas formateadas
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT *
                FROM matrizs
                WHERE id = %s AND active = 1
            """, (matriz_id,))

            matriz = cursor.fetchone()

            if not matriz:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Matriz con id {matriz_id} no encontrada"
                )

            # Formatear cuchillas como en Laravel
            if matriz.get('cuchillas'):
                cuchillas_str = str(matriz['cuchillas']).replace(',', '.')
                try:
                    matriz['cuchillas'] = float(cuchillas_str)
                except ValueError:
                    pass

            return matriz

    except HTTPException:
        raise
    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener datos de matriz: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/cad-by-material/{material_id}")
async def get_cad_by_material(material_id: int):
    """
    Obtiene el material con su CAD asociado.

    Basado en: WorkOrderController@getCadByMaterial (líneas 9991-10000)

    Args:
        material_id: ID del material

    Returns:
        Material con sus relaciones (cad, carton, etc.)
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT
                    m.*,
                    c.cad as cad_codigo,
                    c.id as cad_id,
                    cb.codigo as carton_codigo,
                    cb.id as carton_id,
                    cl.nombre as cliente_nombre,
                    pt.nombre as tipo_producto_nombre,
                    s.nombre as estilo_nombre
                FROM materials m
                LEFT JOIN cads c ON m.cad_id = c.id
                LEFT JOIN cardboards cb ON m.carton_id = cb.id
                LEFT JOIN clients cl ON m.client_id = cl.id
                LEFT JOIN product_types pt ON m.product_type_id = pt.id
                LEFT JOIN styles s ON m.style_id = s.id
                WHERE m.id = %s
            """, (material_id,))

            material = cursor.fetchone()

            if not material:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Material con id {material_id} no encontrado"
                )

            return material

    except HTTPException:
        raise
    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener material: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/carton-color")
async def get_carton_color(carton_color: Optional[int] = None):
    """
    Obtiene cartones filtrados por color.

    Basado en: WorkOrderController@getCartonColor (líneas 10002-10020)

    Args:
        carton_color: 1 = CAFE, 2 = BLANCO, None = todos

    Returns:
        Lista de cartones filtrados por color
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            if carton_color:
                color = 'CAFE' if carton_color == 1 else 'BLANCO'
                cursor.execute("""
                    SELECT id, codigo, descripcion, color_tapa_exterior
                    FROM cardboards
                    WHERE active = 1 AND color_tapa_exterior = %s
                    ORDER BY codigo
                """, (color,))
            else:
                cursor.execute("""
                    SELECT id, codigo, descripcion, color_tapa_exterior
                    FROM cardboards
                    WHERE active = 1
                    ORDER BY codigo
                """)

            cartones = cursor.fetchall()

            return {
                "filtro_color": carton_color,
                "color_nombre": 'CAFE' if carton_color == 1 else ('BLANCO' if carton_color == 2 else 'TODOS'),
                "total": len(cartones),
                "cartones": cartones
            }

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener cartones por color: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/lista-carton")
async def get_lista_carton(
    carton_color: int = Query(..., description="1=CAFE, 2=BLANCO"),
    planta: int = Query(..., description="ID de la planta"),
    impresion: int = Query(..., description="1=Offset, 3=Alta Gráfica, otro=Normal")
):
    """
    Obtiene lista de cartones filtrada por color, planta e impresión.

    Basado en: WorkOrderController@getListaCarton (líneas 10182-10222)

    Args:
        carton_color: 1=CAFE, 2=BLANCO
        planta: ID de la planta
        impresion: 1=Offset, 3=Alta Gráfica, otro=Normal

    Returns:
        Lista de cartones que cumplen los criterios
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            color = 'CAFE' if carton_color == 1 else 'BLANCO'

            # Base query
            base_query = """
                SELECT id, codigo, descripcion, color_tapa_exterior,
                       alta_grafica, offset, planta_id
                FROM cardboards
                WHERE active = 1
                AND color_tapa_exterior = %s
                AND planta_id LIKE %s
            """
            params = [color, f"%{planta}%"]

            # Agregar filtro según impresión
            if impresion == 3:  # Alta Gráfica
                base_query += " AND alta_grafica = 1"
            elif impresion == 1:  # Offset
                base_query += " AND offset = 1"
            # else: Normal (sin filtro adicional)

            base_query += " ORDER BY codigo"

            cursor.execute(base_query, params)
            cartones = cursor.fetchall()

            return {
                "filtros": {
                    "color": color,
                    "planta_id": planta,
                    "impresion": impresion,
                    "tipo_impresion": "Alta Gráfica" if impresion == 3 else ("Offset" if impresion == 1 else "Normal")
                },
                "total": len(cartones),
                "cartones": cartones
            }

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener lista de cartones: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/lista-carton-edit")
async def get_lista_carton_edit(
    carton_color: int = Query(..., description="1=CAFE, 2=BLANCO"),
    planta: int = Query(..., description="ID de la planta"),
    impresion: int = Query(..., description="1=Offset, 3=Alta Gráfica, otro=Normal")
):
    """
    Obtiene lista de cartones para edición (misma lógica que getListaCarton).

    Basado en: WorkOrderController@getListaCartonEdit (líneas 10224-10264)

    Usado en formularios de edición de OT.
    """
    # Reutiliza la misma lógica que get_lista_carton
    return await get_lista_carton(carton_color, planta, impresion)


@router.get("/lista-carton-offset")
async def get_lista_carton_offset(
    impresion: int = Query(..., description="1=Offset, 3=Alta Gráfica, otro=Normal")
):
    """
    Obtiene lista de cartones solo por tipo de impresión (sin filtro de color/planta).

    Basado en: WorkOrderController@getListaCartonOffset (líneas 10266-10290)

    Args:
        impresion: 1=Offset, 3=Alta Gráfica, otro=Normal

    Returns:
        Lista de cartones según tipo de impresión
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            if impresion == 3:  # Alta Gráfica
                cursor.execute("""
                    SELECT id, codigo, descripcion, alta_grafica, offset
                    FROM cardboards
                    WHERE active = 1 AND alta_grafica = 1
                    ORDER BY codigo
                """)
            elif impresion == 1:  # Offset
                cursor.execute("""
                    SELECT id, codigo, descripcion, alta_grafica, offset
                    FROM cardboards
                    WHERE active = 1 AND offset = 1
                    ORDER BY codigo
                """)
            else:  # Normal
                cursor.execute("""
                    SELECT id, codigo, descripcion, alta_grafica, offset
                    FROM cardboards
                    WHERE active = 1 AND alta_grafica = 0 AND offset = 0
                    ORDER BY codigo
                """)

            cartones = cursor.fetchall()

            return {
                "filtro_impresion": impresion,
                "tipo_impresion": "Alta Gráfica" if impresion == 3 else ("Offset" if impresion == 1 else "Normal"),
                "total": len(cartones),
                "cartones": cartones
            }

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener cartones offset: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/maquila-servicio/{maquila_servicio_id}")
async def get_maquila_servicio(maquila_servicio_id: int):
    """
    Obtiene los datos de un servicio de maquila específico.

    Basado en: WorkOrderController@getMaquilaServicio (líneas 10033-10042)

    Args:
        maquila_servicio_id: ID del servicio de maquila

    Returns:
        Datos completos del servicio de maquila
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT *
                FROM maquila_servicios
                WHERE id = %s AND active = 1
            """, (maquila_servicio_id,))

            servicio = cursor.fetchone()

            if not servicio:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Servicio de maquila con id {maquila_servicio_id} no encontrado"
                )

            return servicio

    except HTTPException:
        raise
    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener servicio de maquila: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/design-type/{design_type_id}")
async def get_design_type(design_type_id: int):
    """
    Obtiene los datos de un tipo de diseño específico.

    Basado en: WorkOrderController@getDesignType (líneas 10044-10052)

    Args:
        design_type_id: ID del tipo de diseño

    Returns:
        Datos completos del tipo de diseño
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT *
                FROM design_types
                WHERE id = %s AND active = 1
            """, (design_type_id,))

            design_type = cursor.fetchone()

            if not design_type:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Tipo de diseño con id {design_type_id} no encontrado"
                )

            return design_type

    except HTTPException:
        raise
    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener tipo de diseño: {str(e)}"
        )
    finally:
        connection.close()


# =============================================================================
# SPRINT B GRUPO 2 - ENDPOINTS ADICIONALES (Cascadas y Datos)
# =============================================================================

@router.get("/color-carton")
async def get_color_carton():
    """
    Obtiene todos los cartones activos con su color, impresion_id y planta_id.

    Basado en: WorkOrderController@getColorCarton (líneas 10163-10179)

    Usado para filtrar cartones disponibles según impresión y planta seleccionada
    en el formulario de Ingresos Principales.

    Returns:
        Lista de cartones con: key, descripcion, color, impresion_id, planta_id
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT
                    id,
                    codigo,
                    color_tapa_exterior,
                    impresion_id,
                    planta_id
                FROM cardboards
                WHERE active = 1
                ORDER BY codigo
            """)

            cartons = cursor.fetchall()

            # Formatear según estructura Laravel
            data = []
            for carton in cartons:
                data.append({
                    "key": carton["id"],
                    "descripcion": carton["codigo"],
                    "color": carton["color_tapa_exterior"],
                    "impresion_id": carton["impresion_id"],
                    "planta_id": carton["planta_id"]
                })

            return data

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener colores de cartón: {str(e)}"
        )
    finally:
        connection.close()


@router.post("/verificacion-filtro")
async def post_verificacion_filtro(filtros: List[dict]):
    """
    Verifica si la combinación de filtros de Ingresos Principales tiene plantas activas.

    Basado en: WorkOrderController@postVerificacionFiltro (líneas 10055-10097)

    Esta función permite verificar si el usuario puede seguir llenando el formulario
    según la combinación de filtros seleccionada (impresión, FSC, cinta, recubrimientos).

    Args:
        filtros: Lista de diccionarios con los filtros a verificar. Cada filtro tiene:
            - referencia: tipo de filtro ("impresion_fsc", "cinta", "recubrimiento_interno", "impresion_recubrimiento_externo")
            - Campos específicos según referencia:
                - impresion_fsc: impresion_id, fsc_id
                - cinta: cinta_id
                - recubrimiento_interno: recubrimiento_interno_id
                - impresion_recubrimiento_externo: impresion_id, recubrimiento_externo_id

    Returns:
        {plantas: [...], cantidad_filtro: int} o [] si no hay combinaciones válidas
    """
    if not filtros or len(filtros) == 0:
        return []

    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Construir queries UNION según los filtros recibidos
            union_queries = []
            params = []

            for filtro in filtros:
                referencia = filtro.get("referencia", "")

                if referencia == "impresion_fsc":
                    union_queries.append("""
                        SELECT * FROM relacion_filtro_ingresos_principales AS f
                        WHERE f.filtro_1 = %s AND f.filtro_2 = %s AND f.referencia = 'impresion_fsc'
                    """)
                    params.extend([filtro.get("impresion_id"), filtro.get("fsc_id")])

                elif referencia == "cinta":
                    union_queries.append("""
                        SELECT * FROM relacion_filtro_ingresos_principales AS f
                        WHERE f.filtro_1 = %s AND f.referencia = 'cinta'
                    """)
                    params.append(filtro.get("cinta_id"))

                elif referencia == "recubrimiento_interno":
                    union_queries.append("""
                        SELECT * FROM relacion_filtro_ingresos_principales AS f
                        WHERE f.filtro_1 = %s AND f.referencia = 'recubrimiento_interno'
                    """)
                    params.append(filtro.get("recubrimiento_interno_id"))

                elif referencia == "impresion_recubrimiento_externo":
                    union_queries.append("""
                        SELECT * FROM relacion_filtro_ingresos_principales AS f
                        WHERE f.filtro_1 = %s AND f.filtro_2 = %s AND f.referencia = 'impresion_recubrimiento_externo'
                    """)
                    params.extend([filtro.get("impresion_id"), filtro.get("recubrimiento_externo_id")])

            if not union_queries:
                return []

            # Ejecutar query UNION
            query = " UNION ".join(union_queries)
            cursor.execute(query, params)
            result = cursor.fetchall()

            qty_filters = len(result)
            if qty_filters == 0:
                return []

            # Extraer IDs de plantas de los resultados
            plantas_id = []
            for row in result:
                planta_id_str = row.get("planta_id")
                if planta_id_str:
                    # El campo planta_id puede contener múltiples IDs separados por coma
                    for p in str(planta_id_str).split(","):
                        p = p.strip()
                        if p and p not in plantas_id:
                            plantas_id.append(p)

            return {
                "plantas": plantas_id,
                "cantidad_filtro": qty_filters
            }

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al verificar filtros: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/recubrimiento-interno")
async def get_recubrimiento_interno():
    """
    Obtiene los recubrimientos internos con sus plantas asociadas.

    Basado en: WorkOrderController@getRecubrimientoInterno (líneas 10100-10113)

    Usado para la cascada de selección en el formulario de Ingresos Principales.

    Returns:
        Lista de recubrimientos internos con: key, descripcion, planta_id
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT
                    id,
                    descripcion,
                    planta_id
                FROM coverage_internal
                WHERE status = 1
                ORDER BY descripcion
            """)

            recubrimientos = cursor.fetchall()

            # Formatear según estructura Laravel
            data = []
            for rec in recubrimientos:
                data.append({
                    "key": rec["id"],
                    "descripcion": rec["descripcion"],
                    "planta_id": rec["planta_id"]
                })

            return data

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener recubrimientos internos: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/recubrimiento-externo")
async def get_recubrimiento_externo(
    impresion_id: int = Query(..., description="ID del tipo de impresión"),
    referencia: str = Query(default="impresion_recubrimiento_externo", description="Referencia del filtro")
):
    """
    Obtiene los recubrimientos externos filtrados por tipo de impresión.

    Basado en: WorkOrderController@getRecubrimientoExterno (líneas 10117-10142)

    Usado para la cascada de selección en el formulario de Ingresos Principales.

    Args:
        impresion_id: ID del tipo de impresión seleccionado
        referencia: Referencia del filtro (default: impresion_recubrimiento_externo)

    Returns:
        Lista de recubrimientos externos con: key, filtro_1, filtro_2, planta_id, descripcion
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT
                    f.id,
                    f.filtro_1,
                    f.filtro_2,
                    f.planta_id,
                    e.descripcion
                FROM relacion_filtro_ingresos_principales AS f
                INNER JOIN coverage_external AS e ON e.id = f.filtro_2
                WHERE f.filtro_1 = %s AND f.referencia = %s
                ORDER BY e.descripcion
            """, (impresion_id, referencia))

            recubrimientos = cursor.fetchall()

            # Formatear según estructura Laravel
            data = []
            for rec in recubrimientos:
                data.append({
                    "key": rec["id"],
                    "filtro_1": rec["filtro_1"],
                    "filtro_2": rec["filtro_2"],
                    "planta_id": rec["planta_id"],
                    "descripcion": rec["descripcion"]
                })

            return data

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener recubrimientos externos: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/planta-objetivo")
async def get_planta_objetivo():
    """
    Obtiene todas las plantas disponibles para selección.

    Basado en: WorkOrderController@getPlantaObjetivo (líneas 10146-10159)

    Usado para la cascada de selección en el formulario de Ingresos Principales.

    Returns:
        Lista de plantas con: key, descripcion, planta_id
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT
                    id,
                    nombre
                FROM plantas
                WHERE active = 1
                ORDER BY nombre
            """)

            plantas = cursor.fetchall()

            # Formatear según estructura Laravel
            data = []
            for planta in plantas:
                data.append({
                    "key": planta["id"],
                    "descripcion": planta["nombre"],
                    "planta_id": planta["id"]
                })

            return data

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener plantas objetivo: {str(e)}"
        )
    finally:
        connection.close()


# =============================================
# SPRINT B GRUPO 1: OTs ESPECIALES - CARGA ESTUDIO BENCHMARKING
# =============================================

@router.post("/carga-detalles-estudio")
async def carga_detalles_estudio(
    archivo_detalles: UploadFile = File(...)
):
    """
    Carga un archivo Excel con detalles de Estudio Benchmarking.

    Basado en: WorkOrderController@cargaDetallesEstudio (líneas 11495-11560)

    El Excel debe tener las columnas:
    - identificacion_muestra
    - cliente
    - descripcion

    Args:
        archivo_detalles: Archivo Excel (.xlsx, .xls) con los detalles

    Returns:
        {
            "mensaje": "Archivo cargado Exitosamente",
            "detalles": [...],  # Array de {identificacion_muestra, cliente, descripcion}
            "cantidad": int,    # Número de filas
            "archivo": str      # Path del archivo guardado
        }
    """
    # Validar extensión del archivo
    filename = archivo_detalles.filename
    if not filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Nombre de archivo no válido"
        )

    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato de archivo no válido. Use .xlsx o .xls"
        )

    try:
        # Leer contenido del archivo
        content = await archivo_detalles.read()

        # Cargar Excel con openpyxl
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        # Obtener headers de la primera fila
        headers = []
        for cell in sheet[1]:
            if cell.value:
                # Normalizar headers: minúsculas, sin espacios, reemplazar espacios por _
                header = str(cell.value).lower().strip().replace(" ", "_")
                headers.append(header)
            else:
                headers.append(None)

        # Verificar columnas requeridas
        required_columns = ["identificacion_muestra", "cliente", "descripcion"]
        header_indices = {}
        for col in required_columns:
            if col in headers:
                header_indices[col] = headers.index(col)
            else:
                # Buscar variantes comunes
                for i, h in enumerate(headers):
                    if h and col.replace("_", "") in h.replace("_", ""):
                        header_indices[col] = i
                        break

        # Si faltan columnas, usar orden por posición (A, B, C)
        if len(header_indices) < 3:
            header_indices = {
                "identificacion_muestra": 0,
                "cliente": 1,
                "descripcion": 2
            }

        # Leer datos
        detalles = []
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue  # Saltar filas vacías

            detalle = {
                "identificacion_muestra": str(row[header_indices.get("identificacion_muestra", 0)] or ""),
                "cliente": str(row[header_indices.get("cliente", 1)] or ""),
                "descripcion": str(row[header_indices.get("descripcion", 2)] or "")
            }

            # Solo agregar si tiene al menos un campo con datos
            if any(v for v in detalle.values()):
                detalles.append(detalle)

        if not detalles:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Archivo vacío no contiene información"
            )

        # Guardar archivo (simular comportamiento Laravel)
        # En producción esto iría a un storage configurado
        safe_filename = filename.replace("%", "")
        base_name = safe_filename.rsplit(".", 1)[0] if "." in safe_filename else safe_filename

        # Path virtual (en producción sería path real)
        path_archivo = f"/files/{base_name}.{extension}"

        return {
            "mensaje": "Archivo cargado Exitosamente",
            "detalles": detalles,
            "cantidad": len(detalles),
            "archivo": path_archivo
        }

    except openpyxl.utils.exceptions.InvalidFileException:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Archivo Excel inválido o corrupto"
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al procesar archivo: {str(e)}"
        )


# =============================================
# SPRINT C: APROBACIÓN, RECHAZO Y GESTIÓN OTs
# =============================================

@router.put("/{ot_id}/aprobar")
async def aprobar_ot(
    ot_id: int,
    user: dict = Depends(get_current_user_with_role)
):
    """
    Aprueba una Orden de Trabajo.

    Basado en: WorkOrderController@aprobarOt (líneas 10550-10558)

    - Si el usuario es Jefe de Ventas: actualiza aprobacion_jefe_venta = 2
    - Si el usuario es Jefe de Desarrollo: actualiza aprobacion_jefe_desarrollo = 2

    Args:
        ot_id: ID de la orden de trabajo

    Returns:
        Mensaje de éxito
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Verificar que la OT existe
            cursor.execute("SELECT id FROM work_orders WHERE id = %s", (ot_id,))
            if not cursor.fetchone():
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Orden de trabajo no encontrada"
                )

            role_id = user.get("role_id")

            # Rol 3 = Jefe Ventas, Rol 5 = Jefe Desarrollo (verificar según BD)
            if role_id == 3:  # Jefe de Ventas
                cursor.execute(
                    "UPDATE work_orders SET aprobacion_jefe_venta = 2 WHERE id = %s",
                    (ot_id,)
                )
            elif role_id == 5:  # Jefe de Desarrollo
                cursor.execute(
                    "UPDATE work_orders SET aprobacion_jefe_desarrollo = 2 WHERE id = %s",
                    (ot_id,)
                )
            else:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="No tiene permisos para aprobar OTs"
                )

            connection.commit()

            return {
                "success": True,
                "message": "Orden de Trabajo Aprobada correctamente",
                "ot_id": ot_id
            }

    except HTTPException:
        raise
    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al aprobar OT: {str(e)}"
        )
    finally:
        connection.close()


@router.put("/{ot_id}/rechazar")
async def rechazar_ot(
    ot_id: int,
    user: dict = Depends(get_current_user_with_role)
):
    """
    Rechaza una Orden de Trabajo.

    Basado en: WorkOrderController@rechazarOt (líneas 10560-10568)

    - Si el usuario es Jefe de Ventas: actualiza aprobacion_jefe_venta = 3
    - Si el usuario es Jefe de Desarrollo: actualiza aprobacion_jefe_desarrollo = 3

    Args:
        ot_id: ID de la orden de trabajo

    Returns:
        Mensaje de éxito
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Verificar que la OT existe
            cursor.execute("SELECT id FROM work_orders WHERE id = %s", (ot_id,))
            if not cursor.fetchone():
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Orden de trabajo no encontrada"
                )

            role_id = user.get("role_id")

            # Rol 3 = Jefe Ventas, Rol 5 = Jefe Desarrollo
            if role_id == 3:  # Jefe de Ventas
                cursor.execute(
                    "UPDATE work_orders SET aprobacion_jefe_venta = 3 WHERE id = %s",
                    (ot_id,)
                )
            elif role_id == 5:  # Jefe de Desarrollo
                cursor.execute(
                    "UPDATE work_orders SET aprobacion_jefe_desarrollo = 3 WHERE id = %s",
                    (ot_id,)
                )
            else:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="No tiene permisos para rechazar OTs"
                )

            connection.commit()

            return {
                "success": True,
                "message": "Orden de Trabajo Rechazada correctamente",
                "ot_id": ot_id
            }

    except HTTPException:
        raise
    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al rechazar OT: {str(e)}"
        )
    finally:
        connection.close()


# =============================================
# SPRINT C: CREAR CAD Y MATERIAL
# Basado en: WorkOrderController@createCadMaterial (líneas 10302-10476)
# =============================================

class CreateCadMaterialRequest(BaseModel):
    """Request para crear CAD y Material."""
    cad: Optional[str] = None
    cad_id: Optional[int] = None
    descripcion: str
    material: str
    maquila: int


@router.put("/{ot_id}/crear-cad-material")
async def create_cad_material(
    ot_id: int,
    data: CreateCadMaterialRequest,
    user: dict = Depends(get_current_user_with_role)
):
    """
    Crea CAD y Material para una OT.

    Basado en: WorkOrderController@createCadMaterial (líneas 10302-10476)

    Proceso completo:
    1. Valida dimensiones (rayados, medidas internas/externas, áreas)
    2. Crea o asigna CAD
    3. Crea Material con todos los datos de la OT
    4. Actualiza OT con cad_id, material_id, material_asignado

    Args:
        ot_id: ID de la orden de trabajo
        data: Datos del CAD y material

    Returns:
        CAD y Material creados, OT actualizada
    """
    # Validar que se proporcione cad o cad_id
    if not data.cad and not data.cad_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Debe proporcionar 'cad' o 'cad_id'"
        )

    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Obtener OT completa con todos los campos necesarios
            cursor.execute("""
                SELECT wo.*,
                       ssh.jerarquia_sap
                FROM work_orders wo
                LEFT JOIN sub_sub_hierarchies ssh ON wo.subsubhiearchy_id = ssh.id
                WHERE wo.id = %s
            """, (ot_id,))

            ot = cursor.fetchone()
            if not ot:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Orden de trabajo no encontrada"
                )

            errores = []
            tipo_solicitud = ot.get("tipo_solicitud")
            process_id = ot.get("process_id")

            # Validación 1: Suma de rayados vs anchura_hm (proceso Flexo)
            if tipo_solicitud in [1, 5, 7] and process_id in [1, 5]:
                anchura_hm = ot.get("anchura_hm")
                if anchura_hm:
                    suma_rayado = (
                        (ot.get("rayado_c1r1") or 0) +
                        (ot.get("rayado_r1_r2") or 0) +
                        (ot.get("rayado_r2_c2") or 0)
                    )
                    if anchura_hm != suma_rayado:
                        errores.append(1)  # Error código 1

            # Validación 2: Medidas internas > externas sin CAD
            if tipo_solicitud in [1, 5, 7]:
                if (
                    (ot.get("interno_largo") or 0) > (ot.get("externo_largo") or 0) or
                    (ot.get("interno_ancho") or 0) > (ot.get("externo_ancho") or 0) or
                    (ot.get("interno_alto") or 0) > (ot.get("externo_alto") or 0)
                ):
                    if not ot.get("cad_id"):
                        errores.append(2)  # Error código 2

            # Validación 3: Áreas (area_hc >= area_hm >= area_producto)
            area_hc = ot.get("area_hc") or 0
            area_hm = ot.get("area_hm") or 0
            area_producto = ot.get("area_producto") or 0
            if tipo_solicitud in [1, 5, 7]:
                if area_hc < area_hm or area_hm < area_producto:
                    errores.append(3)  # Error código 3

            # Validación 4: Recorte vs Area HM (según Laravel líneas 10340-10347)
            if tipo_solicitud in [1, 5, 7]:
                recorte_caracteristico = ot.get("recorte_caracteristico")
                # Verificar que no sea "N/A"
                if recorte_caracteristico and str(recorte_caracteristico).strip().upper() != "N/A":
                    try:
                        # Convertir a número (manejar formato con comas como en Laravel)
                        if isinstance(recorte_caracteristico, str):
                            recorte_val = float(
                                recorte_caracteristico.replace(".", "").replace(",", ".")
                            )
                        else:
                            recorte_val = float(recorte_caracteristico or 0)

                        recorte_adicional = float(ot.get("recorte_adicional") or 0)

                        if recorte_val > area_hm or recorte_adicional > area_hm:
                            errores.append(4)  # Error código 4
                    except (ValueError, TypeError):
                        pass  # Si no se puede convertir, no aplica validación

            if errores:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail={
                        "errores": errores,
                        "mensaje": "Validación de dimensiones fallida",
                        "validacion_campos": ",".join(map(str, errores))
                    }
                )

            cad_id_final = data.cad_id
            cad_codigo = None

            # Si se proporciona código de CAD nuevo, crearlo
            if data.cad:
                cursor.execute("""
                    INSERT INTO cads (
                        cad, externo_largo, externo_ancho, externo_alto,
                        interno_largo, interno_ancho, interno_alto,
                        area_producto, recorte_adicional,
                        largura_hm, anchura_hm, largura_hc, anchura_hc,
                        area_hm, rayado_c1r1, rayado_r1_r2, rayado_r2_c2,
                        recorte_caracteristico, veces_item,
                        active, created_at, updated_at
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, 1, NOW(), NOW()
                    )
                """, (
                    data.cad,
                    ot.get("externo_largo"),
                    ot.get("externo_ancho"),
                    ot.get("externo_alto"),
                    ot.get("interno_largo"),
                    ot.get("interno_ancho"),
                    ot.get("interno_alto"),
                    ot.get("area_producto"),
                    ot.get("recorte_adicional"),
                    ot.get("largura_hm"),
                    ot.get("anchura_hm"),
                    ot.get("largura_hc"),
                    ot.get("anchura_hc"),
                    ot.get("area_hm"),
                    ot.get("rayado_c1r1"),
                    ot.get("rayado_r1_r2"),
                    ot.get("rayado_r2_c2"),
                    ot.get("recorte_caracteristico"),
                    ot.get("veces_item")
                ))
                cad_id_final = cursor.lastrowid
                cad_codigo = data.cad
            else:
                # Obtener código del CAD existente
                cursor.execute("SELECT cad FROM cads WHERE id = %s", (data.cad_id,))
                cad_row = cursor.fetchone()
                cad_codigo = cad_row.get("cad") if cad_row else None

            # Crear Material completo (según Laravel líneas 10400-10463)
            cursor.execute("""
                INSERT INTO materials (
                    codigo, descripcion, client_id, vendedor_id, carton_id,
                    product_type_id, style_id, numero_colores, cad_id,
                    pallet_type_id, pallet_box_quantity, placas_por_pallet,
                    pallet_patron_id, patron_zuncho_pallet, pallet_protection_id,
                    boxes_per_package, patron_zuncho_paquete, paquetes_por_unitizado,
                    unitizado_por_pallet, pallet_tag_format_id, fecha_creacion,
                    creador_id, pallet_qa_id, numero_etiquetas,
                    bct_min_lb, bct_min_kg, pallet_treatment, sap_hiearchy_id,
                    tipo_camion, restriccion_especial, horario_recepcion,
                    codigo_producto_cliente, etiquetas_dsc, orientacion_placa,
                    recubrimiento, work_order_id,
                    cinta, corte_liner, tipo_cinta,
                    distancia_cinta_1, distancia_cinta_2, distancia_cinta_3,
                    distancia_cinta_4, distancia_cinta_5, distancia_cinta_6,
                    gramaje, ect, flexion_aleta, peso,
                    incision_rayado_longitudinal, incision_rayado_vertical,
                    fct, cobb_interior, cobb_exterior, espesor,
                    active, created_at, updated_at
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, NOW(), %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, 2, NOW(), NOW()
                )
            """, (
                data.material,  # codigo
                data.descripcion,  # descripcion
                ot.get("client_id"),
                ot.get("creador_id"),  # vendedor_id
                ot.get("carton_id"),
                ot.get("product_type_id"),
                ot.get("style_id"),
                ot.get("numero_colores"),
                cad_id_final,
                ot.get("pallet_type_id"),
                ot.get("cajas_por_pallet"),  # pallet_box_quantity
                ot.get("placas_por_pallet"),
                ot.get("pallet_patron_id"),
                ot.get("patron_zuncho"),  # patron_zuncho_pallet
                ot.get("pallet_protection_id"),
                ot.get("pallet_box_quantity_id"),  # boxes_per_package
                ot.get("patron_zuncho_paquete"),
                ot.get("paquetes_por_unitizado"),
                ot.get("unitizado_por_pallet"),
                ot.get("pallet_tag_format_id"),
                user.get("id"),  # creador_id
                ot.get("pallet_qa_id"),
                ot.get("numero_etiquetas"),
                ot.get("bct_min_lb"),
                ot.get("bct_min_kg"),
                ot.get("pallet_treatment"),
                ot.get("jerarquia_sap"),  # sap_hiearchy_id
                ot.get("tipo_camion"),
                ot.get("restriccion_especial"),
                ot.get("horario_recepcion"),
                ot.get("codigo_producto_cliente"),
                ot.get("etiquetas_dsc"),
                ot.get("orientacion_placa"),
                ot.get("recubrimiento"),
                ot_id,  # work_order_id
                ot.get("cinta"),
                ot.get("corte_liner"),
                ot.get("tipo_cinta"),
                ot.get("distancia_cinta_1"),
                ot.get("distancia_cinta_2"),
                ot.get("distancia_cinta_3"),
                ot.get("distancia_cinta_4"),
                ot.get("distancia_cinta_5"),
                ot.get("distancia_cinta_6"),
                ot.get("gramaje"),
                ot.get("ect"),
                ot.get("flexion_aleta"),
                ot.get("peso"),
                ot.get("incision_rayado_longitudinal"),
                ot.get("incision_rayado_vertical"),
                ot.get("fct"),
                ot.get("cobb_interior"),
                ot.get("cobb_exterior"),
                ot.get("espesor")
            ))
            material_id = cursor.lastrowid

            # Actualizar OT con CAD y Material (según Laravel líneas 10466-10473)
            cursor.execute("""
                UPDATE work_orders
                SET cad = %s,
                    cad_id = %s,
                    material_id = %s,
                    material_asignado = %s,
                    descripcion_material = %s,
                    maquila = %s,
                    updated_at = NOW()
                WHERE id = %s
            """, (
                cad_codigo,
                cad_id_final,
                material_id,
                data.material,
                data.descripcion,
                data.maquila,
                ot_id
            ))

            connection.commit()

            return {
                "success": True,
                "message": "CAD y Material creados correctamente",
                "cad_id": cad_id_final,
                "cad_codigo": cad_codigo,
                "material_id": material_id,
                "material_codigo": data.material,
                "ot_id": ot_id
            }

    except HTTPException:
        raise
    except pymysql.Error as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al crear CAD y Material: {str(e)}"
        )
    finally:
        connection.close()


class CreateCodigoMaterialRequest(BaseModel):
    """
    Request para crear código de material.

    Basado en: WorkOrderController@createCodigoMaterial (líneas 10479-10522)
    Formato código: prefijo_ot + codigo_material + "-" + sufijo_id
    """
    codigo_material: str  # Código base del material (obligatorio según Laravel)
    sufijo_id: str  # Sufijo del código (obligatorio según Laravel)
    prefijo_ot: str  # Prefijo principal (obligatorio según Laravel)
    descripcion: Optional[str] = None  # Descripción para tipo_solicitud = 5
    prefijos_adicionales: Optional[List[str]] = None  # Prefijos adicionales para clonar


@router.put("/{ot_id}/crear-codigo-material")
async def create_codigo_material(
    ot_id: int,
    data: CreateCodigoMaterialRequest
):
    """
    Crea código de material para una OT.

    Basado en: WorkOrderController@createCodigoMaterial (líneas 10479-10522)

    Proceso:
    1. Genera código: prefijo_ot + codigo_material + "-" + sufijo_id
    2. Actualiza material.codigo y material.descripcion
    3. Actualiza ot.material_asignado y ot.codigo_sap_final
    4. Clona materiales con prefijos adicionales si se proporcionan

    Args:
        ot_id: ID de la orden de trabajo
        data: Datos del código de material

    Returns:
        Código de material generado y materiales clonados
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Obtener OT con material
            cursor.execute("""
                SELECT wo.id, wo.tipo_solicitud, wo.material_id,
                       wo.descripcion_material, m.id as mat_id
                FROM work_orders wo
                LEFT JOIN materials m ON wo.material_id = m.id
                WHERE wo.id = %s
            """, (ot_id,))

            ot = cursor.fetchone()
            if not ot:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Orden de trabajo no encontrada"
                )

            material_id = ot.get("material_id")
            if not material_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="La OT no tiene un material asignado. Debe crear CAD y Material primero."
                )

            # Generar código principal: prefijo_ot + codigo_material + "-" + sufijo_id
            codigo_final = f"{data.prefijo_ot}{data.codigo_material}-{data.sufijo_id}"

            # Determinar descripción según tipo_solicitud (Laravel líneas 10494-10499)
            tipo_solicitud = ot.get("tipo_solicitud")
            if tipo_solicitud in [1, 7]:
                descripcion_material = ot.get("descripcion_material")
            elif tipo_solicitud == 5:
                descripcion_material = data.descripcion
            else:
                descripcion_material = ot.get("descripcion_material")

            # Actualizar material principal (Laravel líneas 10492-10502)
            cursor.execute("""
                UPDATE materials
                SET codigo = %s,
                    descripcion = %s,
                    updated_at = NOW()
                WHERE id = %s
            """, (codigo_final, descripcion_material, material_id))

            # Actualizar OT (Laravel líneas 10500-10503)
            cursor.execute("""
                UPDATE work_orders
                SET material_asignado = %s,
                    descripcion_material = %s,
                    codigo_sap_final = 1,
                    updated_at = NOW()
                WHERE id = %s
            """, (codigo_final, descripcion_material, ot_id))

            materiales_clonados = []

            # Clonar materiales con prefijos adicionales (Laravel líneas 10506-10519)
            if data.prefijos_adicionales:
                # Obtener datos del material para clonar
                cursor.execute("SELECT * FROM materials WHERE id = %s", (material_id,))
                material_original = cursor.fetchone()

                for prefijo in data.prefijos_adicionales:
                    codigo_clonado = f"{prefijo}{data.codigo_material}-{data.sufijo_id}"

                    # Clonar material con nuevo código
                    cursor.execute("""
                        INSERT INTO materials (
                            codigo, descripcion, client_id, vendedor_id, carton_id,
                            product_type_id, style_id, numero_colores, cad_id,
                            pallet_type_id, pallet_box_quantity, placas_por_pallet,
                            pallet_patron_id, patron_zuncho_pallet, pallet_protection_id,
                            boxes_per_package, patron_zuncho_paquete, paquetes_por_unitizado,
                            unitizado_por_pallet, pallet_tag_format_id, fecha_creacion,
                            creador_id, pallet_qa_id, numero_etiquetas,
                            bct_min_lb, bct_min_kg, pallet_treatment, sap_hiearchy_id,
                            tipo_camion, restriccion_especial, horario_recepcion,
                            codigo_producto_cliente, etiquetas_dsc, orientacion_placa,
                            recubrimiento, work_order_id,
                            cinta, corte_liner, tipo_cinta,
                            distancia_cinta_1, distancia_cinta_2, distancia_cinta_3,
                            distancia_cinta_4, distancia_cinta_5, distancia_cinta_6,
                            gramaje, ect, flexion_aleta, peso,
                            incision_rayado_longitudinal, incision_rayado_vertical,
                            fct, cobb_interior, cobb_exterior, espesor,
                            active, created_at, updated_at
                        )
                        SELECT
                            %s, %s, client_id, vendedor_id, carton_id,
                            product_type_id, style_id, numero_colores, cad_id,
                            pallet_type_id, pallet_box_quantity, placas_por_pallet,
                            pallet_patron_id, patron_zuncho_pallet, pallet_protection_id,
                            boxes_per_package, patron_zuncho_paquete, paquetes_por_unitizado,
                            unitizado_por_pallet, pallet_tag_format_id, NOW(),
                            creador_id, pallet_qa_id, numero_etiquetas,
                            bct_min_lb, bct_min_kg, pallet_treatment, sap_hiearchy_id,
                            tipo_camion, restriccion_especial, horario_recepcion,
                            codigo_producto_cliente, etiquetas_dsc, orientacion_placa,
                            recubrimiento, work_order_id,
                            cinta, corte_liner, tipo_cinta,
                            distancia_cinta_1, distancia_cinta_2, distancia_cinta_3,
                            distancia_cinta_4, distancia_cinta_5, distancia_cinta_6,
                            gramaje, ect, flexion_aleta, peso,
                            incision_rayado_longitudinal, incision_rayado_vertical,
                            fct, cobb_interior, cobb_exterior, espesor,
                            active, NOW(), NOW()
                        FROM materials WHERE id = %s
                    """, (codigo_clonado, descripcion_material, material_id))

                    materiales_clonados.append({
                        "id": cursor.lastrowid,
                        "codigo": codigo_clonado,
                        "prefijo": prefijo
                    })

            connection.commit()

            return {
                "success": True,
                "message": "Código de material editado correctamente",
                "codigo_material": codigo_final,
                "descripcion": descripcion_material,
                "material_id": material_id,
                "ot_id": ot_id,
                "materiales_clonados": materiales_clonados
            }

    except HTTPException:
        raise
    except pymysql.Error as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al crear código de material: {str(e)}"
        )
    finally:
        connection.close()


# =============================================
# SPRINT C: IMPORTAR MUESTRAS DESDE EXCEL
# =============================================

@router.post("/importar-muestras-excel")
async def importar_muestras_desde_excel(
    ot_muestra: int = Query(..., description="ID de la OT"),
    licitacion_file: UploadFile = File(...)
):
    """
    Importa muestras desde un archivo Excel.

    Basado en: WorkOrderController@importarMuestrasDesdeExcel (líneas 4724-4950)

    Columnas obligatorias del Excel:
    - descripcion, largo_int., ancho_int., alto_int., carton,
    - destinatario, cantidad, tipo_pegado, planta_corte

    Args:
        ot_muestra: ID de la OT origen
        licitacion_file: Archivo Excel con las muestras

    Returns:
        Resumen de muestras importadas
    """
    # Validar extensión
    filename = licitacion_file.filename or ""
    extension = filename.split(".")[-1].lower() if "." in filename else ""
    if extension not in ["xlsx", "xls"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato de archivo no válido. Use .xlsx o .xls"
        )

    connection = get_mysql_connection()
    try:
        # Verificar OT existe
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT id, descripcion FROM work_orders WHERE id = %s",
                (ot_muestra,)
            )
            ot = cursor.fetchone()
            if not ot:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="No se encontró el WorkOrder indicado"
                )

        # Leer Excel
        content = await licitacion_file.read()
        workbook = openpyxl.load_workbook(BytesIO(content))
        sheet = workbook.active

        # Obtener headers
        headers = []
        for cell in sheet[1]:
            if cell.value:
                header = str(cell.value).lower().strip().replace(" ", "_")
                # Normalizar headers con alias
                header = header.replace("largo_int", "largo_int.")
                header = header.replace("ancho_int", "ancho_int.")
                header = header.replace("alto_int", "alto_int.")
                headers.append(header)
            else:
                headers.append(None)

        # Columnas obligatorias
        obligatorias = [
            "descripcion", "largo_int.", "ancho_int.", "alto_int.",
            "carton", "destinatario", "cantidad", "tipo_pegado", "planta_corte"
        ]

        # Verificar columnas faltantes
        columnas_faltantes = [
            col for col in obligatorias
            if col not in headers and col.replace(".", "") not in headers
        ]

        if columnas_faltantes:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Columnas faltantes en el Excel: {', '.join(columnas_faltantes)}"
            )

        # Mapear índices
        header_map = {}
        for i, h in enumerate(headers):
            if h:
                header_map[h] = i
                # También mapear versión sin punto
                header_map[h.replace(".", "")] = i

        # Procesar filas
        muestras_creadas = []
        errores = []

        with connection.cursor() as cursor:
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None for cell in row):
                    continue

                try:
                    # Extraer valores
                    descripcion = row[header_map.get("descripcion", 0)] or ""
                    largo_int = row[header_map.get("largo_int.", header_map.get("largo_int", 1))]
                    ancho_int = row[header_map.get("ancho_int.", header_map.get("ancho_int", 2))]
                    alto_int = row[header_map.get("alto_int.", header_map.get("alto_int", 3))]
                    carton = row[header_map.get("carton", 4)] or ""
                    destinatario_texto = str(row[header_map.get("destinatario", 5)] or "").strip()
                    cantidad = int(row[header_map.get("cantidad", 6)] or 1)
                    tipo_pegado_texto = str(row[header_map.get("tipo_pegado", 7)] or "").strip()
                    planta_corte_texto = str(row[header_map.get("planta_corte", 8)] or "").strip()
                    forma_entrega = row[header_map.get("forma_de_entrega", header_map.get("forma_entrega", -1))]
                    forma_entrega = str(forma_entrega) if forma_entrega and forma_entrega != -1 else ""

                    # Inicializar variables según Laravel (líneas 4938-4944)
                    destinatarios_id = []
                    cantidad_disenador = 0
                    cantidad_laboratorio = 0
                    sala_corte_disenador = 0
                    sala_corte_laboratorio = 0
                    comentario_disenador = ""
                    comentario_laboratorio = ""

                    # Switch planta_corte (Laravel líneas 4947-4957)
                    planta = 0
                    if planta_corte_texto == "Osorno":
                        planta = 1
                    elif planta_corte_texto == "Puente Alto":
                        planta = 2

                    # Switch destinatario (Laravel líneas 4959-4975)
                    if destinatario_texto == "Retira Diseñador":
                        destinatarios_id.append(2)
                        cantidad_disenador = cantidad
                        sala_corte_disenador = planta
                        comentario_disenador = forma_entrega
                    elif destinatario_texto == "Envío Laboratorio":
                        destinatarios_id.append(3)
                        cantidad_laboratorio = cantidad
                        sala_corte_laboratorio = planta
                        comentario_laboratorio = forma_entrega
                    else:
                        destinatarios_id.append(0)

                    # Switch tipo_pegado (Laravel líneas 4979-4994)
                    pegado_id = 1  # default
                    if tipo_pegado_texto == "Pegado Flexo Interior":
                        pegado_id = 2
                    elif tipo_pegado_texto == "Pegado Flexo Exterior":
                        pegado_id = 3
                    elif tipo_pegado_texto == "Pegado Diecutter":
                        pegado_id = 4
                    elif tipo_pegado_texto == "Pegado Cajas Fruta":
                        pegado_id = 5

                    # Convertir destinatarios_id a JSON
                    destinatarios_json = json.dumps(destinatarios_id)

                    # Crear muestra vinculada a la OT (según Laravel líneas 4999-5033)
                    cursor.execute("""
                        INSERT INTO muestras (
                            work_order_id, descripcion_muestra,
                            largo_int, ancho_int, alto_int,
                            destinatarios_id, pegado_id,
                            cantidad_disenador, cantidad_laboratorio,
                            sala_corte_disenador, sala_corte_laboratorio,
                            comentario_disenador, comentario_laboratorio,
                            forma_entrega, muestra_excel,
                            created_at, updated_at
                        ) VALUES (
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1,
                            NOW(), NOW()
                        )
                    """, (
                        ot_muestra, descripcion,
                        largo_int, ancho_int, alto_int,
                        destinatarios_json, pegado_id,
                        cantidad_disenador, cantidad_laboratorio,
                        sala_corte_disenador, sala_corte_laboratorio,
                        comentario_disenador, comentario_laboratorio,
                        forma_entrega
                    ))

                    muestras_creadas.append({
                        "fila": row_num,
                        "descripcion": descripcion,
                        "id": cursor.lastrowid
                    })

                except Exception as e:
                    errores.append({
                        "fila": row_num,
                        "error": str(e)
                    })

            connection.commit()

        return {
            "success": len(errores) == 0,
            "mensaje": f"Se importaron {len(muestras_creadas)} muestras",
            "ot_id": ot_muestra,
            "total_importadas": len(muestras_creadas),
            "total_errores": len(errores),
            "muestras": muestras_creadas,
            "errores": errores
        }

    except HTTPException:
        raise
    except Exception as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al importar muestras: {str(e)}"
        )
    finally:
        connection.close()


# =============================================
# SPRINT C: COTIZAR MÚLTIPLES OTs
# =============================================

@router.get("/cotizar-multiples")
async def cotizar_multiples_ot(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    client_id: Optional[int] = None,
    estado_id: Optional[int] = None
):
    """
    Lista OTs disponibles para cotizar en múltiples.

    Basado en: WorkOrderController@cotizarMultiplesOt (líneas 11247-11380)

    Devuelve las OTs que pueden ser cotizadas en grupo.

    Args:
        page: Página actual
        per_page: Registros por página
        client_id: Filtrar por cliente
        estado_id: Filtrar por estado

    Returns:
        Lista paginada de OTs para cotizar
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Base query
            where_clauses = ["wo.active = 1"]
            params = []

            if client_id:
                where_clauses.append("wo.client_id = %s")
                params.append(client_id)

            if estado_id:
                where_clauses.append("wo.estado_id = %s")
                params.append(estado_id)

            where_sql = " AND ".join(where_clauses)

            # Contar total
            cursor.execute(f"""
                SELECT COUNT(*) as total
                FROM work_orders wo
                WHERE {where_sql}
            """, params)
            total = cursor.fetchone()["total"]

            # Obtener registros paginados
            offset = (page - 1) * per_page
            params.extend([per_page, offset])

            cursor.execute(f"""
                SELECT
                    wo.id,
                    wo.codigo,
                    wo.descripcion,
                    wo.cantidad,
                    wo.created_at,
                    c.nombre as cliente_nombre,
                    e.descripcion as estado_nombre,
                    ca.descripcion as canal_nombre
                FROM work_orders wo
                LEFT JOIN clients c ON wo.client_id = c.id
                LEFT JOIN estados e ON wo.estado_id = e.id
                LEFT JOIN canales ca ON wo.canal_id = ca.id
                WHERE {where_sql}
                ORDER BY wo.created_at DESC
                LIMIT %s OFFSET %s
            """, params)

            ots = cursor.fetchall()

            return {
                "success": True,
                "data": ots,
                "pagination": {
                    "total": total,
                    "page": page,
                    "per_page": per_page,
                    "total_pages": (total + per_page - 1) // per_page
                }
            }

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener OTs para cotizar: {str(e)}"
        )
    finally:
        connection.close()


# =============================================
# SPRINT H: ENDPOINTS ADICIONALES DE OT
# Fuente Laravel: WorkOrderController.php
# =============================================

class UpdateDescripcionRequest(BaseModel):
    """Schema para actualizar descripción de OT."""
    descripcion: Optional[str] = Field(None, max_length=100)
    oc: Optional[int] = None
    type_edit: str = Field(..., description="'description' o 'oc'")
    # Datos McKee opcionales
    aplicar_mckee: Optional[int] = None
    carton_id_mckee: Optional[int] = None
    largo_mckee: Optional[int] = None
    ancho_mckee: Optional[int] = None
    alto_mckee: Optional[int] = None
    perimetro_mckee: Optional[int] = None
    espesor_mckee: Optional[float] = None
    ect_mckee: Optional[float] = None
    bct_lib_mckee: Optional[float] = None
    bct_kilos_mckee: Optional[float] = None


class UpdateDescripcionResponse(BaseModel):
    """Respuesta de actualización de descripción."""
    success: bool
    message: str
    recordatorio_fsc: bool = False


@router.put("/{ot_id}/descripcion", response_model=UpdateDescripcionResponse)
async def update_descripcion(
    ot_id: int,
    data: UpdateDescripcionRequest,
    user: dict = Depends(get_current_user_with_role)
):
    """
    Actualiza la descripción o OC de una OT con registro en bitácora.

    Fuente Laravel: WorkOrderController.php líneas 9670-9909

    Lógica:
    - type_edit='description': Actualiza descripción de OT y material
    - type_edit='oc': Actualiza campo OC y crea gestión automática
    - Registra cambios en bitacora_work_orders
    - Si aplicar_mckee=1, registra datos de fórmula McKee
    """
    user_id = user["user_id"]
    role_id = user["role_id"]
    connection = get_mysql_connection()

    try:
        with connection.cursor() as cursor:
            # Obtener OT actual
            cursor.execute("""
                SELECT id, descripcion, cad, material_id, fsc, oc
                FROM work_orders
                WHERE id = %s AND active = 1
            """, (ot_id,))
            ot = cursor.fetchone()

            if not ot:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"OT {ot_id} no encontrada"
                )

            # Obtener datos del usuario
            cursor.execute("""
                SELECT nombre, apellido, rut, role_id
                FROM users WHERE id = %s
            """, (user_id,))
            user_data = cursor.fetchone()

            datos_modificados = {}
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            if data.type_edit == 'description':
                # Actualizar descripción
                if data.descripcion and data.descripcion.strip():
                    nueva_desc = data.descripcion.strip()

                    # Registrar cambio en bitácora si es diferente
                    if ot['descripcion'] != nueva_desc:
                        datos_modificados['descripcion'] = {
                            'texto': 'Descripción',
                            'antiguo_valor': {'id': None, 'descripcion': ot['descripcion']},
                            'nuevo_valor': {'id': None, 'descripcion': nueva_desc}
                        }

                    # Actualizar OT
                    cursor.execute("""
                        UPDATE work_orders
                        SET descripcion = %s, descripcion_material = %s, updated_at = %s
                        WHERE id = %s
                    """, (nueva_desc, nueva_desc, now, ot_id))

                    # Si tiene material, actualizar también
                    if ot['material_id']:
                        cursor.execute("""
                            UPDATE materials
                            SET descripcion = %s, updated_at = %s
                            WHERE id = %s
                        """, (nueva_desc, now, ot['material_id']))

            elif data.type_edit == 'oc':
                # Actualizar OC
                if data.oc is not None and str(ot['oc']) != str(data.oc):
                    datos_modificados['oc'] = {
                        'texto': 'OC',
                        'antiguo_valor': {
                            'id': ot['oc'],
                            'descripcion': 'Sí' if ot['oc'] == 1 else ('No' if ot['oc'] == 0 else None)
                        },
                        'nuevo_valor': {
                            'id': data.oc,
                            'descripcion': 'Sí' if data.oc == 1 else ('No' if data.oc == 0 else None)
                        }
                    }

                    # Actualizar OC
                    cursor.execute("""
                        UPDATE work_orders SET oc = %s, updated_at = %s WHERE id = %s
                    """, (data.oc, now, ot_id))

                    # Crear gestión automática de modificación OC
                    cursor.execute("""
                        INSERT INTO managements
                        (observacion, management_type_id, user_id, work_order_id, work_space_id,
                         duracion_segundos, state_id, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        "Modificación de orden de compra",
                        5,  # Tipo modificacion OC
                        user_id,
                        ot_id,
                        1,  # work_space ventas
                        0,
                        19,  # Estado modificacion OC
                        now,
                        now
                    ))

            # Registrar en bitácora si hubo cambios
            if datos_modificados:
                user_data_json = json.dumps({
                    'nombre': user_data['nombre'],
                    'apellido': user_data['apellido'],
                    'rut': user_data['rut'],
                    'role_id': user_data['role_id']
                }, ensure_ascii=False)

                cursor.execute("""
                    INSERT INTO bitacora_work_orders
                    (observacion, operacion, work_order_id, user_id, user_data,
                     datos_modificados, ip_solicitud, url, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    "Modificación de datos de OT",
                    "Modificación",
                    ot_id,
                    user_id,
                    user_data_json,
                    json.dumps(datos_modificados, ensure_ascii=False),
                    "API",
                    f"/work-orders/{ot_id}/descripcion",
                    now,
                    now
                ))

            # Registrar datos McKee si aplica
            if data.aplicar_mckee == 1:
                datos_mckee = {}

                if data.carton_id_mckee:
                    cursor.execute("SELECT codigo FROM cartons WHERE id = %s", (data.carton_id_mckee,))
                    carton = cursor.fetchone()
                    if carton:
                        datos_mckee['carton'] = {'texto': 'Carton', 'valor': carton['codigo']}

                if data.largo_mckee:
                    datos_mckee['largo'] = {'texto': 'Largo', 'valor': data.largo_mckee}
                if data.ancho_mckee:
                    datos_mckee['ancho'] = {'texto': 'Ancho', 'valor': data.ancho_mckee}
                if data.alto_mckee:
                    datos_mckee['alto'] = {'texto': 'Alto', 'valor': data.alto_mckee}
                if data.perimetro_mckee:
                    datos_mckee['perimetro'] = {'texto': 'Perimetro', 'valor': data.perimetro_mckee}
                if data.espesor_mckee:
                    datos_mckee['espesor'] = {'texto': 'Espesor', 'valor': data.espesor_mckee}
                if data.ect_mckee:
                    datos_mckee['ect'] = {'texto': 'Ect', 'valor': data.ect_mckee}
                if data.bct_lib_mckee:
                    datos_mckee['bct_lb'] = {'texto': 'Bct_lb', 'valor': data.bct_lib_mckee}
                if data.bct_kilos_mckee:
                    datos_mckee['bct_kilos'] = {'texto': 'Bct_kilos', 'valor': data.bct_kilos_mckee}

                if datos_mckee:
                    user_data_json = json.dumps({
                        'nombre': user_data['nombre'],
                        'apellido': user_data['apellido'],
                        'rut': user_data['rut'],
                        'role_id': user_data['role_id']
                    }, ensure_ascii=False)

                    cursor.execute("""
                        INSERT INTO bitacora_work_orders
                        (observacion, operacion, work_order_id, user_id, user_data,
                         datos_modificados, ip_solicitud, url, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        "Aplicacion Formula Mckee",
                        "Mckee",
                        ot_id,
                        user_id,
                        user_data_json,
                        json.dumps(datos_mckee, ensure_ascii=False),
                        "API",
                        f"/work-orders/{ot_id}/descripcion",
                        now,
                        now
                    ))

            connection.commit()

            # Recordatorio FSC para diseñadores
            recordatorio_fsc = (ot['fsc'] == 1 and role_id in [7, 8])

            return UpdateDescripcionResponse(
                success=True,
                message="Datos actualizados correctamente",
                recordatorio_fsc=recordatorio_fsc
            )

    except pymysql.Error as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al actualizar descripción: {str(e)}"
        )
    finally:
        connection.close()


class AplicarMckeeResponse(BaseModel):
    """Respuesta de aplicar fórmula McKee."""
    success: bool
    datos_insertados: dict
    message: str


@router.post("/aplicar-mckee", response_model=AplicarMckeeResponse)
async def aplicar_mckee(
    carton: Optional[int] = Query(None),
    largo: Optional[int] = Query(None),
    ancho: Optional[int] = Query(None),
    alto: Optional[int] = Query(None),
    perimetro: Optional[int] = Query(None),
    espesor: Optional[float] = Query(None),
    ect: Optional[float] = Query(None),
    bct_lb: Optional[float] = Query(None),
    bct_kilos: Optional[float] = Query(None),
    user: dict = Depends(get_current_user_with_role)
):
    """
    Registra aplicación de fórmula McKee en bitácora.

    Fuente Laravel: WorkOrderController.php líneas 11365-11455

    Esta función registra los datos ingresados en la fórmula McKee
    para trazabilidad y auditoría.
    """
    user_id = user["user_id"]
    connection = get_mysql_connection()

    try:
        with connection.cursor() as cursor:
            datos_insertados = {}

            # Obtener código de cartón si se proporciona ID
            if carton:
                cursor.execute("SELECT codigo FROM cartons WHERE id = %s AND active = 1", (carton,))
                carton_data = cursor.fetchone()
                if carton_data:
                    datos_insertados['carton'] = {'texto': 'Carton', 'valor': carton_data['codigo']}

            if largo:
                datos_insertados['largo'] = {'texto': 'Largo', 'valor': largo}
            if ancho:
                datos_insertados['ancho'] = {'texto': 'Ancho', 'valor': ancho}
            if alto:
                datos_insertados['alto'] = {'texto': 'Alto', 'valor': alto}
            if perimetro:
                datos_insertados['perimetro'] = {'texto': 'Perimetro', 'valor': perimetro}
            if espesor:
                datos_insertados['espesor'] = {'texto': 'Espesor', 'valor': espesor}
            if ect:
                datos_insertados['ect'] = {'texto': 'Ect', 'valor': ect}
            if bct_lb:
                datos_insertados['bct_lb'] = {'texto': 'Bct_lb', 'valor': bct_lb}
            if bct_kilos:
                datos_insertados['bct_kilos'] = {'texto': 'Bct_kilos', 'valor': bct_kilos}

            if datos_insertados:
                # Obtener datos del usuario
                cursor.execute("""
                    SELECT nombre, apellido, rut, role_id
                    FROM users WHERE id = %s
                """, (user_id,))
                user_data = cursor.fetchone()

                user_data_json = json.dumps({
                    'nombre': user_data['nombre'],
                    'apellido': user_data['apellido'],
                    'rut': user_data['rut'],
                    'role_id': user_data['role_id']
                }, ensure_ascii=False)

                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                # Registrar en bitácora con work_order_id=909999 como en Laravel
                cursor.execute("""
                    INSERT INTO bitacora_work_orders
                    (observacion, operacion, work_order_id, user_id, user_data,
                     datos_modificados, ip_solicitud, url, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    "Aplicacion Formula Mckee",
                    "Mckee",
                    909999,  # ID especial para fórmula McKee sin OT específica
                    user_id,
                    user_data_json,
                    json.dumps(datos_insertados, ensure_ascii=False),
                    "API",
                    "/work-orders/aplicar-mckee",
                    now,
                    now
                ))

                connection.commit()

            return AplicarMckeeResponse(
                success=True,
                datos_insertados=datos_insertados,
                message="Fórmula McKee aplicada y registrada correctamente"
            )

    except pymysql.Error as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al aplicar fórmula McKee: {str(e)}"
        )
    finally:
        connection.close()


class OTDataCompleteResponse(BaseModel):
    """Respuesta con datos completos de OT para edición."""
    success: bool
    data: dict


@router.get("/{ot_id}/data-complete")
async def get_ot_data_complete(ot_id: int):
    """
    Obtiene todos los datos de una OT para edición completa.

    Fuente Laravel: WorkOrderController.php getOtData() implícito en edit()

    Retorna:
    - Datos de la OT
    - Relaciones (cliente, instalación, contacto, cartón, CAD, etc.)
    - Datos calculados
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Obtener OT con todas las relaciones
            cursor.execute("""
                SELECT
                    wo.*,
                    c.nombre as cliente_nombre,
                    c.codigo as cliente_codigo,
                    c.rut as cliente_rut,
                    inst.nombre as instalacion_nombre,
                    inst.direccion as instalacion_direccion,
                    cart.codigo as carton_codigo,
                    cart.peso as carton_gramaje,
                    cart.espesor as carton_espesor,
                    cart.ect as carton_ect,
                    cart.tipo as carton_tipo,
                    cad.cad as cad_codigo,
                    cad.largura_hm as cad_largura_hm,
                    cad.anchura_hm as cad_anchura_hm,
                    cad.area_producto as cad_area_producto,
                    mat.codigo as material_codigo,
                    mat.descripcion as material_descripcion,
                    cn.nombre as canal_nombre,
                    pt.descripcion as product_type_nombre,
                    stl.glosa as estilo_nombre,
                    imp.descripcion as impresion_nombre,
                    proc.descripcion as proceso_nombre,
                    arm.descripcion as armado_nombre,
                    cov_int.descripcion as recubrimiento_interno_nombre,
                    cov_ext.descripcion as recubrimiento_externo_nombre,
                    pl.nombre as planta_nombre,
                    CONCAT(u.nombre, ' ', u.apellido) as creador_nombre
                FROM work_orders wo
                LEFT JOIN clients c ON wo.client_id = c.id
                LEFT JOIN instalaciones inst ON wo.instalacion_cliente = inst.id
                LEFT JOIN cartons cart ON wo.carton_id = cart.id
                LEFT JOIN cads cad ON wo.cad_id = cad.id
                LEFT JOIN materials mat ON wo.material_id = mat.id
                LEFT JOIN canals cn ON wo.canal_id = cn.id
                LEFT JOIN product_types pt ON wo.product_type_id = pt.id
                LEFT JOIN styles stl ON wo.style_id = stl.id
                LEFT JOIN impresion imp ON wo.impresion = imp.id
                LEFT JOIN processes proc ON wo.process_id = proc.id
                LEFT JOIN armados arm ON wo.armado_id = arm.id
                LEFT JOIN coverage_internals cov_int ON wo.coverage_internal_id = cov_int.id
                LEFT JOIN coverage_externals cov_ext ON wo.coverage_external_id = cov_ext.id
                LEFT JOIN plantas pl ON wo.planta_id = pl.id
                LEFT JOIN users u ON wo.creador_id = u.id
                WHERE wo.id = %s
            """, (ot_id,))

            ot = cursor.fetchone()

            if not ot:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"OT {ot_id} no encontrada"
                )

            # Convertir datetime a string
            for key, value in ot.items():
                if isinstance(value, datetime):
                    ot[key] = value.isoformat()

            # Obtener colores si existen
            colores = []
            for i in range(1, 6):
                color_id = ot.get(f'color_{i}_id')
                if color_id:
                    cursor.execute("SELECT id, descripcion FROM colors WHERE id = %s", (color_id,))
                    color = cursor.fetchone()
                    if color:
                        colores.append(color)
            ot['colores'] = colores

            # Obtener jerarquía si existe
            if ot.get('subsubhierarchy_id'):
                cursor.execute("""
                    SELECT
                        h3.id as subsubhierarchy_id,
                        h3.descripcion as jerarquia_3,
                        h2.descripcion as jerarquia_2,
                        h1.descripcion as jerarquia_1
                    FROM subsubhierarchies h3
                    LEFT JOIN subhierarchies h2 ON h3.subhierarchy_id = h2.id
                    LEFT JOIN hierarchies h1 ON h2.hierarchy_id = h1.id
                    WHERE h3.id = %s
                """, (ot['subsubhierarchy_id'],))
                jerarquia = cursor.fetchone()
                if jerarquia:
                    ot['jerarquia'] = jerarquia

            # Obtener último estado
            cursor.execute("""
                SELECT s.nombre as estado_nombre, s.abreviatura as estado_abrev
                FROM managements m
                JOIN states s ON m.state_id = s.id
                WHERE m.work_order_id = %s
                ORDER BY m.id DESC
                LIMIT 1
            """, (ot_id,))
            estado = cursor.fetchone()
            if estado:
                ot['estado_nombre'] = estado['estado_nombre']
                ot['estado_abrev'] = estado['estado_abrev']

            return {"success": True, "data": ot}

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener datos de OT: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/search-matriz-cad")
async def search_matriz_cad(cad: int = Query(..., description="ID del CAD")):
    """
    Busca matrices asociadas a un CAD.

    Fuente Laravel: WorkOrderController.php líneas 11853-11862

    Args:
        cad: ID del CAD a buscar

    Returns:
        Lista de matrices asociadas al CAD
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            # Obtener el código del CAD
            cursor.execute("SELECT cad FROM cads WHERE id = %s", (cad,))
            cad_data = cursor.fetchone()

            if not cad_data:
                return {"success": True, "data": []}

            # Buscar matrices con ese plano_cad
            cursor.execute("""
                SELECT * FROM matrices WHERE plano_cad = %s
            """, (cad_data['cad'],))

            matrices = cursor.fetchall()

            # Convertir datetime a string
            for matriz in matrices:
                for key, value in matriz.items():
                    if isinstance(value, datetime):
                        matriz[key] = value.isoformat()

            return {"success": True, "data": matrices}

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al buscar matriz por CAD: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/secuencias-operacionales-planta")
async def get_secuencias_operacionales_planta(
    planta_id: int = Query(..., description="ID de la planta")
):
    """
    Obtiene secuencias operacionales disponibles para una planta.

    Fuente Laravel: WorkOrderController.php líneas 11597-11618

    Args:
        planta_id: ID de la planta

    Returns:
        Lista de secuencias operacionales
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, descripcion
                FROM secuencia_operacionales
                WHERE active = 1 AND planta_id = %s
                ORDER BY descripcion
            """, (planta_id,))

            secuencias = cursor.fetchall()

            return {
                "success": True,
                "data": secuencias
            }

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener secuencias operacionales: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/secuencias-operacionales-ot")
async def get_secuencias_operacionales_ot(
    ot_id: int = Query(..., description="ID de la OT")
):
    """
    Obtiene datos de secuencia operacional de una OT.

    Fuente Laravel: WorkOrderController.php líneas 11620-11628

    Args:
        ot_id: ID de la OT

    Returns:
        Datos de la OT relacionados a secuencia operacional
    """
    connection = get_mysql_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT
                    wo.id,
                    wo.so_planta_original,
                    wo.so_planta_original_select_values,
                    wo.so_planta_alt1,
                    wo.so_planta_alt1_select_values,
                    wo.so_planta_alt2,
                    wo.so_planta_alt2_select_values,
                    wo.check_planta_alt1,
                    wo.check_planta_alt2,
                    pl.nombre as planta_nombre
                FROM work_orders wo
                LEFT JOIN plantas pl ON wo.so_planta_original = pl.id
                WHERE wo.id = %s
            """, (ot_id,))

            ot = cursor.fetchone()

            if not ot:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"OT {ot_id} no encontrada"
                )

            # Parsear JSON de select_values si existe
            if ot.get('so_planta_original_select_values'):
                try:
                    ot['so_planta_original_select_values'] = json.loads(
                        ot['so_planta_original_select_values']
                    )
                except json.JSONDecodeError:
                    pass

            if ot.get('so_planta_alt1_select_values'):
                try:
                    ot['so_planta_alt1_select_values'] = json.loads(
                        ot['so_planta_alt1_select_values']
                    )
                except json.JSONDecodeError:
                    pass

            if ot.get('so_planta_alt2_select_values'):
                try:
                    ot['so_planta_alt2_select_values'] = json.loads(
                        ot['so_planta_alt2_select_values']
                    )
                except json.JSONDecodeError:
                    pass

            return {"success": True, "data": ot}

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener secuencias de OT: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/filtro-multiples-ot")
async def filtro_multiples_ot(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    # Filtros básicos
    id: Optional[int] = None,
    material: Optional[str] = None,
    cad: Optional[str] = None,
    carton: Optional[str] = None,
    descripcion: Optional[str] = None,
    # Filtros de listas
    client_id: Optional[List[int]] = Query(None),
    canal_id: Optional[List[int]] = Query(None),
    vendedor_id: Optional[List[int]] = Query(None),
    estado_id: Optional[List[int]] = Query(None),
    planta_id: Optional[List[int]] = Query(None),
    impresion_id: Optional[List[int]] = Query(None),
    proceso_id: Optional[List[int]] = Query(None),
    style_id: Optional[List[int]] = Query(None),
    solicitud_id: Optional[List[int]] = Query(None),
    cinta_id: Optional[List[int]] = Query(None),
    fsc_codigo: Optional[List[str]] = Query(None),
    # Filtros de fechas
    date_desde: Optional[str] = None,
    date_hasta: Optional[str] = None,
    # Usuario
    current_user: dict = Depends(get_current_user_with_role)
):
    """
    Filtra OTs con múltiples criterios para cotización múltiple.

    Fuente Laravel: WorkOrderController.php método filtro() líneas 70-310
    y cotizarMultiplesOt() líneas 11247-11362

    Este endpoint replica la lógica completa del filtro de Laravel,
    incluyendo filtrado por rol, tiempos por área, y estados activos.

    Returns:
        Lista paginada de OTs filtradas con tiempos calculados
    """
    user_id = current_user["user_id"]
    role_id = current_user["role_id"]
    connection = get_mysql_connection()

    try:
        with connection.cursor() as cursor:
            # Query base
            base_query = """
                SELECT
                    wo.id,
                    wo.descripcion,
                    wo.tipo_solicitud,
                    wo.cad,
                    wo.created_at,
                    wo.current_area_id,
                    c.nombre as client_name,
                    cn.nombre as canal_nombre,
                    CONCAT(u.nombre, ' ', u.apellido) as creador_nombre,
                    mat.codigo as material_codigo,
                    cart.codigo as carton_codigo,
                    cad_t.cad as cad_codigo,
                    s.nombre as estado_nombre,
                    s.abreviatura as estado_abrev,
                    -- Tiempos por área
                    COALESCE((SELECT SUM(duracion_segundos) FROM managements
                        WHERE work_order_id = wo.id AND management_type_id = 1 AND work_space_id = 1 AND mostrar = 1), 0) / 3600.0 as tiempo_venta,
                    COALESCE((SELECT SUM(duracion_segundos) FROM managements
                        WHERE work_order_id = wo.id AND management_type_id = 1 AND work_space_id = 2 AND mostrar = 1), 0) / 3600.0 as tiempo_desarrollo,
                    COALESCE((SELECT SUM(duracion_segundos) FROM managements
                        WHERE work_order_id = wo.id AND management_type_id = 1 AND work_space_id = 3 AND mostrar = 1), 0) / 3600.0 as tiempo_diseno,
                    COALESCE((SELECT SUM(duracion_segundos) FROM managements
                        WHERE work_order_id = wo.id AND management_type_id = 1 AND work_space_id = 4 AND mostrar = 1), 0) / 3600.0 as tiempo_catalogacion,
                    COALESCE((SELECT SUM(duracion_segundos) FROM managements
                        WHERE work_order_id = wo.id AND management_type_id = 1 AND work_space_id = 5 AND mostrar = 1), 0) / 3600.0 as tiempo_precatalogacion,
                    COALESCE((SELECT SUM(duracion_segundos) FROM managements
                        WHERE work_order_id = wo.id AND management_type_id = 1 AND work_space_id = 6 AND mostrar = 1), 0) / 3600.0 as tiempo_muestra,
                    COALESCE((SELECT SUM(duracion_segundos) FROM managements
                        WHERE work_order_id = wo.id AND management_type_id = 1 AND mostrar = 1), 0) / 3600.0 as tiempo_total
                FROM work_orders wo
                LEFT JOIN clients c ON wo.client_id = c.id
                LEFT JOIN canals cn ON wo.canal_id = cn.id
                LEFT JOIN users u ON wo.creador_id = u.id
                LEFT JOIN materials mat ON wo.material_id = mat.id
                LEFT JOIN cartons cart ON wo.carton_id = cart.id
                LEFT JOIN cads cad_t ON wo.cad_id = cad_t.id
                LEFT JOIN (
                    SELECT work_order_id, state_id
                    FROM managements
                    WHERE id IN (SELECT MAX(id) FROM managements GROUP BY work_order_id)
                ) m ON wo.id = m.work_order_id
                LEFT JOIN states s ON m.state_id = s.id
                WHERE wo.active = 1
            """

            params = []

            # Filtro por ID
            if id:
                base_query += " AND wo.id = %s"
                params.append(id)

            # Filtro por material
            if material:
                base_query += " AND mat.codigo LIKE %s"
                params.append(f"%{material}%")

            # Filtro por CAD
            if cad:
                base_query += " AND cad_t.cad LIKE %s"
                params.append(f"%{cad}%")

            # Filtro por cartón
            if carton:
                base_query += " AND cart.codigo LIKE %s"
                params.append(f"%{carton}%")

            # Filtro por descripción
            if descripcion:
                base_query += " AND wo.descripcion LIKE %s"
                params.append(f"%{descripcion}%")

            # Filtro por vendedor (según rol)
            if vendedor_id:
                placeholders = ','.join(['%s'] * len(vendedor_id))
                base_query += f" AND wo.creador_id IN ({placeholders})"
                params.extend(vendedor_id)
            elif role_id == 4:  # Vendedor
                base_query += " AND wo.creador_id = %s"
                params.append(user_id)
            elif role_id == 19:  # VendedorExterno
                base_query += " AND wo.creador_id = %s"
                params.append(user_id)

            # Filtros de listas
            if client_id:
                placeholders = ','.join(['%s'] * len(client_id))
                base_query += f" AND wo.client_id IN ({placeholders})"
                params.extend(client_id)

            if canal_id:
                placeholders = ','.join(['%s'] * len(canal_id))
                base_query += f" AND wo.canal_id IN ({placeholders})"
                params.extend(canal_id)

            if planta_id:
                placeholders = ','.join(['%s'] * len(planta_id))
                base_query += f" AND wo.planta_id IN ({placeholders})"
                params.extend(planta_id)

            if impresion_id:
                placeholders = ','.join(['%s'] * len(impresion_id))
                base_query += f" AND wo.impresion IN ({placeholders})"
                params.extend(impresion_id)

            if proceso_id:
                placeholders = ','.join(['%s'] * len(proceso_id))
                base_query += f" AND wo.process_id IN ({placeholders})"
                params.extend(proceso_id)

            if style_id:
                placeholders = ','.join(['%s'] * len(style_id))
                base_query += f" AND wo.style_id IN ({placeholders})"
                params.extend(style_id)

            if solicitud_id:
                placeholders = ','.join(['%s'] * len(solicitud_id))
                base_query += f" AND wo.tipo_solicitud IN ({placeholders})"
                params.extend(solicitud_id)

            if cinta_id and len(cinta_id) < 2:
                placeholders = ','.join(['%s'] * len(cinta_id))
                base_query += f" AND wo.cinta IN ({placeholders})"
                params.extend(cinta_id)

            if fsc_codigo:
                placeholders = ','.join(['%s'] * len(fsc_codigo))
                base_query += f" AND wo.fsc IN ({placeholders})"
                params.extend(fsc_codigo)

            # Filtro por estado
            if estado_id:
                placeholders = ','.join(['%s'] * len(estado_id))
                base_query += f" AND m.state_id IN ({placeholders})"
                params.extend(estado_id)
            else:
                # Estados activos por defecto
                if role_id in [3, 4]:  # Jefe Ventas o Vendedor
                    estados_activos = [1, 2, 3, 4, 5, 6, 7, 10, 12, 13, 14, 15, 16, 17, 18, 20, 21, 22]
                elif role_id in [13, 14]:  # JefeMuestras o TecnicoMuestras
                    estados_activos = [17]
                else:
                    estados_activos = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]

                placeholders = ','.join(['%s'] * len(estados_activos))
                base_query += f" AND m.state_id IN ({placeholders})"
                params.extend(estados_activos)

            # Filtros de fechas
            if date_desde:
                base_query += " AND wo.created_at >= %s"
                params.append(date_desde)

            if date_hasta:
                base_query += " AND wo.created_at <= %s"
                params.append(date_hasta + " 23:59:59")

            # Contar total
            count_query = f"SELECT COUNT(*) as total FROM ({base_query}) as subq"
            cursor.execute(count_query, params)
            total = cursor.fetchone()["total"]

            # Ordenar y paginar
            base_query += " ORDER BY wo.id DESC LIMIT %s OFFSET %s"
            offset = (page - 1) * per_page
            params.extend([per_page, offset])

            cursor.execute(base_query, params)
            ots = cursor.fetchall()

            # Convertir datetime a string
            for ot in ots:
                for key, value in ot.items():
                    if isinstance(value, datetime):
                        ot[key] = value.isoformat()

            return {
                "success": True,
                "data": ots,
                "pagination": {
                    "total": total,
                    "page": page,
                    "per_page": per_page,
                    "total_pages": (total + per_page - 1) // per_page
                }
            }

    except pymysql.Error as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al filtrar OTs: {str(e)}"
        )
    finally:
        connection.close()


class ValidarExcelResponse(BaseModel):
    """Respuesta de validación de Excel."""
    validado: bool
    errores: List[dict]


@router.post("/validar-excel", response_model=ValidarExcelResponse)
async def validar_excel(archivo: UploadFile = File(...)):
    """
    Valida estructura de archivo Excel antes de importar muestras.

    Fuente Laravel: WorkOrderController.php líneas 4619-4722

    Valida que:
    - El archivo tenga las columnas obligatorias
    - Cada fila tenga valores en los campos requeridos

    Columnas obligatorias:
    - descripcion, largo_int., ancho_int., alto_int., carton
    - destinatario, cantidad, tipo_pegado, planta_corte

    Args:
        archivo: Archivo Excel (.xls o .xlsx)

    Returns:
        validado: True si pasa todas las validaciones
        errores: Lista de errores encontrados
    """
    errores = []
    validado = True

    # Verificar extensión
    ext = archivo.filename.split('.')[-1].lower() if archivo.filename else ''
    if ext not in ['xls', 'xlsx']:
        # Si no es Excel, se considera validado (como Laravel)
        return ValidarExcelResponse(validado=True, errores=[])

    try:
        # Leer archivo Excel
        contents = await archivo.read()
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
        ws = wb.active

        # Columnas obligatorias
        obligatorios = [
            'descripcion', 'largo_int.', 'ancho_int.', 'alto_int.',
            'carton', 'destinatario', 'cantidad', 'tipo_pegado', 'planta_corte'
        ]

        # Alias para variantes sin punto
        alias = {
            'largo_int': 'largo_int.',
            'ancho_int': 'ancho_int.',
            'alto_int': 'alto_int.',
            'largo_ext': 'largo_ext.',
            'ancho_ext': 'ancho_ext.',
            'alto_ext': 'alto_ext.',
        }

        # Obtener encabezados de primera fila
        encabezados = []
        columna_a_indice = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                header = str(cell.value).lower().strip()
                # Aplicar alias
                header = alias.get(header, header)
                encabezados.append(header)
                columna_a_indice[header] = col_idx
            else:
                encabezados.append(None)

        # Verificar columnas faltantes
        faltantes = [c for c in obligatorios if c not in encabezados]
        if faltantes:
            validado = False
            errores.append({
                'tipo': 'faltantes',
                'mensaje': f"Columnas faltantes: {', '.join(faltantes)}"
            })

        # Validar filas
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Verificar si la fila tiene contenido
            hay_contenido = any(
                v is not None and str(v).strip() != ''
                for v in row
            )

            if not hay_contenido:
                continue

            # Validar campos obligatorios vacíos
            for campo in obligatorios:
                if campo in columna_a_indice:
                    col_idx = columna_a_indice[campo] - 1
                    if col_idx < len(row):
                        valor = row[col_idx]
                        if valor is None or str(valor).strip() == '':
                            validado = False
                            errores.append({
                                'tipo': 'campo_vacio',
                                'mensaje': f"Fila {row_idx}: campo vacío en '{campo}'"
                            })

        return ValidarExcelResponse(validado=validado, errores=errores)

    except Exception as e:
        return ValidarExcelResponse(
            validado=False,
            errores=[{'tipo': 'error', 'mensaje': f"Error al procesar archivo: {str(e)}"}]
        )


# =============================================================================
# SPRINT M: BITACORA DE CAMBIOS
# Fuente Laravel: WorkOrderController.php lineas 517-522, 5127-5132
# =============================================================================

from app.services.bitacora_service import get_bitacora_service
from app.models.bitacora import HistorialCambiosResponse, BitacoraResponse


@router.get("/{ot_id}/historial-cambios", response_model=HistorialCambiosResponse)
async def get_historial_cambios(
    ot_id: int,
    limite: int = Query(default=100, ge=1, le=500, description="Máximo de registros a retornar")
):
    """
    Obtiene el historial de cambios de una OT.

    Fuente Laravel: WorkOrderController.php linea 519
        $bitacora_ot = BitacoraWorkOrder::where('work_order_id', $ot->id)->get()

    Retorna:
    - work_order_id: ID de la OT
    - total_cambios: Cantidad total de cambios registrados
    - registros: Lista de cambios con detalles
    - tiene_mckee: Si la OT tiene fórmula McKee aplicada
    """
    bitacora_service = get_bitacora_service()
    return bitacora_service.obtener_historial(ot_id, limite)


@router.get("/{ot_id}/tiene-cambios")
async def verificar_tiene_cambios(ot_id: int):
    """
    Verifica si una OT tiene cambios registrados en bitácora.

    Fuente Laravel: WorkOrderController.php lineas 517-522
        $bitacora_ot = BitacoraWorkOrder::where('work_order_id', $ot->id)->get()->count();

    Retorna:
    - tiene_cambios: True si tiene cambios registrados
    - tiene_mckee: True si tiene McKee aplicado
    """
    bitacora_service = get_bitacora_service()
    tiene_cambios = bitacora_service.verificar_tiene_cambios(ot_id)
    tiene_mckee = bitacora_service.verificar_tiene_mckee(ot_id)

    return {
        "work_order_id": ot_id,
        "tiene_cambios": tiene_cambios,
        "tiene_mckee": tiene_mckee
    }
