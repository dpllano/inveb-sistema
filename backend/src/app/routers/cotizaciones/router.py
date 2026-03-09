"""
Router principal de Cotizaciones.
Endpoints para CRUD, aprobaciones, versionamiento y exportacion PDF.

Basado en: PLAN_MIGRACION_COTIZACIONES.md
Sprint G: Agregado middleware de roles (FASE 2)
"""
from fastapi import APIRouter, HTTPException, Query, Depends
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from typing import List, Optional
from datetime import datetime, timedelta
from pydantic import BaseModel
import pymysql
import pymysql.cursors
import os
import logging
import io
import jwt

# Sprint G: Middleware de roles y constantes
from ...constants import (
    Roles, ROLES_COTIZADOR, ROLES_APROBAR_COTIZACION,
    can_use_cotizador, can_approve_cotizacion
)
from ...middleware.roles import require_roles, require_cotizador, require_aprobar_cotizacion

from ...schemas.cotizacion import (
    CotizacionCreate,
    CotizacionUpdate,
    CotizacionResponse,
    CotizacionListResponse,
    CotizacionConDetalles,
    CotizacionFilters,
    CotizacionApprovalResponse,
    SolicitarAprobacionRequest,
    GestionarAprobacionRequest,
    CotizacionResumen,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/cotizaciones", tags=["Cotizaciones"])
security = HTTPBearer()

# Modelo para crear cotización desde OT
class CotizacionFromOTRequest(BaseModel):
    observacion_interna: Optional[str] = None
    observacion_cliente: Optional[str] = None


def get_current_user_optional(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    """Obtiene el usuario actual del token JWT"""
    try:
        token = credentials.credentials
        payload = jwt.decode(
            token,
            os.getenv("JWT_SECRET_KEY", "d1ff1cult_s3cr3t_k3y_f0r_jwt_t0k3n"),
            algorithms=["HS256"]
        )
        return {
            "id": int(payload.get("sub", 0)),
            "rut": payload.get("rut"),
            "email": payload.get("email"),
            "role_id": payload.get("role_id"),
            "role_nombre": payload.get("role_nombre")
        }
    except Exception as e:
        logger.warning(f"Error decodificando token: {e}")
        raise HTTPException(status_code=401, detail="Token inválido")


# =============================================
# CONEXIÓN A BASE DE DATOS
# =============================================

def get_db_connection():
    """Obtiene conexión a MySQL con DictCursor"""
    return pymysql.connect(
        host=os.getenv("LARAVEL_MYSQL_HOST", os.getenv("MYSQL_HOST", "127.0.0.1")),
        port=int(os.getenv("LARAVEL_MYSQL_PORT", os.getenv("MYSQL_PORT", "3306"))),
        user=os.getenv("LARAVEL_MYSQL_USER", os.getenv("MYSQL_USER", "envases")),
        password=os.getenv("LARAVEL_MYSQL_PASSWORD", os.getenv("MYSQL_PASSWORD", "secret")),
        database=os.getenv("LARAVEL_MYSQL_DATABASE", os.getenv("MYSQL_DATABASE", "envases_ot")),
        cursorclass=pymysql.cursors.DictCursor
    )


# =============================================
# ENDPOINTS CRUD
# =============================================

@router.get("/", response_model=CotizacionListResponse)
async def list_cotizaciones(
    estado_id: Optional[List[int]] = Query(None),
    client_id: Optional[List[int]] = Query(None),
    user_id: Optional[int] = None,
    cotizacion_id: Optional[int] = None,
    date_desde: Optional[str] = None,
    date_hasta: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    """
    Lista cotizaciones con filtros y paginación.

    Filtros disponibles:
    - estado_id: Lista de IDs de estado
    - client_id: Lista de IDs de cliente
    - user_id: ID del usuario creador
    - cotizacion_id: ID específico
    - date_desde/date_hasta: Rango de fechas (dd/mm/yyyy)
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Construir query base
        query = """
            SELECT
                c.id, c.client_id, c.nombre_contacto, c.email_contacto,
                c.telefono_contacto, c.moneda_id, c.dias_pago, c.comision,
                c.observacion_interna, c.observacion_cliente, c.user_id,
                c.estado_id, c.role_can_show, c.nivel_aprobacion,
                c.previous_version_id, c.original_version_id, c.version_number,
                c.active, c.created_at, c.updated_at,
                ce.nombre as estado_nombre,
                cl.nombre as cliente_nombre,
                u.nombre as usuario_nombre,
                (SELECT COUNT(*) FROM detalle_cotizacions dc WHERE dc.cotizacion_id = c.id) as total_detalles
            FROM cotizacions c
            LEFT JOIN cotizacion_estados ce ON c.estado_id = ce.id
            LEFT JOIN clients cl ON c.client_id = cl.id
            LEFT JOIN users u ON c.user_id = u.id
            WHERE c.active = 1
        """
        params = []

        # Aplicar filtros
        if estado_id:
            placeholders = ",".join(["%s"] * len(estado_id))
            query += f" AND c.estado_id IN ({placeholders})"
            params.extend(estado_id)

        if client_id:
            placeholders = ",".join(["%s"] * len(client_id))
            query += f" AND c.client_id IN ({placeholders})"
            params.extend(client_id)

        if user_id:
            query += " AND c.user_id = %s"
            params.append(user_id)

        if cotizacion_id:
            query += " AND c.id = %s"
            params.append(cotizacion_id)

        # Filtro de fechas
        if date_desde:
            try:
                from_date = datetime.strptime(date_desde, "%d/%m/%Y")
                query += " AND c.created_at >= %s"
                params.append(from_date)
            except ValueError:
                pass

        if date_hasta:
            try:
                to_date = datetime.strptime(date_hasta, "%d/%m/%Y") + timedelta(days=1)
                query += " AND c.created_at < %s"
                params.append(to_date)
            except ValueError:
                pass

        # Contar total
        count_query = f"SELECT COUNT(*) as total FROM ({query}) as subq"
        cursor.execute(count_query, params)
        total = cursor.fetchone()["total"]

        # Aplicar paginación
        query += " ORDER BY c.id DESC LIMIT %s OFFSET %s"
        offset = (page - 1) * page_size
        params.extend([page_size, offset])

        cursor.execute(query, params)
        rows = cursor.fetchall()

        # Convertir a response
        items = []
        for row in rows:
            items.append(CotizacionResponse(
                id=row["id"],
                client_id=row["client_id"],
                nombre_contacto=row["nombre_contacto"],
                email_contacto=row["email_contacto"],
                telefono_contacto=row["telefono_contacto"],
                moneda_id=row["moneda_id"],
                dias_pago=row["dias_pago"],
                comision=row["comision"],
                observacion_interna=row["observacion_interna"],
                observacion_cliente=row["observacion_cliente"],
                user_id=row["user_id"],
                estado_id=row["estado_id"],
                role_can_show=row["role_can_show"],
                nivel_aprobacion=row["nivel_aprobacion"],
                previous_version_id=row["previous_version_id"],
                original_version_id=row["original_version_id"],
                version_number=row["version_number"],
                active=row["active"],
                created_at=row["created_at"],
                updated_at=row["updated_at"],
                cliente_nombre=row["cliente_nombre"],
                usuario_nombre=row["usuario_nombre"],
                total_detalles=row["total_detalles"],
            ))

        total_pages = (total + page_size - 1) // page_size

        return CotizacionListResponse(
            items=items,
            total=total,
            page=page,
            page_size=page_size,
            total_pages=total_pages,
        )

    finally:
        cursor.close()
        conn.close()


@router.get("/{id}", response_model=CotizacionConDetalles)
async def get_cotizacion(id: int):
    """
    Obtiene una cotización con todos sus detalles.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Obtener cotización
        cursor.execute("""
            SELECT
                c.*, ce.nombre as estado_nombre,
                cl.nombre as cliente_nombre,
                u.nombre as usuario_nombre
            FROM cotizacions c
            LEFT JOIN cotizacion_estados ce ON c.estado_id = ce.id
            LEFT JOIN clients cl ON c.client_id = cl.id
            LEFT JOIN users u ON c.user_id = u.id
            WHERE c.id = %s AND c.active = 1
        """, (id,))

        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        # Obtener detalles
        cursor.execute("""
            SELECT dc.*, p.nombre as planta_nombre,
                   cb.codigo as carton_codigo,
                   pr.descripcion as proceso_nombre,
                   r.descripcion as rubro_nombre
            FROM detalle_cotizacions dc
            LEFT JOIN plantas p ON dc.planta_id = p.id
            LEFT JOIN cardboards cb ON dc.carton_id = cb.id
            LEFT JOIN processes pr ON dc.process_id = pr.id
            LEFT JOIN rubros r ON dc.rubro_id = r.id
            WHERE dc.cotizacion_id = %s
            ORDER BY dc.id
        """, (id,))

        detalles = cursor.fetchall()

        # Contar ganados/perdidos (estado: 1=ganado, 2=perdido)
        cursor.execute("""
            SELECT
                SUM(CASE WHEN estado = 1 THEN 1 ELSE 0 END) as ganados,
                SUM(CASE WHEN estado = 2 THEN 1 ELSE 0 END) as perdidos
            FROM detalle_cotizacions
            WHERE cotizacion_id = %s
        """, (id,))
        counts = cursor.fetchone()

        return CotizacionConDetalles(
            id=row["id"],
            client_id=row["client_id"],
            nombre_contacto=row["nombre_contacto"],
            email_contacto=row["email_contacto"],
            telefono_contacto=row["telefono_contacto"],
            moneda_id=row["moneda_id"],
            dias_pago=row["dias_pago"],
            comision=row["comision"],
            observacion_interna=row["observacion_interna"],
            observacion_cliente=row["observacion_cliente"],
            user_id=row["user_id"],
            estado_id=row["estado_id"],
            role_can_show=row["role_can_show"],
            nivel_aprobacion=row["nivel_aprobacion"],
            previous_version_id=row["previous_version_id"],
            original_version_id=row["original_version_id"],
            version_number=row["version_number"],
            active=row["active"],
            created_at=row["created_at"],
            updated_at=row["updated_at"],
            cliente_nombre=row["cliente_nombre"],
            usuario_nombre=row["usuario_nombre"],
            total_detalles=len(detalles),
            detalles=detalles,
            detalles_ganados_count=counts["ganados"] or 0,
            detalles_perdidos_count=counts["perdidos"] or 0,
        )

    finally:
        cursor.close()
        conn.close()


@router.get("/{id}/costos-resumen")
async def get_cotizacion_costos_resumen(id: int):
    """
    Obtiene el resumen de costos de una cotización con los datos calculados
    de historial_resultados para mostrar las tablas de:
    - Parametros Por Producto
    - Nuevos Detalles Cotizacion
    - Costos Productos (USD/MM2)
    - Costos Servicios (USD/MM2)
    """
    import json

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Obtener cotización
        cursor.execute("""
            SELECT c.*, ce.nombre as estado_nombre,
                   cl.nombre_sap as cliente_nombre
            FROM cotizacions c
            LEFT JOIN cotizacion_estados ce ON c.estado_id = ce.id
            LEFT JOIN clients cl ON c.client_id = cl.id
            WHERE c.id = %s AND c.active = 1
        """, (id,))

        cotizacion = cursor.fetchone()
        if not cotizacion:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        # Obtener detalles con historial_resultados
        cursor.execute("""
            SELECT
                dc.id, dc.descripcion, dc.cantidad, dc.flete,
                dc.carton_id, dc.cad_material_id,
                dc.tipo_detalle_id, dc.planta_id,
                dc.historial_resultados,
                dc.maquila, dc.armado, dc.clisse_valor, dc.matriz_valor,
                dc.mano_obra, dc.flete as flete_valor,
                cb.codigo as carton_codigo,
                cad.nombre as cad_nombre,
                p.nombre as planta_nombre,
                td.nombre as tipo_producto_nombre
            FROM detalle_cotizacions dc
            LEFT JOIN cardboards cb ON dc.carton_id = cb.id
            LEFT JOIN cads cad ON dc.cad_material_id = cad.id
            LEFT JOIN plantas p ON dc.planta_id = p.id
            LEFT JOIN tipo_detalles td ON dc.tipo_detalle_id = td.id
            WHERE dc.cotizacion_id = %s
            ORDER BY dc.id
        """, (id,))

        detalles = cursor.fetchall()

        # Procesar cada detalle y extraer costos
        parametros_producto = []
        nuevos_detalles = []
        costos_productos = []
        costos_servicios = []

        for idx, det in enumerate(detalles, 1):
            # Parse historial_resultados JSON
            resultados = {}
            if det.get('historial_resultados'):
                try:
                    if isinstance(det['historial_resultados'], str):
                        resultados = json.loads(det['historial_resultados'])
                    else:
                        resultados = det['historial_resultados']
                except (json.JSONDecodeError, TypeError):
                    resultados = {}

            # Common fields
            base_info = {
                "numero": idx,
                "descripcion": det.get('descripcion', ''),
                "cad": det.get('cad_nombre', '-'),
                "tipo_producto": det.get('tipo_producto_nombre', '-'),
                "item": f"Item-{det['id']}",
                "carton": det.get('carton_codigo', '-'),
            }

            # 1. Parametros Por Producto
            costo_directo = resultados.get('costo_directo', {})
            parametros_producto.append({
                **base_info,
                "planta": det.get('planta_nombre', '-'),
                "flete": det.get('flete', 0),
                "margen_papeles": resultados.get('margen_papeles', 0),
                "margen": resultados.get('margen', 0),
                "margen_minimo": resultados.get('margen_minimo', 0),
                "precio_usd_mm2": resultados.get('precio_usd_mm2', 0),
                "precio_usd_ton": resultados.get('precio_usd_ton', 0),
                "precio_usd_un": resultados.get('precio_usd_un', 0),
                "precio_clp_un": resultados.get('precio_clp_un', 0),
                "cantidad": det.get('cantidad', 0),
                "precio_total_musd": resultados.get('precio_total_musd', 0),
            })

            # 2. Nuevos Detalles Cotizacion
            nuevos_detalles.append({
                **base_info,
                "mc_usd_mm2": resultados.get('mc_usd_mm2', 0),
                "margen_bruto_sin_flete": resultados.get('margen_bruto_sin_flete', 0),
                "margen_bruto_sin_flete_pct": resultados.get('margen_bruto_sin_flete_pct', 0),
                "margen_servir": resultados.get('margen_servir', 0),
                "margen_servir_pct": resultados.get('margen_servir_pct', 0),
                "ebitda_usd_mm2": resultados.get('ebitda_usd_mm2', 0),
                "mg_ebitda": resultados.get('mg_ebitda', 0),
                "diferencia_margen": resultados.get('diferencia_margen', 0),
            })

            # 3. Costos Productos (USD/MM2)
            costo_indirecto = resultados.get('costo_indirecto', {})
            costo_gvv = resultados.get('costo_gvv', {})
            costo_fijo = resultados.get('costo_fijo_total', {})
            costo_total = resultados.get('costo_total', {})

            costos_productos.append({
                **base_info,
                "costo_directo": costo_directo.get('usd_mm2', 0) if isinstance(costo_directo, dict) else costo_directo,
                "costo_indirecto": costo_indirecto.get('usd_mm2', 0) if isinstance(costo_indirecto, dict) else costo_indirecto,
                "gvv": costo_gvv.get('usd_mm2', 0) if isinstance(costo_gvv, dict) else costo_gvv,
                "costo_fijo": costo_fijo.get('usd_mm2', 0) if isinstance(costo_fijo, dict) else costo_fijo,
                "costo_total": costo_total.get('usd_mm2', 0) if isinstance(costo_total, dict) else costo_total,
            })

            # 4. Costos Servicios (USD/MM2)
            costos_servicios.append({
                **base_info,
                "maquila": det.get('maquila', 0) or 0,
                "armado": det.get('armado', 0) or 0,
                "clisses": det.get('clisse_valor', 0) or 0,
                "matriz": det.get('matriz_valor', 0) or 0,
                "mano_obra": det.get('mano_obra', 0) or 0,
                "flete": det.get('flete_valor', 0) or 0,
            })

        return {
            "cotizacion_id": id,
            "estado_id": cotizacion['estado_id'],
            "estado_nombre": cotizacion['estado_nombre'],
            "cliente_nombre": cotizacion['cliente_nombre'],
            "tiene_resultados": any(det.get('historial_resultados') for det in detalles),
            "parametros_producto": parametros_producto,
            "nuevos_detalles": nuevos_detalles,
            "costos_productos": costos_productos,
            "costos_servicios": costos_servicios,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error obteniendo resumen de costos: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cursor.close()
        conn.close()


@router.post("/", response_model=CotizacionResponse)
async def create_cotizacion(data: CotizacionCreate):
    """
    Crea una nueva cotización en estado borrador.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            INSERT INTO cotizacions (
                client_id, nombre_contacto, email_contacto, telefono_contacto,
                moneda_id, dias_pago, comision, observacion_interna,
                observacion_cliente, user_id, estado_id, active, check_nombre_contacto, created_at, updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1, 1, 0, NOW(), NOW())
        """, (
            data.client_id, data.nombre_contacto, data.email_contacto,
            data.telefono_contacto, data.moneda_id, data.dias_pago,
            data.comision, data.observacion_interna, data.observacion_cliente,
            data.user_id
        ))

        conn.commit()
        cotizacion_id = cursor.lastrowid

        # Obtener la cotización creada
        cursor.execute("SELECT * FROM cotizacions WHERE id = %s", (cotizacion_id,))
        row = cursor.fetchone()

        return CotizacionResponse(**row)

    except Exception as e:
        conn.rollback()
        logger.error(f"Error creando cotización: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cursor.close()
        conn.close()


@router.post("/from-ot/{work_order_id}")
async def create_cotizacion_from_ot(
    work_order_id: int,
    request: CotizacionFromOTRequest = None,
    current_user: dict = Depends(get_current_user_optional)
):
    """
    Crea una nueva cotización a partir de una OT existente.

    Copia la información del cliente, contacto, etc. desde la OT
    y crea un detalle inicial con los datos del producto.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # 1. Obtener datos de la OT
        cursor.execute("""
            SELECT
                wo.id, wo.client_id, wo.nombre_contacto, wo.email_contacto,
                wo.telefono_contacto, wo.product_type_id, wo.codigo_producto,
                wo.cantidad, wo.interno_largo, wo.interno_ancho, wo.interno_alto,
                c.nombre as cliente_nombre,
                pt.descripcion as product_type_nombre
            FROM work_orders wo
            LEFT JOIN clients c ON wo.client_id = c.id
            LEFT JOIN product_types pt ON wo.product_type_id = pt.id
            WHERE wo.id = %s
        """, (work_order_id,))

        ot = cursor.fetchone()
        if not ot:
            raise HTTPException(status_code=404, detail=f"OT {work_order_id} no encontrada")

        # 2. Crear la cotización
        observacion_interna = request.observacion_interna if request else None
        observacion_cliente = request.observacion_cliente if request else None

        cursor.execute("""
            INSERT INTO cotizacions (
                client_id, nombre_contacto, email_contacto, telefono_contacto,
                moneda_id, dias_pago, comision, observacion_interna,
                observacion_cliente, user_id, estado_id, active,
                check_nombre_contacto, created_at, updated_at
            ) VALUES (%s, %s, %s, %s, 1, 30, 0, %s, %s, %s, 1, 1, 0, NOW(), NOW())
        """, (
            ot['client_id'],
            ot['nombre_contacto'] or '',
            ot['email_contacto'] or '',
            ot['telefono_contacto'] or '',
            observacion_interna,
            observacion_cliente,
            current_user['id']
        ))

        conn.commit()
        cotizacion_id = cursor.lastrowid

        # 3. Crear detalle inicial si la OT tiene producto
        if ot.get('product_type_id'):
            cursor.execute("""
                INSERT INTO detalle_cotizacions (
                    cotizacion_id, tipo_detalle_id, product_type_id,
                    cantidad, largo, ancho, alto,
                    created_at, updated_at
                ) VALUES (%s, 1, %s, %s, %s, %s, %s, NOW(), NOW())
            """, (
                cotizacion_id,
                ot.get('product_type_id'),
                ot.get('cantidad') or 1000,
                ot.get('interno_largo'),
                ot.get('interno_ancho'),
                ot.get('interno_alto')
            ))
            conn.commit()

        logger.info(f"Cotización {cotizacion_id} creada desde OT {work_order_id} por usuario {current_user['id']}")

        return {
            "success": True,
            "cotizacion_id": cotizacion_id,
            "message": f"Cotización #{cotizacion_id} creada exitosamente desde OT #{work_order_id}",
            "work_order_id": work_order_id
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error creando cotización desde OT: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cursor.close()
        conn.close()


@router.put("/{id}", response_model=CotizacionResponse)
async def update_cotizacion(id: int, data: CotizacionUpdate):
    """
    Actualiza una cotización existente.
    Solo se puede actualizar si está en estado borrador (estado_id = 1).
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Verificar que existe y está en borrador
        cursor.execute(
            "SELECT estado_id FROM cotizacions WHERE id = %s AND active = 1",
            (id,)
        )
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")
        if row["estado_id"] != 1:
            raise HTTPException(
                status_code=400,
                detail="Solo se pueden editar cotizaciones en estado borrador"
            )

        # Construir UPDATE dinámico
        updates = []
        params = []
        update_data = data.model_dump(exclude_unset=True)

        for field, value in update_data.items():
            if value is not None:
                updates.append(f"{field} = %s")
                params.append(value)

        if updates:
            updates.append("updated_at = NOW()")
            query = f"UPDATE cotizacions SET {', '.join(updates)} WHERE id = %s"
            params.append(id)
            cursor.execute(query, params)
            conn.commit()

        # Obtener cotización actualizada
        cursor.execute("SELECT * FROM cotizacions WHERE id = %s", (id,))
        row = cursor.fetchone()

        return CotizacionResponse(**row)

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error actualizando cotización: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cursor.close()
        conn.close()


@router.delete("/{id}")
async def delete_cotizacion(id: int):
    """
    Elimina (desactiva) una cotización.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute(
            "UPDATE cotizacions SET active = 0, updated_at = NOW() WHERE id = %s",
            (id,)
        )

        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        conn.commit()
        return {"message": "Cotización eliminada exitosamente"}

    finally:
        cursor.close()
        conn.close()


# =============================================
# ENDPOINTS DE APROBACIÓN
# =============================================


def calcular_margen_cotizacion(cursor, cotizacion_id: int) -> float:
    """
    Calcula el margen de una cotización según la lógica de Laravel.
    Fuente: CotizacionController.php líneas 1199-1216

    Fórmula:
    - margen_total = suma de margen de todos los detalles
    - margen_sugerido_total = suma de margen_sugerido de todos los detalles
    - Si margen_total >= margen_sugerido_total → retorna 0 (dentro del rango)
    - Si no → retorna abs((1 - margen_total / margen_sugerido_total) * 100)

    Returns:
        float: Porcentaje de diferencia entre margen y margen sugerido.
               0 = dentro del rango, >0 = por debajo del sugerido
    """
    cursor.execute("""
        SELECT
            COALESCE(SUM(margen), 0) as margen_total,
            COALESCE(SUM(margen_sugerido), 0) as margen_sugerido_total
        FROM detalle_cotizacions
        WHERE cotizacion_id = %s
    """, (cotizacion_id,))

    result = cursor.fetchone()
    margen_total = float(result["margen_total"] or 0)
    margen_sugerido_total = float(result["margen_sugerido_total"] or 0)

    # Si el margen está dentro del margen sugerido, retornamos 0
    if margen_total >= margen_sugerido_total:
        return 0.0

    # Evitar división por cero
    if margen_sugerido_total == 0:
        return 0.0

    # Calcular porcentaje absoluto de la diferencia
    margen = abs((1 - margen_total / margen_sugerido_total) * 100)
    return margen


def determinar_nivel_aprobacion(margen: float, es_jefe_venta: bool) -> dict:
    """
    Determina el nivel de aprobación según el margen y el rol del solicitante.
    Fuente: CotizacionController.php líneas 1104-1130

    Reglas:
    - Si margen <= 0 OR (es JefeVenta AND margen < 10) → Aprobación automática (estado=3)
    - Si margen < 10 → Nivel 1 (solo Jefe Ventas)
    - Si margen < 25 → Nivel 2 (Jefe Ventas + Gerente Comercial)
    - Si margen >= 25 → Nivel 3 (Jefe Ventas + Gerente Comercial + Gerente General)

    Args:
        margen: Porcentaje de diferencia entre margen y margen sugerido
        es_jefe_venta: True si el solicitante es Jefe de Ventas

    Returns:
        dict: {
            "estado": 2 (pendiente) o 3 (aprobado automático),
            "nivel_aprobacion": 1, 2, 3 o None,
            "role_can_show": ID del rol que debe aprobar primero
        }
    """
    # Si el margen está dentro del rango, o es JefeVenta con margen < 10
    # → Aprobación automática
    if margen <= 0 or (es_jefe_venta and margen < 10):
        return {
            "estado": 3,  # Aprobado
            "nivel_aprobacion": None,
            "role_can_show": None
        }

    # Determinar nivel según margen
    if margen < 10:
        # Nivel 1: Solo Jefe Ventas
        return {
            "estado": 2,  # Pendiente
            "nivel_aprobacion": 1,
            "role_can_show": 3  # Jefe Ventas
        }
    elif margen < 25:
        # Nivel 2: Jefe Ventas + Gerente Comercial
        # Si es JefeVenta, va directo al Gerente Comercial
        return {
            "estado": 2,
            "nivel_aprobacion": 2,
            "role_can_show": 15 if es_jefe_venta else 3
        }
    else:
        # Nivel 3: Jefe Ventas + Gerente Comercial + Gerente General
        # Si es JefeVenta, va directo al Gerente Comercial
        return {
            "estado": 2,
            "nivel_aprobacion": 3,
            "role_can_show": 15 if es_jefe_venta else 3
        }


@router.post("/{id}/solicitar-aprobacion")
async def solicitar_aprobacion(
    id: int,
    data: Optional[SolicitarAprobacionRequest] = None,
    user: dict = Depends(get_current_user_optional)
):
    """
    Solicita aprobación para una cotización.

    Sprint K+: Implementación completa de lógica de nivel de aprobación según margen.
    Fuente: CotizacionController.php líneas 1080-1141

    Lógica de determinación de nivel:
    - Calcula el % de diferencia entre margen y margen_sugerido
    - Según el % y si es JefeVenta, determina nivel (1, 2 o 3) o aprobación automática

    | Margen (% diferencia) | Nivel | Flujo de aprobación |
    |-----------------------|-------|---------------------|
    | ≤ 0% o (JefeVenta y < 10%) | - | Aprobación automática |
    | < 10% | 1 | Solo Jefe Ventas |
    | < 25% | 2 | Jefe Ventas → Gerente Comercial |
    | ≥ 25% | 3 | Jefe Ventas → Gerente Comercial → Gerente General |
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Verificar cotización
        cursor.execute(
            "SELECT * FROM cotizacions WHERE id = %s AND active = 1",
            (id,)
        )
        cotizacion = cursor.fetchone()
        if not cotizacion:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        if cotizacion["estado_id"] != 1:
            raise HTTPException(
                status_code=400,
                detail="Solo se puede solicitar aprobación de cotizaciones en borrador"
            )

        # Verificar que tenga al menos un detalle
        cursor.execute(
            "SELECT COUNT(*) as count FROM detalle_cotizacions WHERE cotizacion_id = %s",
            (id,)
        )
        count = cursor.fetchone()["count"]
        if count == 0:
            raise HTTPException(
                status_code=400,
                detail="La cotización debe tener al menos un detalle"
            )

        # Sprint K+: Calcular margen y determinar nivel de aprobación
        margen = calcular_margen_cotizacion(cursor, id)
        es_jefe_venta = user.get("role_id") == Roles.JefeVenta

        resultado_nivel = determinar_nivel_aprobacion(margen, es_jefe_venta)

        estado = resultado_nivel["estado"]
        nivel_aprobacion = resultado_nivel["nivel_aprobacion"]
        role_can_show = resultado_nivel["role_can_show"]

        # Actualizar cotización con estado, nivel y quién debe aprobar
        cursor.execute("""
            UPDATE cotizacions
            SET estado_id = %s,
                nivel_aprobacion = %s,
                role_can_show = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (estado, nivel_aprobacion, role_can_show, id))

        # Guardar historial de resultados en cada detalle
        cursor.execute("""
            SELECT id FROM detalle_cotizacions WHERE cotizacion_id = %s
        """, (id,))
        detalles = cursor.fetchall()

        conn.commit()

        # Preparar mensaje según resultado
        if estado == 3:
            message = "Cotización aprobada automáticamente (margen dentro del rango)"
        else:
            message = f"Aprobación solicitada exitosamente (Nivel {nivel_aprobacion})"

        return {
            "message": message,
            "cotizacion_id": id,
            "estado": estado,
            "nivel_aprobacion": nivel_aprobacion,
            "role_can_show": role_can_show,
            "margen_calculado": round(margen, 2),
            "aprobacion_automatica": estado == 3
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error solicitando aprobación: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cursor.close()
        conn.close()


@router.post("/{id}/gestionar-aprobacion")
async def gestionar_aprobacion(
    id: int,
    data: GestionarAprobacionRequest,
    user: dict = Depends(require_aprobar_cotizacion())
):
    """
    Aprueba o rechaza una cotización.

    Sprint G: Requiere rol autorizado para aprobar cotizaciones.
    Roles permitidos: Admin, SuperAdmin, Gerente, GerenteComercial, JefeVenta

    FLUJO DE APROBACIÓN MULTINIVEL (replicado de Laravel CotizacionApprovalController.php):
    - Nivel 1: Solo Jefe Ventas → Aprobación Total
    - Nivel 2: Jefe Ventas → Aprobación Parcial → Gerente Comercial → Aprobación Total
    - Nivel 3: Jefe Ventas → Gerente Comercial → Gerente General → Aprobación Total

    Actions:
    - aprobar: Según nivel y rol del usuario
    - rechazar: Cambia estado a rechazado (6)
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Verificar cotización
        cursor.execute(
            "SELECT * FROM cotizacions WHERE id = %s AND active = 1",
            (id,)
        )
        cotizacion = cursor.fetchone()
        if not cotizacion:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        if cotizacion["estado_id"] != 2:
            raise HTTPException(
                status_code=400,
                detail="Solo se pueden gestionar cotizaciones pendientes de aprobación"
            )

        role_id = user["role_id"]
        nivel_aprobacion = cotizacion.get("nivel_aprobacion", 1)
        accion = None
        nuevo_estado = None
        nuevo_role_can_show = None

        # LÓGICA DE APROBACIÓN MULTINIVEL (igual que Laravel líneas 17-90)
        if data.action == "aprobar":
            if nivel_aprobacion == 1:
                # Nivel 1: Solo Jefe Ventas - Aprobación Total
                nuevo_estado = 3
                accion = "Aprobación Total"

            elif nivel_aprobacion == 2:
                # Nivel 2: Jefe Ventas + Gerente Comercial
                if role_id == 3:  # Jefe Ventas
                    accion = "Aprobación Parcial"
                    nuevo_role_can_show = 15  # Pasa a Gerente Comercial
                else:
                    # Gerente Comercial o superior
                    nuevo_estado = 3
                    accion = "Aprobación Total"

            elif nivel_aprobacion == 3:
                # Nivel 3: Jefe Ventas + Gerente Comercial + Gerente General
                if role_id == 3:  # Jefe Ventas
                    accion = "Aprobación Parcial"
                    nuevo_role_can_show = 15  # Pasa a Gerente Comercial
                elif role_id == 15:  # Gerente Comercial
                    accion = "Aprobación Parcial"
                    nuevo_role_can_show = 2  # Pasa a Gerente General
                else:
                    # Gerente General o superior
                    nuevo_estado = 3
                    accion = "Aprobación Total"
            else:
                # Nivel no definido - Aprobación Total por defecto
                nuevo_estado = 3
                accion = "Aprobación Total"
        else:
            # Rechazo
            nuevo_estado = 6
            accion = "Rechazo"

        # Actualizar cotización
        if nuevo_estado:
            cursor.execute("""
                UPDATE cotizacions
                SET estado_id = %s, updated_at = NOW()
                WHERE id = %s
            """, (nuevo_estado, id))
        elif nuevo_role_can_show:
            cursor.execute("""
                UPDATE cotizacions
                SET role_can_show = %s, updated_at = NOW()
                WHERE id = %s
            """, (nuevo_role_can_show, id))

        # Registrar aprobación
        cursor.execute("""
            INSERT INTO cotizacion_approvals
            (cotizacion_id, user_id, role_do_action, action_made, motivo, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, NOW(), NOW())
        """, (
            id,
            user["id"],
            role_id,
            accion,
            data.motivo
        ))

        conn.commit()

        return {
            "message": f"Cotización gestionada exitosamente: {accion}",
            "cotizacion_id": id,
            "nuevo_estado": nuevo_estado,
            "accion": accion,
            "role_can_show": nuevo_role_can_show
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error gestionando aprobación: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cursor.close()
        conn.close()


@router.get("/{id}/aprobaciones", response_model=List[CotizacionApprovalResponse])
async def get_aprobaciones(id: int):
    """
    Obtiene historial de aprobaciones de una cotización.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT ca.*, u.nombre as usuario_nombre
            FROM cotizacion_approvals ca
            LEFT JOIN users u ON ca.user_id = u.id
            WHERE ca.cotizacion_id = %s
            ORDER BY ca.created_at DESC
        """, (id,))

        rows = cursor.fetchall()
        return [CotizacionApprovalResponse(**row) for row in rows]

    finally:
        cursor.close()
        conn.close()


# =============================================
# ENDPOINTS DE VERSIONAMIENTO
# =============================================

@router.post("/{id}/duplicar")
async def duplicar_cotizacion(id: int):
    """
    Duplica una cotización existente.
    Crea una copia con estado borrador y sin versión asociada.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Obtener cotización original
        cursor.execute(
            "SELECT * FROM cotizacions WHERE id = %s AND active = 1",
            (id,)
        )
        original = cursor.fetchone()
        if not original:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        # Crear copia
        cursor.execute("""
            INSERT INTO cotizacions (
                client_id, nombre_contacto, email_contacto, telefono_contacto,
                moneda_id, dias_pago, comision, observacion_interna,
                observacion_cliente, user_id, estado_id, active, created_at, updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1, 1, NOW(), NOW())
        """, (
            original["client_id"], original["nombre_contacto"],
            original["email_contacto"], original["telefono_contacto"],
            original["moneda_id"], original["dias_pago"],
            original["comision"], original["observacion_interna"],
            original["observacion_cliente"], original["user_id"]
        ))

        nueva_id = cursor.lastrowid

        # Copiar detalles
        cursor.execute("""
            INSERT INTO detalle_cotizacions (
                cotizacion_id, tipo_detalle_id, cantidad, product_type_id,
                numero_colores, area_hc, anchura, largura, carton_id, impresion,
                golpes_largo, golpes_ancho, process_id, rubro_id, planta_id,
                margen, ciudad_id, pallet, zuncho, funda, stretch_film,
                matriz, clisse, maquila, armado_automatico, variable_cotizador_id,
                created_at, updated_at
            )
            SELECT
                %s, tipo_detalle_id, cantidad, product_type_id,
                numero_colores, area_hc, anchura, largura, carton_id, impresion,
                golpes_largo, golpes_ancho, process_id, rubro_id, planta_id,
                margen, ciudad_id, pallet, zuncho, funda, stretch_film,
                matriz, clisse, maquila, armado_automatico, variable_cotizador_id,
                NOW(), NOW()
            FROM detalle_cotizacions
            WHERE cotizacion_id = %s
        """, (nueva_id, id))

        conn.commit()

        return {
            "message": "Cotización duplicada exitosamente",
            "original_id": id,
            "nueva_id": nueva_id
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error duplicando cotización: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cursor.close()
        conn.close()


@router.post("/{id}/versionar")
async def versionar_cotizacion(id: int):
    """
    Crea una nueva versión de una cotización.
    Mantiene referencia a la versión original.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Obtener cotización original
        cursor.execute(
            "SELECT * FROM cotizacions WHERE id = %s AND active = 1",
            (id,)
        )
        original = cursor.fetchone()
        if not original:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        # Determinar número de versión
        original_id = original["original_version_id"] or id
        cursor.execute(
            "SELECT MAX(version_number) as max_version FROM cotizacions WHERE original_version_id = %s",
            (original_id,)
        )
        max_version = cursor.fetchone()["max_version"] or 0
        nueva_version = max_version + 1

        # Crear nueva versión
        cursor.execute("""
            INSERT INTO cotizacions (
                client_id, nombre_contacto, email_contacto, telefono_contacto,
                moneda_id, dias_pago, comision, observacion_interna,
                observacion_cliente, user_id, estado_id, active,
                previous_version_id, original_version_id, version_number,
                created_at, updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1, 1, %s, %s, %s, NOW(), NOW())
        """, (
            original["client_id"], original["nombre_contacto"],
            original["email_contacto"], original["telefono_contacto"],
            original["moneda_id"], original["dias_pago"],
            original["comision"], original["observacion_interna"],
            original["observacion_cliente"], original["user_id"],
            id, original_id, nueva_version
        ))

        nueva_id = cursor.lastrowid

        # Copiar detalles (similar a duplicar)
        cursor.execute("""
            INSERT INTO detalle_cotizacions (
                cotizacion_id, tipo_detalle_id, cantidad, product_type_id,
                numero_colores, area_hc, anchura, largura, carton_id, impresion,
                golpes_largo, golpes_ancho, process_id, rubro_id, planta_id,
                margen, ciudad_id, pallet, zuncho, funda, stretch_film,
                matriz, clisse, maquila, armado_automatico, variable_cotizador_id,
                created_at, updated_at
            )
            SELECT
                %s, tipo_detalle_id, cantidad, product_type_id,
                numero_colores, area_hc, anchura, largura, carton_id, impresion,
                golpes_largo, golpes_ancho, process_id, rubro_id, planta_id,
                margen, ciudad_id, pallet, zuncho, funda, stretch_film,
                matriz, clisse, maquila, armado_automatico, variable_cotizador_id,
                NOW(), NOW()
            FROM detalle_cotizacions
            WHERE cotizacion_id = %s
        """, (nueva_id, id))

        conn.commit()

        return {
            "message": "Nueva versión creada exitosamente",
            "original_id": id,
            "nueva_id": nueva_id,
            "version_number": nueva_version
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error versionando cotización: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cursor.close()
        conn.close()


@router.post("/{id}/retomar")
async def retomar_cotizacion(id: int):
    """
    Retoma una cotización rechazada.
    Crea una nueva versión en estado borrador.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Verificar que está rechazada
        cursor.execute(
            "SELECT * FROM cotizacions WHERE id = %s AND active = 1",
            (id,)
        )
        original = cursor.fetchone()
        if not original:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        if original["estado_id"] != 6:  # Rechazado
            raise HTTPException(
                status_code=400,
                detail="Solo se pueden retomar cotizaciones rechazadas"
            )

        # Usar el endpoint de versionar
        return await versionar_cotizacion(id)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error retomando cotización: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cursor.close()
        conn.close()


# =============================================
# ENDPOINTS DE ESTADOS
# =============================================

@router.get("/estados/")
async def list_estados():
    """
    Lista todos los estados de cotización disponibles.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT * FROM cotizacion_estados ORDER BY id")
        return cursor.fetchall()

    finally:
        cursor.close()
        conn.close()


@router.get("/pendientes-aprobacion/")
async def list_pendientes_aprobacion(
    user_id: Optional[int] = None,
    role_id: Optional[int] = None,
):
    """
    Lista cotizaciones pendientes de aprobación.
    Filtra según rol del usuario aprobador.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        query = """
            SELECT
                c.*, cl.nombre as cliente_nombre, u.nombre as usuario_nombre,
                (SELECT COUNT(*) FROM detalle_cotizacions WHERE cotizacion_id = c.id) as total_detalles
            FROM cotizacions c
            LEFT JOIN clients cl ON c.client_id = cl.id
            LEFT JOIN users u ON c.user_id = u.id
            WHERE c.estado_id = 2 AND c.active = 1
        """
        params = []

        if role_id:
            # Filtrar según nivel de aprobación y rol
            query += " AND c.nivel_aprobacion >= %s"
            params.append(role_id)

        query += " ORDER BY c.created_at DESC"

        cursor.execute(query, params)
        return cursor.fetchall()

    finally:
        cursor.close()
        conn.close()


# =============================================
# ENDPOINTS DE APROBACIÓN EXTERNA
# =============================================

@router.get("/pendientes-aprobacion-externo/")
async def list_pendientes_aprobacion_externo(
    user_id: Optional[int] = None,
    date_desde: Optional[str] = None,
    date_hasta: Optional[str] = None,
    cotizacion_id: Optional[int] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    """
    Lista cotizaciones pendientes de aprobación por vendedor externo.
    Filtra cotizaciones con role_can_show = 4 (vendedor externo) y estado = 2.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        query = """
            SELECT
                c.id, c.client_id, c.nombre_contacto, c.email_contacto,
                c.telefono_contacto, c.user_id, c.estado_id, c.role_can_show,
                c.nivel_aprobacion, c.version_number, c.created_at, c.updated_at,
                ce.nombre as estado_nombre,
                cl.nombre as cliente_nombre,
                u.nombre as creador_nombre,
                DATEDIFF(NOW(), c.created_at) as dias_pendiente,
                (SELECT COUNT(*) FROM detalle_cotizacions dc WHERE dc.cotizacion_id = c.id) as total_detalles,
                (SELECT COUNT(*) FROM detalle_cotizacions dc WHERE dc.cotizacion_id = c.id AND dc.ganado = 1) as detalles_ganados,
                (SELECT COUNT(*) FROM detalle_cotizacions dc WHERE dc.cotizacion_id = c.id AND dc.perdido = 1) as detalles_perdidos,
                (SELECT dc2.descripcion FROM detalle_cotizacions dc2 WHERE dc2.cotizacion_id = c.id ORDER BY dc2.id LIMIT 1) as primer_detalle_descripcion,
                (SELECT cad.codigo FROM detalle_cotizacions dc3 LEFT JOIN cads cad ON dc3.cad_id = cad.id WHERE dc3.cotizacion_id = c.id ORDER BY dc3.id LIMIT 1) as primer_detalle_cad
            FROM cotizacions c
            LEFT JOIN cotizacion_estados ce ON c.estado_id = ce.id
            LEFT JOIN clients cl ON c.client_id = cl.id
            LEFT JOIN users u ON c.user_id = u.id
            WHERE c.active = 1 AND c.estado_id = 2 AND c.role_can_show = 4
        """
        params = []

        # Filtro por usuario externo (si aplica)
        if user_id:
            # TODO: Filtrar por vendedor externo responsable
            pass

        if cotizacion_id:
            query += " AND c.id = %s"
            params.append(cotizacion_id)

        # Filtros de fecha
        if date_desde:
            try:
                from_date = datetime.strptime(date_desde, "%d/%m/%Y")
                query += " AND c.created_at >= %s"
                params.append(from_date)
            except ValueError:
                pass

        if date_hasta:
            try:
                to_date = datetime.strptime(date_hasta, "%d/%m/%Y") + timedelta(days=1)
                query += " AND c.created_at < %s"
                params.append(to_date)
            except ValueError:
                pass

        # Contar total
        count_query = f"SELECT COUNT(*) as total FROM ({query}) as subq"
        cursor.execute(count_query, params)
        total = cursor.fetchone()["total"]

        # Aplicar paginación y orden
        query += " ORDER BY c.created_at DESC LIMIT %s OFFSET %s"
        offset = (page - 1) * page_size
        params.extend([page_size, offset])

        cursor.execute(query, params)
        rows = cursor.fetchall()

        total_pages = (total + page_size - 1) // page_size

        return {
            "items": rows,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages
        }

    finally:
        cursor.close()
        conn.close()


@router.post("/{id}/solicitar-aprobacion-externo")
async def solicitar_aprobacion_externo(id: int, user_id: int = Query(...)):
    """
    Solicita aprobación de cotización por vendedor externo.
    Cambia role_can_show a 4 (vendedor externo) y estado a 2 (pendiente).
    Almacena historial de resultados en cada detalle.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Verificar cotización
        cursor.execute(
            "SELECT * FROM cotizacions WHERE id = %s AND active = 1",
            (id,)
        )
        cotizacion = cursor.fetchone()
        if not cotizacion:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        if cotizacion["estado_id"] not in [1, 6]:  # Borrador o Rechazada
            raise HTTPException(
                status_code=400,
                detail="Solo se puede solicitar aprobación externa de cotizaciones en borrador o rechazadas"
            )

        # Verificar que tenga al menos un detalle
        cursor.execute(
            "SELECT COUNT(*) as count FROM detalle_cotizacions WHERE cotizacion_id = %s",
            (id,)
        )
        count = cursor.fetchone()["count"]
        if count == 0:
            raise HTTPException(
                status_code=400,
                detail="La cotización debe tener al menos un detalle"
            )

        # Actualizar estado y role_can_show
        cursor.execute("""
            UPDATE cotizacions
            SET estado_id = 2, role_can_show = 4, updated_at = NOW()
            WHERE id = %s
        """, (id,))

        # Guardar historial de resultados en cada detalle (si existe columna)
        # cursor.execute("""
        #     UPDATE detalle_cotizacions
        #     SET historial_resultados = JSON_OBJECT('fecha', NOW(), 'accion', 'solicitar_aprobacion_externo')
        #     WHERE cotizacion_id = %s
        # """, (id,))

        conn.commit()

        return {
            "success": True,
            "message": "Aprobación externa solicitada exitosamente",
            "cotizacion_id": id,
            "redirect_url": f"/cotizador-externo"
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error solicitando aprobación externa: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cursor.close()
        conn.close()


@router.post("/{id}/gestionar-aprobacion-externo")
async def gestionar_aprobacion_externo(id: int, data: GestionarAprobacionRequest):
    """
    Aprueba o rechaza una cotización como vendedor externo.

    Actions:
    - aprobar: Cambia estado a aprobado (3)
    - rechazar: Cambia estado a rechazado (6)

    Registra la acción en cotizacion_approvals con role_do_action = 4 (externo).
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Verificar cotización
        cursor.execute(
            "SELECT * FROM cotizacions WHERE id = %s AND active = 1",
            (id,)
        )
        cotizacion = cursor.fetchone()
        if not cotizacion:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        if cotizacion["estado_id"] != 2:
            raise HTTPException(
                status_code=400,
                detail="Solo se pueden gestionar cotizaciones pendientes de aprobación"
            )

        if cotizacion["role_can_show"] != 4:
            raise HTTPException(
                status_code=400,
                detail="Esta cotización no está pendiente de aprobación externa"
            )

        # Determinar nuevo estado y tipo de acción
        if data.action == "aprobar":
            nuevo_estado = 3
            action_made = "Aprobación Total"
        elif data.action == "aprobar_parcial":
            nuevo_estado = 3
            action_made = "Aprobación Parcial"
        else:
            nuevo_estado = 6
            action_made = "Rechazo"

        # Actualizar estado de la cotización
        cursor.execute("""
            UPDATE cotizacions
            SET estado_id = %s, role_can_show = NULL, updated_at = NOW()
            WHERE id = %s
        """, (nuevo_estado, id))

        # Registrar aprobación en historial
        cursor.execute("""
            INSERT INTO cotizacion_approvals
            (cotizacion_id, user_id, role_do_action, action_made, motivo, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, NOW(), NOW())
        """, (
            id,
            data.user_id if hasattr(data, 'user_id') and data.user_id else 1,
            4,  # role_do_action = 4 (vendedor externo)
            action_made,
            data.motivo
        ))

        conn.commit()

        estado_texto = "aprobada" if nuevo_estado == 3 else "rechazada"

        return {
            "success": True,
            "message": f"Cotización {estado_texto} exitosamente",
            "cotizacion_id": id,
            "nuevo_estado": nuevo_estado,
            "action_made": action_made,
            "redirect_url": "/cotizador-externo"
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error gestionando aprobación externa: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cursor.close()
        conn.close()


@router.get("/{id}/historial-aprobaciones")
async def get_historial_aprobaciones(id: int):
    """
    Obtiene historial completo de aprobaciones de una cotización,
    incluyendo aprobaciones internas y externas.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT
                ca.id,
                ca.cotizacion_id,
                ca.user_id,
                ca.role_do_action,
                ca.action_made,
                ca.motivo,
                ca.created_at,
                u.nombre as usuario_nombre,
                CASE ca.role_do_action
                    WHEN 1 THEN 'Administrador'
                    WHEN 2 THEN 'Supervisor'
                    WHEN 3 THEN 'Vendedor'
                    WHEN 4 THEN 'Vendedor Externo'
                    ELSE 'Otro'
                END as rol_nombre
            FROM cotizacion_approvals ca
            LEFT JOIN users u ON ca.user_id = u.id
            WHERE ca.cotizacion_id = %s
            ORDER BY ca.created_at DESC
        """, (id,))

        rows = cursor.fetchall()
        return {
            "cotizacion_id": id,
            "total_aprobaciones": len(rows),
            "historial": rows
        }

    finally:
        cursor.close()
        conn.close()


# =============================================
# ENDPOINT DE EXPORTACION PDF
# =============================================

def generate_pdf_content(cotizacion: dict, detalles: list, cliente_info: dict) -> bytes:
    """
    Genera el contenido PDF de una cotizacion.
    Usa reportlab si esta disponible, de lo contrario genera HTML simple.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)

        elements = []
        styles = getSampleStyleSheet()

        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=12,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1a56db')
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            textColor=colors.HexColor('#6b7280')
        )

        # Titulo
        elements.append(Paragraph("INVEB - Cotizacion", title_style))
        elements.append(Paragraph(f"Cotizacion #{cotizacion.get('id', 'N/A')}", subtitle_style))
        elements.append(Spacer(1, 20))

        # Informacion del cliente
        elements.append(Paragraph("<b>Informacion del Cliente</b>", styles['Heading2']))
        cliente_data = [
            ["Cliente:", cliente_info.get('nombre', 'N/A')],
            ["Contacto:", cotizacion.get('nombre_contacto', 'N/A')],
            ["Email:", cotizacion.get('email_contacto', 'N/A')],
            ["Telefono:", cotizacion.get('telefono_contacto', 'N/A')],
        ]
        cliente_table = Table(cliente_data, colWidths=[3*cm, 10*cm])
        cliente_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(cliente_table)
        elements.append(Spacer(1, 20))

        # Condiciones comerciales
        elements.append(Paragraph("<b>Condiciones Comerciales</b>", styles['Heading2']))
        comercial_data = [
            ["Dias de Pago:", f"{cotizacion.get('dias_pago', 0)} dias"],
            ["Comision:", f"{cotizacion.get('comision', 0)}%"],
            ["Observacion:", cotizacion.get('observacion_cliente', 'Sin observaciones')],
        ]
        comercial_table = Table(comercial_data, colWidths=[3*cm, 10*cm])
        comercial_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(comercial_table)
        elements.append(Spacer(1, 20))

        # Detalles de la cotizacion
        if detalles:
            elements.append(Paragraph("<b>Detalle de Productos</b>", styles['Heading2']))

            # Encabezados de tabla
            header = ["#", "Cantidad", "Dimensiones", "Carton", "Proceso"]
            table_data = [header]

            for idx, det in enumerate(detalles, 1):
                row = [
                    str(idx),
                    f"{det.get('cantidad', 0):,}",
                    f"{det.get('largura', 0)} x {det.get('anchura', 0)} cm",
                    det.get('carton_codigo', 'N/A'),
                    det.get('proceso_nombre', 'N/A'),
                ]
                table_data.append(row)

            col_widths = [1*cm, 2*cm, 4*cm, 3*cm, 4*cm]
            detail_table = Table(table_data, colWidths=col_widths)
            detail_table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a56db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                # Body
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                # Grid
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                # Alternating rows
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ]))
            elements.append(detail_table)
            elements.append(Spacer(1, 20))

        # Pie de pagina
        fecha_str = datetime.now().strftime("%d/%m/%Y %H:%M")
        elements.append(Spacer(1, 30))
        elements.append(Paragraph(
            f"<i>Documento generado el {fecha_str}</i>",
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray)
        ))

        doc.build(elements)
        return buffer.getvalue()

    except ImportError:
        # Si reportlab no esta instalado, generar HTML simple
        logger.warning("reportlab no disponible, generando HTML")
        html_content = f"""
        <html>
        <head>
            <title>Cotizacion #{cotizacion.get('id', 'N/A')}</title>
            <style>
                body {{ font-family: Arial, sans-serif; padding: 20px; }}
                h1 {{ color: #1a56db; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #1a56db; color: white; }}
            </style>
        </head>
        <body>
            <h1>INVEB - Cotizacion #{cotizacion.get('id', 'N/A')}</h1>
            <h2>Cliente: {cliente_info.get('nombre', 'N/A')}</h2>
            <p>Contacto: {cotizacion.get('nombre_contacto', 'N/A')}</p>
            <p>Email: {cotizacion.get('email_contacto', 'N/A')}</p>
            <h3>Detalles ({len(detalles)} items)</h3>
            <table>
                <tr><th>#</th><th>Cantidad</th><th>Dimensiones</th></tr>
                {''.join(f"<tr><td>{i+1}</td><td>{d.get('cantidad', 0)}</td><td>{d.get('largura', 0)} x {d.get('anchura', 0)}</td></tr>" for i, d in enumerate(detalles))}
            </table>
            <p><i>Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}</i></p>
        </body>
        </html>
        """
        return html_content.encode('utf-8')


@router.get("/{id}/export-pdf")
async def export_cotizacion_pdf(id: int):
    """
    Exporta una cotizacion a PDF.
    Retorna el archivo PDF para descarga.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Obtener cotizacion
        cursor.execute("""
            SELECT c.*, cl.nombre as cliente_nombre, cl.rut as cliente_rut
            FROM cotizacions c
            LEFT JOIN clients cl ON c.client_id = cl.id
            WHERE c.id = %s AND c.active = 1
        """, (id,))

        cotizacion = cursor.fetchone()
        if not cotizacion:
            raise HTTPException(status_code=404, detail="Cotizacion no encontrada")

        # Obtener detalles
        cursor.execute("""
            SELECT dc.*,
                   cb.codigo as carton_codigo,
                   pr.descripcion as proceso_nombre
            FROM detalle_cotizacions dc
            LEFT JOIN cardboards cb ON dc.carton_id = cb.id
            LEFT JOIN processes pr ON dc.process_id = pr.id
            WHERE dc.cotizacion_id = %s
            ORDER BY dc.id
        """, (id,))

        detalles = cursor.fetchall()

        # Informacion del cliente
        cliente_info = {
            "nombre": cotizacion.get("cliente_nombre", "Sin cliente"),
            "rut": cotizacion.get("cliente_rut", ""),
        }

        # Generar PDF
        pdf_content = generate_pdf_content(cotizacion, detalles, cliente_info)

        # Determinar tipo de contenido
        try:
            from reportlab.lib import colors
            content_type = "application/pdf"
            filename = f"cotizacion_{id}.pdf"
        except ImportError:
            content_type = "text/html"
            filename = f"cotizacion_{id}.html"

        return StreamingResponse(
            io.BytesIO(pdf_content),
            media_type=content_type,
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    finally:
        cursor.close()
        conn.close()


# =============================================
# ENDPOINT DETALLE A OT (Crear OT desde Cotización)
# =============================================

@router.get("/detalle/{detalle_id}/para-ot")
async def get_detalle_para_ot(detalle_id: int, tipo_solicitud: int = Query(..., ge=1, le=9)):
    """
    Obtiene los datos de un detalle de cotización transformados para crear una OT.

    Tipos de solicitud soportados:
    - 1: Desarrollo Completo
    - 2: Cotiza con CAD
    - 3: Muestra con CAD
    - 5: Arte con Material

    Retorna los datos del detalle y cotización mapeados a campos de OT.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Obtener detalle con todas sus relaciones
        cursor.execute("""
            SELECT
                dc.*,
                cb.codigo as carton_codigo,
                cb.id as carton_id,
                pr.descripcion as proceso_nombre,
                pt.descripcion as product_type_nombre,
                cad.codigo as cad_codigo,
                mat.codigo as material_codigo,
                ms.descripcion as maquila_servicio_nombre,
                ssh.id as subsubhierarchy_id,
                ssh.nombre as jerarquia_3_nombre,
                sh.id as subhierarchy_id,
                sh.nombre as jerarquia_2_nombre,
                h.id as hierarchy_id,
                h.nombre as jerarquia_1_nombre
            FROM detalle_cotizacions dc
            LEFT JOIN cardboards cb ON dc.carton_id = cb.id
            LEFT JOIN processes pr ON dc.process_id = pr.id
            LEFT JOIN product_types pt ON dc.product_type_id = pt.id
            LEFT JOIN cads cad ON dc.cad_material_id = cad.id
            LEFT JOIN materials mat ON dc.material_id = mat.id
            LEFT JOIN maquila_servicios ms ON dc.maquila_servicio_id = ms.id
            LEFT JOIN subsubhierarchies ssh ON dc.subsubhierarchy_id = ssh.id
            LEFT JOIN subhierarchies sh ON ssh.subhierarchy_id = sh.id
            LEFT JOIN hierarchies h ON sh.hierarchy_id = h.id
            WHERE dc.id = %s
        """, (detalle_id,))

        detalle = cursor.fetchone()
        if not detalle:
            raise HTTPException(status_code=404, detail="Detalle de cotización no encontrado")

        # Obtener cotización padre
        cursor.execute("""
            SELECT
                c.*,
                cl.id as cliente_id,
                cl.nombre as cliente_nombre,
                cl.rut as cliente_rut,
                u.nombre as creador_nombre
            FROM cotizacions c
            LEFT JOIN clients cl ON c.client_id = cl.id
            LEFT JOIN users u ON c.user_id = u.id
            WHERE c.id = %s AND c.active = 1
        """, (detalle["cotizacion_id"],))

        cotizacion = cursor.fetchone()
        if not cotizacion:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        # Construir datos base para OT (comunes a todos los tipos)
        datos_ot = {
            # Datos de origen
            "detalle_cotizacion_id": detalle_id,
            "cotizacion_id": cotizacion["id"],
            "tipo_solicitud": tipo_solicitud,

            # Datos comerciales
            "client_id": cotizacion["client_id"],
            "cliente_nombre": cotizacion["cliente_nombre"],
            "nombre_contacto": cotizacion.get("nombre_contacto"),
            "email_contacto": cotizacion.get("email_contacto"),
            "telefono_contacto": cotizacion.get("telefono_contacto"),

            # Descripción del producto
            "descripcion": detalle.get("descripcion_material_detalle") or detalle.get("descripcion"),

            # Cartón
            "carton_id": detalle.get("carton_id"),
            "carton_codigo": detalle.get("carton_codigo"),

            # Jerarquías
            "hierarchy_id": detalle.get("hierarchy_id"),
            "subhierarchy_id": detalle.get("subhierarchy_id"),
            "subsubhierarchy_id": detalle.get("subsubhierarchy_id"),
            "jerarquia_1_nombre": detalle.get("jerarquia_1_nombre"),
            "jerarquia_2_nombre": detalle.get("jerarquia_2_nombre"),
            "jerarquia_3_nombre": detalle.get("jerarquia_3_nombre"),

            # Medidas
            "largo_externo": detalle.get("largura"),
            "ancho_externo": detalle.get("anchura"),
            "alto_externo": detalle.get("altura"),
            "largo_interno": detalle.get("largo_interno"),
            "ancho_interno": detalle.get("ancho_interno"),
            "alto_interno": detalle.get("alto_interno"),

            # Cantidad
            "cantidad": detalle.get("cantidad"),
        }

        # Campos específicos según tipo de solicitud
        if tipo_solicitud == 1:  # Desarrollo Completo
            datos_ot.update({
                "reference_id": detalle.get("material_id"),
                "material_codigo": detalle.get("material_codigo"),
                "cad_id": detalle.get("cad_material_id"),
                "cad_codigo": detalle.get("cad_codigo"),
                "bct_min_lb": detalle.get("bct_min_lb"),
                "bct_min_kg": detalle.get("bct_min_kg"),
                "product_type_id": detalle.get("product_type_id"),
                "product_type_nombre": detalle.get("product_type_nombre"),
                "maquila": detalle.get("maquila"),
                "maquila_servicio_id": detalle.get("maquila_servicio_id"),
                "maquila_servicio_nombre": detalle.get("maquila_servicio_nombre"),
                "golpes_ancho": detalle.get("golpes_ancho"),
                "golpes_largo": detalle.get("golpes_largo"),
                "numero_colores": detalle.get("numero_colores"),
                "process_id": detalle.get("process_id"),
                "proceso_nombre": detalle.get("proceso_nombre"),
            })

        elif tipo_solicitud == 2:  # Cotiza con CAD
            datos_ot.update({
                "cad_id": detalle.get("cad_material_id"),
                "cad_codigo": detalle.get("cad_codigo"),
            })

        elif tipo_solicitud == 5:  # Arte con Material
            datos_ot.update({
                "reference_id": detalle.get("material_id"),
                "material_codigo": detalle.get("material_codigo"),
                "cad_id": detalle.get("cad_material_id"),
                "cad_codigo": detalle.get("cad_codigo"),
                "bct": detalle.get("bct"),
                "unidad_medida_bct": detalle.get("unidad_medida_bct"),
                "maquila": detalle.get("maquila"),
                "maquila_servicio_id": detalle.get("maquila_servicio_id"),
                "maquila_servicio_nombre": detalle.get("maquila_servicio_nombre"),
                "golpes_ancho": detalle.get("golpes_ancho"),
                "golpes_largo": detalle.get("golpes_largo"),
            })

        # Tipo 3 (Muestra con CAD) no tiene campos específicos adicionales

        return {
            "success": True,
            "detalle_id": detalle_id,
            "cotizacion_id": cotizacion["id"],
            "tipo_solicitud": tipo_solicitud,
            "datos_ot": datos_ot,
            "mensaje": f"Datos preparados para crear OT tipo {tipo_solicitud}"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error obteniendo detalle para OT: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cursor.close()
        conn.close()


@router.get("/detalles-para-ot/")
async def list_detalles_para_ot(
    cotizacion_id: Optional[int] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    """
    Lista detalles de cotizaciones que pueden convertirse en OT.
    Solo muestra detalles de cotizaciones aprobadas (estado_id = 3)
    que no tienen OT asociada.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        query = """
            SELECT
                dc.id as detalle_id,
                dc.cotizacion_id,
                dc.descripcion,
                dc.descripcion_material_detalle,
                dc.cantidad,
                dc.largura,
                dc.anchura,
                dc.altura,
                c.client_id,
                cl.nombre as cliente_nombre,
                c.estado_id,
                ce.nombre as estado_nombre,
                c.created_at as cotizacion_fecha,
                cad.codigo as cad_codigo,
                cb.codigo as carton_codigo,
                CASE WHEN dc.work_order_id IS NOT NULL THEN 1 ELSE 0 END as tiene_ot,
                dc.work_order_id
            FROM detalle_cotizacions dc
            INNER JOIN cotizacions c ON dc.cotizacion_id = c.id
            LEFT JOIN clients cl ON c.client_id = cl.id
            LEFT JOIN cotizacion_estados ce ON c.estado_id = ce.id
            LEFT JOIN cads cad ON dc.cad_material_id = cad.id
            LEFT JOIN cardboards cb ON dc.carton_id = cb.id
            WHERE c.active = 1
            AND c.estado_id IN (3, 4, 5)
            AND (dc.work_order_id IS NULL OR dc.work_order_id = 0)
        """
        params = []

        if cotizacion_id:
            query += " AND dc.cotizacion_id = %s"
            params.append(cotizacion_id)

        # Contar total
        count_query = f"SELECT COUNT(*) as total FROM ({query}) as subq"
        cursor.execute(count_query, params)
        total = cursor.fetchone()["total"]

        # Aplicar paginación
        query += " ORDER BY c.created_at DESC, dc.id ASC LIMIT %s OFFSET %s"
        offset = (page - 1) * page_size
        params.extend([page_size, offset])

        cursor.execute(query, params)
        rows = cursor.fetchall()

        return {
            "items": rows,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size
        }

    finally:
        cursor.close()
        conn.close()


# =============================================
# SPRINT A: ENDPOINTS ADICIONALES PARA 100% PARIDAD
# Basado en: CotizacionController.php líneas 1199-1950
# =============================================

@router.get("/calcular-margen/{cotizacion_id}")
async def calcular_margen(cotizacion_id: int):
    """
    Calcula la diferencia porcentual de margen de una cotización.

    Basado en: CotizacionController@calcular_margen (líneas 1199-1216)

    Fórmula: ((margen_nuevo - margen_anterior) / margen_anterior) * 100

    Returns:
        margen_nuevo: Suma del margen actual de todos los detalles
        margen_anterior: Margen de la versión anterior (si existe)
        diferencia_porcentual: Porcentaje de cambio
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Obtener cotización actual con su margen
        cursor.execute("""
            SELECT c.id, c.version_padre_id,
                   COALESCE(SUM(dc.margen), 0) as margen_nuevo
            FROM cotizacions c
            LEFT JOIN detalle_cotizacions dc ON c.id = dc.cotizacion_id
            WHERE c.id = %s AND c.active = 1
            GROUP BY c.id, c.version_padre_id
        """, (cotizacion_id,))

        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        margen_nuevo = float(row["margen_nuevo"]) if row["margen_nuevo"] else 0
        version_padre_id = row["version_padre_id"]

        # Si tiene versión anterior, obtener su margen
        margen_anterior = 0
        if version_padre_id:
            cursor.execute("""
                SELECT COALESCE(SUM(dc.margen), 0) as margen_anterior
                FROM detalle_cotizacions dc
                WHERE dc.cotizacion_id = %s
            """, (version_padre_id,))

            padre_row = cursor.fetchone()
            margen_anterior = float(padre_row["margen_anterior"]) if padre_row and padre_row["margen_anterior"] else 0

        # Calcular diferencia porcentual
        diferencia_porcentual = 0
        if margen_anterior > 0:
            diferencia_porcentual = ((margen_nuevo - margen_anterior) / margen_anterior) * 100

        return {
            "cotizacion_id": cotizacion_id,
            "margen_nuevo": round(margen_nuevo, 4),
            "margen_anterior": round(margen_anterior, 4),
            "diferencia_porcentual": round(diferencia_porcentual, 2),
            "tiene_version_anterior": version_padre_id is not None
        }

    finally:
        cursor.close()
        conn.close()


@router.get("/calcular-mc-bruto/{cotizacion_id}")
async def calcular_mc_bruto(cotizacion_id: int):
    """
    Verifica si la cotización tiene margen de contribución bruto negativo.

    Basado en: CotizacionController@calcular_mc_bruto (líneas 1933-1950)

    Retorna alertas si algún detalle tiene MC bruto < 0
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Obtener detalles con sus cálculos de MC bruto
        cursor.execute("""
            SELECT
                dc.id,
                dc.descripcion,
                dc.cantidad,
                dc.precio_unitario,
                dc.costo_total,
                (dc.precio_unitario * dc.cantidad) - dc.costo_total as mc_bruto,
                CASE
                    WHEN dc.costo_total > 0
                    THEN (((dc.precio_unitario * dc.cantidad) - dc.costo_total) / dc.costo_total) * 100
                    ELSE 0
                END as margen_porcentual
            FROM detalle_cotizacions dc
            WHERE dc.cotizacion_id = %s
        """, (cotizacion_id,))

        detalles = cursor.fetchall()

        alertas = []
        mc_bruto_total = 0

        for d in detalles:
            mc_bruto = float(d["mc_bruto"]) if d["mc_bruto"] else 0
            mc_bruto_total += mc_bruto

            if mc_bruto < 0:
                alertas.append({
                    "detalle_id": d["id"],
                    "descripcion": d["descripcion"],
                    "mc_bruto": round(mc_bruto, 2),
                    "mensaje": f"El detalle '{d['descripcion']}' tiene MC bruto negativo: {round(mc_bruto, 2)}"
                })

        return {
            "cotizacion_id": cotizacion_id,
            "mc_bruto_total": round(mc_bruto_total, 2),
            "tiene_alertas": len(alertas) > 0,
            "cantidad_alertas": len(alertas),
            "alertas": alertas
        }

    finally:
        cursor.close()
        conn.close()


@router.get("/detalles-corrugados/{cotizacion_id}/excel")
async def export_detalles_corrugados(cotizacion_id: int):
    """
    Exporta los detalles tipo corrugado de una cotización a Excel.

    Basado en: CotizacionController@detalles_corrugados (líneas 1694-1760)

    Solo incluye detalles donde tipo_detalle_id = 1 (corrugados)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Obtener cotización
        cursor.execute("""
            SELECT c.*, cl.nombre as cliente_nombre
            FROM cotizacions c
            LEFT JOIN clients cl ON c.client_id = cl.id
            WHERE c.id = %s AND c.active = 1
        """, (cotizacion_id,))

        cotizacion = cursor.fetchone()
        if not cotizacion:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        # Obtener detalles corrugados (tipo_detalle_id = 1)
        cursor.execute("""
            SELECT
                dc.*,
                cb.codigo as carton_codigo,
                cb.descripcion as carton_descripcion,
                p.nombre as planta_nombre,
                cad.cad as cad_codigo
            FROM detalle_cotizacions dc
            LEFT JOIN cardboards cb ON dc.carton_id = cb.id
            LEFT JOIN plantas p ON dc.planta_id = p.id
            LEFT JOIN cads cad ON dc.cad_material_id = cad.id
            WHERE dc.cotizacion_id = %s
            AND dc.tipo_detalle_id = 1
            ORDER BY dc.id
        """, (cotizacion_id,))

        detalles = cursor.fetchall()

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Corrugados"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws.merge_cells('A1:K1')
        ws['A1'] = f"Cotización #{cotizacion_id} - Detalles Corrugados"
        ws['A1'].font = Font(bold=True, size=14)

        ws.merge_cells('A2:K2')
        ws['A2'] = f"Cliente: {cotizacion['cliente_nombre']}"

        # Headers
        headers = ['ID', 'Descripción', 'CAD', 'Cartón', 'Cantidad', 'Largura', 'Anchura',
                   'Área HC', 'Precio Unit.', 'Costo Total', 'Planta']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Datos
        for row_idx, d in enumerate(detalles, 5):
            ws.cell(row=row_idx, column=1, value=d['id']).border = border
            ws.cell(row=row_idx, column=2, value=d['descripcion']).border = border
            ws.cell(row=row_idx, column=3, value=d['cad_codigo']).border = border
            ws.cell(row=row_idx, column=4, value=d['carton_codigo']).border = border
            ws.cell(row=row_idx, column=5, value=d['cantidad']).border = border
            ws.cell(row=row_idx, column=6, value=d['largura']).border = border
            ws.cell(row=row_idx, column=7, value=d['anchura']).border = border
            ws.cell(row=row_idx, column=8, value=float(d['area_hc'] or 0)).border = border
            ws.cell(row=row_idx, column=9, value=float(d['precio_unitario'] or 0)).border = border
            ws.cell(row=row_idx, column=10, value=float(d['costo_total'] or 0)).border = border
            ws.cell(row=row_idx, column=11, value=d['planta_nombre']).border = border

        # Ajustar anchos
        for col in range(1, 12):
            ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = 15

        # Guardar a buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"cotizacion_{cotizacion_id}_corrugados.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


@router.get("/detalles-esquineros/{cotizacion_id}/excel")
async def export_detalles_esquineros(cotizacion_id: int):
    """
    Exporta los detalles tipo esquinero de una cotización a Excel.

    Basado en: CotizacionController@detalles_esquineros (líneas 1762-1803)

    Solo incluye detalles donde tipo_detalle_id = 2 (esquineros)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Obtener cotización
        cursor.execute("""
            SELECT c.*, cl.nombre as cliente_nombre
            FROM cotizacions c
            LEFT JOIN clients cl ON c.client_id = cl.id
            WHERE c.id = %s AND c.active = 1
        """, (cotizacion_id,))

        cotizacion = cursor.fetchone()
        if not cotizacion:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        # Obtener detalles esquineros (tipo_detalle_id = 2)
        cursor.execute("""
            SELECT
                dc.*,
                p.nombre as planta_nombre
            FROM detalle_cotizacions dc
            LEFT JOIN plantas p ON dc.planta_id = p.id
            WHERE dc.cotizacion_id = %s
            AND dc.tipo_detalle_id = 2
            ORDER BY dc.id
        """, (cotizacion_id,))

        detalles = cursor.fetchall()

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Esquineros"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws.merge_cells('A1:H1')
        ws['A1'] = f"Cotización #{cotizacion_id} - Detalles Esquineros"
        ws['A1'].font = Font(bold=True, size=14)

        ws.merge_cells('A2:H2')
        ws['A2'] = f"Cliente: {cotizacion['cliente_nombre']}"

        # Headers específicos para esquineros
        headers = ['ID', 'Descripción', 'Cantidad', 'Largo', 'Ala', 'Espesor',
                   'Precio Unit.', 'Planta']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Datos
        for row_idx, d in enumerate(detalles, 5):
            ws.cell(row=row_idx, column=1, value=d['id']).border = border
            ws.cell(row=row_idx, column=2, value=d['descripcion']).border = border
            ws.cell(row=row_idx, column=3, value=d['cantidad']).border = border
            ws.cell(row=row_idx, column=4, value=d['largo_esquinero']).border = border
            ws.cell(row=row_idx, column=5, value=d['ala_esquinero']).border = border
            ws.cell(row=row_idx, column=6, value=d['espesor_esquinero']).border = border
            ws.cell(row=row_idx, column=7, value=float(d['precio_unitario'] or 0)).border = border
            ws.cell(row=row_idx, column=8, value=d['planta_nombre']).border = border

        # Ajustar anchos
        for col in range(1, 9):
            ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = 15

        # Guardar a buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"cotizacion_{cotizacion_id}_esquineros.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cursor.close()
        conn.close()


@router.get("/obtener-margen-papeles/{cotizacion_id}")
async def obtener_margen_papeles(cotizacion_id: int):
    """
    Obtiene el margen calculado desde los papeles (liners/médiums) de la cotización.

    Basado en: CotizacionController@obtenerMargenPapeles (líneas 1810-1857)

    Calcula el margen basado en los costos de papel del cartón utilizado.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Obtener detalles con información de cartón
        cursor.execute("""
            SELECT
                dc.id,
                dc.cantidad,
                dc.area_hc,
                dc.carton_id,
                cb.codigo as carton_codigo,
                cb.gramaje_liner_1,
                cb.gramaje_medium_1,
                cb.gramaje_liner_2,
                cb.gramaje_medium_2,
                cb.gramaje_liner_3
            FROM detalle_cotizacions dc
            LEFT JOIN cardboards cb ON dc.carton_id = cb.id
            WHERE dc.cotizacion_id = %s
            AND dc.tipo_detalle_id = 1
        """, (cotizacion_id,))

        detalles = cursor.fetchall()

        # Obtener precios de papeles
        cursor.execute("""
            SELECT id, gramaje, precio_usd_ton
            FROM papel_liners
            WHERE active = 1
        """)
        liners = {r['gramaje']: float(r['precio_usd_ton']) for r in cursor.fetchall()}

        cursor.execute("""
            SELECT id, gramaje, precio_usd_ton
            FROM papel_mediums
            WHERE active = 1
        """)
        mediums = {r['gramaje']: float(r['precio_usd_ton']) for r in cursor.fetchall()}

        resultados = []
        margen_papeles_total = 0

        for d in detalles:
            area_m2 = float(d['area_hc'] or 0) / 1000000  # mm2 a m2
            cantidad = d['cantidad'] or 0

            # Calcular costo de papeles para este detalle
            costo_liner = 0
            costo_medium = 0

            # Liner 1
            if d['gramaje_liner_1'] and d['gramaje_liner_1'] in liners:
                costo_liner += liners[d['gramaje_liner_1']] * area_m2 * d['gramaje_liner_1'] / 1000

            # Medium 1
            if d['gramaje_medium_1'] and d['gramaje_medium_1'] in mediums:
                costo_medium += mediums[d['gramaje_medium_1']] * area_m2 * d['gramaje_medium_1'] / 1000

            # Liner 2
            if d['gramaje_liner_2'] and d['gramaje_liner_2'] in liners:
                costo_liner += liners[d['gramaje_liner_2']] * area_m2 * d['gramaje_liner_2'] / 1000

            # Medium 2
            if d['gramaje_medium_2'] and d['gramaje_medium_2'] in mediums:
                costo_medium += mediums[d['gramaje_medium_2']] * area_m2 * d['gramaje_medium_2'] / 1000

            # Liner 3
            if d['gramaje_liner_3'] and d['gramaje_liner_3'] in liners:
                costo_liner += liners[d['gramaje_liner_3']] * area_m2 * d['gramaje_liner_3'] / 1000

            costo_papeles = (costo_liner + costo_medium) * cantidad
            margen_papeles_total += costo_papeles

            resultados.append({
                "detalle_id": d['id'],
                "carton_codigo": d['carton_codigo'],
                "costo_liner": round(costo_liner * cantidad, 4),
                "costo_medium": round(costo_medium * cantidad, 4),
                "costo_papeles_total": round(costo_papeles, 4)
            })

        return {
            "cotizacion_id": cotizacion_id,
            "margen_papeles_total": round(margen_papeles_total, 4),
            "detalles": resultados
        }

    finally:
        cursor.close()
        conn.close()


@router.post("/{cotizacion_id}/retomar-externo")
async def retomar_cotizacion_externo(cotizacion_id: int):
    """
    Retoma una cotización externa para volver a editarla.

    Basado en: CotizacionController@retomarCotizacionExterno (líneas 1291-1303)

    Cambia el estado a borrador (1) y limpia los permisos de aprobación.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Verificar que existe
        cursor.execute(
            "SELECT id, estado_id FROM cotizacions WHERE id = %s AND active = 1",
            (cotizacion_id,)
        )
        cotizacion = cursor.fetchone()

        if not cotizacion:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        # Cambiar estado a borrador y limpiar permisos
        cursor.execute("""
            UPDATE cotizacions
            SET estado_id = 1,
                role_can_show = NULL,
                nivel_aprobacion = NULL,
                updated_at = NOW()
            WHERE id = %s
        """, (cotizacion_id,))

        conn.commit()

        return {
            "success": True,
            "message": "Cotización retomada exitosamente",
            "cotizacion_id": cotizacion_id,
            "estado_nuevo": 1
        }

    finally:
        cursor.close()
        conn.close()


@router.post("/{cotizacion_id}/editar-externo")
async def editar_cotizacion_externo(cotizacion_id: int):
    """
    Permite editar una cotización externa volviendo a estado borrador.

    Basado en: CotizacionController@editarCotizacionExterno (líneas 1305-1317)

    Similar a retomar pero para el flujo de edición.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Verificar que existe
        cursor.execute(
            "SELECT id, estado_id FROM cotizacions WHERE id = %s AND active = 1",
            (cotizacion_id,)
        )
        cotizacion = cursor.fetchone()

        if not cotizacion:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        # Cambiar estado a borrador y limpiar permisos
        cursor.execute("""
            UPDATE cotizacions
            SET estado_id = 1,
                role_can_show = NULL,
                nivel_aprobacion = NULL,
                updated_at = NOW()
            WHERE id = %s
        """, (cotizacion_id,))

        conn.commit()

        return {
            "success": True,
            "message": "Cotización lista para edición",
            "cotizacion_id": cotizacion_id,
            "estado_nuevo": 1
        }

    finally:
        cursor.close()
        conn.close()


@router.get("/carga-materiales")
async def carga_materiales(
    codigo_material: Optional[str] = None,
    descripcion_material: Optional[str] = None,
    style_id: Optional[int] = None,
    cad: Optional[str] = None,
    limit: int = Query(100, le=500),
):
    """
    Busca materiales para cargar en cotización.

    Basado en: CotizacionController@cargaMateriales (líneas 1319-1346)

    Permite filtrar por código, descripción, estilo y CAD.
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        query = """
            SELECT
                m.id,
                m.codigo,
                m.descripcion,
                m.cad_id,
                m.client_id,
                m.carton_id,
                m.product_type_id,
                m.style_id,
                cad.cad as cad_codigo,
                cl.nombre as cliente_nombre,
                cb.codigo as carton_codigo,
                pt.nombre as tipo_producto_nombre,
                s.nombre as estilo_nombre
            FROM materials m
            LEFT JOIN cads cad ON m.cad_id = cad.id
            LEFT JOIN clients cl ON m.client_id = cl.id
            LEFT JOIN cardboards cb ON m.carton_id = cb.id
            LEFT JOIN product_types pt ON m.product_type_id = pt.id
            LEFT JOIN styles s ON m.style_id = s.id
            WHERE m.active = 1
        """
        params = []

        if codigo_material:
            query += " AND m.codigo LIKE %s"
            params.append(f"%{codigo_material}%")

        if descripcion_material:
            query += " AND m.descripcion LIKE %s"
            params.append(f"%{descripcion_material}%")

        if style_id:
            query += " AND m.style_id = %s"
            params.append(style_id)

        if cad:
            query += " AND cad.cad LIKE %s"
            params.append(f"%{cad}%")

        query += f" LIMIT {limit}"

        cursor.execute(query, params)
        materiales = cursor.fetchall()

        return {
            "mensaje": "Materiales encontrados exitosamente",
            "total": len(materiales),
            "materiales": materiales
        }

    finally:
        cursor.close()
        conn.close()


class SolicitarAprobacionNewRequest(BaseModel):
    """Request para solicitar aprobación (nuevo flujo)."""
    guardar_historial: bool = True


@router.post("/{cotizacion_id}/solicitar-aprobacion-new")
async def solicitar_aprobacion_new(
    cotizacion_id: int,
    data: Optional[SolicitarAprobacionNewRequest] = None,
    current_user: dict = Depends(get_current_user_optional)
):
    """
    Solicita aprobación de cotización con lógica de margen y rechazo automático.

    Basado en: CotizacionController@solicitarAprobacionNew (líneas 1859-1931)

    Flujo:
    1. Guarda historial de precios en cada detalle
    2. Calcula margen de la cotización
    3. Si cliente es clasificación 4 y margen != 0: rechazo automático
    4. Si margen <= 0: aprobación automática (estado 3)
    5. Si margen < 4%: requiere aprobación de Jefe de Área (role 3)
    6. Si margen >= 4%: requiere aprobación de Gerente Comercial (role 15)
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Obtener cotización con cliente
        cursor.execute("""
            SELECT c.*, cl.clasificacion as cliente_clasificacion
            FROM cotizacions c
            LEFT JOIN clients cl ON c.client_id = cl.id
            WHERE c.id = %s AND c.active = 1
        """, (cotizacion_id,))

        cotizacion = cursor.fetchone()
        if not cotizacion:
            raise HTTPException(status_code=404, detail="Cotización no encontrada")

        if cotizacion["estado_id"] != 1:
            raise HTTPException(
                status_code=400,
                detail="Solo se puede solicitar aprobación de cotizaciones en borrador"
            )

        # Guardar historial de precios en cada detalle
        if data is None or data.guardar_historial:
            cursor.execute("""
                UPDATE detalle_cotizacions
                SET historial_resultados = (
                    SELECT JSON_OBJECT(
                        'precio_unitario', precio_unitario,
                        'costo_total', costo_total,
                        'margen', margen,
                        'fecha', NOW()
                    )
                )
                WHERE cotizacion_id = %s
            """, (cotizacion_id,))

        # Calcular margen
        cursor.execute("""
            SELECT
                COALESCE(SUM(margen), 0) as margen_total,
                COUNT(*) as num_detalles
            FROM detalle_cotizacions
            WHERE cotizacion_id = %s
        """, (cotizacion_id,))

        margen_result = cursor.fetchone()
        margen = float(margen_result["margen_total"]) if margen_result else 0

        # Verificar MC bruto negativo
        cursor.execute("""
            SELECT COUNT(*) as negativos
            FROM detalle_cotizacions
            WHERE cotizacion_id = %s
            AND (precio_unitario * cantidad) - costo_total < 0
        """, (cotizacion_id,))

        mc_result = cursor.fetchone()
        mc_bruto_negativo = mc_result["negativos"] > 0 if mc_result else False

        # Determinar estado y aprobador
        estado_nuevo = 2  # Pendiente aprobación
        role_can_show = None
        nivel_aprobacion = None
        enviar_a_comite = 0
        rechazo_automatico = False
        mensaje = ""

        clasificacion_cliente = cotizacion.get("cliente_clasificacion")

        # Rechazo automático: Presencia Rentable (clasificación 4) con margen != 0
        if clasificacion_cliente == 4 and margen != 0:
            estado_nuevo = 6  # Rechazado
            rechazo_automatico = True
            mensaje = "Rechazo automático: Presencia Rentable requiere margen igual a 0"

            # Crear registro de aprobación (rechazo automático)
            cursor.execute("""
                INSERT INTO cotizacion_approvals (
                    cotizacion_id, user_id, role_do_action, action_made,
                    motivo, created_at, updated_at
                ) VALUES (%s, %s, 1, 'Rechazo Automático',
                    'Presencia Rentable y no se pone un margen igual o mayor al minimo',
                    NOW(), NOW())
            """, (cotizacion_id, current_user.get("id", 1)))

        elif margen <= 0:
            # Aprobación automática
            estado_nuevo = 3  # Aprobado
            mensaje = "Cotización aprobada automáticamente (margen <= 0)"

        elif margen < 4:
            # Requiere aprobación de Jefe de Área
            role_can_show = 3
            nivel_aprobacion = 1
            mensaje = "Cotización enviada a Jefe de Área para aprobación"

        else:
            # Requiere aprobación de Gerente Comercial
            role_can_show = 15
            nivel_aprobacion = 1
            if mc_bruto_negativo:
                enviar_a_comite = 1
            mensaje = "Cotización enviada a Gerente Comercial para aprobación"

        # Actualizar cotización
        cursor.execute("""
            UPDATE cotizacions
            SET estado_id = %s,
                role_can_show = %s,
                nivel_aprobacion = %s,
                enviar_a_comite = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (estado_nuevo, role_can_show, nivel_aprobacion, enviar_a_comite, cotizacion_id))

        conn.commit()

        return {
            "success": True,
            "cotizacion_id": cotizacion_id,
            "estado_nuevo": estado_nuevo,
            "margen_calculado": round(margen, 4),
            "rechazo_automatico": rechazo_automatico,
            "mc_bruto_negativo": mc_bruto_negativo,
            "enviar_a_comite": enviar_a_comite == 1,
            "role_aprobador": role_can_show,
            "mensaje": mensaje
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error en solicitar_aprobacion_new: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        cursor.close()
        conn.close()
