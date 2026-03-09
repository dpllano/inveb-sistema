"""
Router de Gestiones de OT - INVEB Cascade Service
==================================================

Endpoints para gestión de órdenes de trabajo (cambios de estado, consultas, respuestas).

Fuente Laravel: ManagementController.php (~1,500 líneas)
Rutas Laravel (web.php líneas 224-245):
- GET /gestionarOt/{id}
- GET /reactivarOt/{id}
- GET /detalleLogOt/{id}
- GET /descargar-detalle-log-excel/{id}
- POST /crear-gestion/{id}
- POST /respuesta/{id}
- GET /retomarOt/{id}
"""
from datetime import datetime
from typing import Optional, List, Dict, Any
from io import BytesIO
from fastapi import APIRouter, HTTPException, Depends, status, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import pymysql

from ..config import get_settings
from .auth import get_current_user

router = APIRouter(prefix="/gestiones", tags=["Gestiones OT"])
settings = get_settings()


# =============================================
# SCHEMAS
# =============================================

class GestionBase(BaseModel):
    observacion: Optional[str] = None
    management_type_id: int
    state_id: Optional[int] = None
    work_space_id: Optional[int] = None
    consulta_area_id: Optional[int] = None
    motivo_id: Optional[int] = None


class GestionCreate(GestionBase):
    pass


class GestionResponse(BaseModel):
    id: int
    observacion: Optional[str]
    management_type_id: int
    management_type_nombre: Optional[str]
    state_id: Optional[int]
    state_nombre: Optional[str]
    work_space_id: Optional[int]
    work_space_nombre: Optional[str]
    user_id: int
    user_nombre: Optional[str]
    work_order_id: int
    duracion_segundos: Optional[int]
    created_at: datetime

    class Config:
        from_attributes = True


class RespuestaCreate(BaseModel):
    respuesta: str


class RespuestaResponse(BaseModel):
    id: int
    respuesta: str
    user_id: int
    user_nombre: Optional[str]
    management_id: int
    created_at: datetime

    class Config:
        from_attributes = True


class GestionarOTResponse(BaseModel):
    ot: Dict[str, Any]
    gestiones: List[Dict[str, Any]]
    states_by_area: List[Dict[str, Any]]
    work_spaces: List[Dict[str, Any]]
    management_types: List[Dict[str, Any]]
    motivos: List[Dict[str, Any]]
    usuario_asignado: bool


class DetalleLogResponse(BaseModel):
    ot_id: int
    ot_numero: Optional[str]
    gestiones: List[Dict[str, Any]]
    total_gestiones: int


# =============================================
# HELPERS
# =============================================

def get_mysql_connection():
    """Crea conexión a MySQL de Laravel."""
    return pymysql.connect(
        host=settings.LARAVEL_MYSQL_HOST,
        port=settings.LARAVEL_MYSQL_PORT,
        user=settings.LARAVEL_MYSQL_USER,
        password=settings.LARAVEL_MYSQL_PASSWORD,
        database=settings.LARAVEL_MYSQL_DATABASE,
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )


def get_states_by_area(area_id: int, ot: Dict, connection) -> List[int]:
    """
    Retorna los estados permitidos según el área del usuario.

    Fuente Laravel: ManagementController@gestionarOt líneas 69-285
    """
    cursor = connection.cursor()

    # Constantes de áreas (de Laravel Constants.php)
    AREA_VENTA = 1
    AREA_DESARROLLO = 2
    AREA_DISEÑO = 3
    AREA_CATALOGACION = 4
    AREA_MUESTRAS = 6

    states = []

    if area_id == AREA_VENTA:
        # Verificar si ha sido enviado a desarrollo
        cursor.execute("""
            SELECT COUNT(*) as count FROM managements
            WHERE work_order_id = %s AND management_type_id = 1 AND state_id = 2
        """, (ot['id'],))
        enviado_desarrollo = cursor.fetchone()['count'] > 0

        if enviado_desarrollo:
            if ot.get('tipo_solicitud') == 6:
                states = [2, 10, 11, 15, 8]
            else:
                states = [2, 5, 6, 7, 9, 10, 11, 14, 15, 16, 20, 21]
        else:
            if ot.get('tipo_solicitud') == 6:
                states = [2, 10, 11, 15, 8]
            else:
                states = [2, 9, 10, 11, 14, 15, 16, 20, 21]

    elif area_id == AREA_DESARROLLO:
        if ot.get('tipo_solicitud') == 6:
            if ot.get('ajuste_area_desarrollo') == 3:
                states = [1, 3, 12, 13]
            else:
                states = [1, 3, 17, 12, 13]
        else:
            states = [1, 3, 5, 6, 7, 12, 16, 17]
            if ot.get('tipo_solicitud') != 1:
                states.append(13)

    elif area_id == AREA_DISEÑO:
        states = [1, 2, 7, 12, 16]
        if ot.get('tipo_solicitud') != 1:
            states.append(13)

    elif area_id == AREA_CATALOGACION:
        if ot.get('current_area_id') == 4:
            states = [1, 2, 5, 8, 12]
        elif ot.get('current_area_id') == 5:
            states = [1, 2, 5, 7, 12]

    elif area_id == AREA_MUESTRAS:
        states = [12, 18, 22]

    cursor.close()
    return states


def get_management_types_by_area(area_id: int, ot: Dict, connection) -> List[int]:
    """
    Retorna los tipos de gestión permitidos según área y estado de la OT.

    Fuente Laravel: ManagementController@gestionarOt líneas 326-388
    """
    if ot.get('current_area_id') == area_id:
        if ot.get('current_area_id') == 2:
            if ot.get('tipo_solicitud') == 6:
                return [1, 2, 3]
            else:
                return [1, 2, 3, 6]
        else:
            return [1, 2, 3, 8]
    else:
        return [2, 3]


# =============================================
# ENDPOINTS
# =============================================

@router.get("/{ot_id}", response_model=GestionarOTResponse)
async def gestionar_ot(
    ot_id: int,
    current_user: dict = Depends(get_current_user)
):
    """
    Obtiene datos completos para gestionar una OT.

    Fuente Laravel: ManagementController@gestionarOt (líneas 47-646)
    Ruta Laravel: GET /gestionarOt/{id}

    Retorna:
    - Datos de la OT
    - Historial de gestiones
    - Estados permitidos según área del usuario
    - Tipos de gestión permitidos
    - Motivos de rechazo
    """
    connection = get_mysql_connection()
    try:
        cursor = connection.cursor()

        # Obtener OT con relaciones
        cursor.execute("""
            SELECT wo.*,
                   c.nombre as cliente_nombre,
                   c.codigo as cliente_codigo,
                   s.nombre as state_nombre,
                   ws.nombre as area_nombre
            FROM work_orders wo
            LEFT JOIN clients c ON wo.client_id = c.id
            LEFT JOIN states s ON wo.state_id = s.id
            LEFT JOIN work_spaces ws ON wo.current_area_id = ws.id
            WHERE wo.id = %s AND wo.active = 1
        """, (ot_id,))
        ot = cursor.fetchone()

        if not ot:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"OT {ot_id} no encontrada"
            )

        # Obtener gestiones con relaciones
        cursor.execute("""
            SELECT m.*,
                   u.name as user_nombre,
                   mt.nombre as management_type_nombre,
                   s.nombre as state_nombre,
                   ws.nombre as work_space_nombre,
                   wsc.nombre as consulta_area_nombre
            FROM managements m
            LEFT JOIN users u ON m.user_id = u.id
            LEFT JOIN management_types mt ON m.management_type_id = mt.id
            LEFT JOIN states s ON m.state_id = s.id
            LEFT JOIN work_spaces ws ON m.work_space_id = ws.id
            LEFT JOIN work_spaces wsc ON m.consulta_area_id = wsc.id
            WHERE m.work_order_id = %s
            ORDER BY m.created_at DESC
        """, (ot_id,))
        gestiones = cursor.fetchall()

        # Obtener respuestas de cada gestión
        for gestion in gestiones:
            cursor.execute("""
                SELECT a.*, u.name as user_nombre
                FROM answers a
                LEFT JOIN users u ON a.user_id = u.id
                WHERE a.management_id = %s
                ORDER BY a.created_at ASC
            """, (gestion['id'],))
            gestion['respuestas'] = cursor.fetchall()

        # Verificar si usuario está asignado a la OT
        cursor.execute("""
            SELECT COUNT(*) as count FROM user_work_orders
            WHERE work_order_id = %s AND user_id = %s
        """, (ot_id, current_user.get('id', 0)))
        usuario_asignado = cursor.fetchone()['count'] > 0

        # Obtener área del usuario
        cursor.execute("""
            SELECT r.area_id FROM users u
            JOIN roles r ON u.role_id = r.id
            WHERE u.id = %s
        """, (current_user.get('id', 0),))
        user_area = cursor.fetchone()
        area_id = user_area['area_id'] if user_area else None

        # Estados según área
        state_ids = get_states_by_area(area_id, ot, connection) if area_id else []

        if state_ids:
            placeholders = ','.join(['%s'] * len(state_ids))
            cursor.execute(f"""
                SELECT id, nombre FROM states WHERE id IN ({placeholders})
            """, state_ids)
            states_by_area = cursor.fetchall()
        else:
            states_by_area = []

        # Tipos de gestión según área
        if area_id:
            type_ids = get_management_types_by_area(area_id, ot, connection)
            placeholders = ','.join(['%s'] * len(type_ids))
            cursor.execute(f"""
                SELECT id, nombre FROM management_types WHERE id IN ({placeholders})
            """, type_ids)
            management_types = cursor.fetchall()
        else:
            management_types = []

        # Áreas de trabajo
        cursor.execute("""
            SELECT id, nombre FROM work_spaces WHERE status = 'active'
        """)
        work_spaces = cursor.fetchall()

        # Motivos de rechazo
        motivos = [
            {"id": 1, "nombre": "Falta Información"},
            {"id": 2, "nombre": "Información Erronea"},
            {"id": 3, "nombre": "Falta Muestra Física"},
            {"id": 4, "nombre": "Formato Imagen Inadecuado"},
            {"id": 5, "nombre": "Medida Erronea"},
            {"id": 6, "nombre": "Descripción de Producto"},
            {"id": 7, "nombre": "Plano mal Acotado"},
            {"id": 8, "nombre": "Error de Digitación"},
            {"id": 9, "nombre": "Error tipo Sustrato"},
            {"id": 10, "nombre": "No viable por restricciones"},
            {"id": 11, "nombre": "Falta CAD para corte"},
            {"id": 12, "nombre": "Falta OT Chileexpress"},
            {"id": 13, "nombre": "Falta OT Laboratorio"}
        ]

        # Convertir fechas a string para JSON
        for key in ot:
            if isinstance(ot[key], datetime):
                ot[key] = ot[key].isoformat()

        for gestion in gestiones:
            for key in gestion:
                if isinstance(gestion[key], datetime):
                    gestion[key] = gestion[key].isoformat()
            for resp in gestion.get('respuestas', []):
                for key in resp:
                    if isinstance(resp[key], datetime):
                        resp[key] = resp[key].isoformat()

        return GestionarOTResponse(
            ot=ot,
            gestiones=gestiones,
            states_by_area=states_by_area,
            work_spaces=work_spaces,
            management_types=management_types,
            motivos=motivos,
            usuario_asignado=usuario_asignado
        )

    finally:
        connection.close()


@router.post("/{ot_id}/crear", response_model=GestionResponse)
async def crear_gestion(
    ot_id: int,
    gestion: GestionCreate,
    current_user: dict = Depends(get_current_user)
):
    """
    Crea una nueva gestión para una OT.

    Fuente Laravel: ManagementController@store (líneas 761-1000+)
    Ruta Laravel: POST /crear-gestion/{id}

    Tipos de gestión:
    - 1: Cambio de Estado
    - 2: Consulta a Área
    - 3: Comentario
    - 6: Datos cargados con lector PDF
    - 8: Solicitud de Muestra
    - 9: Envío a Diseñador Externo
    """
    connection = get_mysql_connection()
    try:
        cursor = connection.cursor()

        # Verificar OT existe
        cursor.execute("""
            SELECT id, current_area_id, ultimo_cambio_area
            FROM work_orders WHERE id = %s AND active = 1
        """, (ot_id,))
        ot = cursor.fetchone()

        if not ot:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"OT {ot_id} no encontrada"
            )

        # Calcular duración desde último cambio
        ultimo_cambio = ot.get('ultimo_cambio_area')
        duracion_segundos = 0
        if ultimo_cambio:
            diff = datetime.now() - ultimo_cambio
            duracion_segundos = int(diff.total_seconds())

        # Preparar observación
        observacion = gestion.observacion or "sin observaciones por usuario"
        if gestion.management_type_id == 6:
            observacion = f"Datos OT cargados con lector PDF. {observacion}"

        # Insertar gestión
        cursor.execute("""
            INSERT INTO managements
            (observacion, management_type_id, user_id, work_order_id, work_space_id,
             state_id, consulta_area_id, duracion_segundos, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
        """, (
            observacion,
            gestion.management_type_id,
            current_user.get('id'),
            ot_id,
            gestion.work_space_id,
            gestion.state_id,
            gestion.consulta_area_id,
            duracion_segundos
        ))
        gestion_id = cursor.lastrowid

        # Si es cambio de estado (type 1), actualizar OT
        if gestion.management_type_id == 1 and gestion.state_id:
            cursor.execute("""
                UPDATE work_orders
                SET state_id = %s, ultimo_cambio_area = NOW(), updated_at = NOW()
                WHERE id = %s
            """, (gestion.state_id, ot_id))

            # Si cambia de área
            if gestion.work_space_id:
                cursor.execute("""
                    UPDATE work_orders
                    SET current_area_id = %s
                    WHERE id = %s
                """, (gestion.work_space_id, ot_id))

        connection.commit()

        # Obtener gestión creada con relaciones
        cursor.execute("""
            SELECT m.*,
                   u.name as user_nombre,
                   mt.nombre as management_type_nombre,
                   s.nombre as state_nombre,
                   ws.nombre as work_space_nombre
            FROM managements m
            LEFT JOIN users u ON m.user_id = u.id
            LEFT JOIN management_types mt ON m.management_type_id = mt.id
            LEFT JOIN states s ON m.state_id = s.id
            LEFT JOIN work_spaces ws ON m.work_space_id = ws.id
            WHERE m.id = %s
        """, (gestion_id,))
        nueva_gestion = cursor.fetchone()

        return GestionResponse(
            id=nueva_gestion['id'],
            observacion=nueva_gestion['observacion'],
            management_type_id=nueva_gestion['management_type_id'],
            management_type_nombre=nueva_gestion.get('management_type_nombre'),
            state_id=nueva_gestion.get('state_id'),
            state_nombre=nueva_gestion.get('state_nombre'),
            work_space_id=nueva_gestion.get('work_space_id'),
            work_space_nombre=nueva_gestion.get('work_space_nombre'),
            user_id=nueva_gestion['user_id'],
            user_nombre=nueva_gestion.get('user_nombre'),
            work_order_id=nueva_gestion['work_order_id'],
            duracion_segundos=nueva_gestion.get('duracion_segundos'),
            created_at=nueva_gestion['created_at']
        )

    except Exception as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al crear gestión: {str(e)}"
        )
    finally:
        connection.close()


@router.post("/{gestion_id}/respuesta", response_model=RespuestaResponse)
async def crear_respuesta(
    gestion_id: int,
    respuesta: RespuestaCreate,
    current_user: dict = Depends(get_current_user)
):
    """
    Crea una respuesta a una gestión/consulta.

    Fuente Laravel: ManagementController@storeRespuesta (líneas 648-662)
    Ruta Laravel: POST /respuesta/{id}
    """
    connection = get_mysql_connection()
    try:
        cursor = connection.cursor()

        # Verificar gestión existe
        cursor.execute("""
            SELECT id, work_order_id, user_id FROM managements WHERE id = %s
        """, (gestion_id,))
        gestion = cursor.fetchone()

        if not gestion:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Gestión {gestion_id} no encontrada"
            )

        # Insertar respuesta
        cursor.execute("""
            INSERT INTO answers (respuesta, user_id, management_id, created_at, updated_at)
            VALUES (%s, %s, %s, NOW(), NOW())
        """, (respuesta.respuesta, current_user.get('id'), gestion_id))
        respuesta_id = cursor.lastrowid

        # Crear notificación para el usuario que hizo la consulta
        mensaje = respuesta.respuesta[:180] + "..." if len(respuesta.respuesta) > 180 else respuesta.respuesta
        cursor.execute("""
            INSERT INTO notifications
            (user_id, work_order_id, titulo, mensaje, active, created_at, updated_at)
            VALUES (%s, %s, %s, %s, 1, NOW(), NOW())
        """, (gestion['user_id'], gestion['work_order_id'], "Respuesta a Consulta", mensaje))

        connection.commit()

        # Obtener respuesta creada
        cursor.execute("""
            SELECT a.*, u.name as user_nombre
            FROM answers a
            LEFT JOIN users u ON a.user_id = u.id
            WHERE a.id = %s
        """, (respuesta_id,))
        nueva_respuesta = cursor.fetchone()

        return RespuestaResponse(
            id=nueva_respuesta['id'],
            respuesta=nueva_respuesta['respuesta'],
            user_id=nueva_respuesta['user_id'],
            user_nombre=nueva_respuesta.get('user_nombre'),
            management_id=nueva_respuesta['management_id'],
            created_at=nueva_respuesta['created_at']
        )

    except Exception as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al crear respuesta: {str(e)}"
        )
    finally:
        connection.close()


@router.get("/{ot_id}/log", response_model=DetalleLogResponse)
async def detalle_log_ot(
    ot_id: int,
    current_user: dict = Depends(get_current_user)
):
    """
    Obtiene el historial completo de gestiones de una OT.

    Fuente Laravel: ManagementController@detalleLogOt
    Ruta Laravel: GET /detalleLogOt/{id}
    """
    connection = get_mysql_connection()
    try:
        cursor = connection.cursor()

        # Verificar OT existe
        cursor.execute("""
            SELECT id FROM work_orders WHERE id = %s AND active = 1
        """, (ot_id,))
        ot = cursor.fetchone()

        if not ot:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"OT {ot_id} no encontrada"
            )

        # Obtener gestiones con relaciones
        cursor.execute("""
            SELECT m.*,
                   u.name as user_nombre,
                   mt.nombre as management_type_nombre,
                   s.nombre as state_nombre,
                   ws.nombre as work_space_nombre,
                   wsc.nombre as consulta_area_nombre
            FROM managements m
            LEFT JOIN users u ON m.user_id = u.id
            LEFT JOIN management_types mt ON m.management_type_id = mt.id
            LEFT JOIN states s ON m.state_id = s.id
            LEFT JOIN work_spaces ws ON m.work_space_id = ws.id
            LEFT JOIN work_spaces wsc ON m.consulta_area_id = wsc.id
            WHERE m.work_order_id = %s
            ORDER BY m.created_at DESC
        """, (ot_id,))
        gestiones = cursor.fetchall()

        # Convertir fechas
        for gestion in gestiones:
            for key in gestion:
                if isinstance(gestion[key], datetime):
                    gestion[key] = gestion[key].isoformat()

        return DetalleLogResponse(
            ot_id=ot_id,
            ot_numero=str(ot_id),
            gestiones=gestiones,
            total_gestiones=len(gestiones)
        )

    finally:
        connection.close()


@router.get("/{ot_id}/log/excel")
async def descargar_log_excel(
    ot_id: int,
    current_user: dict = Depends(get_current_user)
):
    """
    Descarga el historial de gestiones en formato Excel.

    Fuente Laravel: ManagementController@descargarDetalleLogExcel
    Ruta Laravel: GET /descargar-detalle-log-excel/{id}
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")

    connection = get_mysql_connection()
    try:
        cursor = connection.cursor()

        # Verificar OT existe
        cursor.execute("""
            SELECT id FROM work_orders WHERE id = %s AND active = 1
        """, (ot_id,))
        if not cursor.fetchone():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"OT {ot_id} no encontrada"
            )

        # Obtener gestiones
        cursor.execute("""
            SELECT m.id, m.created_at, u.name as usuario,
                   mt.nombre as tipo_gestion, s.nombre as estado,
                   ws.nombre as area, m.observacion, m.duracion_segundos
            FROM managements m
            LEFT JOIN users u ON m.user_id = u.id
            LEFT JOIN management_types mt ON m.management_type_id = mt.id
            LEFT JOIN states s ON m.state_id = s.id
            LEFT JOIN work_spaces ws ON m.work_space_id = ws.id
            WHERE m.work_order_id = %s
            ORDER BY m.created_at DESC
        """, (ot_id,))
        gestiones = cursor.fetchall()

        # Crear Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Log OT {ot_id}"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = ["ID", "Fecha", "Usuario", "Tipo Gestión", "Estado", "Área", "Observación", "Duración (seg)"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Datos
        for row_idx, gestion in enumerate(gestiones, start=2):
            ws.cell(row=row_idx, column=1, value=gestion['id']).border = thin_border
            ws.cell(row=row_idx, column=2, value=str(gestion['created_at'])).border = thin_border
            ws.cell(row=row_idx, column=3, value=gestion['usuario']).border = thin_border
            ws.cell(row=row_idx, column=4, value=gestion['tipo_gestion']).border = thin_border
            ws.cell(row=row_idx, column=5, value=gestion['estado']).border = thin_border
            ws.cell(row=row_idx, column=6, value=gestion['area']).border = thin_border
            ws.cell(row=row_idx, column=7, value=gestion['observacion']).border = thin_border
            ws.cell(row=row_idx, column=8, value=gestion['duracion_segundos']).border = thin_border

        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 50
        ws.column_dimensions['H'].width = 15

        # Guardar en buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Log_OT_{ot_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        connection.close()


@router.post("/{ot_id}/reactivar")
async def reactivar_ot(
    ot_id: int,
    current_user: dict = Depends(get_current_user)
):
    """
    Reactiva una OT que estaba hibernada o anulada.

    Fuente Laravel: ManagementController@reactivarOT (líneas 686-705)
    Ruta Laravel: GET /reactivarOt/{id}
    """
    connection = get_mysql_connection()
    try:
        cursor = connection.cursor()

        # Verificar OT existe
        cursor.execute("""
            SELECT id, current_area_id FROM work_orders WHERE id = %s AND active = 1
        """, (ot_id,))
        ot = cursor.fetchone()

        if not ot:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"OT {ot_id} no encontrada"
            )

        # Crear gestión de reactivación
        cursor.execute("""
            INSERT INTO managements
            (observacion, management_type_id, user_id, work_order_id, work_space_id,
             duracion_segundos, state_id, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
        """, (
            "Órden de Trabajo Reactivada",
            1,  # Cambio de Estado
            current_user.get('id'),
            ot_id,
            ot['current_area_id'],
            0,
            1  # Estado: Venta
        ))

        # Actualizar OT
        cursor.execute("""
            UPDATE work_orders
            SET ultimo_cambio_area = NOW(), updated_at = NOW()
            WHERE id = %s
        """, (ot_id,))

        connection.commit()

        return {"message": "Órden de Trabajo Reactivada Correctamente", "ot_id": ot_id}

    except Exception as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al reactivar OT: {str(e)}"
        )
    finally:
        connection.close()


@router.post("/{ot_id}/retomar")
async def retomar_ot(
    ot_id: int,
    current_user: dict = Depends(get_current_user)
):
    """
    Retoma una OT desde el área actual del usuario.

    Fuente Laravel: ManagementController@retomarOT (líneas 707-759)
    Ruta Laravel: GET /retomarOt/{id}
    """
    connection = get_mysql_connection()
    try:
        cursor = connection.cursor()

        # Verificar OT existe
        cursor.execute("""
            SELECT wo.id, wo.ultimo_cambio_area,
                   (SELECT created_at FROM managements
                    WHERE work_order_id = wo.id AND management_type_id = 1
                    ORDER BY created_at DESC LIMIT 1) as ultimo_cambio_estado
            FROM work_orders wo
            WHERE wo.id = %s AND wo.active = 1
        """, (ot_id,))
        ot = cursor.fetchone()

        if not ot:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"OT {ot_id} no encontrada"
            )

        # Obtener área del usuario
        cursor.execute("""
            SELECT r.area_id FROM users u
            JOIN roles r ON u.role_id = r.id
            WHERE u.id = %s
        """, (current_user.get('id'),))
        user_area = cursor.fetchone()
        area_id = user_area['area_id'] if user_area else 2

        # Determinar estado según área
        estado_map = {
            2: 2,   # Desarrollo -> En Desarrollo
            3: 5,   # Diseño -> En Diseño Gráfico
            4: 7,   # Catalogación -> En Catalogación
        }
        estado = estado_map.get(area_id, 2)

        # Calcular duración
        ultimo_cambio = ot.get('ultimo_cambio_area') or ot.get('ultimo_cambio_estado')
        duracion_segundos = 0
        if ultimo_cambio:
            diff = datetime.now() - ultimo_cambio
            duracion_segundos = int(diff.total_seconds())

        # Obtener nombre del área
        cursor.execute("SELECT nombre FROM work_spaces WHERE id = %s", (area_id,))
        area_result = cursor.fetchone()
        area_nombre = area_result['nombre'] if area_result else "Área"

        # Crear gestión
        cursor.execute("""
            INSERT INTO managements
            (observacion, management_type_id, user_id, work_order_id, work_space_id,
             duracion_segundos, state_id, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
        """, (
            f"Órden de Trabajo Retomada por {area_nombre}",
            1,
            current_user.get('id'),
            ot_id,
            1,  # Venta
            duracion_segundos,
            estado
        ))

        # Actualizar OT
        cursor.execute("""
            UPDATE work_orders
            SET ultimo_cambio_area = NOW(), current_area_id = %s, updated_at = NOW()
            WHERE id = %s
        """, (area_id, ot_id))

        connection.commit()

        return {"message": "Órden de Trabajo Retomada Correctamente", "ot_id": ot_id}

    except Exception as e:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al retomar OT: {str(e)}"
        )
    finally:
        connection.close()
