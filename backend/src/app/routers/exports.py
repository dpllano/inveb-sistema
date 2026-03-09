"""
Router de Exportaciones - INVEB Cascade Service
Genera archivos Excel y SAP para descarga de OTs.
FASE 6.23
"""
from datetime import datetime, timedelta
from typing import Optional, List
from io import BytesIO
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import pymysql
import jwt

from app.config import get_settings

router = APIRouter(prefix="/exports", tags=["Exportaciones"])
security = HTTPBearer(auto_error=False)
settings = get_settings()


def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    """Extrae info del usuario del token JWT."""
    if not credentials:
        raise HTTPException(status_code=401, detail="Token no proporcionado")
    try:
        payload = jwt.decode(
            credentials.credentials,
            settings.JWT_SECRET_KEY,
            algorithms=[settings.JWT_ALGORITHM]
        )
        return {"id": int(payload.get("sub")), "role_id": payload.get("role_id", 0)}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalido")


def get_db_connection():
    """Conexion a MySQL usando configuracion de settings."""
    return pymysql.connect(
        host=settings.LARAVEL_MYSQL_HOST,
        port=settings.LARAVEL_MYSQL_PORT,
        user=settings.LARAVEL_MYSQL_USER,
        password=settings.LARAVEL_MYSQL_PASSWORD,
        database=settings.LARAVEL_MYSQL_DATABASE,
        cursorclass=pymysql.cursors.DictCursor,
        charset='utf8mb4'
    )


def get_ot_data(ot_id: int) -> dict:
    """Obtiene todos los datos de una OT para exportacion."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            sql = """
            SELECT
                ot.*,
                c.nombre as client_name,
                c.codigo_sap as client_sap_code,
                ca.nombre as canal_nombre,
                pt.nombre as product_type_nombre,
                st.nombre as estilo_nombre,
                st.codigo as estilo_codigo,
                ct.nombre as carton_nombre,
                ct.codigo as carton_codigo,
                ry.nombre as rayado_nombre,
                m.codigo as material_codigo,
                m.descripcion as material_descripcion,
                ssh.nombre as subsubhierarchy_nombre,
                ssh.codigo as subsubhierarchy_codigo,
                p.nombre as planta_nombre,
                p.codigo as planta_codigo,
                tp.nombre as tipo_palet_nombre,
                fsc.nombre as fsc_nombre,
                proc.nombre as proceso_nombre,
                proc.codigo as proceso_codigo,
                arm.nombre as armado_nombre,
                peg.nombre as pegado_nombre,
                adh.nombre as adhesivo_nombre,
                adh.codigo as adhesivo_codigo,
                CONCAT(u.name, ' ', u.last_name) as creador_nombre
            FROM work_orders ot
            LEFT JOIN clients c ON ot.client_id = c.id
            LEFT JOIN canals ca ON ot.canal_id = ca.id
            LEFT JOIN product_types pt ON ot.product_type_id = pt.id
            LEFT JOIN styles st ON ot.style_id = st.id
            LEFT JOIN cartons ct ON ot.carton_id = ct.id
            LEFT JOIN rayados ry ON ot.rayado_id = ry.id
            LEFT JOIN materials m ON ot.material_id = m.id
            LEFT JOIN subsubhierarchies ssh ON ot.subsubhierarchy_id = ssh.id
            LEFT JOIN plantas p ON ot.planta_id = p.id
            LEFT JOIN pallet_types tp ON ot.tipo_pallet_id = tp.id
            LEFT JOIN fsc ON ot.fsc_id = fsc.id
            LEFT JOIN processes proc ON ot.process_id = proc.id
            LEFT JOIN armados arm ON ot.armado_id = arm.id
            LEFT JOIN pegados peg ON ot.pegado_id = peg.id
            LEFT JOIN adhesivos adh ON ot.adhesivo_id = adh.id
            LEFT JOIN users u ON ot.creador_id = u.id
            WHERE ot.id = %s
            """
            cursor.execute(sql, (ot_id,))
            ot = cursor.fetchone()
            if not ot:
                raise HTTPException(status_code=404, detail="OT no encontrada")

            # Obtener colores
            cursor.execute("""
                SELECT c.nombre, c.codigo, wc.consumo
                FROM work_order_colors wc
                JOIN colors c ON wc.color_id = c.id
                WHERE wc.work_order_id = %s
                ORDER BY wc.id
            """, (ot_id,))
            ot['colores'] = cursor.fetchall()

            return ot
    finally:
        conn.close()


def format_number(value, decimals=2):
    """Formatea un numero con precision especifica."""
    if value is None:
        return ""
    try:
        return f"{float(value):.{decimals}f}"
    except (ValueError, TypeError):
        return str(value)


@router.get("/ot/{ot_id}/excel")
async def export_ot_excel(
    ot_id: int,
    current_user: dict = Depends(get_current_user)
):
    """
    Exporta una OT a formato Excel estandar.
    Retorna archivo .xlsx con todos los datos de catalogacion.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Libreria openpyxl no instalada"
        )

    ot = get_ot_data(ot_id)

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"OT_{ot_id}"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Datos a exportar (Campo, Valor)
    data = [
        ("DATOS DE LA ORDEN DE TRABAJO", ""),
        ("Numero OT", ot.get('id')),
        ("Fecha Creacion", str(ot.get('created_at', ''))[:10]),
        ("Estado", ot.get('estado_nombre', '')),
        ("Creador", ot.get('creador_nombre', '')),
        ("", ""),
        ("DATOS COMERCIALES", ""),
        ("Cliente", ot.get('client_name', '')),
        ("Codigo SAP Cliente", ot.get('client_sap_code', '')),
        ("Descripcion", ot.get('descripcion', '')),
        ("Canal", ot.get('canal_nombre', '')),
        ("Tipo Producto", ot.get('product_type_nombre', '')),
        ("", ""),
        ("DIMENSIONES INTERIORES", ""),
        ("Largo Interior (mm)", format_number(ot.get('largo'))),
        ("Ancho Interior (mm)", format_number(ot.get('ancho'))),
        ("Alto Interior (mm)", format_number(ot.get('alto'))),
        ("", ""),
        ("DIMENSIONES SUSTRATO", ""),
        ("Largura (mm)", format_number(ot.get('largura'))),
        ("Anchura (mm)", format_number(ot.get('anchura'))),
        ("", ""),
        ("DIMENSIONES EXTERIORES", ""),
        ("Largo Exterior (mm)", format_number(ot.get('largo_exterior'))),
        ("Ancho Exterior (mm)", format_number(ot.get('ancho_exterior'))),
        ("Alto Exterior (mm)", format_number(ot.get('alto_exterior'))),
        ("", ""),
        ("ESPECIFICACIONES TECNICAS", ""),
        ("Carton", ot.get('carton_nombre', '')),
        ("Codigo Carton", ot.get('carton_codigo', '')),
        ("Estilo", ot.get('estilo_nombre', '')),
        ("Codigo Estilo", ot.get('estilo_codigo', '')),
        ("Rayado", ot.get('rayado_nombre', '')),
        ("Proceso", ot.get('proceso_nombre', '')),
        ("Armado", ot.get('armado_nombre', '')),
        ("Pegado", ot.get('pegado_nombre', '')),
        ("Adhesivo", ot.get('adhesivo_nombre', '')),
        ("", ""),
        ("MATERIAL SAP", ""),
        ("Codigo Material", ot.get('material_codigo', '')),
        ("Descripcion Material", ot.get('material_descripcion', '')),
        ("Jerarquia", ot.get('subsubhierarchy_nombre', '')),
        ("Codigo Jerarquia", ot.get('subsubhierarchy_codigo', '')),
        ("", ""),
        ("CALIDAD", ""),
        ("Gramaje (g/m2)", format_number(ot.get('gramaje'))),
        ("ECT Minimo", format_number(ot.get('ect_minimo'))),
        ("FCT Minimo", format_number(ot.get('fct_minimo'))),
        ("BCT Minimo", format_number(ot.get('bct_minimo'))),
        ("Mullen", format_number(ot.get('mullen'))),
        ("", ""),
        ("IMPRESION", ""),
        ("Numero de Colores", ot.get('numero_colores', 0)),
        ("Tipo Impresion", ot.get('tipo_impresion', '')),
    ]

    # Agregar colores
    for i, color in enumerate(ot.get('colores', []), 1):
        data.append((f"Color {i}", color.get('nombre', '')))
        data.append((f"Consumo Color {i} (g)", format_number(color.get('consumo'))))

    # Agregar mas datos
    data.extend([
        ("", ""),
        ("PALETIZADO", ""),
        ("Planta", ot.get('planta_nombre', '')),
        ("Tipo Palet", ot.get('tipo_palet_nombre', '')),
        ("Cajas por Palet", ot.get('cajas_por_palet', '')),
        ("", ""),
        ("CERTIFICACIONES", ""),
        ("FSC", ot.get('fsc_nombre', '')),
    ])

    # Escribir datos
    for row_idx, (campo, valor) in enumerate(data, 1):
        ws.cell(row=row_idx, column=1, value=campo)
        ws.cell(row=row_idx, column=2, value=valor)

        # Estilo para headers de seccion
        if valor == "" and campo != "":
            ws.cell(row=row_idx, column=1).font = header_font
            ws.cell(row=row_idx, column=1).fill = header_fill
            ws.cell(row=row_idx, column=2).fill = header_fill

        # Bordes
        ws.cell(row=row_idx, column=1).border = thin_border
        ws.cell(row=row_idx, column=2).border = thin_border

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"OT_{ot_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/ot/{ot_id}/sap")
async def export_ot_sap(
    ot_id: int,
    current_user: dict = Depends(get_current_user)
):
    """
    Exporta una OT a formato SAP (Excel vertical).
    Formato: Campo SAP | Descripcion | Valor
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Libreria openpyxl no instalada"
        )

    ot = get_ot_data(ot_id)

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SAP"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["Campo SAP", "Descripcion", "Valor"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Generar codigo material SAP
    material_codigo = ot.get('material_codigo', '')
    material_ge1 = f"GE1{material_codigo}" if material_codigo else ""
    material_ge2 = f"GE2{material_codigo}" if material_codigo else ""

    # Datos SAP (Campo, Descripcion, Valor)
    sap_data = [
        ("MATNR", "Numero de Material", material_ge1),
        ("EN_PLANCHA_SEMI", "Material Semielaborado", material_ge2),
        ("MAKTX", "Descripcion Comercial", ot.get('descripcion', '')),
        ("PRDHA", "Jerarquia SAP", ot.get('subsubhierarchy_codigo', '')),
        ("PRODH", "Jerarquia Producto", ot.get('subsubhierarchy_nombre', '')),
        ("", "", ""),
        ("WERKS", "Centro", ot.get('planta_codigo', '')),
        ("LGORT", "Almacen", "001"),
        ("VKORG", "Organizacion de Ventas", ot.get('org_venta_id', '')),
        ("", "", ""),
        ("EN_LARGO", "Largo Interior (MM)", format_number(ot.get('largo'))),
        ("EN_ANCHO", "Ancho Interior (MM)", format_number(ot.get('ancho'))),
        ("EN_ALTO", "Alto Interior (MM)", format_number(ot.get('alto'))),
        ("EN_LARGURA", "Largura Sustrato (MM)", format_number(ot.get('largura'))),
        ("EN_ANCHURA", "Anchura Sustrato (MM)", format_number(ot.get('anchura'))),
        ("EN_LARGOEXT", "Largo Exterior (MM)", format_number(ot.get('largo_exterior'))),
        ("EN_ANCHOEXT", "Ancho Exterior (MM)", format_number(ot.get('ancho_exterior'))),
        ("EN_ALTOEXT", "Alto Exterior (MM)", format_number(ot.get('alto_exterior'))),
        ("", "", ""),
        ("EN_CARTON", "Tipo Carton", ot.get('carton_codigo', '')),
        ("EN_TIPO", "Tipo Producto", ot.get('product_type_nombre', '')),
        ("EN_ESTILO", "Estilo Producto", ot.get('estilo_codigo', '')),
        ("EN_GRAMAJE_G_M2", "Gramaje (G/m2)", format_number(ot.get('gramaje'))),
        ("EN_ESPESOR_MM", "Espesor Caja (mm)", format_number(ot.get('espesor'))),
        ("EN_ECT_MINIMO", "ECT Minimo (LB/PULG2)", format_number(ot.get('ect_minimo'))),
        ("EN_FCT_MINIMO", "FCT Minimo (LB/PULG2)", format_number(ot.get('fct_minimo'))),
        ("EN_MULLEN", "Mullen (LB/PULG2)", format_number(ot.get('mullen'))),
        ("EN_RESISTENCIA_MT", "BCT Minimo (LB)", format_number(ot.get('bct_minimo'))),
        ("", "", ""),
        ("EN_TIPOIMPRESION", "Tipo Impresion", ot.get('tipo_impresion', '')),
        ("EN_COLORES", "Numero de Colores", str(ot.get('numero_colores', 0))),
        ("EN_RAYADO", "Tipo Rayado", ot.get('rayado_nombre', '')),
        ("", "", ""),
    ]

    # Agregar colores
    for i, color in enumerate(ot.get('colores', [])[:7], 1):
        sap_data.append((f"EN_COLOR_COMP_{i}", f"Color Componente {i}", color.get('codigo', '')))
        sap_data.append((f"EN_CONSUMO_{i}", f"Consumo Color {i} (G)", format_number(color.get('consumo'), 4)))

    # Mas datos SAP
    sap_data.extend([
        ("", "", ""),
        ("EN_ADHESIVO", "Codigo Adhesivo", ot.get('adhesivo_codigo', '')),
        ("EN_PROCESO", "Codigo Proceso", ot.get('proceso_codigo', '')),
        ("", "", ""),
        ("EN_TIPO_PALET_GE", "Tipo Palet", ot.get('tipo_palet_nombre', '')),
        ("EN_CAJAS_POR_PALET", "Cajas por Palet", str(ot.get('cajas_por_palet', ''))),
        ("", "", ""),
        ("EN_SELLO", "Certificacion FSC", ot.get('fsc_nombre', '')),
    ])

    # Escribir datos
    for row_idx, (campo, descripcion, valor) in enumerate(sap_data, 2):
        ws.cell(row=row_idx, column=1, value=campo).border = thin_border
        ws.cell(row=row_idx, column=2, value=descripcion).border = thin_border
        ws.cell(row=row_idx, column=3, value=valor).border = thin_border

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 40

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"SAP_OT_{ot_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/ots/excel")
async def export_ots_list_excel(
    estado_id: Optional[int] = Query(None),
    area_id: Optional[int] = Query(None),
    canal_id: Optional[int] = Query(None),
    client_id: Optional[int] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    current_user: dict = Depends(get_current_user)
):
    """
    Exporta la lista de OTs filtrada a Excel.
    Incluye todos los campos principales.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Libreria openpyxl no instalada"
        )

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            # Query base
            sql = """
            SELECT
                ot.id,
                ot.created_at,
                c.nombre as cliente,
                ot.descripcion,
                ca.nombre as canal,
                pt.descripcion as tipo_producto,
                s.nombre as estado,
                CONCAT(u.nombre, ' ', u.apellido) as creador,
                m.codigo as material_codigo,
                ct.codigo as carton,
                st.glosa as estilo,
                ot.interno_largo as largo,
                ot.interno_ancho as ancho,
                ot.interno_alto as alto,
                ot.numero_colores
            FROM work_orders ot
            LEFT JOIN clients c ON ot.client_id = c.id
            LEFT JOIN canals ca ON ot.canal_id = ca.id
            LEFT JOIN product_types pt ON ot.product_type_id = pt.id
            LEFT JOIN states s ON ot.current_area_id = s.id
            LEFT JOIN users u ON ot.creador_id = u.id
            LEFT JOIN materials m ON ot.material_id = m.id
            LEFT JOIN cartons ct ON ot.carton_id = ct.id
            LEFT JOIN styles st ON ot.style_id = st.id
            WHERE 1=1
            """
            params = []

            if estado_id:
                sql += " AND ot.current_area_id = %s"
                params.append(estado_id)
            if area_id:
                sql += " AND ot.current_area_id = %s"
                params.append(area_id)
            if canal_id:
                sql += " AND ot.canal_id = %s"
                params.append(canal_id)
            if client_id:
                sql += " AND ot.client_id = %s"
                params.append(client_id)
            if fecha_desde:
                sql += " AND DATE(ot.created_at) >= %s"
                params.append(fecha_desde)
            if fecha_hasta:
                sql += " AND DATE(ot.created_at) <= %s"
                params.append(fecha_hasta)

            sql += " ORDER BY ot.id DESC LIMIT 5000"

            cursor.execute(sql, params)
            rows = cursor.fetchall()
    finally:
        conn.close()

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OTs"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "OT", "Fecha", "Cliente", "Descripcion", "Canal",
        "Tipo Producto", "Estado", "Creador", "Material",
        "Carton", "Estilo", "Largo", "Ancho", "Alto", "Colores"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Datos
    for row_idx, row in enumerate(rows, 2):
        values = [
            row.get('id'),
            str(row.get('created_at', ''))[:10],
            row.get('cliente', ''),
            row.get('descripcion', ''),
            row.get('canal', ''),
            row.get('tipo_producto', ''),
            row.get('estado', ''),
            row.get('creador', ''),
            row.get('material_codigo', ''),
            row.get('carton', ''),
            row.get('estilo', ''),
            row.get('largo'),
            row.get('ancho'),
            row.get('alto'),
            row.get('numero_colores', 0)
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    # Ajustar ancho de columnas
    widths = [8, 12, 30, 40, 15, 20, 15, 25, 15, 20, 20, 10, 10, 10, 10]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Lista_OTs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =============================================
# FASE 6.26 - DETALLE LOG OT (Bitácora)
# =============================================

@router.get("/ot/{ot_id}/log")
async def get_ot_log(
    ot_id: int,
    date_desde: Optional[str] = Query(None, description="Fecha inicio dd/mm/yyyy"),
    date_hasta: Optional[str] = Query(None, description="Fecha fin dd/mm/yyyy"),
    user_id: Optional[int] = Query(None, description="Filtrar por usuario"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: dict = Depends(get_current_user)
):
    """
    Obtiene el log de modificaciones de una OT (bitácora).
    Devuelve historial de cambios con detalles de campos modificados.
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            # Calcular fechas por defecto (último mes)
            if date_desde and date_hasta:
                try:
                    from_date = datetime.strptime(date_desde, '%d/%m/%Y').strftime('%Y-%m-%d')
                    to_date = datetime.strptime(date_hasta, '%d/%m/%Y').strftime('%Y-%m-%d')
                except ValueError:
                    from_date = (datetime.now().replace(day=1) - timedelta(days=1)).replace(day=1).strftime('%Y-%m-%d')
                    to_date = datetime.now().strftime('%Y-%m-%d')
            else:
                from_date = (datetime.now().replace(day=1) - timedelta(days=1)).replace(day=1).strftime('%Y-%m-%d')
                to_date = datetime.now().strftime('%Y-%m-%d')

            # Query base
            sql = """
                SELECT
                    b.id,
                    b.work_order_id,
                    b.operacion,
                    b.observacion,
                    b.datos_modificados,
                    b.user_data,
                    DATE_FORMAT(b.created_at, '%%Y-%%m-%%d %%H:%%i:%%s') as created_at,
                    b.user_id
                FROM bitacora_work_orders b
                WHERE b.work_order_id = %s
                  AND b.operacion IN ('Mckee', 'Modificación')
                  AND b.datos_modificados IS NOT NULL
                  AND b.created_at BETWEEN %s AND %s
            """
            params = [ot_id, from_date, to_date + ' 23:59:59']

            if user_id:
                sql += " AND b.user_id = %s"
                params.append(user_id)

            sql += " ORDER BY b.id DESC"

            # Contar total
            count_sql = sql.replace(
                "SELECT\n                    b.id,",
                "SELECT COUNT(*) as total FROM (SELECT b.id,"
            ) + ") as sub"
            cursor.execute(count_sql.replace("ORDER BY b.id DESC", ""), params)
            total_result = cursor.fetchone()
            total = total_result['total'] if total_result else 0

            # Aplicar paginación
            offset = (page - 1) * page_size
            sql += f" LIMIT {page_size} OFFSET {offset}"
            cursor.execute(sql, params)
            rows = cursor.fetchall()

            # Procesar datos
            import json
            items = []
            for row in rows:
                datos_mod = row.get('datos_modificados')
                user_data = row.get('user_data')

                # Parsear JSON si es string
                if isinstance(datos_mod, str):
                    try:
                        datos_mod = json.loads(datos_mod)
                    except:
                        datos_mod = {}

                if isinstance(user_data, str):
                    try:
                        user_data = json.loads(user_data)
                    except:
                        user_data = {}

                usuario_nombre = ""
                if user_data:
                    usuario_nombre = f"{user_data.get('nombre', '')} {user_data.get('apellido', '')}".strip()

                items.append({
                    "id": row['id'],
                    "work_order_id": row['work_order_id'],
                    "operacion": row['operacion'],
                    "observacion": row['observacion'],
                    "datos_modificados": datos_mod if datos_mod else {},
                    "usuario": usuario_nombre,
                    "created_at": row['created_at']
                })

            # Obtener usuarios que han modificado esta OT
            cursor.execute("""
                SELECT DISTINCT u.id, CONCAT(u.nombre, ' ', u.apellido) as nombre
                FROM users u
                INNER JOIN bitacora_work_orders b ON u.id = b.user_id
                WHERE b.work_order_id = %s
                ORDER BY nombre
            """, (ot_id,))
            usuarios = cursor.fetchall()

            total_pages = (total + page_size - 1) // page_size if total > 0 else 1

            return {
                "items": items,
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": total_pages,
                "usuarios_filtro": usuarios
            }

    finally:
        conn.close()


@router.get("/ot/{ot_id}/log/excel")
async def export_ot_log_excel(
    ot_id: int,
    date_desde: Optional[str] = Query(None, description="Fecha inicio dd/mm/yyyy"),
    date_hasta: Optional[str] = Query(None, description="Fecha fin dd/mm/yyyy"),
    user_id: Optional[int] = Query(None, description="Filtrar por usuario"),
    current_user: dict = Depends(get_current_user)
):
    """
    Exporta el log de modificaciones de una OT a Excel.
    Incluye: OT, ID Cambio, Fecha, Descripción, Campo Modificado, Valor Antiguo, Valor Nuevo, Usuario.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Libreria openpyxl no instalada"
        )

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            # Calcular fechas
            if date_desde and date_hasta:
                try:
                    from_date = datetime.strptime(date_desde, '%d/%m/%Y').strftime('%Y-%m-%d')
                    to_date = datetime.strptime(date_hasta, '%d/%m/%Y').strftime('%Y-%m-%d')
                except ValueError:
                    from_date = (datetime.now().replace(day=1) - timedelta(days=1)).replace(day=1).strftime('%Y-%m-%d')
                    to_date = datetime.now().strftime('%Y-%m-%d')
            else:
                from_date = (datetime.now().replace(day=1) - timedelta(days=1)).replace(day=1).strftime('%Y-%m-%d')
                to_date = datetime.now().strftime('%Y-%m-%d')

            # Query
            sql = """
                SELECT
                    b.id,
                    b.work_order_id,
                    b.operacion,
                    b.observacion,
                    b.datos_modificados,
                    b.user_data,
                    DATE_FORMAT(b.created_at, '%%Y-%%m-%%d %%H:%%i:%%s') as created_at
                FROM bitacora_work_orders b
                WHERE b.work_order_id = %s
                  AND b.operacion IN ('Mckee', 'Modificación')
                  AND b.datos_modificados IS NOT NULL
                  AND b.created_at BETWEEN %s AND %s
            """
            params = [ot_id, from_date, to_date + ' 23:59:59']

            if user_id:
                sql += " AND b.user_id = %s"
                params.append(user_id)

            sql += " ORDER BY b.id DESC"
            cursor.execute(sql, params)
            rows = cursor.fetchall()

    finally:
        conn.close()

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Log OT {ot_id}"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["OT", "ID Cambio", "Fecha", "Descripción", "Campo Modificado", "Valor Antiguo", "Valor Nuevo", "Usuario"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Datos
    import json
    row_idx = 2
    for row in rows:
        datos_mod = row.get('datos_modificados')
        user_data = row.get('user_data')

        # Parsear JSON
        if isinstance(datos_mod, str):
            try:
                datos_mod = json.loads(datos_mod)
            except:
                datos_mod = {}

        if isinstance(user_data, str):
            try:
                user_data = json.loads(user_data)
            except:
                user_data = {}

        usuario = ""
        if user_data:
            usuario = f"{user_data.get('nombre', '')} {user_data.get('apellido', '')}".strip()

        operacion = row.get('operacion', '')

        # Si hay datos modificados, crear una fila por cada campo
        if datos_mod and isinstance(datos_mod, dict):
            for campo, valor in datos_mod.items():
                if isinstance(valor, dict):
                    texto_campo = valor.get('texto', campo)

                    if operacion == 'Modificación':
                        antiguo = valor.get('antiguo_valor', {})
                        nuevo = valor.get('nuevo_valor', {})
                        val_antiguo = antiguo.get('descripcion', '') if isinstance(antiguo, dict) else str(antiguo)
                        val_nuevo = nuevo.get('descripcion', '') if isinstance(nuevo, dict) else str(nuevo)
                    else:
                        # Mckee u otra operación
                        val_antiguo = 'N/A'
                        val_info = valor.get('valor', {})
                        val_nuevo = val_info.get('descripcion', '') if isinstance(val_info, dict) else str(val_info)

                    values = [
                        row['work_order_id'],
                        row['id'],
                        row['created_at'],
                        row.get('observacion', ''),
                        texto_campo,
                        val_antiguo or 'Campo Vacío',
                        val_nuevo or 'Campo Vacío',
                        usuario
                    ]

                    for col, value in enumerate(values, 1):
                        cell = ws.cell(row=row_idx, column=col, value=value)
                        cell.border = thin_border

                    row_idx += 1
        else:
            # Si no hay datos estructurados, crear una fila simple
            values = [
                row['work_order_id'],
                row['id'],
                row['created_at'],
                row.get('observacion', ''),
                '-',
                '-',
                '-',
                usuario
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

            row_idx += 1

    # Ajustar ancho de columnas
    widths = [8, 12, 20, 35, 25, 30, 30, 25]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Log_OT_{ot_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
